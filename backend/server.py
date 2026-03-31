"""
QR Inventory Management System - Backend API
Mobile-first inventory tracking with QR codes for kitchen operations
Features: Auth, Invoice OCR, Requisitions, GRN, Issue, Transfer, Alerts
"""
import os
import io
import base64
import asyncio
import json
import re
import httpx  # For cross-app API calls
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from fastapi import FastAPI, HTTPException, Query, Depends, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, EmailStr
from pymongo import MongoClient
from bson import ObjectId
import qrcode
from passlib.context import CryptContext
from jose import jwt, JWTError
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

load_dotenv()

# Environment variables
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "dreamoven_db")
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY")
JWT_SECRET = os.environ.get("JWT_SECRET", "qr-inventory-secret-key-2024")
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_HOURS = 24
KITCHEN_APP_URL = os.environ.get("KITCHEN_APP_URL")  # URL of the Kitchen App for cross-app photo fetching

# Cloudflare R2 Configuration
R2_ACCOUNT_ID = os.environ.get("R2_ACCOUNT_ID")
R2_ACCESS_KEY = os.environ.get("R2_ACCESS_KEY")
R2_SECRET_KEY = os.environ.get("R2_SECRET_KEY")
R2_BUCKET_NAME = os.environ.get("R2_BUCKET_NAME", "dreamoven-storage")
R2_ENDPOINT_URL = f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com" if R2_ACCOUNT_ID else None
R2_PUBLIC_URL = os.environ.get("R2_PUBLIC_URL", "")  # Public URL for viewing images

# Initialize R2 client
r2_client = None
if R2_ACCOUNT_ID and R2_ACCESS_KEY and R2_SECRET_KEY:
    try:
        import boto3
        from botocore.config import Config
        r2_client = boto3.client(
            's3',
            endpoint_url=R2_ENDPOINT_URL,
            aws_access_key_id=R2_ACCESS_KEY,
            aws_secret_access_key=R2_SECRET_KEY,
            config=Config(signature_version='s3v4')
        )
        print("✅ Cloudflare R2 client initialized")
    except Exception as e:
        print(f"⚠️ R2 initialization failed: {e}")

# MongoDB connection with timeout settings optimized for Atlas
client = MongoClient(
    MONGO_URL,
    serverSelectionTimeoutMS=10000,   # 10 seconds to select server
    connectTimeoutMS=10000,           # 10 seconds to connect
    socketTimeoutMS=30000,            # 30 seconds for socket operations (reduced from 60)
    maxPoolSize=50,                   # Max connections in pool
    retryWrites=True,
    retryReads=True,                  # Enable retry for read operations
    w="majority",                     # Write concern
    maxIdleTimeMS=45000               # Close idle connections after 45 seconds
)
db = client[DB_NAME]

# Collections
items_collection = db["items"]
locations_collection = db["locations"]
vendors_collection = db["vendors"]
lots_collection = db["lots"]
transactions_collection = db["transactions"]
categories_collection = db["categories"]
users_collection = db["users"]
requisitions_collection = db["requisitions"]
invoices_collection = db["invoices"]
purchase_orders_collection = db["purchase_orders"]
company_settings_collection = db["company_settings"]

# ==================== CREATE INDEXES FOR PERFORMANCE ====================
# These indexes are critical for Atlas performance - created in background thread
def ensure_indexes():
    """Create indexes if they don't exist - runs in background to not block startup"""
    import threading
    
    def create_indexes_background():
        try:
            # ==================== PURCHASE ORDERS INDEXES ====================
            # For vendor ledger queries (Main Store filter + date range)
            try:
                purchase_orders_collection.create_index(
                    [("created_by_location_name", 1), ("vendor_id", 1), ("created_at", -1)],
                    background=True, name="idx_location_vendor_date"
                )
            except: pass
            
            # For vendor-specific queries with date
            try:
                purchase_orders_collection.create_index(
                    [("vendor_id", 1), ("created_at", -1)],
                    background=True, name="idx_vendor_date"
                )
            except: pass
            
            # For date range queries
            try:
                purchase_orders_collection.create_index(
                    [("created_at", -1)],
                    background=True, name="idx_created_at"
                )
            except: pass
            
            # For status filtering
            try:
                purchase_orders_collection.create_index("status", background=True)
            except: pass
            
            # ==================== TRANSACTIONS INDEXES ====================
            # For consumption analysis (type='issue' with date range)
            try:
                transactions_collection.create_index(
                    [("type", 1), ("created_at", -1)],
                    background=True, name="idx_type_date"
                )
            except: pass
            
            # For lot lookups
            try:
                transactions_collection.create_index("lot_id", background=True)
            except: pass
            
            # ==================== LOTS INDEXES ====================
            # For vendor + date queries
            try:
                lots_collection.create_index(
                    [("vendor_id", 1), ("created_at", -1)],
                    background=True, name="idx_vendor_date"
                )
            except: pass
            
            # For item + location queries (stock)
            try:
                lots_collection.create_index(
                    [("item_id", 1), ("location_id", 1)],
                    background=True, name="idx_item_location"
                )
            except: pass
            
            # For current quantity queries
            try:
                lots_collection.create_index(
                    [("location_id", 1), ("current_quantity", 1)],
                    background=True, name="idx_location_qty"
                )
            except: pass
            
            # ==================== KITCHEN RECEIVABLES INDEXES ====================
            try:
                db["kitchen_receivables"].create_index(
                    [("vendor_id", 1), ("created_at", -1)],
                    background=True, name="idx_vendor_date"
                )
            except: pass
            try:
                db["kitchen_receivables"].create_index(
                    [("kitchen_id", 1), ("created_at", -1)],
                    background=True, name="idx_kitchen_date"
                )
            except: pass
            
            # ==================== ITEMS INDEXES ====================
            try:
                items_collection.create_index("category", background=True)
            except: pass
            try:
                items_collection.create_index("name", background=True)
            except: pass
            
            # ==================== REQUISITIONS INDEXES ====================
            try:
                requisitions_collection.create_index(
                    [("kitchen_id", 1), ("created_at", -1)],
                    background=True, name="idx_kitchen_date"
                )
            except: pass
            try:
                requisitions_collection.create_index(
                    [("status", 1), ("created_at", -1)],
                    background=True, name="idx_status_date"
                )
            except: pass
            
            print("✅ Database indexes created/verified")
        except Exception as e:
            print(f"ℹ️ Index creation: {type(e).__name__}")
    
    # Run in background thread so startup is not blocked
    thread = threading.Thread(target=create_indexes_background, daemon=True)
    thread.start()

# Trigger background index creation
ensure_indexes()

# PAR Stock Update History Collection
par_stock_history_collection = db["par_stock_history"]

def update_par_stock_from_consumption(triggered_by: str = "system"):
    """
    Update PAR stock for all items based on previous month's consumption.
    Formula: PAR Stock = (Monthly OUT Qty / 30 days) × 10
    
    Only updates items that had consumption in the previous month.
    """
    from calendar import monthrange
    
    # Calculate previous month date range
    today = datetime.now(timezone.utc)
    first_day_current_month = today.replace(day=1)
    last_day_prev_month = first_day_current_month - timedelta(days=1)
    first_day_prev_month = last_day_prev_month.replace(day=1)
    
    start_date = first_day_prev_month.strftime("%Y-%m-%d")
    end_date = last_day_prev_month.strftime("%Y-%m-%d")
    days_in_month = (last_day_prev_month - first_day_prev_month).days + 1
    
    print(f"📊 Starting PAR Stock Update: {start_date} to {end_date} ({days_in_month} days)")
    print(f"   Triggered by: {triggered_by}")
    
    # Get all OUT transactions for previous month (dispatch, issue, transfer)
    txn_query = {
        "type": {"$in": ["dispatch", "issue", "transfer"]},
        "created_at": {
            "$gte": start_date + "T00:00:00",
            "$lte": end_date + "T23:59:59"
        }
    }
    
    # Aggregate consumption by item
    item_consumption = {}
    
    for txn in transactions_collection.find(txn_query):
        lot_id = txn.get("lot_id")
        qty = float(txn.get("quantity", 0) or 0)
        
        if not lot_id or qty <= 0:
            continue
        
        # Get lot to find item_id
        lot = None
        try:
            lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
        except:
            pass
        
        if not lot:
            continue
        
        item_id = lot.get("item_id", "")
        if not item_id:
            continue
        
        if item_id not in item_consumption:
            item_consumption[item_id] = {
                "total_out_qty": 0,
                "item_name": lot.get("item_name", "Unknown")
            }
        item_consumption[item_id]["total_out_qty"] += qty
    
    # Calculate and update PAR stock for each item
    updates_made = 0
    update_log = []
    
    for item_id, data in item_consumption.items():
        total_out = data["total_out_qty"]
        daily_avg = total_out / days_in_month
        new_par_stock = round(daily_avg * 10, 2)  # 10-day PAR
        
        if new_par_stock <= 0:
            continue
        
        # Get current item data
        try:
            item = items_collection.find_one({"_id": ObjectId(item_id)})
            if not item:
                continue
            
            old_par_stock = item.get("par_stock", 0) or 0
            
            # Update item's PAR stock
            items_collection.update_one(
                {"_id": ObjectId(item_id)},
                {"$set": {
                    "par_stock": new_par_stock,
                    "par_stock_updated_at": datetime.now(timezone.utc).isoformat(),
                    "par_stock_updated_by": triggered_by
                }}
            )
            
            updates_made += 1
            update_log.append({
                "item_id": item_id,
                "item_name": item.get("name", data["item_name"]),
                "category": item.get("category", "Unknown"),
                "old_par_stock": old_par_stock,
                "new_par_stock": new_par_stock,
                "monthly_consumption": round(total_out, 2),
                "daily_avg": round(daily_avg, 2)
            })
            
        except Exception as e:
            print(f"   Error updating item {item_id}: {e}")
    
    # Save history record
    history_record = {
        "run_date": datetime.now(timezone.utc).isoformat(),
        "triggered_by": triggered_by,
        "period_start": start_date,
        "period_end": end_date,
        "days_analyzed": days_in_month,
        "items_updated": updates_made,
        "update_details": update_log
    }
    par_stock_history_collection.insert_one(history_record)
    
    print(f"✅ PAR Stock Update Complete: {updates_made} items updated")
    
    return {
        "success": True,
        "items_updated": updates_made,
        "period": f"{start_date} to {end_date}",
        "days_analyzed": days_in_month,
        "triggered_by": triggered_by
    }

# ==================== SCHEDULER FOR AUTO PAR STOCK UPDATE ====================
scheduler = BackgroundScheduler(timezone="Asia/Kolkata")

def scheduled_par_stock_update():
    """Scheduled job to update PAR stock on last day of month"""
    try:
        print("🕐 Scheduled PAR Stock Update triggered")
        update_par_stock_from_consumption(triggered_by="scheduled_monthly")
    except Exception as e:
        print(f"❌ Scheduled PAR Stock Update failed: {e}")

# Schedule to run at 11:59 PM on the last day of every month
# Using day='last' to run on last day of each month
scheduler.add_job(
    scheduled_par_stock_update,
    CronTrigger(day='last', hour=23, minute=59),
    id='monthly_par_stock_update',
    replace_existing=True
)

# Start scheduler
scheduler.start()
print("📅 PAR Stock Scheduler started - runs on last day of every month at 11:59 PM IST")

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer(auto_error=False)

app = FastAPI(title="QR Inventory API", version="2.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add cache-control headers to prevent browser caching of API responses
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response

class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/api/"):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        return response

app.add_middleware(NoCacheMiddleware)

# ============ Health Check Endpoint ============
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes liveness/readiness probes"""
    return {"status": "healthy", "service": "qr-inventory-backend"}

@app.get("/api/health")
async def api_health_check():
    """API health check endpoint"""
    return {"status": "healthy", "service": "qr-inventory-backend", "version": "v12-po-totals-fixed"}

@app.get("/api/debug/db-config")
async def debug_db_config():
    """Debug endpoint to check database configuration"""
    return {
        "mongo_url": MONGO_URL,
        "db_name": DB_NAME,
        "collections": db.list_collection_names(),
        "lots_count": lots_collection.count_documents({})
    }

@app.get("/api/debug/po-received/{po_id}")
async def debug_po_received(po_id: str):
    """Debug endpoint to check received quantities for a PO"""
    try:
        po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
        if not po:
            return {"error": "PO not found"}
        
        po_number = po.get("po_number", "")
        
        # Check lots
        lot_conditions = [{"po_id": po_id}]
        if po_number:
            lot_conditions.append({"po_number": po_number})
        
        lots = list(lots_collection.find({"$or": lot_conditions}))
        
        # Check kitchen_receivables
        kr_conditions = [{"po_id": po_id}]
        if po_number:
            kr_conditions.append({"po_number": po_number})
        
        kitchen_receivables = db["kitchen_receivables"]
        krs = list(kitchen_receivables.find({"$or": kr_conditions}))
        
        return {
            "po_id": po_id,
            "po_number": po_number,
            "po_status": po.get("status"),
            "po_items_count": len(po.get("items", [])),
            "lots_found": len(lots),
            "lots_sample": [{"item_name": l.get("item_name"), "qty": l.get("initial_quantity")} for l in lots[:3]],
            "kitchen_receivables_found": len(krs),
            "kr_sample": [{"item_name": k.get("item_name"), "qty": k.get("quantity"), "po_id": k.get("po_id")} for k in krs[:3]]
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/debug/stock-export-test")
async def debug_stock_export():
    """Debug endpoint to test stock export data"""
    try:
        main_store = locations_collection.find_one({"type": "main_store"})
        if not main_store:
            return {"error": "Main Store not found"}
        
        main_store_id = str(main_store["_id"])
        
        # Get sample lots
        lots = list(lots_collection.find({
            "location_id": main_store_id,
            "current_quantity": {"$gt": 0}
        }).limit(5))
        
        results = []
        for lot in lots:
            try:
                item_id = lot.get("item_id")
                item = None
                if item_id:
                    try:
                        item = items_collection.find_one({"_id": ObjectId(item_id)})
                    except Exception as e:
                        results.append({"lot_id": str(lot["_id"]), "error": f"ObjectId error: {str(e)}", "item_id": item_id})
                        continue
                
                results.append({
                    "lot_id": str(lot["_id"]),
                    "item_id": item_id,
                    "item_found": item is not None,
                    "item_name": item.get("name") if item else lot.get("item_name"),
                    "qty": lot.get("current_quantity")
                })
            except Exception as e:
                results.append({"lot_id": str(lot.get("_id", "unknown")), "error": str(e)})
        
        return {
            "main_store_id": main_store_id,
            "main_store_name": main_store.get("name"),
            "total_lots_at_mainstore": lots_collection.count_documents({"location_id": main_store_id, "current_quantity": {"$gt": 0}}),
            "sample_lots": results
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/stock-export-full-test")
async def debug_stock_export_full():
    """Debug endpoint that runs the full export logic to find errors"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import traceback
    
    try:
        main_store = locations_collection.find_one({"type": "main_store"})
        if not main_store:
            return {"error": "Main Store not found"}
        
        main_store_id = str(main_store["_id"])
        
        lots = list(lots_collection.find({
            "location_id": main_store_id,
            "current_quantity": {"$gt": 0}
        }))
        
        all_vendors = {str(v["_id"]): v.get("name", "") for v in vendors_collection.find({})}
        
        po_vendors = {}
        for po in purchase_orders_collection.find({}, {"_id": 1, "vendor_id": 1, "vendor_name": 1}):
            po_id = str(po["_id"])
            vendor_name = po.get("vendor_name", "")
            if not vendor_name and po.get("vendor_id"):
                vendor_name = all_vendors.get(str(po["vendor_id"]), "")
            po_vendors[po_id] = vendor_name
        
        item_aggregates = {}
        errors = []
        
        for i, lot in enumerate(lots):
            try:
                item_id = str(lot.get("item_id", ""))
                if not item_id:
                    continue
                
                try:
                    item = items_collection.find_one({"_id": ObjectId(item_id)})
                except Exception as e:
                    errors.append(f"Lot {i}: ObjectId error for item_id={item_id}: {str(e)}")
                    continue
                    
                if not item:
                    continue
                
                item_name = item.get("name", "Unknown")
                category = item.get("category", "Uncategorized") or "Uncategorized"
                unit = item.get("unit", "")
                standard_price = float(item.get("standard_price", 0) or 0)
                
                qty = float(lot.get("current_quantity", 0) or 0)
                lot_rate = float(lot.get("purchase_rate", 0) or 0)
                
                rate = standard_price if standard_price > 0 else lot_rate
                
                if item_id not in item_aggregates:
                    item_aggregates[item_id] = {
                        "item_id": item_id,
                        "item_name": item_name,
                        "category": category,
                        "unit": unit,
                        "quantity": 0,
                        "price": rate,
                        "lot_count": 0,
                        "vendors": set()
                    }
                
                item_aggregates[item_id]["quantity"] += qty
                item_aggregates[item_id]["lot_count"] += 1
            except Exception as e:
                errors.append(f"Lot {i}: {str(e)}")
                continue
        
        # Try creating Excel
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test"
            ws['A1'] = "Test"
            
            # Try converting vendors sets to lists (common issue)
            for item in item_aggregates.values():
                item["vendors"] = list(item.get("vendors", set()))
            
            return {
                "success": True,
                "lots_processed": len(lots),
                "items_aggregated": len(item_aggregates),
                "errors": errors[:10],
                "sample_item": list(item_aggregates.values())[0] if item_aggregates else None
            }
        except Exception as e:
            return {"error": f"Excel creation failed: {str(e)}", "traceback": traceback.format_exc()}
            
    except Exception as e:
        return {"error": str(e), "traceback": traceback.format_exc()}

# ============ Pydantic Models ============

# Auth Models
class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "kitchen"  # "admin", "main_store", "kitchen"
    location_id: Optional[str] = None  # Required for kitchen users

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    location_id: Optional[str] = None
    location_name: Optional[str] = None
    created_at: str

# Existing Models
class ItemCreate(BaseModel):
    name: str
    category: str
    unit: str
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = 0
    vendor: Optional[str] = None
    standard_price: Optional[float] = None  # For price variance detection
    par_stock: Optional[float] = None  # Minimum stock level for alerts

class ItemUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = None
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None
    vendor: Optional[str] = None
    standard_price: Optional[float] = None
    par_stock: Optional[float] = None

class LocationCreate(BaseModel):
    name: str
    type: str = "kitchen"

class VendorCreate(BaseModel):
    name: str
    contact: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    payment_terms: Optional[str] = None  # e.g., "Net 30", "COD", etc.
    supply_categories: List[str] = []  # Categories this vendor supplies

class VendorUpdate(BaseModel):
    name: Optional[str] = None
    contact: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    payment_terms: Optional[str] = None
    supply_categories: Optional[List[str]] = None

# Purchase Order Models
class POItem(BaseModel):
    item_id: str
    quantity: float
    rate: float
    notes: Optional[str] = None

class PurchaseOrderCreate(BaseModel):
    vendor_id: str
    items: List[POItem]
    delivery_date: Optional[str] = None
    delivery_address: Optional[str] = None
    payment_terms: Optional[str] = None
    notes: Optional[str] = None

class GRNCreate(BaseModel):
    item_id: str
    quantity: float
    expiry_date: str
    location_id: str
    vendor_id: Optional[str] = None
    purchase_rate: Optional[float] = None
    notes: Optional[str] = None

class IssueCreate(BaseModel):
    lot_id: str
    quantity: float
    destination_id: str
    notes: Optional[str] = None

class TransferCreate(BaseModel):
    lot_id: str
    quantity: float
    destination_id: str
    notes: Optional[str] = None

# Requisition Models
class RequisitionItem(BaseModel):
    item_id: str
    quantity: float
    notes: Optional[str] = None

class RequisitionCreate(BaseModel):
    items: List[RequisitionItem]
    priority: str = "normal"  # "urgent", "normal", "low"
    notes: Optional[str] = None

class DispatchItem(BaseModel):
    item_id: str
    quantity_sent: float
    lot_id: Optional[str] = None  # Which lot to issue from
    remark: str = "ok"  # "ok" or "short"

class DispatchCreate(BaseModel):
    items: List[DispatchItem]
    notes: Optional[str] = None

# Invoice OCR Models
class InvoiceItem(BaseModel):
    item_name: str
    quantity: float
    unit: str
    rate: float
    amount: float
    matched_item_id: Optional[str] = None
    price_variance: Optional[float] = None  # Difference from standard price

class InvoiceData(BaseModel):
    vendor_name: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    total_amount: float
    items: List[InvoiceItem]
    price_variances: List[dict] = []  # Items with price differences

# ============ Auth Helper Functions ============

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_token(user_id: str, email: str, role: str, location_id: Optional[str] = None) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "location_id": location_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except JWTError:
        return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        return None
    
    payload = decode_token(credentials.credentials)
    if not payload:
        return None
    
    user = users_collection.find_one({"_id": ObjectId(payload["sub"])})
    if not user:
        return None
    
    return {
        "id": str(user["_id"]),
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "location_id": user.get("location_id")
    }

def require_auth(current_user = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    return current_user

def require_role(allowed_roles: List[str]):
    def role_checker(current_user = Depends(require_auth)):
        if current_user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return current_user
    return role_checker

# ============ Helper Functions ============

def generate_lot_number():
    now = datetime.now(timezone.utc)
    count = lots_collection.count_documents({})
    return f"LOT-{now.strftime('%Y%m%d')}-{count + 1:04d}"

def generate_qr_code(lot_id: str, lot_number: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(f"LOT:{lot_id}:{lot_number}")
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode()

def parse_expiry_date(expiry_str: str) -> datetime:
    # Handle None or empty expiry dates - return a far future date to allow dispatch
    if not expiry_str:
        return datetime(2099, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
    if "T" in expiry_str:
        return datetime.fromisoformat(expiry_str.replace("Z", "+00:00"))
    else:
        return datetime.strptime(expiry_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)

def get_lot_status(lot: dict) -> str:
    current_qty = lot.get("current_quantity", 0) or 0
    if current_qty <= 0:
        return "exhausted"
    expiry = parse_expiry_date(lot.get("expiry_date"))
    if expiry < datetime.now(timezone.utc):
        return "expired"
    return "active"

def serialize_lot(lot: dict) -> dict:
    item = None
    try:
        item_id = lot.get("item_id")
        if item_id:
            item = items_collection.find_one({"_id": ObjectId(item_id)})
    except:
        pass
    
    location = None
    try:
        loc_id = lot.get("location_id")
        if loc_id:
            location = locations_collection.find_one({"_id": ObjectId(loc_id)})
    except:
        pass
    
    vendor_name = None
    if lot.get("vendor_id"):
        try:
            vendor = vendors_collection.find_one({"_id": ObjectId(lot["vendor_id"])})
            vendor_name = vendor["name"] if vendor else None
        except:
            pass
    
    # Use stored item_name as fallback if item was deleted
    item_name = item["name"] if item else (lot.get("item_name") or "Unknown")
    item_category = item["category"] if item else (lot.get("category") or "Unknown")
    item_unit = item["unit"] if item else (lot.get("unit") or "Unknown")
    
    return {
        "id": str(lot["_id"]),
        "lot_number": lot.get("lot_number", ""),
        "item_id": str(lot.get("item_id", "")) if lot.get("item_id") else "",
        "item_name": item_name,
        "category": item_category,
        "unit": item_unit,
        "initial_quantity": lot.get("initial_quantity", 0),
        "current_quantity": lot.get("current_quantity", 0),
        "expiry_date": lot.get("expiry_date", ""),
        "location_id": str(lot.get("location_id", "")) if lot.get("location_id") else "",
        "location_name": location["name"] if location else "Unknown",
        "vendor_id": str(lot.get("vendor_id")) if lot.get("vendor_id") else None,
        "vendor_name": vendor_name,
        "purchase_rate": lot.get("purchase_rate"),
        "qr_code": lot.get("qr_code", ""),
        "status": get_lot_status(lot),
        "created_at": str(lot.get("created_at", "")) if lot.get("created_at") else ""
    }

def serialize_transaction(txn: dict) -> dict:
    lot = None
    item = None
    
    try:
        lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
        if lot:
            try:
                item = items_collection.find_one({"_id": ObjectId(lot["item_id"])})
            except:
                pass
    except:
        pass
    
    source_loc = None
    dest_loc = None
    
    if txn.get("source_location_id"):
        try:
            loc = locations_collection.find_one({"_id": ObjectId(txn["source_location_id"])})
            source_loc = loc["name"] if loc else None
        except:
            pass
    
    if txn.get("destination_location_id"):
        try:
            loc = locations_collection.find_one({"_id": ObjectId(txn["destination_location_id"])})
            dest_loc = loc["name"] if loc else None
        except:
            pass
    
    # Use fallbacks for item name
    item_name = "Unknown"
    if item:
        item_name = item["name"]
    elif lot and lot.get("item_name"):
        item_name = lot["item_name"]
    
    return {
        "id": str(txn["_id"]),
        "type": txn["type"],
        "lot_id": txn["lot_id"],
        "lot_number": lot["lot_number"] if lot else "Unknown",
        "item_name": item_name,
        "quantity": txn["quantity"],
        "source_location": source_loc,
        "destination_location": dest_loc,
        "notes": txn.get("notes"),
        "created_at": txn["created_at"]
    }

def generate_requisition_number():
    """Generate a unique requisition serial number like REQ-20260115-0001"""
    now = datetime.now(timezone.utc)
    date_str = now.strftime('%Y%m%d')
    
    # Count requisitions for today
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    count = requisitions_collection.count_documents({
        "created_at": {"$gte": today_start.isoformat()}
    })
    
    return f"REQ-{date_str}-{count + 1:04d}"

def serialize_requisition(req: dict) -> dict:
    # Try to find kitchen by ID
    kitchen = None
    kitchen_id = req.get("kitchen_id") or req.get("location_id")
    if kitchen_id:
        try:
            kitchen = locations_collection.find_one({"_id": ObjectId(kitchen_id)})
        except:
            pass
    
    # Use stored name as fallback
    stored_kitchen_name = req.get("kitchen_name") or req.get("location_name", "")
    stored_kitchen_address = req.get("kitchen_address") or req.get("location_address", "")
    
    items_detail = []
    for item in req.get("items", []):
        item_id = item.get("item_id")
        item_doc = None
        if item_id:
            try:
                item_doc = items_collection.find_one({"_id": ObjectId(item_id)})
            except:
                pass
        
        qty_requested = item.get("quantity") or item.get("quantity_requested") or 0
        qty_sent = item.get("quantity_sent", 0)
        
        items_detail.append({
            "item_id": item_id or "",
            "item_name": item.get("item_name") or (item_doc["name"] if item_doc else "Unknown"),
            "category": item.get("category") or (item_doc["category"] if item_doc else ""),
            "unit": item.get("unit") or (item_doc["unit"] if item_doc else ""),
            "quantity_requested": qty_requested,
            "quantity_sent": qty_sent,
            "shortage": qty_requested - qty_sent,
            "remark": item.get("remark", ""),
            "notes": item.get("notes")
        })
    
    return {
        "id": str(req["_id"]),
        "serial_number": req.get("serial_number", ""),
        "kitchen_id": kitchen_id or "",
        "kitchen_name": kitchen["name"] if kitchen else (stored_kitchen_name or "Unknown"),
        "kitchen_address": kitchen.get("address", "") if kitchen else stored_kitchen_address,
        "items": items_detail,
        "status": req["status"],
        "priority": req.get("priority", "normal"),
        "notes": req.get("notes"),
        "dispatch_notes": req.get("dispatch_notes"),
        "dispatched_at": req.get("dispatched_at"),
        "challan_number": req.get("challan_number"),
        "created_at": req["created_at"],
        "updated_at": req.get("updated_at")
    }

# ============ Auth Endpoints ============

@app.post("/api/auth/register")
async def register_user(user: UserCreate, current_user = Depends(get_current_user)):
    # Only admin can create users (or if no users exist)
    user_count = users_collection.count_documents({})
    if user_count > 0 and (not current_user or current_user["role"] != "admin"):
        raise HTTPException(status_code=403, detail="Only admin can create users")
    
    # Check if email exists
    if users_collection.find_one({"email": user.email.lower()}):
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate location for kitchen users
    if user.role == "kitchen" and not user.location_id:
        raise HTTPException(status_code=400, detail="Kitchen users must have a location assigned")
    
    if user.location_id:
        location = locations_collection.find_one({"_id": ObjectId(user.location_id)})
        if not location:
            raise HTTPException(status_code=404, detail="Location not found")
    
    # Create user
    user_doc = {
        "email": user.email.lower(),
        "password": hash_password(user.password),
        "name": user.name,
        "role": user.role,
        "location_id": user.location_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = users_collection.insert_one(user_doc)
    
    return {
        "message": "User created successfully",
        "user_id": str(result.inserted_id)
    }

@app.post("/api/auth/admin-reset")
async def reset_admin_user():
    """
    Emergency admin reset endpoint - creates or resets all users
    This should be removed or secured after initial setup
    """
    # All users to create/reset
    users_to_setup = [
        {"email": "parveenkatyal2312@gmail.com", "password": "admin@123", "name": "Admin", "role": "admin", "location_name": None},
        {"email": "adreamoven@gmail.com", "password": "store@123", "name": "Main Store", "role": "main_store", "location_name": "Main Store"},
        {"email": "stickyricereceiving@gmail.com", "password": "kitchen@123", "name": "Sticky Rice GCR", "role": "kitchen", "location_name": "Sticky Rice GCR"},
        {"email": "asianbbqtanyan@gmail.com", "password": "kitchen@123", "name": "Tanyan", "role": "kitchen", "location_name": "Tanyan"},
        {"email": "malaproject@kinfolk.com", "password": "kitchen@123", "name": "Mala Project", "role": "kitchen", "location_name": "Mala Project"},
        {"email": "kalaunjibestech@gmail.com", "password": "kitchen@123", "name": "Kalaunji Bestech", "role": "kitchen", "location_name": "Kalaunji Bestech"},
        {"email": "kalaunji43@gmail.com", "password": "kitchen@123", "name": "Kalaunji GCR", "role": "kitchen", "location_name": "Kalaunji GCR"},
        {"email": "sstickyricebestech@gmail.com", "password": "kitchen@123", "name": "Sticky Rice Bestech", "role": "kitchen", "location_name": "Sticky Rice Bestech"},
        {"email": "supersalsa5325@gmail.com", "password": "kitchen@123", "name": "Super Salsa", "role": "kitchen", "location_name": "Super Salsa"},
        {"email": "stickyriceelan@gmail.com", "password": "kitchen@123", "name": "Sticky Rice Elan Epic", "role": "kitchen", "location_name": "Sticky Rice Elan Epic"},
    ]
    
    results = []
    
    for user_data in users_to_setup:
        email = user_data["email"].lower()
        
        # Find location_id if location_name is provided
        location_id = None
        if user_data["location_name"]:
            location = locations_collection.find_one({"name": {"$regex": f"^{user_data['location_name']}$", "$options": "i"}})
            if location:
                location_id = str(location["_id"])
        
        # Check if user exists
        existing_user = users_collection.find_one({"email": email})
        
        if existing_user:
            # Update password
            users_collection.update_one(
                {"email": email},
                {"$set": {
                    "password": hash_password(user_data["password"]),
                    "role": user_data["role"],
                    "location_id": location_id
                }}
            )
            results.append({"email": email, "status": "password_reset"})
        else:
            # Create new user
            user_doc = {
                "email": email,
                "password": hash_password(user_data["password"]),
                "name": user_data["name"],
                "role": user_data["role"],
                "location_id": location_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            users_collection.insert_one(user_doc)
            results.append({"email": email, "status": "created"})
    
    return {"message": "All users setup completed", "results": results}

@app.post("/api/auth/login")
async def login(credentials: UserLogin):
    user = users_collection.find_one({"email": credentials.email.lower()})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    location_name = None
    location_address = None
    location_code = None
    if user.get("location_id"):
        location = locations_collection.find_one({"_id": ObjectId(user["location_id"])})
        if location:
            location_name = location.get("name")
            location_address = location.get("address", "")
            location_code = location.get("code", "")
    
    token = create_token(
        str(user["_id"]), 
        user["email"], 
        user["role"],
        user.get("location_id")
    )
    
    return {
        "token": token,
        "user": {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "role": user["role"],
            "location_id": user.get("location_id"),
            "location_name": location_name,
            "location_address": location_address,
            "location_code": location_code
        }
    }

@app.get("/api/auth/me")
async def get_me(current_user = Depends(require_auth)):
    location_name = None
    location_address = None
    location_code = None
    if current_user.get("location_id"):
        location = locations_collection.find_one({"_id": ObjectId(current_user["location_id"])})
        if location:
            location_name = location.get("name")
            location_address = location.get("address", "")
            location_code = location.get("code", "")
    
    return {
        **current_user,
        "location_name": location_name,
        "location_address": location_address,
        "location_code": location_code
    }

@app.get("/api/users")
async def get_users(current_user = Depends(require_role(["admin", "main_store"]))):
    users = list(users_collection.find({}, {"password": 0}))
    result = []
    for user in users:
        location_name = None
        if user.get("location_id"):
            location = locations_collection.find_one({"_id": ObjectId(user["location_id"])})
            location_name = location["name"] if location else None
        result.append({
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "role": user["role"],
            "location_id": user.get("location_id"),
            "location_name": location_name,
            "created_at": user["created_at"]
        })
    return result

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: str, current_user = Depends(require_role(["admin"]))):
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = users_collection.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    location_id: Optional[str] = None
    password: Optional[str] = None  # Optional password reset

@app.put("/api/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdate, current_user = Depends(require_role(["admin"]))):
    """Update user details including location assignment"""
    
    # Check if user exists
    existing_user = users_collection.find_one({"_id": ObjectId(user_id)})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Build update dict
    update_dict = {}
    
    if user_data.name is not None:
        update_dict["name"] = user_data.name
    
    if user_data.email is not None:
        # Check if email is already taken by another user
        email_exists = users_collection.find_one({
            "email": user_data.email,
            "_id": {"$ne": ObjectId(user_id)}
        })
        if email_exists:
            raise HTTPException(status_code=400, detail="Email already in use by another user")
        update_dict["email"] = user_data.email
    
    if user_data.role is not None:
        if user_data.role not in ["admin", "main_store", "kitchen"]:
            raise HTTPException(status_code=400, detail="Invalid role. Must be 'admin', 'main_store', or 'kitchen'")
        update_dict["role"] = user_data.role
    
    if user_data.location_id is not None:
        # Verify location exists
        if user_data.location_id:
            location = locations_collection.find_one({"_id": ObjectId(user_data.location_id)})
            if not location:
                raise HTTPException(status_code=400, detail="Location not found")
        update_dict["location_id"] = user_data.location_id if user_data.location_id else None
    
    if user_data.password is not None and user_data.password.strip():
        # Hash the new password using the existing hash_password function
        update_dict["password"] = hash_password(user_data.password)
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # Update the user
    users_collection.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": update_dict}
    )
    
    # Fetch updated user
    updated_user = users_collection.find_one({"_id": ObjectId(user_id)})
    location_name = None
    if updated_user.get("location_id"):
        location = locations_collection.find_one({"_id": ObjectId(updated_user["location_id"])})
        location_name = location["name"] if location else None
    
    return {
        "message": "User updated successfully",
        "user": {
            "id": str(updated_user["_id"]),
            "email": updated_user["email"],
            "name": updated_user["name"],
            "role": updated_user["role"],
            "location_id": updated_user.get("location_id"),
            "location_name": location_name
        }
    }

# ============ Invoice OCR Endpoints ============

@app.post("/api/invoice/scan")
async def scan_invoice(
    file: UploadFile = File(...),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Scan invoice image and extract items using GPT-4 Vision"""
    
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="LLM API key not configured")
    
    # Read and encode image
    contents = await file.read()
    image_base64 = base64.b64encode(contents).decode()
    
    # Import LLM chat
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    
    # Create prompt for invoice extraction
    extraction_prompt = """Extract data from this invoice image as JSON:
    {
        "vendor_name": "supplier name",
        "invoice_number": "invoice number",
        "invoice_date": "YYYY-MM-DD",
        "total_amount": total_numeric,
        "items": [
            {
                "item_name": "EXACT article name from invoice",
                "quantity": qty_number,
                "total_amount": total_incl_tax_for_line,
                "rate_per_unit_incl_gst": per_unit_rate_incl_gst
            }
        ]
    }
    
    RULES:
    - For "rate_per_unit_incl_gst": Use "Rate p.u" column if exists, else calculate total_amount/quantity
    - Copy item names EXACTLY including sizes (500 G, 1 KG, 300 ML, PK24, etc)
    - Extract ALL items from the invoice
    - Return ONLY valid JSON"""
    
    # Initialize chat with a vision-capable model
    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"invoice-{datetime.now().timestamp()}",
        system_message="You are an expert at extracting structured data from invoice images. Return only valid JSON."
    ).with_model("openai", "gpt-4o")
    
    # Create message with image using ImageContent for base64
    image_content = ImageContent(image_base64=image_base64)
    user_message = UserMessage(
        text=extraction_prompt,
        file_contents=[image_content]
    )
    
    # Get response
    try:
        response = await chat.send_message(user_message)
        
        # Parse JSON from response
        json_match = re.search(r'\{[\s\S]*\}', response)
        if not json_match:
            raise HTTPException(status_code=422, detail="Could not extract invoice data")
        
        invoice_data = json.loads(json_match.group())
        
        # Match items with database and check price variance
        items_with_match = []
        price_variances = []
        
        # Common OCR spelling corrections
        ocr_corrections = {
            'OYESTER': 'OYSTER',
            'M.S.G': 'MSG',
            'LOTOUS': 'LOTUS', 
            'FLOOR': 'FLOUR',
            'TROTILA': 'TORTILLA',
            'TORTILA': 'TORTILLA',
            'CHESSE': 'CHEESE',
            'SAUSE': 'SAUCE',
            'SEASNING': 'SEASONING',
        }
        
        def correct_ocr(name):
            """Apply common OCR corrections"""
            corrected = name.upper()
            for wrong, right in ocr_corrections.items():
                corrected = corrected.replace(wrong, right)
            return corrected
        
        # Helper function for fuzzy matching
        def find_best_match(invoice_item_name):
            """Try multiple matching strategies to find the best database match"""
            invoice_name = invoice_item_name.upper().strip()
            corrected_name = correct_ocr(invoice_name)
            
            # Try with both original and corrected name
            for search_name in [invoice_name, corrected_name]:
                # Strategy 1: Exact match (case-insensitive)
                db_item = items_collection.find_one({
                    "name": {"$regex": f"^{re.escape(search_name)}$", "$options": "i"}
                })
                if db_item:
                    return db_item, 100
                
                # Strategy 2: Contains match - invoice name in db name or vice versa
                db_item = items_collection.find_one({
                    "name": {"$regex": re.escape(search_name), "$options": "i"}
                })
                if db_item:
                    return db_item, 95
            
            # Strategy 3: Tokenized matching - match key words
            # Clean up the name: remove numbers, special chars, common suffixes
            clean_name = re.sub(r'\d+\s*\*\s*\d+', '', corrected_name)  # Remove patterns like 1*12
            clean_name = re.sub(r'\b\d+\s*(ML|G|KG|L|GM|GRM|PCS|PKT|BTL|CAN|CASE|U|TRAY|PK\d*)s?\b', '', clean_name, flags=re.I)
            clean_name = re.sub(r'\s+', ' ', clean_name).strip()
            
            # Get key words (at least 3 chars, skip common words)
            skip_words = {'THE', 'AND', 'FOR', 'WITH'}
            key_words = [w for w in clean_name.split() if len(w) >= 3 and w not in skip_words]
            
            if len(key_words) >= 2:
                # Build pattern with first 2-3 significant words that must all be present
                search_words = key_words[:3]
                search_pattern = '(?=.*' + ')(?=.*'.join(re.escape(w) for w in search_words) + ')'
                candidates = list(items_collection.find({
                    "name": {"$regex": search_pattern, "$options": "i"}
                }).limit(10))
                
                # Score candidates by word overlap
                best_match = None
                best_score = 0
                for candidate in candidates:
                    cand_name = candidate["name"].upper()
                    cand_words = set(w for w in re.split(r'\W+', cand_name) if len(w) >= 3)
                    invoice_words = set(key_words)
                    
                    # Calculate Jaccard similarity
                    intersection = len(cand_words & invoice_words)
                    union = len(cand_words | invoice_words)
                    if union > 0:
                        score = (intersection / union) * 80
                        if score > best_score:
                            best_score = score
                            best_match = candidate
                
                if best_match and best_score >= 40:
                    return best_match, best_score
            
            # Strategy 4: Match on first unique word (brand name)
            if key_words and len(key_words[0]) >= 4:
                main_word = key_words[0]
                candidates = list(items_collection.find({
                    "name": {"$regex": f"\\b{re.escape(main_word)}\\b", "$options": "i"}
                }).limit(10))
                
                if len(candidates) == 1:
                    return candidates[0], 70
                elif candidates:
                    # Score by word overlap
                    best_match = None
                    best_score = 0
                    for candidate in candidates:
                        cand_name = candidate["name"].upper()
                        match_count = sum(1 for w in key_words if w in cand_name)
                        score = (match_count / max(len(key_words), 1)) * 65
                        if score > best_score:
                            best_score = score
                            best_match = candidate
                    if best_match and best_score >= 35:
                        return best_match, best_score
            
            return None, 0
        
        for item in invoice_data.get("items", []):
            item_name = item.get("item_name", "")
            quantity = item.get("quantity", 1) or 1
            
            # Get GST-inclusive per-unit rate using multiple methods
            rate_with_gst = None
            
            # Method 1: Direct "rate_per_unit_incl_gst" field (from Rate PU column)
            if item.get("rate_per_unit_incl_gst"):
                rate_with_gst = float(item.get("rate_per_unit_incl_gst"))
            
            # Method 2: Calculate from total_amount / quantity
            elif item.get("total_amount") and quantity:
                rate_with_gst = float(item.get("total_amount")) / float(quantity)
            
            # Method 3: Calculate from base_rate + GST%
            elif item.get("base_rate"):
                base_rate = float(item.get("base_rate", 0))
                gst_rate_percent = float(item.get("gst_rate", 0) or 0)
                rate_with_gst = base_rate * (1 + gst_rate_percent / 100)
            
            # Method 4: Fallback to any rate field
            elif item.get("rate"):
                base_rate = float(item.get("rate", 0))
                gst_rate_percent = float(item.get("gst_rate", 0) or 0)
                if gst_rate_percent:
                    rate_with_gst = base_rate * (1 + gst_rate_percent / 100)
                else:
                    rate_with_gst = base_rate
            
            if not rate_with_gst:
                rate_with_gst = 0
            
            # Find matching item using improved matching
            db_item, match_score = find_best_match(item_name)
            
            matched_item_id = None
            matched_item_name = None
            variance = None
            
            if db_item:
                matched_item_id = str(db_item["_id"])
                matched_item_name = db_item["name"]
                standard_price = db_item.get("standard_price", 0)  # This is GST INCL price
                
                # Compare GST-inclusive prices (round to 2 decimal places for comparison)
                rate_with_gst_rounded = round(rate_with_gst, 2)
                standard_price_rounded = round(standard_price, 2) if standard_price else 0
                
                # Check price variance (allow small tolerance of ₹0.50 for rounding differences)
                if standard_price and abs(rate_with_gst_rounded - standard_price_rounded) > 0.5:
                    variance = round(rate_with_gst_rounded - standard_price_rounded, 2)
                    price_variances.append({
                        "item_name": item_name,
                        "matched_db_item": db_item["name"],
                        "invoice_rate_gst_incl": rate_with_gst_rounded,
                        "standard_price_gst_incl": standard_price_rounded,
                        "variance": variance,
                        "variance_percent": round((variance / standard_price_rounded) * 100, 2) if standard_price_rounded else 0
                    })
            
            items_with_match.append({
                **item,
                "rate_gst_inclusive": round(rate_with_gst, 2) if rate_with_gst else None,
                "matched_item_id": matched_item_id,
                "matched_item_name": matched_item_name,
                "matched_standard_price": db_item.get("standard_price") if db_item else None,
                "match_score": match_score,
                "price_variance": variance
            })
        
        result = {
            "vendor_name": invoice_data.get("vendor_name"),
            "invoice_number": invoice_data.get("invoice_number"),
            "invoice_date": invoice_data.get("invoice_date"),
            "total_amount": invoice_data.get("total_amount", 0),
            "items": items_with_match,
            "price_variances": price_variances,
            "has_price_variances": len(price_variances) > 0
        }
        
        # Save to invoices collection
        invoice_doc = {
            **result,
            "image_base64": image_base64[:100] + "...",  # Store truncated for reference
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        invoices_collection.insert_one(invoice_doc)
        
        return result
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="Could not parse invoice data")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing invoice: {str(e)}")

class InvoiceGRNRequest(BaseModel):
    items: List[dict]
    vendor_id: Optional[str] = None
    location_id: str
    expiry_date: str

@app.post("/api/invoice/confirm-grn")
async def confirm_invoice_grn(
    request: InvoiceGRNRequest,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Create GRNs from confirmed invoice items"""
    
    created_lots = []
    
    for item in request.items:
        if not item.get("matched_item_id"):
            continue
        
        # Fetch item details for storing in lot
        db_item = items_collection.find_one({"_id": ObjectId(item["matched_item_id"])})
        
        lot_number = generate_lot_number()
        
        lot_doc = {
            "lot_number": lot_number,
            "item_id": item["matched_item_id"],
            "item_name": db_item["name"] if db_item else item.get("item_name", ""),
            "category": db_item.get("category", "") if db_item else "",
            "unit": db_item.get("unit", "") if db_item else "",
            "initial_quantity": item["quantity"],
            "current_quantity": item["quantity"],
            "expiry_date": request.expiry_date,
            "location_id": request.location_id,
            "vendor_id": request.vendor_id,
            "purchase_rate": item.get("rate"),
            "qr_code": "",
            "source": "invoice_scan",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = lots_collection.insert_one(lot_doc)
        lot_id = str(result.inserted_id)
        
        qr_code = generate_qr_code(lot_id, lot_number)
        lots_collection.update_one(
            {"_id": ObjectId(lot_id)},
            {"$set": {"qr_code": qr_code}}
        )
        
        # Create transaction
        transactions_collection.insert_one({
            "type": "grn",
            "lot_id": lot_id,
            "quantity": item["quantity"],
            "destination_location_id": request.location_id,
            "notes": "Invoice scan GRN",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        created_lots.append({
            "lot_number": lot_number,
            "item_name": item.get("item_name"),
            "quantity": item["quantity"]
        })
    
    return {
        "message": f"Created {len(created_lots)} GRN entries",
        "lots": created_lots
    }

# ============ Requisition Endpoints ============

@app.post("/api/requisitions")
async def create_requisition(
    requisition: RequisitionCreate,
    current_user = Depends(require_role(["kitchen"]))
):
    """Kitchen creates a requisition for items needed"""
    
    if not current_user.get("location_id"):
        raise HTTPException(status_code=400, detail="User has no assigned location")
    
    # DUPLICATE PREVENTION: Check if similar requisition was created in last 30 seconds
    recent_cutoff = datetime.now(timezone.utc) - timedelta(seconds=30)
    recent_req = requisitions_collection.find_one({
        "kitchen_id": current_user["location_id"],
        "created_at": {"$gte": recent_cutoff},
        "status": "pending"
    })
    
    if recent_req:
        return {
            "message": "Requisition already submitted! Please wait...",
            "requisition_id": str(recent_req["_id"]),
            "serial_number": recent_req.get("serial_number", ""),
            "duplicate": True
        }
    
    # Get the user's location details
    location = locations_collection.find_one({"_id": ObjectId(current_user["location_id"])})
    location_name = location["name"] if location else "Unknown"
    location_address = location.get("address", "") if location else ""
    
    # Validate items
    items_data = []
    for item in requisition.items:
        db_item = items_collection.find_one({"_id": ObjectId(item.item_id)})
        if not db_item:
            raise HTTPException(status_code=404, detail=f"Item {item.item_id} not found")
        items_data.append({
            "item_id": item.item_id,
            "item_name": db_item["name"],  # Store item name for reference
            "category": db_item.get("category", ""),
            "unit": db_item.get("unit", ""),
            "quantity": item.quantity,
            "quantity_sent": 0,
            "remark": "",
            "notes": item.notes
        })
    
    # Generate serial number
    serial_number = generate_requisition_number()
    
    req_doc = {
        "serial_number": serial_number,
        "kitchen_id": current_user["location_id"],
        "location_id": ObjectId(current_user["location_id"]),  # Store as ObjectId for consistency
        "location_name": location_name,  # Store kitchen name for display
        "location_address": location_address,  # Store address for challan
        "created_by": ObjectId(current_user["id"]),  # Store as ObjectId
        "items": items_data,
        "status": "pending",  # pending, dispatched, partial, completed
        "priority": requisition.priority,
        "notes": requisition.notes,
        "created_at": datetime.now(timezone.utc)
    }
    
    result = requisitions_collection.insert_one(req_doc)
    
    return {
        "message": "Requisition created successfully",
        "requisition_id": str(result.inserted_id),
        "serial_number": serial_number
    }

@app.get("/api/requisitions")
async def get_requisitions(
    status: Optional[str] = None,
    kitchen_id: Optional[str] = None,
    current_user = Depends(require_auth)
):
    """Get requisitions - Main Store sees all, Kitchen sees only their own"""
    
    query = {}
    
    # Kitchen users can only see their own requisitions
    if current_user["role"] == "kitchen":
        query["kitchen_id"] = current_user["location_id"]
    elif kitchen_id:
        query["kitchen_id"] = kitchen_id
    
    if status:
        query["status"] = status
    
    requisitions = list(requisitions_collection.find(query).sort("created_at", -1).limit(500))
    
    # Pre-fetch all kitchens and items to avoid N+1 queries
    kitchen_ids = set()
    item_ids = set()
    for req in requisitions:
        kid = req.get("kitchen_id") or req.get("location_id")
        if kid:
            kitchen_ids.add(kid)
        for item in req.get("items", []):
            if item.get("item_id"):
                item_ids.add(item["item_id"])
    
    kitchens_map = {}
    for k in locations_collection.find({"_id": {"$in": [ObjectId(kid) for kid in kitchen_ids if kid]}}):
        kitchens_map[str(k["_id"])] = k
    
    items_map = {}
    for item in items_collection.find({"_id": {"$in": [ObjectId(iid) for iid in item_ids if iid]}}):
        items_map[str(item["_id"])] = item
    
    # Serialize with pre-fetched data
    result = []
    for req in requisitions:
        kitchen_id = req.get("kitchen_id") or req.get("location_id")
        kitchen = kitchens_map.get(kitchen_id) if kitchen_id else None
        
        stored_kitchen_name = req.get("kitchen_name") or req.get("location_name", "")
        stored_kitchen_address = req.get("kitchen_address") or req.get("location_address", "")
        
        items_detail = []
        for item in req.get("items", []):
            item_id = item.get("item_id")
            item_doc = items_map.get(item_id) if item_id else None
            
            qty_requested = item.get("quantity") or item.get("quantity_requested") or 0
            qty_sent = item.get("quantity_sent", 0)
            
            items_detail.append({
                "item_id": item_id or "",
                "item_name": item.get("item_name") or (item_doc["name"] if item_doc else "Unknown"),
                "category": item.get("category") or (item_doc["category"] if item_doc else ""),
                "unit": item.get("unit") or (item_doc["unit"] if item_doc else ""),
                "quantity_requested": qty_requested,
                "quantity_sent": qty_sent,
                "shortage": qty_requested - qty_sent,
                "remark": item.get("remark", ""),
                "notes": item.get("notes")
            })
        
        result.append({
            "id": str(req["_id"]),
            "serial_number": req.get("serial_number", ""),
            "kitchen_id": kitchen_id or "",
            "kitchen_name": kitchen["name"] if kitchen else (stored_kitchen_name or "Unknown"),
            "kitchen_address": kitchen.get("address", "") if kitchen else stored_kitchen_address,
            "items": items_detail,
            "status": req["status"],
            "priority": req.get("priority", "normal"),
            "notes": req.get("notes"),
            "dispatch_notes": req.get("dispatch_notes"),
            "dispatched_at": req.get("dispatched_at"),
            "challan_number": req.get("challan_number"),
            "created_at": req["created_at"],
            "updated_at": req.get("updated_at")
        })
    
    return result

@app.get("/api/requisitions/{req_id}")
async def get_requisition(req_id: str, current_user = Depends(require_auth)):
    req = requisitions_collection.find_one({"_id": ObjectId(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Requisition not found")
    
    # Kitchen users can only see their own
    if current_user["role"] == "kitchen" and req["kitchen_id"] != current_user["location_id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return serialize_requisition(req)

@app.get("/api/requisitions/pending/count")
async def get_pending_count(current_user = Depends(require_role(["admin", "main_store"]))):
    """Get count of pending requisitions for Main Store notification"""
    count = requisitions_collection.count_documents({"status": "pending"})
    return {"pending_count": count}

@app.post("/api/requisitions/{req_id}/dispatch")
async def dispatch_requisition(
    req_id: str,
    dispatch: DispatchCreate,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Main Store dispatches items against a requisition"""
    
    req = requisitions_collection.find_one({"_id": ObjectId(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Requisition not found")
    
    if req["status"] == "completed":
        raise HTTPException(status_code=400, detail="Requisition already completed")
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=500, detail="Main Store location not configured")
    
    main_store_id = str(main_store["_id"])
    destination_id = req["kitchen_id"]
    
    # Generate challan number
    now = datetime.now(timezone.utc)
    challan_count = requisitions_collection.count_documents({
        "challan_number": {"$exists": True}
    })
    challan_number = f"CH-{now.strftime('%Y%m%d')}-{challan_count + 1:04d}"
    
    # Process each dispatch item
    shortages = []
    updated_items = req["items"].copy()
    
    for dispatch_item in dispatch.items:
        # Find matching item in requisition
        req_item_idx = None
        for idx, ri in enumerate(updated_items):
            if ri["item_id"] == dispatch_item.item_id:
                req_item_idx = idx
                break
        
        if req_item_idx is None:
            continue
        
        req_item = updated_items[req_item_idx]
        qty_to_send = dispatch_item.quantity_sent
        
        # Use FEFO to get lots for this item
        lots = list(lots_collection.find({
            "item_id": dispatch_item.item_id,
            "location_id": main_store_id,
            "current_quantity": {"$gt": 0}
        }).sort("expiry_date", 1))
        
        # Filter out expired lots
        now_check = datetime.now(timezone.utc)
        active_lots = []
        for lot in lots:
            try:
                expiry = parse_expiry_date(lot.get("expiry_date"))
                if expiry >= now_check:
                    active_lots.append(lot)
            except Exception as e:
                # If expiry date parsing fails, include the lot (better to allow dispatch than block)
                print(f"Warning: Could not parse expiry date for lot {lot.get('_id')}: {e}")
                active_lots.append(lot)
        
        # Calculate available quantity
        available = sum(l["current_quantity"] for l in active_lots)
        actual_sent = min(qty_to_send, available)
        
        # Determine remark based on quantity
        # Handle both 'quantity' and 'quantity_requested' keys for backward compatibility
        requested_qty = req_item.get("quantity") or req_item.get("quantity_requested") or 0
        remark = dispatch_item.remark
        if actual_sent < requested_qty:
            remark = "short"
            item_doc = items_collection.find_one({"_id": ObjectId(dispatch_item.item_id)})
            shortages.append({
                "item_id": dispatch_item.item_id,
                "item_name": item_doc["name"] if item_doc else "Unknown",
                "requested": requested_qty,
                "sent": req_item.get("quantity_sent", 0) + actual_sent,
                "shortage": requested_qty - (req_item.get("quantity_sent", 0) + actual_sent)
            })
        
        # Deduct from lots using FEFO
        remaining = actual_sent
        for lot in active_lots:
            if remaining <= 0:
                break
            
            deduct = min(lot["current_quantity"], remaining)
            new_qty = lot["current_quantity"] - deduct
            
            lots_collection.update_one(
                {"_id": lot["_id"]},
                {"$set": {"current_quantity": new_qty}}
            )
            
            # Get item for standard price fallback
            item = items_collection.find_one({"_id": ObjectId(lot["item_id"])})
            rate = lot.get("purchase_rate") or (item.get("standard_price") if item else 0) or 0
            
            # Create issue transaction with rate stored
            transactions_collection.insert_one({
                "type": "issue",
                "lot_id": str(lot["_id"]),
                "item_id": lot["item_id"],
                "quantity": deduct,
                "rate": rate,  # Store rate directly in transaction
                "value": deduct * rate,  # Store calculated value
                "source_location_id": main_store_id,
                "destination_location_id": destination_id,
                "requisition_id": req_id,
                "challan_number": challan_number,
                "notes": "Requisition dispatch",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            
            remaining -= deduct
        
        # Update requisition item with quantity sent and remark
        updated_items[req_item_idx]["quantity_sent"] = req_item.get("quantity_sent", 0) + actual_sent
        updated_items[req_item_idx]["remark"] = remark
    
    # Determine new status
    # Handle both 'quantity' and 'quantity_requested' keys
    all_fulfilled = all(
        item.get("quantity_sent", 0) >= (item.get("quantity") or item.get("quantity_requested") or 0)
        for item in updated_items
    )
    any_sent = any(item.get("quantity_sent", 0) > 0 for item in updated_items)
    
    if all_fulfilled:
        new_status = "completed"
    elif any_sent:
        new_status = "partial"
    else:
        new_status = "dispatched"
    
    # Update requisition
    requisitions_collection.update_one(
        {"_id": ObjectId(req_id)},
        {"$set": {
            "items": updated_items,
            "status": new_status,
            "challan_number": challan_number,
            "dispatch_notes": dispatch.notes,
            "dispatched_by": current_user["id"],
            "dispatched_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "shortages": shortages
        }}
    )
    
    return {
        "message": "Dispatch completed",
        "status": new_status,
        "challan_number": challan_number,
        "shortages": shortages,
        "has_shortages": len(shortages) > 0
    }

@app.post("/api/requisitions/{req_id}/confirm-receipt")
async def confirm_requisition_receipt(
    req_id: str,
    current_user = Depends(require_role(["kitchen"]))
):
    """Kitchen confirms receipt of dispatched goods"""
    
    requisition = requisitions_collection.find_one({"_id": ObjectId(req_id)})
    if not requisition:
        raise HTTPException(status_code=404, detail="Requisition not found")
    
    # Verify this requisition belongs to the kitchen
    if requisition.get("kitchen_id") != current_user.get("location_id"):
        raise HTTPException(status_code=403, detail="You can only confirm receipt for your own kitchen's requisitions")
    
    # Check if requisition is in dispatched status
    if requisition["status"] not in ["dispatched", "partial"]:
        raise HTTPException(status_code=400, detail="Requisition must be dispatched before confirming receipt")
    
    # Update requisition status to "received"
    requisitions_collection.update_one(
        {"_id": ObjectId(req_id)},
        {"$set": {
            "status": "received",
            "received_by": current_user["id"],
            "received_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Get kitchen name for notification
    kitchen = locations_collection.find_one({"_id": ObjectId(requisition["kitchen_id"])})
    kitchen_name = kitchen["name"] if kitchen else "Unknown Kitchen"
    
    return {
        "message": f"Receipt confirmed for {kitchen_name}",
        "status": "received",
        "received_at": datetime.now(timezone.utc).isoformat()
    }

@app.get("/api/requisitions/shortages")
async def get_all_shortages(current_user = Depends(require_auth)):
    """Get all requisitions with shortages"""
    
    query = {
        "shortages": {"$exists": True, "$ne": []},
        "status": {"$in": ["partial", "dispatched"]}
    }
    
    # Kitchen sees only their shortages
    if current_user["role"] == "kitchen":
        query["kitchen_id"] = current_user["location_id"]
    
    requisitions = list(requisitions_collection.find(query).sort("updated_at", -1))
    
    result = []
    for req in requisitions:
        kitchen = locations_collection.find_one({"_id": ObjectId(req["kitchen_id"])})
        result.append({
            "requisition_id": str(req["_id"]),
            "kitchen_name": kitchen["name"] if kitchen else "Unknown",
            "shortages": req.get("shortages", []),
            "created_at": req["created_at"],
            "dispatched_at": req.get("dispatched_at")
        })
    
    return result


@app.delete("/api/requisitions/{req_id}")
async def delete_requisition(
    req_id: str,
    current_user = Depends(require_role(["admin"]))
):
    """Delete a Requisition (admin only) - for removing test data"""
    req = requisitions_collection.find_one({"_id": ObjectId(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Requisition not found")
    
    result = requisitions_collection.delete_one({"_id": ObjectId(req_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Failed to delete Requisition")
    
    return {"message": f"Requisition {req.get('serial_number', req_id)} deleted successfully"}


@app.delete("/api/requisitions/bulk")
async def bulk_delete_requisitions(
    month: int = Query(..., description="Month (1-12)"),
    year: int = Query(..., description="Year (e.g., 2026)"),
    status: str = Query(None, description="Filter by status: pending, partial, dispatched, received"),
    current_user = Depends(require_role(["admin"]))
):
    """Bulk delete Requisitions by month/year (admin only) - for cleaning test data"""
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    query = {
        "created_at": {"$gte": start_date, "$lt": end_date}
    }
    if status:
        query["status"] = status
    
    count = requisitions_collection.count_documents(query)
    if count == 0:
        return {"message": "No matching Requisitions found", "deleted_count": 0}
    
    result = requisitions_collection.delete_many(query)
    return {"message": f"Deleted {result.deleted_count} Requisitions from {month:02d}/{year}", "deleted_count": result.deleted_count}


@app.get("/api/requisitions/{req_id}/challan")
async def get_challan_pdf(req_id: str):
    """Generate Delivery Challan PDF for a dispatched requisition"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from fastapi.responses import Response
    
    req = requisitions_collection.find_one({"_id": ObjectId(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Requisition not found")
    
    if not req.get("challan_number"):
        raise HTTPException(status_code=400, detail="Requisition has not been dispatched yet")
    
    # Get kitchen details
    kitchen = locations_collection.find_one({"_id": ObjectId(req["kitchen_id"])})
    
    # Get company settings
    company = company_settings_collection.find_one({}) or {
        "name": "Dreamoven",
        "address": "Main Store Address",
        "phone": "",
        "email": "",
        "gst_number": "07AGSPA1692G1ZB"
    }
    
    # Get main store details
    main_store = locations_collection.find_one({"type": "main_store"})
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER, spaceAfter=20)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9)
    right_style = ParagraphStyle('Right', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT)
    
    elements = []
    
    # Title
    elements.append(Paragraph("DELIVERY CHALLAN", title_style))
    elements.append(Paragraph("(Material Transfer Note)", subtitle_style))
    
    # Helper to format date from various formats (string or datetime)
    def format_date(date_val):
        if not date_val:
            return ''
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            return date_val[:10]
        return str(date_val)[:10]
    
    # Challan Details
    challan_info = [
        [Paragraph(f"<b>Challan No:</b> {req['challan_number']}", normal_style),
         Paragraph(f"<b>Date:</b> {format_date(req.get('dispatched_at'))}", right_style)],
        [Paragraph(f"<b>Requisition No:</b> {req.get('serial_number', '')}", normal_style),
         Paragraph(f"<b>Req. Date:</b> {format_date(req.get('created_at'))}", right_style)]
    ]
    challan_table = Table(challan_info, colWidths=[3*inch, 3*inch])
    challan_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(challan_table)
    elements.append(Spacer(1, 15))
    
    # From/To Section
    from_to = [
        [Paragraph("<b>FROM (Main Store):</b>", header_style), Paragraph("<b>TO (Kitchen/Outlet):</b>", header_style)],
        [Paragraph(f"{company.get('name', 'A Dream Oven')}<br/>{company.get('address', '')}<br/>Phone: {company.get('phone', '')}<br/>GSTIN: {company.get('gst_number', '')}", normal_style),
         Paragraph(f"{kitchen['name'] if kitchen else 'Unknown'}<br/>{kitchen.get('address', '') if kitchen else ''}", normal_style)]
    ]
    from_to_table = Table(from_to, colWidths=[3*inch, 3*inch])
    from_to_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(from_to_table)
    elements.append(Spacer(1, 20))
    
    # Items Table
    items_header = ['#', 'Item Description', 'Qty Req.', 'Qty Sent', 'Unit', 'Remark']
    items_data = [items_header]
    
    for idx, item in enumerate(req['items'], 1):
        item_doc = None
        try:
            item_doc = items_collection.find_one({"_id": ObjectId(item["item_id"])})
        except:
            pass
        
        qty_sent = item.get("quantity_sent", 0)
        # Handle both 'quantity' and 'quantity_requested' keys for backward compatibility
        qty_req = item.get("quantity") or item.get("quantity_requested") or 0
        remark = item.get("remark", "")
        if not remark:
            remark = "OK" if qty_sent >= qty_req else "SHORT"
        
        # Use item_name from requisition as fallback if item not found in items collection
        item_name = item.get("item_name") or (item_doc["name"] if item_doc else "Unknown")
        item_unit = item.get("unit") or (item_doc["unit"] if item_doc else "")
        
        items_data.append([
            str(idx),
            item_name,
            str(int(qty_req)),
            str(int(qty_sent)),
            item_unit,
            remark.upper()
        ])
    
    items_table = Table(items_data, colWidths=[0.4*inch, 2.8*inch, 0.7*inch, 0.7*inch, 0.6*inch, 0.9*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 20))
    
    # Notes
    if req.get("dispatch_notes"):
        elements.append(Paragraph(f"<b>Notes:</b> {req['dispatch_notes']}", normal_style))
        elements.append(Spacer(1, 10))
    
    # Total items
    elements.append(Paragraph(f"<b>Total Items:</b> {len(req['items'])}", normal_style))
    elements.append(Spacer(1, 30))
    
    # Signature Section
    sig_data = [
        [Paragraph("<b>Prepared By</b>", normal_style), 
         Paragraph("<b>Checked By</b>", normal_style),
         Paragraph("<b>Received By</b>", normal_style)],
        ['', '', ''],
        ['_________________', '_________________', '_________________'],
        [Paragraph("(Main Store)", normal_style), 
         Paragraph("(Supervisor)", normal_style),
         Paragraph("(Kitchen)", normal_style)]
    ]
    sig_table = Table(sig_data, colWidths=[2*inch, 2*inch, 2*inch])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 30),
    ]))
    elements.append(sig_table)
    
    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("This is a computer-generated document.", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray)))
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Challan-{req['challan_number']}.pdf"}
    )

# ============ Notifications Endpoint ============

@app.get("/api/notifications")
async def get_notifications(current_user = Depends(require_auth)):
    """Get notifications for current user"""
    
    notifications = []
    
    if current_user["role"] in ["admin", "main_store"]:
        # Pending requisitions
        try:
            pending = requisitions_collection.count_documents({"status": "pending"})
            if pending > 0:
                notifications.append({
                    "type": "requisition",
                    "title": "Pending Requisitions",
                    "message": f"{pending} requisition(s) waiting for dispatch",
                    "count": pending,
                    "priority": "high"
                })
        except:
            pass
        
        # Expiring items - OPTIMIZED: Only check recent lots with expiry dates
        try:
            now = datetime.now(timezone.utc)
            seven_days_later = now + timedelta(days=7)
            
            # Only fetch lots that have expiry_date and current stock
            expiring_count = lots_collection.count_documents({
                "current_quantity": {"$gt": 0},
                "expiry_date": {
                    "$exists": True,
                    "$ne": None,
                    "$lte": seven_days_later.isoformat()[:10],
                    "$gte": now.isoformat()[:10]
                }
            })
            
            if expiring_count > 0:
                notifications.append({
                    "type": "expiry",
                    "title": "Expiring Soon",
                    "message": f"{expiring_count} item(s) expiring within 7 days",
                    "count": expiring_count,
                    "priority": "medium"
                })
        except Exception as e:
            print(f"Expiry check skipped: {e}")
    
    elif current_user["role"] == "kitchen":
        # Check for dispatched requisitions
        dispatched = requisitions_collection.count_documents({
            "kitchen_id": current_user["location_id"],
            "status": {"$in": ["dispatched", "partial", "completed"]},
            "dispatched_at": {"$exists": True}
        })
        
        # Check for shortages
        with_shortages = list(requisitions_collection.find({
            "kitchen_id": current_user["location_id"],
            "shortages": {"$exists": True, "$ne": []}
        }))
        
        if len(with_shortages) > 0:
            total_shortages = sum(len(r.get("shortages", [])) for r in with_shortages)
            notifications.append({
                "type": "shortage",
                "title": "Short Deliveries",
                "message": f"{total_shortages} item(s) were short delivered",
                "count": total_shortages,
                "priority": "high"
            })
    
    return notifications

# ============ Original Endpoints (keeping all existing functionality) ============

# Note: /health and /api/health are defined at the top of the file for Kubernetes health checks

@app.post("/api/admin/import-data")
async def import_data_endpoint(current_user = Depends(require_role(["admin"]))):
    """
    Import data from exported JSON files.
    This endpoint is used after deployment to migrate data from preview to production.
    ADMIN ONLY.
    """
    import subprocess
    import sys
    
    script_path = os.path.join(os.path.dirname(__file__), "import_data.py")
    
    if not os.path.exists(script_path):
        raise HTTPException(status_code=404, detail="Import script not found")
    
    try:
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=300
        )
        
        return {
            "success": result.returncode == 0,
            "output": result.stdout,
            "errors": result.stderr if result.returncode != 0 else None
        }
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="Import timed out")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ONE-TIME SETUP ENDPOINT - No auth required for initial data migration
# This should be removed or disabled after first use
@app.post("/api/setup/migrate-data")
async def migrate_data_setup(secret_key: str = ""):
    """
    One-time data migration endpoint. 
    Use secret key: KINFOLK-MIGRATE-2026
    This endpoint will be disabled after successful migration.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    import subprocess
    import sys
    
    script_path = os.path.join(os.path.dirname(__file__), "import_data.py")
    
    if not os.path.exists(script_path):
        raise HTTPException(status_code=404, detail="Import script not found")
    
    try:
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=300
        )
        
        return {
            "success": result.returncode == 0,
            "output": result.stdout,
            "errors": result.stderr if result.returncode != 0 else None,
            "message": "Migration complete. Please redeploy without this endpoint for security."
        }
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="Import timed out")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ONE-TIME PASSWORD RESET ENDPOINT
@app.post("/api/setup/reset-passwords")
async def reset_passwords_setup(secret_key: str = ""):
    """
    One-time password reset for all users.
    Use secret key: KINFOLK-MIGRATE-2026
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Define user passwords
    user_passwords = {
        "parveenkatyal2312@gmail.com": "admin@123",
        "adreamoven@gmail.com": "store@123",
        "stickyricereceiving@gmail.com": "kitchen@123",
        "asianbbqtanyan@gmail.com": "kitchen@123",
        "malaproject@kinfolk.com": "kitchen@123",
        "kalaunjibestech@gmail.com": "kitchen@123",
        "kalaunji43@gmail.com": "kitchen@123",
        "sstickyricebestech@gmail.com": "kitchen@123",
        "supersalsa5325@gmail.com": "kitchen@123",
        "stickyriceelan@gmail.com": "kitchen@123",
    }
    
    updated = []
    for email, password in user_passwords.items():
        # Use the existing hash_password function
        hashed = hash_password(password)
        result = users_collection.update_one(
            {"email": email},
            {"$set": {"password": hashed}}
        )
        if result.modified_count > 0:
            updated.append(email)
        elif result.matched_count == 0:
            # User doesn't exist, create them
            user_data = {
                "email": email,
                "name": email.split("@")[0],
                "password": hashed,
                "role": "admin" if "parveen" in email.lower() else ("main_store" if "adream" in email.lower() else "kitchen"),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            users_collection.insert_one(user_data)
            updated.append(f"{email} (created)")
    
    return {
        "success": True,
        "updated_users": updated,
        "message": "Passwords reset successfully"
    }

# ONE-TIME REQUISITION IMPORT ENDPOINT
@app.post("/api/setup/import-requisitions")
async def import_requisitions_setup(secret_key: str = ""):
    """
    Import missing requisitions from preview environment.
    Use secret key: KINFOLK-MIGRATE-2026
    """
    import json
    import os
    
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Load missing requisitions
    data_file = os.path.join(os.path.dirname(__file__), '..', 'data_export_latest', 'missing_requisitions.json')
    
    if not os.path.exists(data_file):
        raise HTTPException(status_code=404, detail="Missing requisitions file not found")
    
    with open(data_file, 'r') as f:
        missing_reqs = json.load(f)
    
    # Get location mapping
    locations = {loc["name"]: str(loc["_id"]) for loc in locations_collection.find({})}
    
    imported = 0
    skipped = 0
    
    for req in missing_reqs:
        # Check if already exists - check both field names
        req_num = req.get("requisition_number")
        existing = requisitions_collection.find_one({
            "$or": [
                {"requisition_number": req_num},
                {"serial_number": req_num}
            ]
        })
        if existing:
            skipped += 1
            continue
        
        # Map location
        kitchen_name = req.get("source_location_name", req.get("kitchen_name", ""))
        location_id = locations.get(kitchen_name)
        
        # Create requisition - use production field names
        new_req = {
            "serial_number": req_num or f"REQ-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{imported}",
            "kitchen_id": location_id or req.get("source_location_id"),
            "kitchen_name": kitchen_name,
            "kitchen_address": "",
            "status": req.get("status", "pending"),
            "priority": req.get("priority", "normal"),
            "items": req.get("items", []),
            "notes": req.get("notes", ""),
            "dispatch_notes": req.get("dispatch_notes"),
            "dispatched_at": req.get("dispatched_at"),
            "challan_number": req.get("challan_number"),
            "created_at": req.get("created_at", datetime.now(timezone.utc).isoformat()),
            "updated_at": None
        }
        
        requisitions_collection.insert_one(new_req)
        imported += 1
    
    return {
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "message": f"Imported {imported} requisitions, skipped {skipped} duplicates"
    }

# ONE-TIME CATEGORY CLEANUP ENDPOINT
@app.post("/api/setup/cleanup-categories")
async def cleanup_categories_setup(secret_key: str = ""):
    """
    Merge duplicate categories and update all items.
    Use secret key: KINFOLK-MIGRATE-2026
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Define category mappings (source -> target)
    # All variations will be merged into the target name
    category_mappings = {
        # Beverages variations -> Beverage
        "beverages": "Beverage",
        "beverage": "Beverage",
        
        # Mala variations -> Mala Grocery
        "mala": "Mala Grocery",
        "mala grocery": "Mala Grocery",
        
        # Dairy variations -> Dairy
        "dairy product": "Dairy",
        "dairy": "Dairy",
        
        # Standard categories (normalize case)
        "chinese grocery": "Chinese Grocery",
        "continental grocery": "Continental Grocery",
        "indian grocery": "Indian Grocery",
        "housekeeping": "Housekeeping",
        "packaging": "Packaging",
        "seafood": "Seafood",
        "bakery": "Bakery",
        "dry fruits": "Dry Fruits",
        "fruits & vegetables": "Fruits & Vegetables",
        "non veg": "Non Veg",
        "other": "Other",
        "seasoning": "Seasoning",
    }
    
    # Get all categories
    all_cats = list(categories_collection.find({}))
    
    # Track which category IDs to keep and which to delete
    keep_ids = {}  # canonical_name -> id to keep
    delete_ids = []
    
    # First pass: identify canonical categories
    for cat in all_cats:
        name = cat["name"]
        normalized = name.lower().strip()
        canonical = category_mappings.get(normalized, name.title())
        
        if canonical not in keep_ids:
            keep_ids[canonical] = str(cat["_id"])
        else:
            delete_ids.append(str(cat["_id"]))
    
    # Update all items to use canonical category names
    items_updated = 0
    for cat in all_cats:
        name = cat["name"]
        normalized = name.lower().strip()
        canonical = category_mappings.get(normalized, name.title())
        
        # Update items with this category name
        result = items_collection.update_many(
            {"category": name},
            {"$set": {"category": canonical}}
        )
        items_updated += result.modified_count
        
        # Also update items with category_id
        result2 = items_collection.update_many(
            {"category_id": str(cat["_id"])},
            {"$set": {"category": canonical, "category_id": keep_ids.get(canonical)}}
        )
        items_updated += result2.modified_count
    
    # Delete duplicate categories
    deleted_count = 0
    for cat_id in delete_ids:
        try:
            from bson import ObjectId
            categories_collection.delete_one({"_id": ObjectId(cat_id)})
            deleted_count += 1
        except:
            pass
    
    # Rename remaining categories to canonical names
    renamed_count = 0
    for canonical, cat_id in keep_ids.items():
        try:
            from bson import ObjectId
            result = categories_collection.update_one(
                {"_id": ObjectId(cat_id)},
                {"$set": {"name": canonical}}
            )
            if result.modified_count > 0:
                renamed_count += 1
        except:
            pass
    
    return {
        "success": True,
        "categories_deleted": deleted_count,
        "categories_renamed": renamed_count,
        "items_updated": items_updated,
        "final_categories": list(keep_ids.keys())
    }

# ONE-TIME STOCK UPDATE FROM EXCEL ENDPOINT
@app.post("/api/setup/update-stock-from-excel")
async def update_stock_from_excel(secret_key: str = ""):
    """
    Update current stock from Feb 12 Excel file.
    Use secret key: KINFOLK-MIGRATE-2026
    """
    import json
    import os
    import re
    
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Load stock data
    data_file = os.path.join(os.path.dirname(__file__), '..', 'data_export_latest', 'feb12_stock.json')
    
    if not os.path.exists(data_file):
        raise HTTPException(status_code=404, detail="Stock data file not found")
    
    with open(data_file, 'r') as f:
        stock_data = json.load(f)
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Get all items for matching
    all_items = list(items_collection.find({}))
    
    # Create lookup maps
    def normalize(name):
        return re.sub(r'[^a-z0-9]', '', name.lower())
    
    item_by_name = {item["name"].lower().strip(): item for item in all_items}
    item_by_normalized = {normalize(item["name"]): item for item in all_items}
    
    updated = 0
    not_found = []
    
    for stock_item in stock_data:
        item_name = stock_item["name"]
        quantity = stock_item["quantity"]
        
        # Find matching item
        item = None
        
        # Try exact match
        if item_name.lower().strip() in item_by_name:
            item = item_by_name[item_name.lower().strip()]
        
        # Try normalized match
        if not item:
            normalized = normalize(item_name)
            if normalized in item_by_normalized:
                item = item_by_normalized[normalized]
        
        if item:
            item_id = str(item["_id"])
            
            # Find or create lot
            lot = lots_collection.find_one({"item_id": item_id, "location_id": main_store_id})
            
            if lot:
                # Update existing lot
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"current_quantity": quantity, "quantity": quantity}}
                )
            else:
                # Create new lot
                lots_collection.insert_one({
                    "lot_number": f"FEB12-{updated}",
                    "item_id": item_id,
                    "item_name": item["name"],
                    "quantity": quantity,
                    "current_quantity": quantity,
                    "initial_quantity": quantity,
                    "location_id": main_store_id,
                    "source": "excel_import_feb12",
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
            updated += 1
        else:
            if quantity > 0:  # Only track items with stock that weren't found
                not_found.append(item_name)
    
    return {
        "success": True,
        "updated": updated,
        "not_found_count": len(not_found),
        "not_found_sample": not_found[:20]
    }

# ONE-TIME FIX FOR REQUISITIONS WITH UNKNOWN KITCHEN
@app.post("/api/setup/fix-unknown-requisitions")
async def fix_unknown_requisitions(secret_key: str = ""):
    """
    Fix requisitions that have 'Unknown' kitchen name by looking up the location.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Get all locations for mapping
    locations = {str(loc["_id"]): loc["name"] for loc in locations_collection.find({})}
    
    # Find requisitions with Unknown or missing kitchen_name
    fixed = 0
    requisitions = list(requisitions_collection.find({
        "$or": [
            {"kitchen_name": "Unknown"},
            {"kitchen_name": None},
            {"kitchen_name": ""}
        ]
    }))
    
    for req in requisitions:
        kitchen_id = req.get("kitchen_id")
        if kitchen_id:
            # Try to find the location name
            kitchen_name = locations.get(kitchen_id)
            if kitchen_name:
                requisitions_collection.update_one(
                    {"_id": req["_id"]},
                    {"$set": {"kitchen_name": kitchen_name}}
                )
                fixed += 1
    
    return {
        "success": True,
        "fixed": fixed,
        "total_checked": len(requisitions)
    }


# ONE-TIME FIX FOR POs WITH UNKNOWN LOCATION NAME
@app.post("/api/setup/fix-po-locations")
async def fix_po_locations(secret_key: str = ""):
    """
    Fix purchase orders that have missing created_by_location_name by looking up the creator's location.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Build location lookup
    locations = {str(loc["_id"]): loc for loc in locations_collection.find({})}
    
    # Build user lookup
    users = {str(u["_id"]): u for u in users_collection.find({})}
    
    # Find POs with missing location_name
    fixed = 0
    skipped = 0
    pos = list(purchase_orders_collection.find({
        "$or": [
            {"created_by_location_name": {"$exists": False}},
            {"created_by_location_name": None},
            {"created_by_location_name": ""}
        ]
    }))
    
    for po in pos:
        created_by = po.get("created_by")
        if created_by:
            user = users.get(created_by)
            if user and user.get("location_id"):
                location = locations.get(user["location_id"])
                if location:
                    purchase_orders_collection.update_one(
                        {"_id": po["_id"]},
                        {"$set": {
                            "created_by_location_name": location.get("name"),
                            "created_by_location_code": location.get("code"),
                            "location_id": user["location_id"]
                        }}
                    )
                    fixed += 1
                    continue
        
        # If no user/location found, default to Main Store
        main_store = next((loc for loc in locations.values() if loc.get("type") == "main_store"), None)
        if main_store:
            purchase_orders_collection.update_one(
                {"_id": po["_id"]},
                {"$set": {
                    "created_by_location_name": main_store.get("name", "Main Store"),
                    "created_by_location_code": main_store.get("code", "MS01"),
                    "location_id": str(main_store["_id"])
                }}
            )
            fixed += 1
        else:
            skipped += 1
    
    return {
        "success": True,
        "fixed": fixed,
        "skipped": skipped,
        "total_checked": len(pos)
    }


# ONE-TIME CLEANUP FOR DUPLICATE LOCATIONS
@app.post("/api/setup/cleanup-duplicate-locations")
async def cleanup_duplicate_locations(secret_key: str = ""):
    """
    Remove duplicate locations, keeping only those with proper codes.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    from collections import defaultdict
    
    # Get all locations
    locations = list(locations_collection.find({}))
    
    # Group by name
    by_name = defaultdict(list)
    for loc in locations:
        by_name[loc['name']].append(loc)
    
    deleted = 0
    kept = 0
    
    for name, locs in by_name.items():
        if len(locs) <= 1:
            kept += 1
            continue
        
        # Find the one with a code - that's the one to keep
        with_code = [l for l in locs if l.get('code')]
        without_code = [l for l in locs if not l.get('code')]
        
        if with_code:
            # Keep the one with code, delete others
            kept += 1
            for l in without_code:
                locations_collection.delete_one({"_id": l["_id"]})
                deleted += 1
        else:
            # Keep the first one, delete others
            kept += 1
            for l in locs[1:]:
                locations_collection.delete_one({"_id": l["_id"]})
                deleted += 1
    
    return {
        "success": True,
        "kept": kept,
        "deleted": deleted
    }


# ONE-TIME FIX FOR BEVERAGE STOCK
@app.post("/api/setup/fix-beverage-stock")
async def fix_beverage_stock(secret_key: str = ""):
    """
    Update beverage stock to correct values from Feb 20 Excel.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Beverage stock data from the user's Excel (Feb 20, 2026)
    beverage_data = {
        "AGAVE SYRUP": 8,
        "BISLERI SODA 750 ML PK12": 348,
        "CATCH CLUB SODA 500 ML": 0,
        "Coca Cola Can 300 ML": 600,
        "Coke Zero Can 330 Ml": 360,
        "DIET COKE CAN 300 ML": 504,
        "Elder Flower 1L Monin": 5,
        "GRENADINE SYRUP MONIN": 2,
        "Kinley Packaged Drinking Water 1L Pk 12": 48,
        "LAVENDER SYRUP MONIN": 8,
        "PEACH TEA SYRUP MONIN": 1,
        "RASPBERRY SYRUP MONIN": 1,
        "LEMON ICE TEA 1 LTR-MONIN": 5,
        "LYCHEE PUREE MONIN": 3,
        "MANGO 700 ML MONIN": 0,
        "MOJITO MINT 1 LTR - MONIN": 12,
        "Nestea - Lemon Iced Tea Premix, 1 Kg": 15,
        "PANDAN SYRUP MATHEWW": 0,
        "Perrier - Sparkling Water": 128,
        "Pink Grape Fruit 1L Monin": 6,
        "Blueberry Syrup 1L Monin": 1,
        "Puree Blueberry 1L Monin": 5,
        "PUREE BLUEBERRY 1 LTR-MONIN": 5,
        "Passion Fruits Syrup 1L Monin": 4,
        "Puree Passion Fruits 1L Monin": 1,
        "PUREE PASSION FRUIT MONIN": 1,
        "Peach Syrup 1L Monin": 7,
        "Puree Peach 1L Monin": 8,
        "PUREE PEACH 1 LTR -MONIN": 8,
        "Puree Strawberry 1L Monin": 5,
        "PUREE STARWBERRY 1 LTR -MONIN": 5,
        "Puree Yuzu 1L monin": 0,
        "PUREE YUZU 1 LTR -MONIN": 0,
        "Qua Water 750 ML": 504,
        "QUA 750 ML": 504,
        "REAL CRANBERRY JUICE 1 L": 112,
        "REAL ORANGE JUICE 1 L": 180,
        "ROSE MONIN": 5,
        "Real Active Coconut Water Pet 750 Ml": 0,
        "Real Apple Juice 1L PK24": 48,
        "Real Litchi Juice 1L PK 12": 4,
        "Real Pineapple Juice PK 12": 1,
        "SCHWEPPES GINGER ALE CAN 300 ML": 264,
        "STAR ANISE MONIN": 5,
        "Schweppes Soda Water Can 300 ML": 24,
        "Schweppes Tonic Water Can 300 Ml": 0,
        "Sprite Can 300 mL PK24": 210,
        "THUMS UP CAN 300 ML": 402,
        "THUMS UP CAN 300 ML PK24": 402,
        "WATERMELON 1 LTR -MONIN": 5,
        "Red Bull Energy Drink 250 ML PK24": 72,
        "PUREE LYCHEE MONIN": 3,
    }
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main store not found"}
    
    main_store_id = str(main_store["_id"])
    
    updated = 0
    created = 0
    not_found = []
    
    for item_name, quantity in beverage_data.items():
        # Find the item (case-insensitive)
        item = items_collection.find_one({"name": {"$regex": f"^{item_name}$", "$options": "i"}})
        
        if not item:
            # Try partial match
            item = items_collection.find_one({"name": {"$regex": item_name, "$options": "i"}})
        
        if item:
            item_id = str(item["_id"])
            
            # Find existing lot for this item at Main Store
            lot = lots_collection.find_one({
                "item_id": item_id,
                "location_id": main_store_id
            })
            
            if lot:
                # Update existing lot
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"current_quantity": quantity, "quantity": quantity}}
                )
                updated += 1
            else:
                # Create new lot if quantity > 0
                if quantity > 0:
                    lots_collection.insert_one({
                        "lot_number": f"BEV-FEB20-{created}",
                        "item_id": item_id,
                        "item_name": item["name"],
                        "category": "Beverage",
                        "unit": item.get("unit", ""),
                        "quantity": quantity,
                        "current_quantity": quantity,
                        "initial_quantity": quantity,
                        "location_id": main_store_id,
                        "source": "stock_correction_feb20",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    })
                    created += 1
        else:
            if item_name not in not_found:
                not_found.append(item_name)
    
    return {
        "success": True,
        "updated": updated,
        "created": created,
        "not_found": not_found
    }


# FIX REMAINING BEVERAGE ITEMS WITH EXACT NAMES
@app.post("/api/setup/fix-beverage-stock-v2")
async def fix_beverage_stock_v2(secret_key: str = ""):
    """
    Fix remaining beverage items with exact name matching.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Exact name mappings for remaining items
    exact_fixes = {
        "COCA COLA CAN 300ML": 600,
        "QUA 750 ML": 504,
        "PUREE LYCHEE MONIN": 3,
        "PUREE BLUEBERRY 1 LTR-MONIN": 5,
        "PUREE PASSION FRUIT MONIN": 1,
        "PUREE PEACH 1 LTR -MONIN": 8,
        "PUREE YUZU 1 LTR -MONIN": 0,
        "Pink Graoe Fruit 1L Monin": 6,  # Note: might be misspelled in DB
    }
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main store not found"}
    
    main_store_id = str(main_store["_id"])
    
    updated = 0
    not_found = []
    
    for item_name, quantity in exact_fixes.items():
        # Find exact item
        item = items_collection.find_one({"name": item_name})
        
        if item:
            item_id = str(item["_id"])
            
            # Find existing lot
            lot = lots_collection.find_one({
                "item_id": item_id,
                "location_id": main_store_id
            })
            
            if lot:
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"current_quantity": quantity, "quantity": quantity}}
                )
                updated += 1
            elif quantity > 0:
                lots_collection.insert_one({
                    "lot_number": f"BEV-FEB20-V2-{updated}",
                    "item_id": item_id,
                    "item_name": item["name"],
                    "category": "Beverage",
                    "unit": item.get("unit", ""),
                    "quantity": quantity,
                    "current_quantity": quantity,
                    "initial_quantity": quantity,
                    "location_id": main_store_id,
                    "source": "stock_correction_feb20_v2",
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                updated += 1
        else:
            not_found.append(item_name)
    
    return {
        "success": True,
        "updated": updated,
        "not_found": not_found
    }


# ONE-TIME FIX FOR MALA GROCERY STOCK
@app.post("/api/setup/fix-mala-grocery-stock")
async def fix_mala_grocery_stock(secret_key: str = ""):
    """
    Update Mala Grocery stock to correct values from Feb 20 data.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Mala Grocery stock data from the user's image (Feb 20, 2026)
    mala_data = {
        "GOLD MEDAL PEPPER OIL": 0,
        "TIEGUANYIN": 1,
        "AJIDORAKU": 1,
        "ALLSPICE": 1,
        "BBQ SAUCE": 3,
        "BBQ Seasoning": 24,
        "BLACK PAPPER SAUCE": 3,
        "BOILED FISH SEASONING": 13,
        "BUTTER HOT POT BASE": 2,
        "Bbq Dip": 46,
        "Bbq Dip (Better)": 12,
        "Biluochun Tea": 2,
        "CLERA OIL HOT POT BASE": 2,
        "DELICACIES HOT POT BASE": 9,
        "DIPPING SAUCE": 3,
        "FISH WATER CHOOPED PEPPER": 1,
        "FIVE SPICE POWDER": 2,
        "GARLIC CRAYFISH SEASONING": 2,
        "GOLDEN BROTH SAUERKRANT FISH SEASONING": 3,
        "GRAGON": 7,
        "GREEN PEPPER DIPPED IN WATER": 108,
        "GREEN PEPPER OIL": 20,
        "HOT POT BEAN TENDON": 22,
        "HOT POT DIPPING SAUCE": 2,
        "HOT POT HAIBA DIPPING SAUCE": 11,
        "JINGHI HONGSHUFEN": 0,
        "LIANGFENG HOT POT STEW": 19,
        "MALA CHILI": 1,
        "MALA CHILI POWDER SPICY": 1,
        "MALA HOT POT BASE": 18,
        "MALA SAUCE (HOT POT BASE)": 0,
        "MALA XIANG GUO SEASONING": 53,
        "MUSHROOM HOT POT BASE": 6,
        "Mala Seasoning": 16,
        "OLD HOT POT BASE": 27,
        "Oolong Tea": 4,
        "PAPRIKA HOT": 3,
        "PEPPER OIL (2.5Lit)": 0,
        "SEASONING FOR PICKLED FISH IN BROTH": 47,
        "SESAME SOY SAUCE": 11,
        "SINGLE MOUNTAIN DIPPING IN WATER": 21,
        "SMALL SQUARE HOT POT BASE": 1,
        "SPECIAL ALKALINE SURFACE": 1,
        "SPICY DIP (SICHUAN DRY DISH) 10G": 4,
        "SPICY GRILLED FISH SEASONING": 9,
        "SPICY HOT POT BASE": 1,
        "SPICY HOT POT SAUCE, SPICY SHABU SAUCE": 6,
        "SPICY RED CHILI OIL(2.5 L)": 0,
        "SPICY RED CHILI OIL( 400 G)": 40,
        "STIR FRIED CHILI": 3,
        "STIR FRIED DOG SAUCE": 3,
        "SWEET POTATO NOODLES": 1,
        "Savoring Tea": 9,
        "Sichuan Mala Sauce": 3,
        "Sour Pickled Cabbage": 5,
        "Spicy Dipping Sauce": 9,
        "TENDER FISH SEASONING": 0,
        "THAI TEA MIX": 2,
        "TIEGUANYIN TIE GUAN": 1,
        "Ten Chinese (Tieganyin)": 4,
        "TIEGUANYIN Tea Gift": 5,
        "PORK BELLY SKIN": 2,
        "PORK RIBS": 0,
        "SEABASS FISH": 0,
        "SURMAI FISH": 0,
        "pickled peppers 5 kg": 1,
    }
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main store not found"}
    
    main_store_id = str(main_store["_id"])
    
    updated = 0
    created = 0
    not_found = []
    
    for item_name, quantity in mala_data.items():
        # Find the item (case-insensitive)
        item = items_collection.find_one({"name": {"$regex": f"^{item_name}$", "$options": "i"}})
        
        if not item:
            # Try partial match
            item = items_collection.find_one({"name": {"$regex": item_name, "$options": "i"}})
        
        if item:
            item_id = str(item["_id"])
            
            # Find existing lot for this item at Main Store
            lot = lots_collection.find_one({
                "item_id": item_id,
                "location_id": main_store_id
            })
            
            if lot:
                # Update existing lot
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"current_quantity": quantity, "quantity": quantity}}
                )
                updated += 1
            else:
                # Create new lot if quantity > 0
                if quantity > 0:
                    lots_collection.insert_one({
                        "lot_number": f"MALA-FEB20-{created}",
                        "item_id": item_id,
                        "item_name": item["name"],
                        "category": "Mala Grocery",
                        "unit": item.get("unit", ""),
                        "quantity": quantity,
                        "current_quantity": quantity,
                        "initial_quantity": quantity,
                        "location_id": main_store_id,
                        "source": "stock_correction_feb20",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    })
                    created += 1
        else:
            if item_name not in not_found:
                not_found.append(item_name)
    
    return {
        "success": True,
        "updated": updated,
        "created": created,
        "not_found": not_found
    }


# FIX LOTS LOCATION ID - Update lots with wrong location_id
@app.post("/api/setup/fix-lots-location")
async def fix_lots_location(secret_key: str = ""):
    """
    Fix lots that have wrong location_id by updating to current main store ID.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Get current main store
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main store not found"}
    
    correct_main_store_id = str(main_store["_id"])
    
    # Old/wrong main store IDs that might have been used
    wrong_ids = [
        "6969f51728d83d7fd2b2f5ca",  # From local/preview env
        "698c0f79788678217a287fce",  # Another possible old ID
    ]
    
    # Update all lots with wrong location_id
    total_updated = 0
    for wrong_id in wrong_ids:
        result = lots_collection.update_many(
            {"location_id": wrong_id},
            {"$set": {"location_id": correct_main_store_id}}
        )
        total_updated += result.modified_count
    
    # Also update lots that have null/empty location_id
    result2 = lots_collection.update_many(
        {"$or": [
            {"location_id": None},
            {"location_id": ""},
            {"location_id": {"$exists": False}}
        ]},
        {"$set": {"location_id": correct_main_store_id}}
    )
    total_updated += result2.modified_count
    
    return {
        "success": True,
        "correct_main_store_id": correct_main_store_id,
        "lots_updated": total_updated
    }


# FIX REQUISITIONS WITH UNKNOWN KITCHEN
@app.post("/api/setup/fix-requisitions-kitchen")
async def fix_requisitions_kitchen(secret_key: str = ""):
    """
    Fix requisitions that have missing or incorrect kitchen names.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Build location lookup
    locations = {str(loc["_id"]): loc for loc in locations_collection.find({})}
    locations_by_name = {loc["name"].lower(): loc for loc in locations_collection.find({})}
    
    # Build user lookup
    users = {str(u["_id"]): u for u in users_collection.find({})}
    
    # Find requisitions with missing kitchen info
    reqs = list(requisitions_collection.find({
        "$or": [
            {"kitchen_name": {"$exists": False}},
            {"kitchen_name": None},
            {"kitchen_name": ""},
            {"kitchen_name": "Unknown"},
            {"location_id": {"$exists": False}},
            {"location_id": None}
        ]
    }))
    
    fixed = 0
    
    for req in reqs:
        update_data = {}
        
        # Try to find kitchen from created_by user
        created_by = req.get("created_by") or req.get("requested_by")
        if created_by:
            user = users.get(created_by)
            if user and user.get("location_id"):
                location = locations.get(user["location_id"])
                if location:
                    update_data["location_id"] = user["location_id"]
                    update_data["kitchen_name"] = location.get("name")
                    update_data["kitchen_id"] = user["location_id"]
        
        # If still no location, try to match from existing kitchen_id
        if not update_data and req.get("kitchen_id"):
            location = locations.get(req["kitchen_id"])
            if location:
                update_data["location_id"] = req["kitchen_id"]
                update_data["kitchen_name"] = location.get("name")
        
        # Default to first kitchen if still nothing
        if not update_data:
            first_kitchen = next((loc for loc in locations.values() if loc.get("type") == "kitchen"), None)
            if first_kitchen:
                update_data["location_id"] = str(first_kitchen["_id"])
                update_data["kitchen_name"] = first_kitchen.get("name")
                update_data["kitchen_id"] = str(first_kitchen["_id"])
        
        if update_data:
            requisitions_collection.update_one(
                {"_id": req["_id"]},
                {"$set": update_data}
            )
            fixed += 1
    
    return {
        "success": True,
        "total_checked": len(reqs),
        "fixed": fixed
    }


# COMPREHENSIVE STOCK UPDATE FROM EXCEL FILES (Feb 20, 2026)
@app.post("/api/setup/update-all-stock-feb20")
async def update_all_stock_feb20(secret_key: str = "", as_grn: bool = False):
    """
    Update stock for all categories from Excel data.
    If as_grn=True, reset current stock to 0 and add as Today's GRN.
    """
    import json
    import os
    
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Load stock data from JSON file
    json_path = os.path.join(os.path.dirname(__file__), 'stock_data_feb20.json')
    try:
        with open(json_path, 'r') as f:
            all_stock_data = json.load(f)
    except Exception as e:
        return {"error": f"Failed to load stock data: {str(e)}"}
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main store not found"}
    
    main_store_id = str(main_store["_id"])
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    results = {}
    total_updated = 0
    total_created = 0
    total_not_found = []
    
    for category, items_data in all_stock_data.items():
        updated = 0
        created = 0
        not_found = []
        
        for item_name, quantity in items_data.items():
            # Find the item (case-insensitive exact match first)
            item = items_collection.find_one({"name": {"$regex": f"^{item_name}$", "$options": "i"}})
            
            if not item:
                # Try partial match
                item = items_collection.find_one({"name": {"$regex": item_name, "$options": "i"}})
            
            if item:
                item_id = str(item["_id"])
                
                if as_grn:
                    # Reset existing lots to 0
                    lots_collection.update_many(
                        {"item_id": item_id, "location_id": main_store_id},
                        {"$set": {"current_quantity": 0, "quantity": 0}}
                    )
                    
                    # Create new lot as Today's GRN if qty > 0
                    if quantity > 0:
                        lot_number = f"GRN-{today.replace('-', '')}-{created}"
                        lots_collection.insert_one({
                            "lot_number": lot_number,
                            "item_id": item_id,
                            "item_name": item["name"],
                            "category": item.get("category", category),
                            "unit": item.get("unit", ""),
                            "quantity": quantity,
                            "current_quantity": quantity,
                            "initial_quantity": quantity,
                            "location_id": main_store_id,
                            "source": "opening_stock_grn_feb20",
                            "grn_date": today,
                            "received_date": today,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
                        created += 1
                else:
                    # Just update the quantity
                    lot = lots_collection.find_one({
                        "item_id": item_id,
                        "location_id": main_store_id
                    })
                    
                    if lot:
                        lots_collection.update_one(
                            {"_id": lot["_id"]},
                            {"$set": {"current_quantity": quantity, "quantity": quantity}}
                        )
                        updated += 1
                    elif quantity > 0:
                        lots_collection.insert_one({
                            "lot_number": f"STOCK-FEB20-{created}",
                            "item_id": item_id,
                            "item_name": item["name"],
                            "category": item.get("category", category),
                            "unit": item.get("unit", ""),
                            "quantity": quantity,
                            "current_quantity": quantity,
                            "initial_quantity": quantity,
                            "location_id": main_store_id,
                            "source": "stock_correction_feb20",
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
                        created += 1
            else:
                not_found.append(item_name)
        
        results[category] = {
            "updated": updated,
            "created": created,
            "not_found_count": len(not_found)
        }
        total_updated += updated
        total_created += created
        total_not_found.extend(not_found)
    
    return {
        "success": True,
        "mode": "as_grn" if as_grn else "direct_update",
        "total_updated": total_updated,
        "total_created": total_created,
        "total_not_found": len(total_not_found),
        "by_category": results,
        "not_found_items": total_not_found[:50]  # First 50 not found items
    }


# Items
@app.post("/api/setup/fix-seafood-stock-grn")
async def fix_seafood_stock_grn(secret_key: str = ""):
    """
    Reset Sea Food current stock to 0 and add new quantities as Today's GRN.
    """
    MIGRATION_SECRET = "KINFOLK-MIGRATE-2026"
    
    if secret_key != MIGRATION_SECRET:
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Sea Food stock data from Excel (Feb 20, 2026)
    seafood_data = {
        "CRAB STICK": 5,
        "Frozen Duck, 2 Kg": 2,
        "PORK BACON": 1,
        "Prasuma - Chicken Value Sausage, 1 Kg (Frozen)": 7,
        "Prawns - 13/15 PDTO (Extra Jumbo), 30% Glaze": 39,
        "Prawns - 31/40 PDTO (Medium), 30% Glaze": 21,
        "Prawns - 8/12 PDTO (King size), 30% Glaze": 18,
        "SALMON FILLET 1 KGS": 7,
        "SALMON SMOKED 1 KG": 5,
        "SQUID TUBE": 0,
        "TUNA FISH": 0,
        "Vietnamese Basa Fillets (White), 50% Glaze": 8,
        "MSG CANNEN 1 KG 1*25": 36,
    }
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main store not found"}
    
    main_store_id = str(main_store["_id"])
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    reset_count = 0
    grn_created = 0
    not_found = []
    
    # First, reset ALL Sea Food items to 0
    seafood_items = list(items_collection.find({"category": {"$regex": "sea ?food", "$options": "i"}}))
    for item in seafood_items:
        item_id = str(item["_id"])
        # Set all lots for this item to 0
        lots_collection.update_many(
            {"item_id": item_id, "location_id": main_store_id},
            {"$set": {"current_quantity": 0, "quantity": 0}}
        )
        reset_count += 1
    
    # Now add the Excel quantities as new lots (Today's GRN)
    for item_name, quantity in seafood_data.items():
        if quantity <= 0:
            continue
            
        # Find the item (case-insensitive)
        item = items_collection.find_one({"name": {"$regex": f"^{item_name}$", "$options": "i"}})
        
        if not item:
            # Try partial match
            item = items_collection.find_one({"name": {"$regex": item_name.split()[0], "$options": "i"}})
        
        if item:
            item_id = str(item["_id"])
            
            # Create a new lot as Today's GRN
            lot_number = f"GRN-{today.replace('-', '')}-SF-{grn_created}"
            lots_collection.insert_one({
                "lot_number": lot_number,
                "item_id": item_id,
                "item_name": item["name"],
                "category": item.get("category", "Sea Food"),
                "unit": item.get("unit", ""),
                "quantity": quantity,
                "current_quantity": quantity,
                "initial_quantity": quantity,
                "location_id": main_store_id,
                "source": "opening_stock_grn",
                "grn_date": today,
                "received_date": today,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            grn_created += 1
        else:
            not_found.append(item_name)
    
    return {
        "success": True,
        "seafood_items_reset": reset_count,
        "grn_lots_created": grn_created,
        "not_found": not_found
    }


# Items
@app.get("/api/items")
async def get_items(category: Optional[str] = None, search: Optional[str] = None):
    query = {}
    if category:
        query["category"] = category
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    
    items = list(items_collection.find(query).sort("name", 1))
    return [{
        "id": str(item["_id"]),
        "name": item["name"],
        "category": item["category"],
        "unit": item["unit"],
        "hsn_code": item.get("hsn_code"),
        "gst_rate": item.get("gst_rate", 0),
        "vendor": item.get("vendor"),
        "standard_price": item.get("standard_price"),
        "par_stock": item.get("par_stock"),
        "created_at": item["created_at"]
    } for item in items]

@app.post("/api/items")
async def create_item(item: ItemCreate):
    doc = {
        "name": item.name,
        "category": item.category,
        "unit": item.unit,
        "hsn_code": item.hsn_code,
        "gst_rate": item.gst_rate,
        "vendor": item.vendor,
        "standard_price": item.standard_price,
        "par_stock": item.par_stock,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = items_collection.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Item created successfully"}

@app.put("/api/items/{item_id}")
async def update_item(item_id: str, item: ItemUpdate):
    update_data = {k: v for k, v in item.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = items_collection.update_one(
        {"_id": ObjectId(item_id)},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item updated successfully"}

@app.delete("/api/items/{item_id}")
async def delete_item(item_id: str):
    result = items_collection.delete_one({"_id": ObjectId(item_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted successfully"}

# ============ Current Stock API ============

@app.get("/api/stock/current")
async def get_current_stock(
    location_id: Optional[str] = None,
    category: Optional[str] = None,
    below_par_only: bool = False,
    include_perishables: bool = False
):
    """Get current stock levels for Main Store items only.
    
    Main Store manages: Groceries, Beverages, Packaging, Housekeeping, Seafood
    Daily Perishables (Vegetables, Dairy, Non-Veg) go directly to kitchens.
    """
    
    # Categories that are managed by Main Store
    MAIN_STORE_CATEGORIES = [
        'Indian Grocery', 'Chinese Grocery', 'Continental Grocery', 'Grocery',
        'Beverage', 'Beverages',
        'Packaging',
        'Housekeeping',
        'Seafood',
        'Seasoning',
        'Bakery',
        'MALA GROCERY'
    ]
    
    # Daily Perishables - go directly to kitchens, not main store stock
    PERISHABLE_CATEGORIES = [
        'Vegetables', 'Vegetable',
        'Dairy Product', 'Dairy',
        'Non Veg', 'Non-Veg', 'Non Veg Items',
        'Fruits'
    ]
    
    # Build location filter
    location_filter = {}
    if location_id:
        location_filter["location_id"] = location_id
    else:
        # Default to main store
        main_store = locations_collection.find_one({"type": "main_store"})
        if main_store:
            location_filter["location_id"] = str(main_store["_id"])
    
    # Get today's date range for GRN calculation
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    today_start_str = today_start.isoformat()
    today_end_str = today_end.isoformat()
    
    # Aggregate current stock by item
    pipeline = [
        {"$match": {**location_filter, "current_quantity": {"$gte": 0}}},
        {"$group": {
            "_id": "$item_id",
            "total_quantity": {"$sum": "$current_quantity"},
            "lot_count": {"$sum": 1}
        }}
    ]
    
    stock_by_item = {doc["_id"]: doc for doc in lots_collection.aggregate(pipeline)}
    
    # Get today's GRN (lots created today via GRN)
    todays_grn_pipeline = [
        {"$match": {
            **location_filter,
            "created_at": {"$gte": today_start_str, "$lt": today_end_str},
            "source": {"$in": ["grn", "invoice_grn", "opening_stock", None]}
        }},
        {"$group": {
            "_id": "$item_id",
            "grn_quantity": {"$sum": "$initial_quantity"}
        }}
    ]
    
    todays_grn_by_item = {doc["_id"]: doc["grn_quantity"] for doc in lots_collection.aggregate(todays_grn_pipeline)}
    
    # Get items - filter by category if specified
    item_query = {}
    if category:
        item_query["category"] = category
    elif not include_perishables:
        # Exclude perishable categories from main store stock by default
        item_query["category"] = {"$nin": PERISHABLE_CATEGORIES}
    
    items = list(items_collection.find(item_query).sort("name", 1))
    
    result = []
    for item in items:
        item_id = str(item["_id"])
        item_category = item.get("category", "")
        
        # Skip perishables unless specifically requested
        if not include_perishables and item_category in PERISHABLE_CATEGORIES:
            continue
            
        stock_info = stock_by_item.get(item_id, {"total_quantity": 0, "lot_count": 0})
        current_qty = stock_info["total_quantity"]
        todays_grn = todays_grn_by_item.get(item_id, 0)
        par_stock = item.get("par_stock") or 0
        
        is_below_par = par_stock > 0 and current_qty < par_stock
        
        # Skip if filtering below par only and not below par
        if below_par_only and not is_below_par:
            continue
        
        result.append({
            "item_id": item_id,
            "item_name": item["name"],
            "category": item_category,
            "unit": item.get("unit"),
            "current_stock": current_qty,
            "todays_grn": todays_grn,
            "total_stock": current_qty,  # Current stock already includes today's GRN
            "par_stock": par_stock,
            "lot_count": stock_info["lot_count"],
            "status": "below_par" if is_below_par else "ok",
            "shortage": max(0, par_stock - current_qty) if par_stock > 0 else 0
        })
    
    # Sort: below par items first, then by name
    result.sort(key=lambda x: (0 if x["status"] == "below_par" else 1, x["item_name"]))
    
    return result

@app.get("/api/stock/alerts")
async def get_stock_alerts():
    """Get items that are below par stock level"""
    return await get_current_stock(below_par_only=True)

@app.put("/api/items/{item_id}/par-stock")
async def update_par_stock(item_id: str, par_stock: float):
    """Update par stock for a specific item"""
    result = items_collection.update_one(
        {"_id": ObjectId(item_id)},
        {"$set": {"par_stock": par_stock}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Par stock updated successfully"}

@app.post("/api/items/bulk-par-stock")
async def bulk_update_par_stock(items: List[dict]):
    """Bulk update par stock for multiple items"""
    updated = 0
    for item in items:
        if "item_id" in item and "par_stock" in item:
            result = items_collection.update_one(
                {"_id": ObjectId(item["item_id"])},
                {"$set": {"par_stock": item["par_stock"]}}
            )
            if result.modified_count > 0:
                updated += 1
    return {"message": f"Updated par stock for {updated} items"}

# ============ Opening Stock Upload ============

@app.get("/api/stock/opening/template")
async def download_opening_stock_template():
    """Download Excel template for opening stock upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Opening Stock"
    
    # Define headers
    headers = ["Item Name*", "Quantity*", "Expiry Date (YYYY-MM-DD)", "Notes"]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add sample data
    sample_data = [
        ["COCA COLA 330 ML", 100, "2026-06-30", "Opening stock"],
        ["SPRITE 330 ML", 50, "2026-07-15", ""],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    
    # Instructions sheet
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        ["Opening Stock Upload Instructions"],
        [""],
        ["1. Item Name* (Required): Must match EXACTLY with an existing item in the system"],
        ["2. Quantity* (Required): The quantity of stock to add"],
        ["3. Expiry Date: Optional. Format: YYYY-MM-DD. If not provided, defaults to 1 year from now"],
        ["4. Notes: Optional. Any notes about this stock entry"],
        [""],
        ["Important Notes:"],
        ["- Items must already exist in the Items Master before uploading stock"],
        ["- Stock will be added to the Main Store location"],
        ["- If an item name doesn't match exactly, it will be skipped"],
        ["- Use the Items Master page to add new items first"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=row_data[0] if row_data else "")
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    
    ws2.column_dimensions['A'].width = 80
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=opening_stock_template.xlsx"}
    )

@app.delete("/api/stock/lots/clear-all")
async def clear_all_lots(
    current_user = Depends(require_role(["admin"]))
):
    """
    Clear all lots from the database. Admin only.
    Use this before uploading fresh opening stock.
    """
    try:
        count_before = lots_collection.count_documents({})
        result = lots_collection.delete_many({})
        return {
            "message": f"Successfully cleared all lots",
            "deleted_count": result.deleted_count,
            "lots_before": count_before,
            "lots_after": 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear lots: {str(e)}")

@app.post("/api/stock/opening/upload")
async def upload_opening_stock(
    file: UploadFile = File(...),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Upload opening stock from Excel file.
    Creates lots for items that match by name.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
    
    # Normalize column names
    df.columns = df.columns.str.strip().str.lower().str.replace('*', '', regex=False)
    
    # Check required columns
    required_cols = ['item name', 'quantity']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise HTTPException(
            status_code=400, 
            detail=f"Missing required columns: {', '.join(missing_cols)}. Required: Item Name, Quantity"
        )
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=500, detail="Main Store location not configured")
    
    main_store_id = str(main_store["_id"])
    
    # Build item name to ID mapping (case-insensitive)
    all_items = list(items_collection.find())
    item_map = {item["name"].lower().strip(): item for item in all_items}
    
    # Process each row
    created = 0
    skipped = 0
    errors = []
    created_items = []
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)
        
        item_name = str(row.get('item name', '')).strip()
        quantity = row.get('quantity')
        expiry_date = row.get('expiry date (yyyy-mm-dd)') or row.get('expiry date') or row.get('expiry')
        notes = row.get('notes', '')
        
        # Skip empty rows
        if not item_name or pd.isna(item_name):
            continue
        
        # Validate quantity
        try:
            quantity = float(quantity)
            if quantity <= 0:
                errors.append(f"Row {row_num}: Invalid quantity for '{item_name}'")
                skipped += 1
                continue
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: Invalid quantity for '{item_name}'")
            skipped += 1
            continue
        
        # Find matching item (case-insensitive)
        item_key = item_name.lower().strip()
        matched_item = item_map.get(item_key)
        
        if not matched_item:
            # Try partial match
            matched_item = None
            for key, item in item_map.items():
                if item_key in key or key in item_key:
                    matched_item = item
                    break
        
        if not matched_item:
            errors.append(f"Row {row_num}: Item '{item_name}' not found in Items Master")
            skipped += 1
            continue
        
        # Parse expiry date
        if expiry_date and not pd.isna(expiry_date):
            try:
                if isinstance(expiry_date, datetime):
                    expiry_str = expiry_date.strftime("%Y-%m-%d")
                else:
                    expiry_str = str(expiry_date)[:10]
            except:
                expiry_str = (datetime.now(timezone.utc) + timedelta(days=365)).strftime("%Y-%m-%d")
        else:
            expiry_str = (datetime.now(timezone.utc) + timedelta(days=365)).strftime("%Y-%m-%d")
        
        # Generate lot number
        lot_number = generate_lot_number()
        
        # Create lot
        lot_doc = {
            "lot_number": lot_number,
            "item_id": str(matched_item["_id"]),
            "item_name": matched_item["name"],  # Store item name for reference
            "category": matched_item.get("category", ""),
            "unit": matched_item.get("unit", ""),
            "initial_quantity": quantity,
            "current_quantity": quantity,
            "expiry_date": expiry_str,
            "location_id": main_store_id,
            "vendor_id": None,
            "purchase_rate": matched_item.get("standard_price", 0),
            "qr_code": "",
            "source": "opening_stock",
            "notes": notes or "Opening stock upload",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = lots_collection.insert_one(lot_doc)
        lot_id = str(result.inserted_id)
        
        # Generate QR code
        qr_code = generate_qr_code(lot_id, lot_number)
        lots_collection.update_one(
            {"_id": ObjectId(lot_id)},
            {"$set": {"qr_code": qr_code}}
        )
        
        # Create transaction record
        transactions_collection.insert_one({
            "type": "opening_stock",
            "lot_id": lot_id,
            "item_id": str(matched_item["_id"]),
            "quantity": quantity,
            "destination_location_id": main_store_id,
            "notes": notes or "Opening stock upload",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        created += 1
        created_items.append({
            "item_name": matched_item["name"],
            "quantity": quantity,
            "lot_number": lot_number
        })
    
    return {
        "message": "Opening stock upload complete",
        "created": created,
        "skipped": skipped,
        "errors": errors[:20],  # Limit errors shown
        "total_errors": len(errors),
        "created_items": created_items[:20]  # Limit items shown
    }

# ============ Stock Adjustment (February temporary feature) ============

class StockAdjustment(BaseModel):
    item_id: str
    new_quantity: float
    reason: str = "Physical stock count adjustment"

@app.post("/api/stock/adjust")
async def adjust_stock(
    adjustment: StockAdjustment,
    current_user = Depends(require_role(["admin"]))
):
    """
    Adjust stock to match physical count. ADMIN ONLY.
    This is a temporary feature for February 2026 to sync app stock with physical stock.
    Creates an adjustment transaction to correct the stock level.
    """
    # Validate item exists
    item = items_collection.find_one({"_id": ObjectId(adjustment.item_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=400, detail="Main store location not found")
    
    main_store_id = str(main_store["_id"])
    
    # Calculate current stock for this item
    pipeline = [
        {"$match": {
            "item_id": adjustment.item_id,
            "location_id": main_store_id,
            "current_quantity": {"$gt": 0}
        }},
        {"$group": {
            "_id": None,
            "total_quantity": {"$sum": "$current_quantity"}
        }}
    ]
    
    result = list(lots_collection.aggregate(pipeline))
    current_stock = result[0]["total_quantity"] if result else 0
    
    # Calculate adjustment needed
    adjustment_qty = adjustment.new_quantity - current_stock
    
    if adjustment_qty == 0:
        return {
            "message": "No adjustment needed - stock already matches",
            "item_name": item["name"],
            "current_stock": current_stock,
            "adjustment": 0
        }
    
    # Create adjustment lot if adding stock
    if adjustment_qty > 0:
        # Create a new lot for the positive adjustment
        lot_number = f"ADJ-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
        expiry_date = (datetime.now(timezone.utc) + timedelta(days=365)).isoformat()
        
        lot_doc = {
            "lot_number": lot_number,
            "item_id": adjustment.item_id,
            "item_name": item["name"],
            "quantity": adjustment_qty,
            "current_quantity": adjustment_qty,
            "initial_quantity": adjustment_qty,
            "expiry_date": expiry_date,
            "location_id": main_store_id,
            "source": "stock_adjustment",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        lots_collection.insert_one(lot_doc)
        
        # Record the adjustment transaction
        transactions_collection.insert_one({
            "type": "stock_adjustment_add",
            "item_id": adjustment.item_id,
            "item_name": item["name"],
            "quantity": adjustment_qty,
            "destination_location_id": main_store_id,
            "notes": f"{adjustment.reason} (Added {adjustment_qty} {item.get('unit', 'units')})",
            "created_by": current_user["id"],
            "date": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    else:
        # For negative adjustment, reduce from existing lots (FIFO)
        remaining_to_reduce = abs(adjustment_qty)
        lots = list(lots_collection.find({
            "item_id": adjustment.item_id,
            "location_id": main_store_id,
            "current_quantity": {"$gt": 0}
        }).sort("created_at", 1))  # FIFO - oldest first
        
        for lot in lots:
            if remaining_to_reduce <= 0:
                break
            
            lot_qty = lot.get("current_quantity", lot.get("quantity", 0))
            reduce_from_lot = min(lot_qty, remaining_to_reduce)
            
            new_lot_qty = lot_qty - reduce_from_lot
            lots_collection.update_one(
                {"_id": lot["_id"]},
                {"$set": {"current_quantity": new_lot_qty, "quantity": new_lot_qty}}
            )
            
            remaining_to_reduce -= reduce_from_lot
        
        # Record the adjustment transaction
        transactions_collection.insert_one({
            "type": "stock_adjustment_reduce",
            "item_id": adjustment.item_id,
            "item_name": item["name"],
            "quantity": abs(adjustment_qty),
            "source_location_id": main_store_id,
            "notes": f"{adjustment.reason} (Reduced {abs(adjustment_qty)} {item.get('unit', 'units')})",
            "created_by": current_user["id"],
            "date": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return {
        "message": "Stock adjusted successfully",
        "item_name": item["name"],
        "previous_stock": current_stock,
        "new_stock": adjustment.new_quantity,
        "adjustment": adjustment_qty,
        "adjustment_type": "added" if adjustment_qty > 0 else "reduced"
    }

@app.get("/api/stock/adjustment-history")
async def get_adjustment_history(
    current_user = Depends(require_role(["admin"]))
):
    """Get history of stock adjustments. ADMIN ONLY."""
    adjustments = list(transactions_collection.find({
        "type": {"$in": ["stock_adjustment_add", "stock_adjustment_reduce"]}
    }).sort("created_at", -1).limit(100))
    
    return [{
        "id": str(adj["_id"]),
        "item_name": adj.get("item_name", "Unknown"),
        "quantity": adj["quantity"],
        "type": "Added" if adj["type"] == "stock_adjustment_add" else "Reduced",
        "notes": adj.get("notes", ""),
        "date": adj.get("date") or adj.get("created_at")
    } for adj in adjustments]

# ============ Items Bulk Upload ============

@app.get("/api/items/template/download")
async def download_items_template():
    """Download Excel template for bulk item upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    
    # Define headers
    headers = ["Name*", "Category*", "Unit*", "HSN Code", "GST Rate (%)", "Vendor", "Standard Price"]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [40, 20, 10, 15, 15, 25, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Add sample data rows
    sample_data = [
        ["COCA COLA CAN 300 ML", "Beverage", "CASE", "", "18", "CP", "650"],
        ["TATA TEA AGNI 500 G", "Indian Grocery", "PKT", "", "5", "CP", "180"],
        ["MDH WHITE PEPPER 100 G", "Indian Grocery", "PKT", "", "5", "", "95"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        "BULK ITEM UPLOAD INSTRUCTIONS",
        "",
        "Required Fields (marked with *):",
        "- Name: Item name (must be unique)",
        "- Category: Item category (e.g., Beverage, Indian Grocery, Chinese Grocery, etc.)",
        "- Unit: Unit of measurement (e.g., KG, PKT, CASE, BTL, CAN, PCS, L)",
        "",
        "Optional Fields:",
        "- HSN Code: HSN code for GST",
        "- GST Rate: GST percentage (0, 5, 12, 18, 28)",
        "- Vendor: Default vendor name",
        "- Standard Price: Standard purchase price for variance detection",
        "",
        "Notes:",
        "- Delete the sample data rows before uploading your data",
        "- Duplicate item names will be skipped",
        "- Maximum 500 items per upload",
        "",
        "Categories Available:",
        "Beverage, Indian Grocery, Chinese Grocery, Continental Grocery,",
        "Housekeeping, Dairy, Seafood, Packaging"
    ]
    
    for row_idx, text in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif text.startswith("-"):
            cell.font = Font(color="666666")
    
    ws_info.column_dimensions['A'].width = 70
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items_upload_template.xlsx"}
    )

@app.post("/api/items/bulk-upload")
async def bulk_upload_items(
    file: UploadFile = File(...),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Upload Excel file to bulk create items with vendor association"""
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        
        # Clean column names
        df.columns = df.columns.str.strip().str.upper().str.replace('*', '', regex=False)
        
        # Map various column name formats to standard names
        column_map = {
            'NAME': 'name',
            'ITEM NAME': 'name',
            'ITEM NAME ': 'name',
            'CATEGORY': 'category',
            'CATEGO': 'category',
            'UNIT': 'unit',
            'UNITS': 'unit',
            'HSN CODE': 'hsn_code',
            'HSN': 'hsn_code',
            'GST RATE (%)': 'gst_rate',
            'GST RATE': 'gst_rate',
            'GST': 'gst_rate',
            'VENDOR': 'vendor',
            'VENDOR NAME': 'vendor',
            'STANDARD PRICE': 'standard_price',
            'RATE': 'standard_price',
            'RATE ': 'standard_price',
            'PRICE': 'standard_price',
            'GST INCL': 'standard_price',
            'GST INCL ': 'standard_price'
        }
        df = df.rename(columns=column_map)
        
        # Validate required columns - at minimum need name
        if 'name' not in df.columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required column: name/ITEM NAME. Found columns: {list(df.columns)}"
            )
        
        # Remove empty rows
        df = df.dropna(subset=['name'])
        df = df[df['name'].astype(str).str.strip() != '']
        df = df[df['name'].astype(str).str.lower() != 'nan']
        
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="No valid items found in the file")
        
        if len(df) > 500:
            raise HTTPException(status_code=400, detail="Maximum 500 items allowed per upload")
        
        # Get existing item names for duplicate check
        existing_items = {item['name'].lower() for item in items_collection.find({}, {'name': 1})}
        
        # Get vendors for vendor_id lookup
        vendors_map = {}
        for v in vendors_collection.find({}, {'name': 1}):
            vendors_map[v['name'].lower()] = str(v['_id'])
        
        created = 0
        updated = 0
        skipped = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                name = str(row['name']).strip()
                if not name or name.lower() == 'nan':
                    continue
                
                # Get category with default
                category = str(row.get('category', 'Uncategorized')).strip() if pd.notna(row.get('category')) else 'Uncategorized'
                if category.lower() == 'nan':
                    category = 'Uncategorized'
                
                # Get unit with default
                unit = str(row.get('unit', 'KG')).strip().upper() if pd.notna(row.get('unit')) else 'KG'
                if unit.lower() == 'nan' or unit == 'KGS':
                    unit = 'KG'
                
                # Parse optional fields
                gst_rate = 0
                if pd.notna(row.get('gst_rate')):
                    try:
                        gst_rate = float(row['gst_rate'])
                    except:
                        gst_rate = 0
                
                standard_price = None
                if pd.notna(row.get('standard_price')):
                    try:
                        price_val = row['standard_price']
                        if isinstance(price_val, str):
                            price_val = price_val.replace(',', '').replace('₹', '').strip()
                        standard_price = float(price_val) if price_val else None
                    except:
                        standard_price = None
                
                hsn_code = None
                if pd.notna(row.get('hsn_code')):
                    hsn_code = str(row['hsn_code']).strip()
                
                # Get vendor name and lookup vendor_id
                vendor = None
                vendor_id = None
                if pd.notna(row.get('vendor')):
                    vendor = str(row['vendor']).strip()
                    if vendor.lower() != 'nan' and vendor:
                        vendor_id = vendors_map.get(vendor.lower())
                        # Create vendor if not exists
                        if not vendor_id:
                            new_vendor = {
                                "name": vendor,
                                "contact": "",
                                "phone": "",
                                "email": "",
                                "payment_terms": "Net 7",
                                "created_at": datetime.now(timezone.utc).isoformat()
                            }
                            result = vendors_collection.insert_one(new_vendor)
                            vendor_id = str(result.inserted_id)
                            vendors_map[vendor.lower()] = vendor_id
                
                # Check if item exists
                if name.lower() in existing_items:
                    # Update existing item with vendor association if not set
                    existing = items_collection.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
                    if existing:
                        update_data = {}
                        if vendor_id and not existing.get('vendor_id'):
                            update_data['vendor_id'] = vendor_id
                            update_data['vendor'] = vendor
                        if standard_price and not existing.get('standard_price'):
                            update_data['standard_price'] = standard_price
                        if update_data:
                            items_collection.update_one({"_id": existing['_id']}, {"$set": update_data})
                            updated += 1
                        else:
                            skipped += 1
                    continue
                
                # Create new item
                item_doc = {
                    "name": name,
                    "category": category,
                    "unit": unit,
                    "hsn_code": hsn_code,
                    "gst_rate": gst_rate,
                    "vendor": vendor,
                    "vendor_id": vendor_id,
                    "standard_price": standard_price,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                items_collection.insert_one(item_doc)
                existing_items.add(name.lower())
                created += 1
                
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")
        
        return {
            "message": "Bulk upload completed",
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:10] if errors else [],  # Return first 10 errors
            "total_errors": len(errors)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

# Locations
@app.get("/api/locations")
async def get_locations(type: Optional[str] = None):
    query = {}
    if type:
        query["type"] = type
    
    locations = list(locations_collection.find(query).sort("name", 1))
    return [{
        "id": str(loc["_id"]),
        "name": loc["name"],
        "type": loc["type"],
        "code": loc.get("code", ""),
        "address": loc.get("address", ""),
        "contact_phone": loc.get("contact_phone", ""),
        "contact_person": loc.get("contact_person", ""),
        "created_at": loc["created_at"]
    } for loc in locations]

@app.get("/api/locations/{location_id}")
async def get_location(location_id: str):
    """Get a single location by ID"""
    location = locations_collection.find_one({"_id": ObjectId(location_id)})
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    return {
        "id": str(location["_id"]),
        "name": location["name"],
        "type": location["type"],
        "code": location.get("code", ""),
        "address": location.get("address", ""),
        "contact_phone": location.get("contact_phone", ""),
        "contact_person": location.get("contact_person", ""),
        "created_at": location["created_at"]
    }

@app.post("/api/locations")
async def create_location(location: LocationCreate):
    # Generate code for new location
    if location.type == "kitchen":
        # Get next kitchen number
        kitchen_count = locations_collection.count_documents({"type": "kitchen"})
        code = f"K{kitchen_count + 1:02d}"
    else:
        code = "MS01"
    
    doc = {
        "name": location.name,
        "type": location.type,
        "code": code,
        "address": getattr(location, 'address', '') or '',
        "contact_phone": getattr(location, 'contact_phone', '') or '',
        "contact_person": getattr(location, 'contact_person', '') or '',
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = locations_collection.insert_one(doc)
    return {"id": str(result.inserted_id), "code": code, "message": "Location created successfully"}

class LocationUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_person: Optional[str] = None

@app.put("/api/locations/{location_id}")
async def update_location(location_id: str, location: LocationUpdate, current_user = Depends(require_role(["admin"]))):
    """Update location details (Admin only)"""
    existing = locations_collection.find_one({"_id": ObjectId(location_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Location not found")
    
    update_data = {}
    if location.name:
        update_data["name"] = location.name
    if location.address is not None:
        update_data["address"] = location.address
    if location.contact_phone is not None:
        update_data["contact_phone"] = location.contact_phone
    if location.contact_person is not None:
        update_data["contact_person"] = location.contact_person
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        locations_collection.update_one(
            {"_id": ObjectId(location_id)},
            {"$set": update_data}
        )
    
    return {"message": "Location updated successfully"}

@app.delete("/api/locations/{location_id}")
async def delete_location(location_id: str, current_user = Depends(require_role(["admin"]))):
    """Delete a location (Admin only). Be careful - this won't delete associated data."""
    existing = locations_collection.find_one({"_id": ObjectId(location_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Location not found")
    
    # Check if location is in use by users
    user_count = users_collection.count_documents({"location_id": location_id})
    if user_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete location - {user_count} user(s) are assigned to it"
        )
    
    locations_collection.delete_one({"_id": ObjectId(location_id)})
    return {"message": "Location deleted successfully"}

# Vendors
@app.get("/api/vendors")
async def get_vendors():
    vendors = list(vendors_collection.find().sort("name", 1))
    return [{
        "id": str(v["_id"]),
        "name": v["name"],
        "contact": v.get("contact"),
        "email": v.get("email"),
        "phone": v.get("phone"),
        "address": v.get("address"),
        "gst_number": v.get("gst_number"),
        "payment_terms": v.get("payment_terms"),
        "supply_categories": v.get("supply_categories", []),
        "created_at": v.get("created_at")
    } for v in vendors]

@app.get("/api/vendors/{vendor_id}")
async def get_vendor(vendor_id: str):
    vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {
        "id": str(vendor["_id"]),
        "name": vendor["name"],
        "contact": vendor.get("contact"),
        "email": vendor.get("email"),
        "phone": vendor.get("phone"),
        "address": vendor.get("address"),
        "gst_number": vendor.get("gst_number"),
        "payment_terms": vendor.get("payment_terms"),
        "supply_categories": vendor.get("supply_categories", []),
        "created_at": vendor["created_at"]
    }

@app.post("/api/vendors")
async def create_vendor(vendor: VendorCreate):
    doc = {
        "name": vendor.name,
        "contact": vendor.contact,
        "email": vendor.email,
        "phone": vendor.phone,
        "address": vendor.address,
        "gst_number": vendor.gst_number,
        "payment_terms": vendor.payment_terms,
        "supply_categories": vendor.supply_categories,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = vendors_collection.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Vendor created successfully"}

@app.put("/api/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, vendor: VendorUpdate):
    update_data = {k: v for k, v in vendor.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = vendors_collection.update_one(
        {"_id": ObjectId(vendor_id)},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor updated successfully"}

@app.delete("/api/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str):
    result = vendors_collection.delete_one({"_id": ObjectId(vendor_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor deleted successfully"}

@app.get("/api/vendors/{vendor_id}/items")
async def get_vendor_items(vendor_id: str, search: Optional[str] = None):
    """Get items associated with this vendor (supports multiple vendor associations)"""
    vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    vendor_name = vendor.get("name", "")
    supply_categories = vendor.get("supply_categories", [])
    
    # Build query: items that have this vendor_id in vendor_ids array, OR single vendor_id match, OR vendor name match
    query = {"$or": [
        {"vendor_ids": vendor_id},  # New: check vendor_ids array
        {"vendor_id": vendor_id},   # Legacy: single vendor_id field
        {"vendor": {"$regex": f"^{vendor_name}$", "$options": "i"}},
        {"vendors": vendor_name}    # New: check vendors array
    ]}
    
    # Also include items from supply categories if specified
    if supply_categories:
        query["$or"].append({"category": {"$in": supply_categories}})
    
    # Add search filter if provided
    if search and len(search) >= 2:
        query = {
            "$and": [
                query,
                {"name": {"$regex": search, "$options": "i"}}
            ]
        }
    
    # Increase limit to get all relevant items
    items = list(items_collection.find(query).sort("name", 1).limit(1000))
    
    # If no items found with vendor association, return items from supply categories only
    if not items and supply_categories:
        cat_query = {"category": {"$in": supply_categories}}
        if search and len(search) >= 2:
            cat_query["name"] = {"$regex": search, "$options": "i"}
        items = list(items_collection.find(cat_query).sort("name", 1).limit(1000))
    
    # Final fallback: return limited items (not all)
    if not items:
        fallback_query = {}
        if search and len(search) >= 2:
            fallback_query["name"] = {"$regex": search, "$options": "i"}
        items = list(items_collection.find(fallback_query).sort("name", 1).limit(200))
    
    return [{
        "id": str(item["_id"]),
        "name": item["name"],
        "category": item.get("category"),
        "unit": item.get("unit"),
        "hsn_code": item.get("hsn_code"),
        "gst_rate": item.get("gst_rate", 0),
        "standard_price": item.get("standard_price")
    } for item in items]

# ============ Purchase Orders ============

def generate_po_number():
    """Generate a unique PO number using timestamp + random suffix to avoid DB queries"""
    import random
    now = datetime.now(timezone.utc)
    date_prefix = now.strftime('%Y%m%d')
    time_suffix = now.strftime('%H%M%S')
    random_suffix = random.randint(10, 99)
    return f"PO-{date_prefix}-{time_suffix}{random_suffix}"

def serialize_purchase_order(po: dict, include_images: bool = False) -> dict:
    """Serialize a purchase order. Set include_images=True for full detail view."""
    vendor = vendors_collection.find_one({"_id": ObjectId(po["vendor_id"])}) if po.get("vendor_id") else None
    
    # Get received quantities from lots linked to this PO
    po_id = str(po["_id"])
    po_number = po.get("po_number", "")
    
    # Find lots linked to this PO - check multiple possible field names
    lot_or_conditions = [
        {"po_id": po_id},  # Main field used in GRN
        {"po_reference_id": po_id},  # Alternative field name
    ]
    if po_number:
        lot_or_conditions.append({"po_number": po_number})
    
    received_by_item = {}
    lots = list(lots_collection.find({"$or": lot_or_conditions}))
    
    for lot in lots:
        item_id = lot.get("item_id")
        item_name = (lot.get("item_name") or "").strip().upper()
        qty = lot.get("initial_quantity", 0) or 0
        
        # Match by item_id
        if item_id:
            key = str(item_id)
            if key not in received_by_item:
                received_by_item[key] = 0
            received_by_item[key] += qty
        
        # Also track by name for fallback matching
        if item_name:
            if item_name not in received_by_item:
                received_by_item[item_name] = 0
            received_by_item[item_name] += qty
    
    # Also check kitchen_receivables for kitchen POs
    kr_or_conditions = [{"po_id": po_id}]
    if po_number:
        kr_or_conditions.append({"po_number": po_number})
    
    kitchen_receivables = db["kitchen_receivables"]
    kr_records = list(kitchen_receivables.find({"$or": kr_or_conditions}))
    
    for kr in kr_records:
        item_id = kr.get("item_id")
        item_name = (kr.get("item_name") or "").strip().upper()
        qty = kr.get("quantity", 0) or 0
        
        if item_id:
            key = str(item_id)
            if key not in received_by_item:
                received_by_item[key] = 0
            received_by_item[key] += qty
        
        if item_name:
            if item_name not in received_by_item:
                received_by_item[item_name] = 0
            received_by_item[item_name] += qty
    
    items_detail = []
    total_amount = 0
    for item in po["items"]:
        # Try to get item from items collection
        item_doc = None
        if item.get("item_id"):
            try:
                item_doc = items_collection.find_one({"_id": ObjectId(item["item_id"])})
            except:
                pass
        
        amount = item["quantity"] * item["rate"]
        total_amount += amount
        
        # Use stored item_name as fallback if item not found in collection
        stored_name = item.get("item_name") or item.get("name", "")
        stored_unit = item.get("unit", "")
        
        # Calculate received quantity - try multiple matching methods
        item_id = item.get("item_id")
        received_qty = 0
        
        # Try by item_id first
        if item_id:
            received_qty = received_by_item.get(str(item_id), 0)
        
        # If not found, try by name (case-insensitive)
        if received_qty == 0 and stored_name:
            received_qty = received_by_item.get(stored_name.strip().upper(), 0)
        
        items_detail.append({
            "item_id": item.get("item_id"),
            "item_name": item_doc["name"] if item_doc else (stored_name or "Unknown"),
            "unit": item_doc["unit"] if item_doc else (stored_unit or ""),
            "quantity": item["quantity"],
            "received_quantity": received_qty,
            "short_quantity": max(0, item["quantity"] - received_qty),
            "rate": item["rate"],
            "amount": amount,
            "notes": item.get("notes")
        })
    
    # Get creator's location info (for older POs without location data)
    created_by_location_name = po.get("created_by_location_name")
    created_by_location_code = po.get("created_by_location_code")
    
    if not created_by_location_name and po.get("created_by"):
        # Try to get from user's current location
        creator = users_collection.find_one({"_id": ObjectId(po["created_by"])})
        if creator and creator.get("location_id"):
            location = locations_collection.find_one({"_id": ObjectId(creator["location_id"])})
            if location:
                created_by_location_name = location.get("name")
                created_by_location_code = location.get("code")
        
        # If still no location, check delivery address for clues
        if not created_by_location_name:
            delivery_addr = po.get("delivery_address", "")
            if delivery_addr and delivery_addr != "Main Store":
                # Try to find matching kitchen
                for loc in locations_collection.find({"type": "kitchen"}):
                    if loc.get("name", "").lower() in delivery_addr.lower():
                        created_by_location_name = loc.get("name")
                        created_by_location_code = loc.get("code")
                        break
            
            # Default to Main Store if nothing found
            if not created_by_location_name:
                created_by_location_name = "Main Store"
                created_by_location_code = "MS01"
    
    # For list views, only return if verification exists (not the actual image data)
    grn_verification = None
    if include_images:
        grn_verification = po.get("grn_verification")
    elif po.get("grn_verification"):
        # Just indicate that verification exists, without the huge image data
        grn_verification = {"has_photo": True}
    
    return {
        "id": str(po["_id"]),
        "po_number": po["po_number"],
        "vendor_id": po["vendor_id"],
        "vendor_name": vendor["name"] if vendor else po.get("vendor_name", "Unknown"),
        "vendor_email": vendor.get("email") if vendor else None,
        "vendor_phone": vendor.get("phone") if vendor else None,
        "vendor_address": vendor.get("address") if vendor else None,
        "vendor_gst": vendor.get("gst_number") if vendor else None,
        "items": items_detail,
        "total_amount": total_amount,
        "delivery_date": po.get("delivery_date"),
        "delivery_address": po.get("delivery_address"),
        "payment_terms": po.get("payment_terms"),
        "notes": po.get("notes"),
        "status": po["status"],
        "received_items": po.get("received_items", []),
        "created_by": po.get("created_by"),
        "created_by_location_name": created_by_location_name,
        "created_by_location_code": created_by_location_code,
        "created_at": po["created_at"],
        "updated_at": po.get("updated_at"),
        # GRN details
        "grn_invoice_number": po.get("grn_invoice_number"),
        "grn_invoice_date": po.get("grn_invoice_date"),
        "grn_amount": po.get("grn_amount"),
        "grn_date": po.get("grn_date"),
        "grn_location_type": po.get("grn_location_type"),
        "grn_verification": grn_verification
    }

@app.post("/api/purchase-orders")
async def create_purchase_order(
    po: PurchaseOrderCreate,
    current_user = Depends(require_role(["admin", "main_store", "kitchen"]))
):
    """Create a new Purchase Order"""
    
    try:
        # Validate vendor
        if not po.vendor_id:
            raise HTTPException(status_code=400, detail="Vendor is required")
        
        vendor = vendors_collection.find_one({"_id": ObjectId(po.vendor_id)})
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found. Please select a valid vendor.")
        
        # Validate items
        if not po.items or len(po.items) == 0:
            raise HTTPException(status_code=400, detail="At least one item is required")
        
        items_data = []
        for item in po.items:
            if not item.item_id:
                raise HTTPException(status_code=400, detail="Item ID is required for all items")
            
            db_item = items_collection.find_one({"_id": ObjectId(item.item_id)})
            if not db_item:
                raise HTTPException(status_code=404, detail=f"Item not found: {item.item_id}")
            
            if item.quantity <= 0:
                raise HTTPException(status_code=400, detail=f"Quantity must be greater than 0 for {db_item.get('name', 'item')}")
            
            if item.rate <= 0:
                raise HTTPException(status_code=400, detail=f"Rate must be greater than 0 for {db_item.get('name', 'item')}")
            
            items_data.append({
                "item_id": item.item_id,
                "item_name": db_item.get("name"),  # Store item name for reference
                "category": db_item.get("category"),
                "unit": db_item.get("unit"),
                "quantity": item.quantity,
                "rate": item.rate,
                "notes": item.notes
            })
        
        po_number = generate_po_number()
        
        # Get creator's location info
        created_by_location_id = current_user.get("location_id")
        created_by_location_name = None
        created_by_location_code = None
        
        if created_by_location_id:
            location = locations_collection.find_one({"_id": ObjectId(created_by_location_id)})
            if location:
                created_by_location_name = location.get("name")
                created_by_location_code = location.get("code")
        else:
            # For admin/main_store without specific location
            created_by_location_name = "Main Store"
            created_by_location_code = "MS01"
        
        # Calculate total amount
        total_amount = sum(item["quantity"] * item["rate"] for item in items_data)
        
        po_doc = {
            "po_number": po_number,
            "vendor_id": po.vendor_id,
            "vendor_name": vendor.get("name"),  # Store vendor name for historical reference
            "items": items_data,
            "items_count": len(items_data),  # Store count for fast queries
            "total_amount": total_amount,  # Store total for fast queries
            "delivery_date": po.delivery_date,
            "delivery_address": po.delivery_address or "Main Store",
            "payment_terms": po.payment_terms or vendor.get("payment_terms", "As per agreement"),
            "notes": po.notes,
            "status": "pending",  # pending, partial, received, cancelled
            "received_items": [],
            "created_by": current_user["id"],
            "created_by_location_id": created_by_location_id,
            "created_by_location_name": created_by_location_name,
            "created_by_location_code": created_by_location_code,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = purchase_orders_collection.insert_one(po_doc)
        
        return {
            "message": "Purchase Order created successfully",
            "po_id": str(result.inserted_id),
            "po_number": po_number
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error creating PO: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating Purchase Order: {str(e)}")


# Debug endpoint to test PO loading without auth (temporary)
@app.get("/api/debug/test-po-load")
async def debug_test_po_load():
    """Test PO loading to diagnose the 500 error"""
    try:
        # Count total POs
        total = purchase_orders_collection.count_documents({})
        
        # Try to load just 5 POs
        pos = list(purchase_orders_collection.find({}).limit(5))
        
        results = []
        errors = []
        
        for po in pos:
            try:
                po_data = {
                    "id": str(po.get("_id", "")),
                    "po_number": po.get("po_number", ""),
                    "status": po.get("status", ""),
                    "vendor_id": str(po.get("vendor_id", "")) if po.get("vendor_id") else "",
                    "vendor_name": po.get("vendor_name", ""),
                    "items_count": len(po.get("items", [])),
                    "created_at_type": str(type(po.get("created_at"))),
                    "created_at": str(po.get("created_at", ""))[:50]
                }
                results.append(po_data)
            except Exception as e:
                errors.append({
                    "po_id": str(po.get("_id", "")),
                    "error": str(e)
                })
        
        return {
            "total_pos_in_db": total,
            "tested": len(pos),
            "successful": len(results),
            "errors": errors,
            "sample_pos": results
        }
    except Exception as e:
        import traceback
        return {
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@app.get("/api/purchase-orders/stats")
async def get_purchase_order_stats(
    current_user = Depends(require_role(["admin", "main_store", "kitchen"]))
):
    """Get PO counts by status - fast endpoint for dashboard summary"""
    try:
        user_role = current_user.get("role")
        user_location_id = current_user.get("location_id")
        
        base_query = {}
        if user_role == "kitchen" and user_location_id:
            base_query["$or"] = [
                {"created_by_location_id": user_location_id},
                {"location_id": user_location_id}
            ]
        
        # Use simple count_documents with maxTimeMS to avoid hanging
        # This is more reliable than aggregation on large collections
        try:
            pending = purchase_orders_collection.count_documents(
                {**base_query, "status": "pending"}, maxTimeMS=10000
            )
        except:
            pending = 0
            
        try:
            partial = purchase_orders_collection.count_documents(
                {**base_query, "status": "partial"}, maxTimeMS=10000
            )
        except:
            partial = 0
            
        try:
            received = purchase_orders_collection.count_documents(
                {**base_query, "status": "received"}, maxTimeMS=10000
            )
        except:
            received = 0
            
        try:
            cancelled = purchase_orders_collection.count_documents(
                {**base_query, "status": "cancelled"}, maxTimeMS=10000
            )
        except:
            cancelled = 0
        
        total = pending + partial + received + cancelled
        
        return {
            "pending": pending,
            "partial": partial,
            "received": received,
            "cancelled": cancelled,
            "total": total
        }
    except Exception as e:
        print(f"PO stats error: {e}")
        import traceback
        traceback.print_exc()
        return {"pending": 0, "partial": 0, "received": 0, "cancelled": 0, "total": 0}

@app.get("/api/purchase-orders")
async def get_purchase_orders(
    status: Optional[str] = None,
    vendor_id: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    include_count: bool = False,
    current_user = Depends(require_role(["admin", "main_store", "kitchen"]))
):
    try:
        query = {}
        if status:
            query["status"] = status
        if vendor_id:
            query["vendor_id"] = vendor_id
        
        # Filter POs based on user role
        user_role = current_user.get("role")
        user_location_id = current_user.get("location_id")
        
        if user_role == "kitchen" and user_location_id:
            query["$or"] = [
                {"created_by_location_id": user_location_id},
                {"location_id": user_location_id}
            ]
        
        # Use smaller limit and simpler query to avoid timeout
        skip = (page - 1) * limit
        safe_limit = min(limit, 100)  # Cap at 100
        
        # Only count if explicitly requested
        total_count = None
        if include_count:
            try:
                total_count = purchase_orders_collection.count_documents(query)
            except:
                total_count = 0
        
        # Simpler query - sort by _id (indexed) instead of created_at
        # Do NOT include items array - it's too heavy and causes timeouts
        try:
            pos = list(purchase_orders_collection.find(
                query,
                {"_id": 1, "po_number": 1, "vendor_id": 1, "vendor_name": 1,
                 "status": 1, "created_at": 1, "created_by": 1,
                 "created_by_location_name": 1, "delivery_address": 1, "notes": 1,
                 "total_amount": 1, "items_count": 1}  # NO items array
            ).sort("_id", -1).skip(skip).limit(safe_limit).max_time_ms(30000))
        except Exception as query_error:
            print(f"Query error: {query_error}")
            # Fallback: even simpler
            try:
                pos = list(purchase_orders_collection.find(
                    query,
                    {"_id": 1, "po_number": 1, "vendor_id": 1, "vendor_name": 1,
                     "status": 1, "total_amount": 1, "items_count": 1}
                ).sort("_id", -1).limit(min(safe_limit, 20)).max_time_ms(10000))
            except:
                pos = []
        
        # Serialize results
        result = []
        for po in pos:
            try:
                # Use stored total_amount (don't try to calculate from items)
                total_amount = float(po.get("total_amount") or 0)
                items_count = po.get("items_count") or 0
                
                created_at = po.get("created_at")
                created_at_str = ""
                if created_at:
                    try:
                        created_at_str = created_at.isoformat() if hasattr(created_at, "isoformat") else str(created_at)
                    except:
                        created_at_str = str(created_at)
                
                result.append({
                    "id": str(po["_id"]),
                    "po_number": po.get("po_number", ""),
                    "vendor_id": po.get("vendor_id", ""),
                    "vendor_name": po.get("vendor_name", "Unknown"),
                    "items": [],  # Empty in list - fetch via detail endpoint
                    "items_count": items_count,
                    "total_amount": total_amount,
                    "status": po.get("status", "pending"),
                    "delivery_address": po.get("delivery_address", ""),
                    "notes": po.get("notes", ""),
                    "created_by": str(po.get("created_by", "")) if po.get("created_by") else None,
                    "created_by_location_name": po.get("created_by_location_name", "Main Store"),
                    "created_at": created_at_str,
                    "has_grn_photos": False
                })
            except Exception as po_error:
                print(f"Error processing PO {po.get('_id')}: {po_error}")
                continue
        
        return {
            "purchase_orders": result,
            "page": page,
            "limit": safe_limit,
            "total": total_count
        }
    except Exception as e:
        import traceback
        print(f"PO endpoint error: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error loading purchase orders: {str(e)}")

@app.get("/api/purchase-orders/{po_id}")
async def get_purchase_order(
    po_id: str,
    current_user = Depends(require_role(["admin", "main_store", "kitchen"]))
):
    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    return serialize_purchase_order(po, include_images=True)

@app.get("/api/purchase-orders/{po_id}/pdf")
async def get_po_pdf(po_id: str):
    """Generate PDF for a Purchase Order"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from fastapi.responses import Response
    import base64
    
    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    po_data = serialize_purchase_order(po)
    
    # Get company settings
    company = company_settings_collection.find_one({}) or {
        "name": "Dreamoven",
        "address": "Main Store Address",
        "phone": "",
        "email": "",
        "gst_number": "07AGSPA1692G1ZB"
    }
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=20)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9)
    right_style = ParagraphStyle('Right', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT)
    
    elements = []
    
    # Title
    elements.append(Paragraph("PURCHASE ORDER", title_style))
    elements.append(Spacer(1, 10))
    
    # PO Details
    po_info = [
        [Paragraph(f"<b>PO Number:</b> {po_data['po_number']}", normal_style),
         Paragraph(f"<b>Date:</b> {po_data['created_at'][:10]}", right_style)],
        [Paragraph(f"<b>Delivery Date:</b> {po_data.get('delivery_date', 'ASAP')}", normal_style),
         Paragraph(f"<b>Status:</b> {po_data['status'].upper()}", right_style)]
    ]
    po_table = Table(po_info, colWidths=[3*inch, 3*inch])
    po_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(po_table)
    elements.append(Spacer(1, 15))
    
    # From/To Section
    from_to = [
        [Paragraph("<b>FROM:</b>", header_style), Paragraph("<b>TO (Vendor):</b>", header_style)],
        [Paragraph(f"{company.get('name', 'A Dream Oven')}<br/>{company.get('address', '')}<br/>Phone: {company.get('phone', '')}<br/>GST: {company.get('gst_number', '')}", normal_style),
         Paragraph(f"{po_data['vendor_name']}<br/>{po_data.get('vendor_address', '')}<br/>Phone: {po_data.get('vendor_phone', '')}<br/>GST: {po_data.get('vendor_gst', '')}", normal_style)]
    ]
    from_to_table = Table(from_to, colWidths=[3*inch, 3*inch])
    from_to_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(from_to_table)
    elements.append(Spacer(1, 20))
    
    # Items Table
    items_header = ['#', 'Item', 'Qty', 'Unit', 'Rate', 'Amount']
    items_data = [items_header]
    
    for idx, item in enumerate(po_data['items'], 1):
        items_data.append([
            str(idx),
            item['item_name'],
            str(item['quantity']),
            item['unit'],
            f"₹{item['rate']:.2f}",
            f"₹{item['amount']:.2f}"
        ])
    
    # Add total row
    items_data.append(['', '', '', '', Paragraph('<b>Total:</b>', right_style), f"₹{po_data['total_amount']:.2f}"])
    
    items_table = Table(items_data, colWidths=[0.4*inch, 2.5*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.9*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#d1d5db')),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('FONTNAME', (-1, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 20))
    
    # Payment Terms & Notes
    if po_data.get('payment_terms') or po_data.get('notes'):
        terms_data = []
        if po_data.get('payment_terms'):
            terms_data.append([Paragraph(f"<b>Payment Terms:</b> {po_data['payment_terms']}", normal_style)])
        if po_data.get('notes'):
            terms_data.append([Paragraph(f"<b>Notes:</b> {po_data['notes']}", normal_style)])
        
        terms_table = Table(terms_data, colWidths=[6*inch])
        terms_table.setStyle(TableStyle([
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(terms_table)
        elements.append(Spacer(1, 20))
    
    # Delivery Address
    elements.append(Paragraph(f"<b>Delivery Address:</b> {po_data.get('delivery_address', 'Main Store')}", normal_style))
    elements.append(Spacer(1, 20))
    
    # GRN Verification Section (if PO has been received)
    if po_data.get('status') in ['received', 'partial'] and po_data.get('grn_verification'):
        grn_v = po_data['grn_verification']
        
        # Add GRN Verification Header
        elements.append(Paragraph("<b>GRN VERIFICATION</b>", ParagraphStyle('GRN', parent=styles['Heading2'], fontSize=12, spaceAfter=10)))
        
        # GRN Details table
        grn_info = []
        if grn_v.get('capture_time'):
            grn_info.append([Paragraph(f"<b>Verified On:</b> {grn_v['capture_time'][:19].replace('T', ' ')}", normal_style)])
        if grn_v.get('gps_location'):
            loc = grn_v['gps_location']
            grn_info.append([Paragraph(f"<b>GPS Location:</b> {loc.get('latitude', 'N/A'):.6f}, {loc.get('longitude', 'N/A'):.6f}", normal_style)])
        if po_data.get('grn_invoice_number'):
            grn_info.append([Paragraph(f"<b>Invoice #:</b> {po_data['grn_invoice_number']}", normal_style)])
        if po_data.get('grn_amount'):
            grn_info.append([Paragraph(f"<b>GRN Amount:</b> ₹{po_data['grn_amount']:.2f}", normal_style)])
        
        if grn_info:
            grn_table = Table(grn_info, colWidths=[6*inch])
            grn_table.setStyle(TableStyle([
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(grn_table)
        
        # Add verification photo if available
        if grn_v.get('photo'):
            try:
                photo_data = grn_v['photo']
                # Handle data URI format: "data:image/jpeg;base64,..."
                if ',' in photo_data:
                    photo_data = photo_data.split(',')[1]
                
                image_bytes = base64.b64decode(photo_data)
                image_buffer = io.BytesIO(image_bytes)
                
                # Create image for PDF (max width 4 inches, preserving aspect ratio)
                img = Image(image_buffer, width=4*inch, height=3*inch)
                img.hAlign = 'CENTER'
                
                elements.append(Spacer(1, 10))
                elements.append(Paragraph("<b>Verification Photo:</b>", normal_style))
                elements.append(Spacer(1, 5))
                elements.append(img)
            except Exception as e:
                # If image fails to load, just skip it
                print(f"Failed to add verification photo to PDF: {e}")
        
        elements.append(Spacer(1, 20))
    
    # Signature Section
    sig_data = [
        [Paragraph("<b>Authorized Signature</b>", normal_style), '', Paragraph("<b>Vendor Acknowledgment</b>", normal_style)],
        ['_______________________', '', '_______________________'],
    ]
    sig_table = Table(sig_data, colWidths=[2.5*inch, 1*inch, 2.5*inch])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
    ]))
    elements.append(sig_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=PO-{po_data['po_number']}.pdf"}
    )

@app.post("/api/purchase-orders/{po_id}/receive")
async def receive_purchase_order(
    po_id: str,
    invoice_data: Optional[dict] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Receive items against a Purchase Order.
    Can be called from invoice scan or manually.
    """
    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    if po["status"] == "received":
        raise HTTPException(status_code=400, detail="Purchase Order already fully received")
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=500, detail="Main Store location not configured")
    
    location_id = str(main_store["_id"])
    created_lots = []
    variances = []
    
    for po_item in po["items"]:
        # Check if already received
        already_received = sum(
            r.get("quantity", 0) 
            for r in po.get("received_items", []) 
            if r.get("item_id") == po_item["item_id"]
        )
        
        remaining = po_item["quantity"] - already_received
        if remaining <= 0:
            continue
        
        # If invoice data provided, use it; otherwise receive full quantity
        received_qty = remaining
        received_rate = po_item["rate"]
        
        if invoice_data and invoice_data.get("items"):
            for inv_item in invoice_data["items"]:
                if inv_item.get("matched_item_id") == po_item["item_id"]:
                    received_qty = min(inv_item.get("quantity", remaining), remaining)
                    received_rate = inv_item.get("rate", po_item["rate"])
                    
                    # Check for price variance
                    if received_rate != po_item["rate"]:
                        variances.append({
                            "item_id": po_item["item_id"],
                            "po_rate": po_item["rate"],
                            "invoice_rate": received_rate,
                            "variance": received_rate - po_item["rate"]
                        })
                    break
        
        # Create lot
        lot_number = generate_lot_number()
        expiry_date = invoice_data.get("expiry_date") if invoice_data else (datetime.now(timezone.utc) + timedelta(days=90)).strftime("%Y-%m-%d")
        
        # Get item details for storing in lot
        db_item = items_collection.find_one({"_id": ObjectId(po_item["item_id"])})
        
        lot_doc = {
            "lot_number": lot_number,
            "item_id": po_item["item_id"],
            "item_name": db_item["name"] if db_item else po_item.get("item_name", ""),
            "category": db_item.get("category", "") if db_item else "",
            "unit": db_item.get("unit", "") if db_item else "",
            "initial_quantity": received_qty,
            "current_quantity": received_qty,
            "expiry_date": expiry_date,
            "location_id": location_id,
            "vendor_id": po["vendor_id"],
            "purchase_rate": received_rate,
            "qr_code": "",
            "source": "purchase_order",
            "po_id": po_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = lots_collection.insert_one(lot_doc)
        lot_id = str(result.inserted_id)
        
        qr_code = generate_qr_code(lot_id, lot_number)
        lots_collection.update_one(
            {"_id": ObjectId(lot_id)},
            {"$set": {"qr_code": qr_code}}
        )
        
        # Create transaction
        transactions_collection.insert_one({
            "type": "grn",
            "lot_id": lot_id,
            "item_id": po_item["item_id"],
            "quantity": received_qty,
            "rate": received_rate,
            "vendor_id": po["vendor_id"],
            "destination_location_id": location_id,
            "notes": f"Received against PO {po['po_number']}",
            "po_id": po_id,
            "po_number": po.get("po_number", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        created_lots.append({
            "lot_number": lot_number,
            "item_id": po_item["item_id"],
            "quantity": received_qty
        })
    
    # Update PO received_items and status
    received_items = po.get("received_items", [])
    for lot in created_lots:
        received_items.append({
            "item_id": lot["item_id"],
            "quantity": lot["quantity"],
            "lot_number": lot["lot_number"],
            "received_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Check if all items received
    all_received = True
    for po_item in po["items"]:
        total_received = sum(
            r.get("quantity", 0) 
            for r in received_items 
            if r.get("item_id") == po_item["item_id"]
        )
        if total_received < po_item["quantity"]:
            all_received = False
            break
    
    new_status = "received" if all_received else "partial"
    
    purchase_orders_collection.update_one(
        {"_id": ObjectId(po_id)},
        {"$set": {
            "received_items": received_items,
            "status": new_status,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": f"Received {len(created_lots)} items against PO",
        "lots_created": created_lots,
        "po_status": new_status,
        "price_variances": variances,
        "has_variances": len(variances) > 0
    }

@app.put("/api/purchase-orders/{po_id}/cancel")
async def cancel_purchase_order(
    po_id: str,
    current_user = Depends(require_role(["admin"]))
):
    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    if po["status"] == "received":
        raise HTTPException(status_code=400, detail="Cannot cancel fully received PO")
    
    purchase_orders_collection.update_one(
        {"_id": ObjectId(po_id)},
        {"$set": {
            "status": "cancelled",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Purchase Order cancelled"}


class POVendorUpdate(BaseModel):
    vendor_id: str
    vendor_name: str


@app.patch("/api/purchase-orders/{po_id}/vendor")
async def update_po_vendor(
    po_id: str,
    data: POVendorUpdate,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Update vendor for a Purchase Order"""
    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    # Verify vendor exists
    vendor = vendors_collection.find_one({"_id": ObjectId(data.vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    purchase_orders_collection.update_one(
        {"_id": ObjectId(po_id)},
        {"$set": {
            "vendor_id": data.vendor_id,
            "vendor_name": data.vendor_name,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Vendor updated to {data.vendor_name}"}


@app.delete("/api/purchase-orders/{po_id}")
async def delete_purchase_order(
    po_id: str,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Delete a Purchase Order (admin and main_store) - for removing test data"""
    # Try to find by ObjectId first, then by po_number
    po = None
    try:
        po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    except:
        pass
    
    if not po:
        # Try finding by po_number
        po = purchase_orders_collection.find_one({"po_number": po_id})
    
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    result = purchase_orders_collection.delete_one({"_id": po["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Failed to delete Purchase Order")
    
    return {"message": f"Purchase Order {po.get('po_number', po_id)} deleted successfully"}


@app.delete("/api/purchase-orders/bulk")
async def bulk_delete_purchase_orders(
    month: int = Query(..., description="Month (1-12)"),
    year: int = Query(..., description="Year (e.g., 2026)"),
    status: str = Query(None, description="Filter by status: pending, partial, received, cancelled"),
    current_user = Depends(require_role(["admin"]))
):
    """Bulk delete Purchase Orders by month/year (admin only) - for cleaning test data"""
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    query = {
        "created_at": {"$gte": start_date, "$lt": end_date}
    }
    if status:
        query["status"] = status
    
    count = purchase_orders_collection.count_documents(query)
    if count == 0:
        return {"message": "No matching Purchase Orders found", "deleted_count": 0}
    
    result = purchase_orders_collection.delete_many(query)
    return {"message": f"Deleted {result.deleted_count} Purchase Orders from {month:02d}/{year}", "deleted_count": result.deleted_count}


@app.get("/api/purchase-orders/pending/for-vendor/{vendor_id}")
async def get_pending_pos_for_vendor(
    vendor_id: str,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Get pending POs for a specific vendor - useful for invoice matching"""
    pos = list(purchase_orders_collection.find({
        "vendor_id": vendor_id,
        "status": {"$in": ["pending", "partial"]}
    }).sort("created_at", -1))
    
    return [serialize_purchase_order(po) for po in pos]


# ============ Cross-App Photo Fetching ============

async def get_kitchen_app_token():
    """Get authentication token from Kitchen App using admin credentials"""
    if not KITCHEN_APP_URL:
        return None
    
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            # First check if Kitchen App is online
            health_check = await client.get(f"{KITCHEN_APP_URL}/api/health")
            if health_check.status_code != 200:
                print(f"Kitchen App health check failed: {health_check.status_code}")
                return None
            
            # Use the same admin credentials to authenticate
            response = await client.post(
                f"{KITCHEN_APP_URL}/api/auth/login",
                json={
                    "email": "adreamoven@gmail.com",
                    "password": "Store@123"
                }
            )
            if response.status_code == 200:
                data = response.json()
                return data.get("token")
            else:
                print(f"Kitchen App login failed: {response.status_code} - {response.text}")
    except httpx.ConnectError as e:
        print(f"Kitchen App connection failed: {e}")
    except Exception as e:
        print(f"Failed to get Kitchen App token: {e}")
    return None


@app.get("/api/purchase-orders/{po_number}/fetch-photo")
async def fetch_photo_from_kitchen_app(
    po_number: str,
    current_user = Depends(require_auth)
):
    """
    Fetch GRN photo from the Kitchen App for a specific PO.
    This enables cross-app photo synchronization when photos are stored in a different environment.
    """
    if not KITCHEN_APP_URL:
        raise HTTPException(
            status_code=503, 
            detail="Kitchen App URL not configured. Contact administrator."
        )
    
    try:
        # Get auth token from Kitchen App
        kitchen_token = await get_kitchen_app_token()
        if not kitchen_token:
            raise HTTPException(
                status_code=503,
                detail="Kitchen App is offline or sleeping. Please visit the Kitchen App URL first to wake it up, then try again."
            )
        
        headers = {"Authorization": f"Bearer {kitchen_token}"}
        
        # Make API call to the Kitchen App to fetch the PO with photo
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Fetch all POs from Kitchen App
            response = await client.get(
                f"{KITCHEN_APP_URL}/api/purchase-orders",
                headers=headers
            )
            
            if response.status_code != 200:
                raise HTTPException(
                    status_code=502,
                    detail=f"Kitchen App returned error: {response.status_code}"
                )
            
            kitchen_pos = response.json()
            
            # Find the matching PO
            matching_po = None
            if isinstance(kitchen_pos, list):
                for po in kitchen_pos:
                    if po.get("po_number") == po_number:
                        matching_po = po
                        break
            elif isinstance(kitchen_pos, dict) and kitchen_pos.get("po_number") == po_number:
                matching_po = kitchen_pos
            
            if not matching_po:
                raise HTTPException(
                    status_code=404,
                    detail=f"PO {po_number} not found in Kitchen App"
                )
            
            # Extract GRN verification with photo
            grn_verification = matching_po.get("grn_verification", {})
            
            if not grn_verification:
                raise HTTPException(
                    status_code=404,
                    detail=f"No GRN verification found for PO {po_number} in Kitchen App"
                )
            
            photo = grn_verification.get("photo")
            
            if not photo:
                raise HTTPException(
                    status_code=404,
                    detail=f"Photo not available for PO {po_number} in Kitchen App"
                )
            
            # Return the photo and verification data
            return {
                "success": True,
                "po_number": po_number,
                "photo": photo,
                "gps_location": grn_verification.get("gps_location"),
                "capture_time": grn_verification.get("capture_time"),
                "source": "kitchen_app",
                "kitchen_app_url": KITCHEN_APP_URL
            }
            
    except HTTPException:
        raise
    except httpx.RequestError as e:
        raise HTTPException(
            status_code=503,
            detail=f"Could not connect to Kitchen App: {str(e)}. Please ensure the Kitchen App is online."
        )
    except httpx.TimeoutException:
        raise HTTPException(
            status_code=504,
            detail="Request to Kitchen App timed out. Please try again."
        )


@app.post("/api/purchase-orders/{po_id}/sync-photo")
async def sync_photo_from_kitchen_app(
    po_id: str,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Fetch photo from Kitchen App and save it to the local PO record.
    This permanently syncs the photo to the Main Store database.
    """
    # Get the local PO
    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    po_number = po.get("po_number")
    
    if not KITCHEN_APP_URL:
        raise HTTPException(
            status_code=503, 
            detail="Kitchen App URL not configured. Contact administrator."
        )
    
    try:
        # Get auth token from Kitchen App
        kitchen_token = await get_kitchen_app_token()
        if not kitchen_token:
            raise HTTPException(
                status_code=503,
                detail="Kitchen App is offline or sleeping. Please visit the Kitchen App URL first to wake it up, then try again."
            )
        
        headers = {"Authorization": f"Bearer {kitchen_token}"}
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                f"{KITCHEN_APP_URL}/api/purchase-orders",
                headers=headers
            )
            
            if response.status_code != 200:
                raise HTTPException(
                    status_code=502,
                    detail=f"Kitchen App returned error: {response.status_code}"
                )
            
            kitchen_pos = response.json()
            
            # Find the matching PO
            matching_po = None
            for kpo in kitchen_pos if isinstance(kitchen_pos, list) else [kitchen_pos]:
                if kpo.get("po_number") == po_number:
                    matching_po = kpo
                    break
            
            if not matching_po:
                raise HTTPException(
                    status_code=404,
                    detail=f"PO {po_number} not found in Kitchen App"
                )
            
            grn_verification = matching_po.get("grn_verification", {})
            
            if not grn_verification or not grn_verification.get("photo"):
                raise HTTPException(
                    status_code=404,
                    detail=f"No photo available for PO {po_number} in Kitchen App"
                )
            
            # Update local PO with the photo from Kitchen App
            existing_grn = po.get("grn_verification", {})
            updated_grn = {
                **existing_grn,
                "photo": grn_verification.get("photo"),
                "gps_location": grn_verification.get("gps_location") or existing_grn.get("gps_location"),
                "capture_time": grn_verification.get("capture_time") or existing_grn.get("capture_time"),
                "synced_from_kitchen_app": True,
                "synced_at": datetime.now(timezone.utc).isoformat()
            }
            
            purchase_orders_collection.update_one(
                {"_id": ObjectId(po_id)},
                {"$set": {"grn_verification": updated_grn}}
            )
            
            return {
                "success": True,
                "message": f"Photo synced successfully for PO {po_number}",
                "po_number": po_number
            }
            
    except HTTPException:
        raise
    except httpx.RequestError as e:
        raise HTTPException(
            status_code=503,
            detail=f"Could not connect to Kitchen App: {str(e)}"
        )


# ============ Company Settings ============

@app.get("/api/company-settings")
async def get_company_settings():
    settings = company_settings_collection.find_one({})
    if not settings:
        return {
            "name": "A Dream Oven",
            "address": "Main Store Address",
            "phone": "",
            "email": "",
            "gst_number": "",
            "logo_url": ""
        }
    return {
        "name": settings.get("name", "A Dream Oven"),
        "address": settings.get("address", ""),
        "phone": settings.get("phone", ""),
        "email": settings.get("email", ""),
        "gst_number": settings.get("gst_number", ""),
        "logo_url": settings.get("logo_url", "")
    }

@app.put("/api/company-settings")
async def update_company_settings(
    name: Optional[str] = None,
    address: Optional[str] = None,
    phone: Optional[str] = None,
    email: Optional[str] = None,
    gst_number: Optional[str] = None,
    current_user = Depends(require_role(["admin"]))
):
    update_data = {}
    if name: update_data["name"] = name
    if address: update_data["address"] = address
    if phone: update_data["phone"] = phone
    if email: update_data["email"] = email
    if gst_number: update_data["gst_number"] = gst_number
    
    company_settings_collection.update_one(
        {},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Company settings updated"}

# Categories
@app.get("/api/categories")
async def get_categories():
    categories = list(categories_collection.find().sort("name", 1))
    return [{"id": str(c["_id"]), "name": c["name"]} for c in categories]

@app.post("/api/categories")
async def create_category(name: str = Query(...)):
    doc = {"name": name, "created_at": datetime.now(timezone.utc).isoformat()}
    result = categories_collection.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Category created successfully"}

# ==================== FILE UPLOAD TO R2 ====================

@app.post("/api/upload/image")
async def upload_image_to_r2(
    file: UploadFile = File(...),
    folder: str = Query("grn-photos", description="Folder name in R2 bucket"),
    current_user = Depends(get_current_user)
):
    """Upload an image to Cloudflare R2 and return the URL"""
    if not r2_client:
        raise HTTPException(status_code=503, detail="R2 storage not configured")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Invalid file type. Allowed: {allowed_types}")
    
    # Generate unique filename
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    unique_filename = f"{folder}/{timestamp}_{os.urandom(4).hex()}.{file_ext}"
    
    try:
        # Read file content
        contents = await file.read()
        
        # Upload to R2
        r2_client.put_object(
            Bucket=R2_BUCKET_NAME,
            Key=unique_filename,
            Body=contents,
            ContentType=file.content_type
        )
        
        # Generate URL (use public URL if available, otherwise endpoint URL)
        if R2_PUBLIC_URL:
            file_url = f"{R2_PUBLIC_URL}/{unique_filename}"
        else:
            file_url = f"{R2_ENDPOINT_URL}/{R2_BUCKET_NAME}/{unique_filename}"
        
        return {
            "success": True,
            "url": file_url,
            "key": unique_filename,
            "filename": file.filename,
            "size": len(contents)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/api/upload/base64")
async def upload_base64_to_r2(
    image_data: str = Form(..., description="Base64 encoded image data"),
    folder: str = Form("grn-photos", description="Folder name in R2 bucket"),
    filename: str = Form(None, description="Optional filename"),
    current_user = Depends(get_current_user)
):
    """Upload a base64 encoded image to Cloudflare R2"""
    if not r2_client:
        raise HTTPException(status_code=503, detail="R2 storage not configured")
    
    try:
        # Remove data URL prefix if present
        if "," in image_data:
            image_data = image_data.split(",")[1]
        
        # Decode base64
        image_bytes = base64.b64decode(image_data)
        
        # Generate unique filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        ext = filename.split(".")[-1] if filename and "." in filename else "jpg"
        unique_filename = f"{folder}/{timestamp}_{os.urandom(4).hex()}.{ext}"
        
        # Determine content type
        content_type = "image/jpeg"
        if ext.lower() == "png":
            content_type = "image/png"
        elif ext.lower() == "webp":
            content_type = "image/webp"
        
        # Upload to R2
        r2_client.put_object(
            Bucket=R2_BUCKET_NAME,
            Key=unique_filename,
            Body=image_bytes,
            ContentType=content_type
        )
        
        # Generate URL (use public URL if available)
        if R2_PUBLIC_URL:
            file_url = f"{R2_PUBLIC_URL}/{unique_filename}"
        else:
            file_url = f"{R2_ENDPOINT_URL}/{R2_BUCKET_NAME}/{unique_filename}"
        
        return {
            "success": True,
            "url": file_url,
            "key": unique_filename,
            "size": len(image_bytes)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.delete("/api/upload/{key:path}")
async def delete_from_r2(
    key: str,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Delete a file from R2"""
    if not r2_client:
        raise HTTPException(status_code=503, detail="R2 storage not configured")
    
    try:
        r2_client.delete_object(Bucket=R2_BUCKET_NAME, Key=key)
        return {"success": True, "message": "File deleted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")

@app.get("/api/upload/status")
async def get_r2_status():
    """Check R2 storage status"""
    return {
        "configured": r2_client is not None,
        "bucket": R2_BUCKET_NAME if r2_client else None,
        "endpoint": R2_ENDPOINT_URL if r2_client else None
    }

# GRN
@app.post("/api/grn")
async def create_grn(grn: GRNCreate):
    item = items_collection.find_one({"_id": ObjectId(grn.item_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    location = locations_collection.find_one({"_id": ObjectId(grn.location_id)})
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    
    lot_number = generate_lot_number()
    
    lot_doc = {
        "lot_number": lot_number,
        "item_id": grn.item_id,
        "item_name": item["name"],  # Store item name for reference
        "category": item.get("category", ""),
        "unit": item.get("unit", ""),
        "initial_quantity": grn.quantity,
        "current_quantity": grn.quantity,
        "expiry_date": grn.expiry_date,
        "location_id": grn.location_id,
        "vendor_id": grn.vendor_id,
        "purchase_rate": grn.purchase_rate,
        "qr_code": "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = lots_collection.insert_one(lot_doc)
    lot_id = str(result.inserted_id)
    
    qr_code = generate_qr_code(lot_id, lot_number)
    lots_collection.update_one(
        {"_id": ObjectId(lot_id)},
        {"$set": {"qr_code": qr_code}}
    )
    
    txn_doc = {
        "type": "grn",
        "lot_id": lot_id,
        "quantity": grn.quantity,
        "destination_location_id": grn.location_id,
        "notes": grn.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    transactions_collection.insert_one(txn_doc)
    
    updated_lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
    
    return {
        "message": "GRN created successfully",
        "lot": serialize_lot(updated_lot)
    }

@app.get("/api/grn")
async def get_grn_list(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    location_id: Optional[str] = None
):
    query = {"type": "grn"}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if location_id:
        query["destination_location_id"] = location_id
    
    # Limit results for performance
    transactions = list(transactions_collection.find(query).sort("created_at", -1).limit(200))
    
    # OPTIMIZATION: Batch fetch all required data
    lot_ids = set()
    location_ids = set()
    for txn in transactions:
        if txn.get("lot_id"):
            lot_ids.add(txn["lot_id"])
        if txn.get("source_location_id"):
            location_ids.add(txn["source_location_id"])
        if txn.get("destination_location_id"):
            location_ids.add(txn["destination_location_id"])
    
    # Fetch all lots at once
    lots_map = {}
    item_ids = set()
    if lot_ids:
        try:
            lots = list(lots_collection.find({"_id": {"$in": [ObjectId(lid) for lid in lot_ids]}}))
            for lot in lots:
                lots_map[str(lot["_id"])] = lot
                if lot.get("item_id"):
                    item_ids.add(str(lot["item_id"]))
        except:
            pass
    
    # Fetch all items at once
    items_map = {}
    if item_ids:
        try:
            items = list(items_collection.find({"_id": {"$in": [ObjectId(iid) for iid in item_ids]}}))
            items_map = {str(item["_id"]): item for item in items}
        except:
            pass
    
    # Fetch all locations at once
    locations_map = {}
    if location_ids:
        try:
            locs = list(locations_collection.find({"_id": {"$in": [ObjectId(lid) for lid in location_ids]}}))
            locations_map = {str(loc["_id"]): loc.get("name") for loc in locs}
        except:
            pass
    
    # Serialize with pre-fetched data
    result = []
    for txn in transactions:
        lot = lots_map.get(txn.get("lot_id"))
        item = items_map.get(str(lot.get("item_id"))) if lot and lot.get("item_id") else None
        
        item_name = "Unknown"
        if item:
            item_name = item.get("name", "Unknown")
        elif lot and lot.get("item_name"):
            item_name = lot["item_name"]
        
        result.append({
            "id": str(txn["_id"]),
            "type": txn["type"],
            "lot_id": txn.get("lot_id", ""),
            "lot_number": lot["lot_number"] if lot else "Unknown",
            "item_name": item_name,
            "quantity": txn.get("quantity", 0),
            "source_location": locations_map.get(txn.get("source_location_id")),
            "destination_location": locations_map.get(txn.get("destination_location_id")),
            "notes": txn.get("notes"),
            "created_at": txn.get("created_at")
        })
    
    return result

# Bulk GRN from Excel/Invoice
@app.post("/api/grn/bulk")
async def create_bulk_grn(
    file: UploadFile = File(...),
    vendor_id: str = Form(...),
    location_id: str = Form(...),
    invoice_number: str = Form(None),
    invoice_date: str = Form(None)
):
    """
    Create multiple GRN entries from an Excel file.
    Expected columns: Item Name, Quantity, Rate (GST Incl), Expiry Date (optional)
    """
    import pandas as pd
    from io import BytesIO
    
    # Validate vendor and location
    vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    location = locations_collection.find_one({"_id": ObjectId(location_id)})
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    
    # Read Excel file
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        df.columns = df.columns.str.strip()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
    
    # Map columns - try different possible column names
    item_col = None
    qty_col = None
    rate_col = None
    expiry_col = None
    
    for col in df.columns:
        col_lower = col.lower()
        if 'item' in col_lower or 'article' in col_lower or 'name' in col_lower or 'description' in col_lower:
            item_col = col
        elif 'qty' in col_lower or 'quantity' in col_lower:
            qty_col = col
        elif 'rate' in col_lower or 'price' in col_lower or 'p.u' in col_lower:
            rate_col = col
        elif 'expiry' in col_lower or 'exp' in col_lower:
            expiry_col = col
    
    if not item_col:
        raise HTTPException(status_code=400, detail="Could not find Item Name column. Expected: Item Name, Article Name, or Description")
    if not qty_col:
        raise HTTPException(status_code=400, detail="Could not find Quantity column. Expected: Qty, Quantity")
    
    created_lots = []
    errors = []
    
    # Default expiry date (1 year from now)
    default_expiry = (datetime.now(timezone.utc) + timedelta(days=365)).strftime("%Y-%m-%d")
    
    for idx, row in df.iterrows():
        try:
            item_name = str(row[item_col]).strip()
            quantity = float(row[qty_col]) if pd.notna(row[qty_col]) else 0
            
            if not item_name or quantity <= 0:
                continue
            
            # Find matching item in database
            db_item = items_collection.find_one({
                "name": {"$regex": f"^{re.escape(item_name)}$", "$options": "i"}
            })
            
            if not db_item:
                # Try partial match
                db_item = items_collection.find_one({
                    "name": {"$regex": re.escape(item_name[:30]), "$options": "i"}
                })
            
            if not db_item:
                errors.append(f"Row {idx+2}: Item '{item_name}' not found in database")
                continue
            
            # Get rate
            purchase_rate = None
            if rate_col and pd.notna(row.get(rate_col)):
                try:
                    purchase_rate = float(row[rate_col])
                except:
                    pass
            
            if not purchase_rate:
                purchase_rate = db_item.get("standard_price", 0)
            
            # Get expiry date
            expiry_date = default_expiry
            if expiry_col and pd.notna(row.get(expiry_col)):
                try:
                    exp = row[expiry_col]
                    if isinstance(exp, datetime):
                        expiry_date = exp.strftime("%Y-%m-%d")
                    else:
                        expiry_date = str(exp)[:10]
                except:
                    pass
            
            # Create lot
            lot_number = generate_lot_number()
            lot_doc = {
                "lot_number": lot_number,
                "item_id": str(db_item["_id"]),
                "item_name": db_item["name"],  # Store item name for reference
                "category": db_item.get("category", ""),
                "unit": db_item.get("unit", ""),
                "initial_quantity": quantity,
                "current_quantity": quantity,
                "expiry_date": expiry_date,
                "location_id": location_id,
                "vendor_id": vendor_id,
                "purchase_rate": purchase_rate,
                "invoice_number": invoice_number,
                "qr_code": "",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            result = lots_collection.insert_one(lot_doc)
            lot_id = str(result.inserted_id)
            
            # Generate QR code
            qr_code = generate_qr_code(lot_id, lot_number)
            lots_collection.update_one(
                {"_id": ObjectId(lot_id)},
                {"$set": {"qr_code": qr_code}}
            )
            
            # Create transaction
            txn_doc = {
                "type": "grn",
                "lot_id": lot_id,
                "quantity": quantity,
                "destination_location_id": location_id,
                "notes": f"Bulk GRN - Invoice: {invoice_number or 'N/A'}",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            transactions_collection.insert_one(txn_doc)
            
            created_lots.append({
                "item_name": db_item["name"],
                "quantity": quantity,
                "rate": purchase_rate,
                "lot_number": lot_number
            })
            
        except Exception as e:
            errors.append(f"Row {idx+2}: Error - {str(e)}")
    
    return {
        "message": f"Bulk GRN completed. Created {len(created_lots)} lots.",
        "created_count": len(created_lots),
        "error_count": len(errors),
        "created_lots": created_lots,
        "errors": errors[:20]  # Return first 20 errors
    }

# Download GRN Excel Template
@app.get("/api/grn/template")
async def download_grn_template():
    """Download Excel template for bulk GRN upload"""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Create template with sample data
    data = {
        "Item Name": ["COCA COLA CAN 300 ML PK24", "TATA TEA AGNI 500 G", ""],
        "Quantity": [10, 5, ""],
        "Rate (GST Incl)": [653.26, 88.01, "(optional - auto-fills from master)"],
        "Expiry Date": ["2026-12-31", "2027-06-30", "(optional - YYYY-MM-DD)"]
    }
    
    df = pd.DataFrame(data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='GRN Items')
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=grn_template.xlsx"}
    )

# PO-based GRN - Receive against Purchase Order
class POGRNItem(BaseModel):
    item_id: str
    item_name: str
    ordered_qty: float
    received_qty: float
    po_rate: float
    invoice_rate: float
    final_amount: float
    status: str  # ok, short, rejected
    price_variance: float

class GPSLocation(BaseModel):
    latitude: float
    longitude: float
    accuracy: Optional[float] = None

class POGRNCreate(BaseModel):
    po_id: str
    vendor_id: str
    location_id: str
    invoice_number: str
    invoice_date: str
    total_amount: float
    items: List[POGRNItem]
    # Photo verification fields (supports multiple photos)
    verification_photo: Optional[str] = None  # First photo for backwards compatibility
    verification_photos: Optional[List[dict]] = None  # All photos [{data: base64, timestamp: str}]
    gps_location: Optional[GPSLocation] = None
    capture_time: Optional[str] = None

@app.post("/api/grn/from-po")
async def create_grn_from_po(data: POGRNCreate):
    """
    Create GRN entries from a Purchase Order.
    - For Main Store: Items get added to stock (lots)
    - For Kitchen: Items get added to kitchen_receivables (daily perishables), NOT main stock
    - REQUIRES: Photo verification with GPS location, date & timestamp (up to 6 photos)
    """
    
    # Check if any photos are provided
    has_photos = bool(data.verification_photo or (data.verification_photos and len(data.verification_photos) > 0))
    
    # Photo verification is NOW REQUIRED for all outlets
    from datetime import date
    PHOTO_OPTIONAL_UNTIL = date(2026, 2, 12)
    photo_required = date.today() > PHOTO_OPTIONAL_UNTIL
    
    if photo_required and not has_photos:
        raise HTTPException(
            status_code=400, 
            detail="Photo verification is required to complete GRN. Please capture a photo with GPS location and timestamp."
        )
    
    # GPS and timestamp are also required
    if photo_required:
        if not data.gps_location:
            raise HTTPException(
                status_code=400, 
                detail="GPS location is required for GRN verification. Please enable location services and try again."
            )
        
        if not data.capture_time:
            raise HTTPException(
                status_code=400, 
                detail="Capture timestamp is required for GRN verification."
            )
    
    # Validate PO
    po = purchase_orders_collection.find_one({"_id": ObjectId(data.po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    # Validate vendor
    vendor = vendors_collection.find_one({"_id": ObjectId(data.vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Validate location and check if it's a kitchen
    location = locations_collection.find_one({"_id": ObjectId(data.location_id)})
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    
    is_kitchen = location.get("type") == "kitchen"
    
    created_items = []
    errors = []
    
    # Default expiry date (1 year for main store, 1 day for kitchen perishables)
    if is_kitchen:
        default_expiry = datetime.now(timezone.utc).strftime("%Y-%m-%d")  # Today for perishables
    else:
        default_expiry = (datetime.now(timezone.utc) + timedelta(days=365)).strftime("%Y-%m-%d")
    
    # Kitchen receivables collection
    kitchen_receivables = db["kitchen_receivables"]
    
    for item in data.items:
        # Skip rejected items
        if item.status == 'rejected' or item.received_qty <= 0:
            continue
        
        try:
            # Get item from database
            db_item = items_collection.find_one({"_id": ObjectId(item.item_id)})
            if not db_item:
                # Try to find by name
                db_item = items_collection.find_one({"name": {"$regex": f"^{re.escape(item.item_name)}$", "$options": "i"}})
            
            if not db_item:
                errors.append(f"Item '{item.item_name}' not found in database")
                continue
            
            if is_kitchen:
                # Kitchen GRN - Add to kitchen_receivables (Daily Perishables)
                # NOT added to main stock (lots)
                receivable_doc = {
                    "kitchen_id": data.location_id,
                    "kitchen_name": location.get("name", ""),
                    "item_id": str(db_item["_id"]),
                    "item_name": db_item["name"],
                    "category": db_item.get("category", "Uncategorized"),
                    "quantity": item.received_qty,
                    "unit": db_item.get("unit", ""),
                    "rate": item.invoice_rate,
                    "amount": item.final_amount,
                    "vendor_id": data.vendor_id,
                    "vendor_name": vendor.get("name", ""),
                    "po_id": data.po_id,
                    "po_number": po.get("po_number", ""),
                    "invoice_number": data.invoice_number,
                    "invoice_date": data.invoice_date,
                    "receive_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "status": item.status,
                    "type": "daily_perishable",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                kitchen_receivables.insert_one(receivable_doc)
                
                # Update item's standard_price if there's a price variance
                if item.invoice_rate and item.invoice_rate > 0:
                    current_price = db_item.get("standard_price", 0) or 0
                    if item.invoice_rate != current_price:
                        items_collection.update_one(
                            {"_id": db_item["_id"]},
                            {
                                "$set": {
                                    "standard_price": item.invoice_rate,
                                    "price_updated_at": datetime.now(timezone.utc).isoformat(),
                                    "price_updated_from": "grn_kitchen",
                                    "previous_price": current_price
                                }
                            }
                        )
                
                created_items.append({
                    "item_name": db_item["name"],
                    "ordered_qty": item.ordered_qty,
                    "received_qty": item.received_qty,
                    "rate": item.invoice_rate,
                    "amount": item.final_amount,
                    "status": item.status,
                    "type": "kitchen_perishable",
                    "price_updated": item.invoice_rate != (db_item.get("standard_price", 0) or 0)
                })
            else:
                # Main Store GRN - Add to lots (stock)
                lot_number = generate_lot_number()
                lot_doc = {
                    "lot_number": lot_number,
                    "item_id": str(db_item["_id"]),
                    "item_name": db_item["name"],  # Store item name for reference
                    "category": db_item.get("category", ""),
                    "unit": db_item.get("unit", ""),
                    "initial_quantity": item.received_qty,
                    "current_quantity": item.received_qty,
                    "expiry_date": default_expiry,
                    "location_id": data.location_id,
                    "vendor_id": data.vendor_id,
                    "purchase_rate": item.invoice_rate,
                    "invoice_number": data.invoice_number,
                    "po_id": data.po_id,
                    "po_rate": item.po_rate,
                    "price_variance": item.price_variance,
                    "qr_code": "",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                result = lots_collection.insert_one(lot_doc)
                lot_id = str(result.inserted_id)
                
                # Generate QR code
                qr_code = generate_qr_code(lot_id, lot_number)
                lots_collection.update_one(
                    {"_id": ObjectId(lot_id)},
                    {"$set": {"qr_code": qr_code}}
                )
                
                # Create GRN transaction
                txn_doc = {
                    "type": "grn",
                    "lot_id": lot_id,
                    "quantity": item.received_qty,
                    "destination_location_id": data.location_id,
                    "po_id": data.po_id,
                    "invoice_number": data.invoice_number,
                    "notes": f"GRN from PO {po.get('po_number', '')} - Status: {item.status.upper()}",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                transactions_collection.insert_one(txn_doc)
                
                # Update item's standard_price if there's a price variance
                # This ensures item prices stay current with latest purchase rates
                if item.invoice_rate and item.invoice_rate > 0:
                    current_price = db_item.get("standard_price", 0) or 0
                    if item.invoice_rate != current_price:
                        # Update the item's standard_price to the new GRN rate
                        items_collection.update_one(
                            {"_id": db_item["_id"]},
                            {
                                "$set": {
                                    "standard_price": item.invoice_rate,
                                    "price_updated_at": datetime.now(timezone.utc).isoformat(),
                                    "price_updated_from": "grn",
                                    "previous_price": current_price
                                }
                            }
                        )
                
                created_items.append({
                    "item_name": db_item["name"],
                    "ordered_qty": item.ordered_qty,
                    "received_qty": item.received_qty,
                    "rate": item.invoice_rate,
                    "amount": item.final_amount,
                    "status": item.status,
                    "lot_number": lot_number,
                    "price_updated": item.invoice_rate != (db_item.get("standard_price", 0) or 0)
                })
            
        except Exception as e:
            errors.append(f"Error processing {item.item_name}: {str(e)}")
    
    # Update PO status
    all_received = all(
        item.received_qty >= item.ordered_qty 
        for item in data.items 
        if item.status != 'rejected'
    )
    any_received = any(item.received_qty > 0 for item in data.items)
    
    new_po_status = 'pending'
    if all_received:
        new_po_status = 'received'
    elif any_received:
        new_po_status = 'partial'
    
    # Store verification photos (supports up to 6 photos)
    verification_data = None
    photos_to_store = []
    
    # Handle multiple photos array
    if data.verification_photos and len(data.verification_photos) > 0:
        photos_to_store = data.verification_photos[:6]  # Max 6 photos
    elif data.verification_photo:
        # Backwards compatibility - single photo
        photos_to_store = [{"data": data.verification_photo, "url": data.verification_photo, "timestamp": data.capture_time}]
    
    if photos_to_store:
        # Extract first photo URL - handle both 'data' (legacy base64) and 'url' (R2 URL) formats
        first_photo = photos_to_store[0]
        first_photo_url = first_photo.get("url") or first_photo.get("data") or None
        
        verification_data = {
            "photo": first_photo_url,  # First photo URL for backwards compat
            "photos": photos_to_store,  # All photos
            "photo_count": len(photos_to_store),
            "gps_location": {
                "latitude": data.gps_location.latitude,
                "longitude": data.gps_location.longitude,
                "accuracy": data.gps_location.accuracy
            } if data.gps_location else None,
            "capture_time": data.capture_time,
            "verified_at": datetime.now(timezone.utc).isoformat()
        }
    
    # Update PO status with timeout handling
    try:
        purchase_orders_collection.update_one(
            {"_id": ObjectId(data.po_id)},
            {"$set": {
                "status": new_po_status,
                "grn_invoice_number": data.invoice_number,
                "grn_invoice_date": data.invoice_date,
                "grn_amount": data.total_amount,
                "grn_date": datetime.now(timezone.utc).isoformat(),
                "grn_location_type": "kitchen" if is_kitchen else "main_store",
                "grn_kitchen_id": data.location_id if is_kitchen else None,
                "grn_verification": verification_data
            }}
        )
    except Exception as po_update_error:
        print(f"Warning: PO status update failed: {po_update_error}")
        # Continue anyway - the stock/receivables have been created
    
    # Create vendor ledger entry with verification info
    ledger_entry = {
        "vendor_id": data.vendor_id,
        "type": "kitchen_grn" if is_kitchen else "grn",
        "po_id": data.po_id,
        "po_number": po.get("po_number", ""),
        "invoice_number": data.invoice_number,
        "invoice_date": data.invoice_date,
        "amount": data.total_amount,
        "items_count": len(created_items),
        "location_id": data.location_id,
        "location_name": location.get("name", ""),
        "location_type": location.get("type", ""),
        "notes": f"{'Kitchen ' if is_kitchen else ''}GRN against PO {po.get('po_number', '')}",
        "has_verification_photo": len(photos_to_store) > 0,
        "photo_count": len(photos_to_store),
        "gps_coordinates": f"{data.gps_location.latitude},{data.gps_location.longitude}" if data.gps_location else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Insert into vendor_ledger collection (create if doesn't exist)
    try:
        vendor_ledger_collection = db["vendor_ledger"]
        ledger_result = vendor_ledger_collection.insert_one(ledger_entry)
    except Exception as ledger_error:
        print(f"Warning: Vendor ledger entry failed: {ledger_error}")
        # Continue anyway
    
    msg = f"GRN completed successfully. {len(created_items)} items "
    if is_kitchen:
        msg += "recorded as daily perishables (not added to main store stock)."
    else:
        msg += "added to stock."
    
    # Count how many item prices were updated
    prices_updated = sum(1 for item in created_items if item.get("price_updated", False))
    if prices_updated > 0:
        msg += f" {prices_updated} item price(s) updated to current GRN rate."
    
    return {
        "message": msg,
        "po_status": new_po_status,
        "is_kitchen_grn": is_kitchen,
        "created_lots" if not is_kitchen else "received_items": created_items,
        "prices_updated_count": prices_updated,
        "errors": errors,
        "ledger_entry": {
            "id": str(ledger_result.inserted_id),
            "amount": data.total_amount,
            "invoice_number": data.invoice_number
        }
    }

# Lots
@app.get("/api/lots")
async def get_lots(
    location_id: Optional[str] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None
):
    query = {}
    
    if location_id:
        query["location_id"] = location_id
    
    # Add limit to prevent timeout on large datasets
    lots = list(lots_collection.find(query).sort("created_at", -1).limit(1000))
    result = []
    
    for lot in lots:
        serialized = serialize_lot(lot)
        
        if status and serialized["status"] != status:
            continue
        
        if category and serialized["category"] != category:
            continue
        
        if search:
            search_lower = search.lower()
            if search_lower not in serialized["item_name"].lower() and search_lower not in serialized["lot_number"].lower():
                continue
        
        result.append(serialized)
    
    return result

@app.get("/api/lots/{lot_id}")
async def get_lot(lot_id: str):
    lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    return serialize_lot(lot)

@app.get("/api/lots/scan/{qr_data}")
async def get_lot_by_qr(qr_data: str):
    parts = qr_data.split(":")
    if len(parts) != 3 or parts[0] != "LOT":
        raise HTTPException(status_code=400, detail="Invalid QR code format")
    
    lot_id = parts[1]
    lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    return serialize_lot(lot)

# Issue
@app.post("/api/issue")
async def create_issue(issue: IssueCreate):
    lot = lots_collection.find_one({"_id": ObjectId(issue.lot_id)})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["current_quantity"] < issue.quantity:
        raise HTTPException(
            status_code=400, 
            detail=f"Insufficient quantity. Available: {lot['current_quantity']}"
        )
    
    destination = locations_collection.find_one({"_id": ObjectId(issue.destination_id)})
    if not destination:
        raise HTTPException(status_code=404, detail="Destination location not found")
    
    new_quantity = lot["current_quantity"] - issue.quantity
    lots_collection.update_one(
        {"_id": ObjectId(issue.lot_id)},
        {"$set": {"current_quantity": new_quantity}}
    )
    
    txn_doc = {
        "type": "issue",
        "lot_id": issue.lot_id,
        "quantity": issue.quantity,
        "source_location_id": lot["location_id"],
        "destination_location_id": issue.destination_id,
        "notes": issue.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    transactions_collection.insert_one(txn_doc)
    
    return {
        "message": "Issue recorded successfully",
        "remaining_quantity": new_quantity
    }

@app.get("/api/issue")
async def get_issue_list(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    destination_id: Optional[str] = None
):
    query = {"type": "issue"}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if destination_id:
        query["destination_location_id"] = destination_id
    
    transactions = list(transactions_collection.find(query).sort("created_at", -1))
    return [serialize_transaction(txn) for txn in transactions]

# FEFO Issue
@app.post("/api/issue/fefo")
async def create_fefo_issue(
    item_id: str = Query(...),
    quantity: float = Query(...),
    destination_id: str = Query(...),
    source_location_id: Optional[str] = None,
    notes: Optional[str] = None
):
    query = {
        "item_id": item_id,
        "current_quantity": {"$gt": 0}
    }
    if source_location_id:
        query["location_id"] = source_location_id
    
    lots = list(lots_collection.find(query).sort("expiry_date", 1))
    
    now = datetime.now(timezone.utc)
    active_lots = []
    for lot in lots:
        try:
            expiry = parse_expiry_date(lot.get("expiry_date"))
            if expiry >= now:
                active_lots.append(lot)
        except Exception:
            active_lots.append(lot)
    
    if not active_lots:
        raise HTTPException(status_code=400, detail="No active lots available for this item")
    
    total_available = sum(lot["current_quantity"] for lot in active_lots)
    if total_available < quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient quantity. Available: {total_available}"
        )
    
    remaining_to_issue = quantity
    issued_from = []
    
    for lot in active_lots:
        if remaining_to_issue <= 0:
            break
        
        issue_qty = min(lot["current_quantity"], remaining_to_issue)
        new_quantity = lot["current_quantity"] - issue_qty
        
        lots_collection.update_one(
            {"_id": lot["_id"]},
            {"$set": {"current_quantity": new_quantity}}
        )
        
        txn_doc = {
            "type": "issue",
            "lot_id": str(lot["_id"]),
            "quantity": issue_qty,
            "source_location_id": lot["location_id"],
            "destination_location_id": destination_id,
            "notes": f"FEFO Issue. {notes or ''}",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        transactions_collection.insert_one(txn_doc)
        
        issued_from.append({
            "lot_number": lot["lot_number"],
            "quantity": issue_qty,
            "expiry_date": lot["expiry_date"]
        })
        
        remaining_to_issue -= issue_qty
    
    return {
        "message": "FEFO issue completed successfully",
        "total_issued": quantity,
        "issued_from": issued_from
    }

# Transfer
@app.post("/api/transfer")
async def create_transfer(transfer: TransferCreate):
    lot = lots_collection.find_one({"_id": ObjectId(transfer.lot_id)})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["current_quantity"] < transfer.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient quantity. Available: {lot['current_quantity']}"
        )
    
    destination = locations_collection.find_one({"_id": ObjectId(transfer.destination_id)})
    if not destination:
        raise HTTPException(status_code=404, detail="Destination location not found")
    
    new_quantity = lot["current_quantity"] - transfer.quantity
    
    if new_quantity == 0 or transfer.quantity == lot["current_quantity"]:
        lots_collection.update_one(
            {"_id": ObjectId(transfer.lot_id)},
            {"$set": {
                "current_quantity": new_quantity,
                "location_id": transfer.destination_id
            }}
        )
    else:
        lots_collection.update_one(
            {"_id": ObjectId(transfer.lot_id)},
            {"$set": {"current_quantity": new_quantity}}
        )
        
        new_lot_number = generate_lot_number()
        new_lot_doc = {
            "lot_number": new_lot_number,
            "item_id": lot["item_id"],
            "item_name": lot.get("item_name", ""),  # Preserve item name
            "category": lot.get("category", ""),
            "unit": lot.get("unit", ""),
            "initial_quantity": transfer.quantity,
            "current_quantity": transfer.quantity,
            "expiry_date": lot["expiry_date"],
            "location_id": transfer.destination_id,
            "vendor_id": lot.get("vendor_id"),
            "purchase_rate": lot.get("purchase_rate"),
            "qr_code": "",
            "parent_lot_id": str(lot["_id"]),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        result = lots_collection.insert_one(new_lot_doc)
        new_lot_id = str(result.inserted_id)
        
        qr_code = generate_qr_code(new_lot_id, new_lot_number)
        lots_collection.update_one(
            {"_id": ObjectId(new_lot_id)},
            {"$set": {"qr_code": qr_code}}
        )
    
    txn_doc = {
        "type": "transfer",
        "lot_id": transfer.lot_id,
        "quantity": transfer.quantity,
        "source_location_id": lot["location_id"],
        "destination_location_id": transfer.destination_id,
        "notes": transfer.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    transactions_collection.insert_one(txn_doc)
    
    return {
        "message": "Transfer completed successfully",
        "transferred_quantity": transfer.quantity
    }

@app.get("/api/transfer")
async def get_transfer_list(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {"type": "transfer"}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    transactions = list(transactions_collection.find(query).sort("created_at", -1))
    return [serialize_transaction(txn) for txn in transactions]

# Expiry Alerts
@app.get("/api/alerts/expiry")
async def get_expiry_alerts():
    now = datetime.now(timezone.utc)
    lots = list(lots_collection.find({"current_quantity": {"$gt": 0}}))
    
    expired = []
    expiring_7_days = []
    expiring_15_days = []
    expiring_30_days = []
    
    for lot in lots:
        try:
            expiry = parse_expiry_date(lot.get("expiry_date"))
        except Exception:
            continue
        days_until_expiry = (expiry - now).days
        
        serialized = serialize_lot(lot)
        serialized["days_until_expiry"] = days_until_expiry
        
        if days_until_expiry < 0:
            expired.append(serialized)
        elif days_until_expiry <= 7:
            expiring_7_days.append(serialized)
        elif days_until_expiry <= 15:
            expiring_15_days.append(serialized)
        elif days_until_expiry <= 30:
            expiring_30_days.append(serialized)
    
    return {
        "expired": expired,
        "expiring_7_days": expiring_7_days,
        "expiring_15_days": expiring_15_days,
        "expiring_30_days": expiring_30_days,
        "summary": {
            "total_expired": len(expired),
            "total_expiring_7_days": len(expiring_7_days),
            "total_expiring_15_days": len(expiring_15_days),
            "total_expiring_30_days": len(expiring_30_days)
        }
    }

# Waste
@app.post("/api/waste")
async def record_waste(
    lot_id: str = Query(...),
    quantity: float = Query(...),
    reason: str = Query(...)
):
    lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if lot["current_quantity"] < quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Quantity exceeds available. Available: {lot['current_quantity']}"
        )
    
    new_quantity = lot["current_quantity"] - quantity
    lots_collection.update_one(
        {"_id": ObjectId(lot_id)},
        {"$set": {"current_quantity": new_quantity}}
    )
    
    txn_doc = {
        "type": "waste",
        "lot_id": lot_id,
        "quantity": quantity,
        "source_location_id": lot["location_id"],
        "notes": f"Waste reason: {reason}",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    transactions_collection.insert_one(txn_doc)
    
    return {
        "message": "Waste recorded successfully",
        "remaining_quantity": new_quantity
    }


# Stock in Hand for Main Store - Category wise summary
@app.get("/api/reports/stock-mainstore")
async def get_stock_mainstore_summary():
    """Get category-wise stock summary for Main Store only - shows both quantity and value"""
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"categories": [], "summary": {"total_value": 0, "total_items": 0}}
    
    main_store_id = str(main_store["_id"])
    
    # Get all lots with positive quantity at main store
    lots = list(lots_collection.find({
        "location_id": main_store_id,
        "current_quantity": {"$gt": 0}
    }))
    
    # Group by category
    category_data = {}
    total_value = 0
    total_items = 0
    
    for lot in lots:
        # Get item details for category
        item_id = lot.get("item_id")
        item = items_collection.find_one({"_id": ObjectId(item_id)}) if item_id else None
        category = item.get("category", "Uncategorized") if item else lot.get("category", "Uncategorized")
        
        qty = float(lot.get("current_quantity", 0) or 0)
        rate = float(lot.get("purchase_rate", 0) or 0)
        value = qty * rate
        
        if category not in category_data:
            category_data[category] = {
                "category": category,
                "total_quantity": 0,
                "total_value": 0,
                "item_count": 0
            }
        
        category_data[category]["total_quantity"] += qty
        category_data[category]["total_value"] += value
        category_data[category]["item_count"] += 1
        total_value += value
        total_items += 1
    
    # Sort categories alphabetically
    categories = sorted(category_data.values(), key=lambda x: x["category"])
    
    return {
        "categories": categories,
        "summary": {
            "total_value": round(total_value, 2),
            "total_items": total_items,
            "location_name": main_store.get("name", "Main Store")
        }
    }

# Export Stock in Hand for Main Store as Excel
@app.get("/api/reports/stock-in-hand")
async def get_stock_in_hand_report():
    """Get comprehensive stock in hand report with category-wise summary - AGGREGATED BY ITEM"""
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=404, detail="Main Store not found")
    
    main_store_id = str(main_store["_id"])
    
    # Get all lots with positive quantity at main store
    lots = list(lots_collection.find({
        "location_id": main_store_id,
        "current_quantity": {"$gt": 0}
    }))
    
    # OPTIMIZATION: Fetch all items at once instead of individual queries
    all_item_ids = list(set(str(lot.get("item_id", "")) for lot in lots if lot.get("item_id")))
    all_items = {str(item["_id"]): item for item in items_collection.find({"_id": {"$in": [ObjectId(i) for i in all_item_ids if i]}})}
    
    # AGGREGATE by item_id - combine all lots of same item
    item_aggregates = {}
    
    for lot in lots:
        item_id = str(lot.get("item_id", ""))
        if not item_id:
            continue
            
        item = all_items.get(item_id)
        if not item:
            continue
        
        item_name = item.get("name", "Unknown")
        category = item.get("category", "Uncategorized") or "Uncategorized"
        unit = item.get("unit", "")
        standard_price = float(item.get("standard_price", 0) or 0)
        
        qty = float(lot.get("current_quantity", 0) or 0)
        lot_rate = float(lot.get("purchase_rate", 0) or 0)
        
        # Use standard_price if available, else use lot's purchase_rate
        rate = standard_price if standard_price > 0 else lot_rate
        
        if item_id not in item_aggregates:
            item_aggregates[item_id] = {
                "item_id": item_id,
                "item_name": item_name,
                "category": category,
                "unit": unit,
                "quantity": 0,
                "price": rate,  # Use consistent price
                "lot_count": 0
            }
        
        item_aggregates[item_id]["quantity"] += qty
        item_aggregates[item_id]["lot_count"] += 1
        # Keep the best price (non-zero preferred)
        if item_aggregates[item_id]["price"] == 0 and rate > 0:
            item_aggregates[item_id]["price"] = rate
    
    # Build final items list
    items_data = []
    category_summary = {}
    total_qty = 0
    total_value = 0
    items_with_price = 0
    items_without_price = 0
    
    for item in item_aggregates.values():
        qty = item["quantity"]
        rate = item["price"]
        value = qty * rate
        category = item["category"]
        
        total_qty += qty
        total_value += value
        
        if rate > 0:
            items_with_price += 1
        else:
            items_without_price += 1
        
        # Category summary
        if category not in category_summary:
            category_summary[category] = {
                "category": category,
                "item_count": 0,
                "total_quantity": 0,
                "total_value": 0,
                "items_without_price": 0
            }
        category_summary[category]["item_count"] += 1
        category_summary[category]["total_quantity"] += qty
        category_summary[category]["total_value"] += value
        if rate == 0:
            category_summary[category]["items_without_price"] += 1
        
        items_data.append({
            "item_name": item["item_name"],
            "category": category,
            "quantity": round(qty, 2),
            "unit": item["unit"],
            "price": rate,
            "amount": round(value, 2),
            "has_price": rate > 0,
            "lot_count": item["lot_count"]
        })
    
    # Sort items by category then item name
    items_data.sort(key=lambda x: (x["category"], x["item_name"]))
    
    # Convert category summary to sorted list
    categories = sorted(category_summary.values(), key=lambda x: x["total_value"], reverse=True)
    for cat in categories:
        cat["total_value"] = round(cat["total_value"], 2)
        cat["total_quantity"] = round(cat["total_quantity"], 2)
    
    return {
        "store_name": main_store.get("name", "Main Store"),
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "total_items": len(items_data),
            "total_quantity": round(total_qty, 2),
            "total_value": round(total_value, 2),
            "items_with_price": items_with_price,
            "items_without_price": items_without_price,
            "category_count": len(categories)
        },
        "categories": categories,
        "items": items_data
    }


@app.get("/api/export/stock-mainstore")
async def export_stock_mainstore_excel():
    """Export category-wise stock at Main Store as Excel - AGGREGATED BY ITEM (no duplicates)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=404, detail="Main Store not found")
    
    main_store_id = str(main_store["_id"])
    
    # Get all lots with positive quantity at main store
    lots = list(lots_collection.find({
        "location_id": main_store_id,
        "current_quantity": {"$gt": 0}
    }))
    
    # Build vendor lookup from vendors collection
    all_vendors = {}
    try:
        all_vendors = {str(v["_id"]): v.get("name", "") for v in vendors_collection.find({}).limit(500)}
    except:
        pass
    
    # OPTIMIZATION: Batch fetch all items at once
    all_item_ids = list(set(str(lot.get("item_id", "")) for lot in lots if lot.get("item_id")))
    all_items = {}
    if all_item_ids:
        try:
            items_data = list(items_collection.find({"_id": {"$in": [ObjectId(i) for i in all_item_ids]}}))
            all_items = {str(item["_id"]): item for item in items_data}
        except:
            pass
    
    # AGGREGATE by item_id - combine all lots of same item
    item_aggregates = {}
    
    for lot in lots:
        try:
            item_id = str(lot.get("item_id", ""))
            if not item_id:
                continue
            
            item = all_items.get(item_id)
            if not item:
                continue
            
            item_name = item.get("name", "Unknown")
            category = item.get("category", "Uncategorized") or "Uncategorized"
            unit = item.get("unit", "")
            standard_price = float(item.get("standard_price", 0) or 0)
            
            qty = float(lot.get("current_quantity", 0) or 0)
            lot_rate = float(lot.get("purchase_rate", 0) or 0)
            
            # Get vendor name - try multiple sources
            lot_vendor = ""
            
            # 1. Try vendor_id from lot
            vendor_id = lot.get("vendor_id", "")
            if vendor_id:
                lot_vendor = all_vendors.get(str(vendor_id), "")
            
            # 2. Try item's default vendor (skip PO lookup to avoid timeouts)
            if not lot_vendor:
                item_vendor = item.get("vendor", "")
                if item_vendor:
                    lot_vendor = item_vendor
            
            # Use standard_price if available, else use lot's purchase_rate
            rate = standard_price if standard_price > 0 else lot_rate
            
            if item_id not in item_aggregates:
                item_aggregates[item_id] = {
                    "item_id": item_id,
                    "item_name": item_name,
                    "category": category,
                    "unit": unit,
                    "quantity": 0,
                    "price": rate,
                    "lot_count": 0,
                    "vendors": set()
                }
            
            item_aggregates[item_id]["quantity"] += qty
            item_aggregates[item_id]["lot_count"] += 1
            if lot_vendor:
                item_aggregates[item_id]["vendors"].add(lot_vendor)
            if item_aggregates[item_id]["price"] == 0 and rate > 0:
                item_aggregates[item_id]["price"] = rate
        except Exception as e:
            # Skip problematic lots
            continue
    
    # Build final items list
    items_data = []
    category_summary = {}
    total_value = 0
    total_qty = 0
    
    for item in item_aggregates.values():
        qty = item["quantity"]
        rate = item["price"]
        value = qty * rate
        category = item["category"]
        # Convert vendors set to comma-separated string
        vendors_list = list(item.get("vendors", set()))
        vendor_str = ", ".join(vendors_list) if vendors_list else "-"
        
        total_qty += qty
        total_value += value
        
        # Category summary
        if category not in category_summary:
            category_summary[category] = {"qty": 0, "value": 0, "count": 0, "no_price": 0}
        category_summary[category]["qty"] += qty
        category_summary[category]["value"] += value
        category_summary[category]["count"] += 1
        if rate == 0:
            category_summary[category]["no_price"] += 1
        
        items_data.append({
            "item_name": item["item_name"],
            "category": category,
            "quantity": round(qty, 2),
            "unit": item["unit"],
            "price": rate,
            "amount": round(value, 2),
            "vendor": vendor_str
        })
    
    # Sort by category then item name
    items_data.sort(key=lambda x: (x["category"], x["item_name"]))
    
    # Create Excel
    wb = Workbook()
    
    # ============ SHEET 1: Category Summary ============
    ws_summary = wb.active
    ws_summary.title = "Category Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    black_font = Font(color="000000")
    black_bold_font = Font(bold=True, color="000000")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws_summary['A1'] = f"Stock in Hand - Category Summary"
    ws_summary['A1'].font = Font(bold=True, size=14, color="000000")
    ws_summary.merge_cells('A1:F1')
    
    ws_summary['A2'] = f"Store: {main_store.get('name', 'Main Store')}"
    ws_summary['A2'].font = black_font
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary['A3'].font = black_font
    
    # Category summary headers
    cat_headers = ["#", "Category", "Items", "Total Quantity", "Total Value (₹)", "Items w/o Price"]
    for col, header in enumerate(cat_headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Category data rows
    row_num = 6
    sorted_categories = sorted(category_summary.items(), key=lambda x: x[1]["value"], reverse=True)
    for idx, (cat_name, cat_data) in enumerate(sorted_categories, 1):
        ws_summary.cell(row=row_num, column=1, value=idx).font = black_font
        ws_summary.cell(row=row_num, column=2, value=cat_name).font = black_font
        ws_summary.cell(row=row_num, column=3, value=cat_data["count"]).font = black_font
        
        qty_cell = ws_summary.cell(row=row_num, column=4, value=cat_data["qty"])
        qty_cell.font = black_font
        qty_cell.number_format = '#,##0.00'
        
        val_cell = ws_summary.cell(row=row_num, column=5, value=cat_data["value"])
        val_cell.font = black_font
        val_cell.number_format = '₹#,##0.00'
        
        ws_summary.cell(row=row_num, column=6, value=cat_data["no_price"]).font = black_font
        
        for col in range(1, 7):
            ws_summary.cell(row=row_num, column=col).border = thin_border
        
        row_num += 1
    
    # Grand total row
    ws_summary.cell(row=row_num, column=1, value="").fill = total_fill
    ws_summary.cell(row=row_num, column=2, value="GRAND TOTAL").font = black_bold_font
    ws_summary.cell(row=row_num, column=2).fill = total_fill
    ws_summary.cell(row=row_num, column=3, value=len(items_data)).font = black_bold_font
    ws_summary.cell(row=row_num, column=3).fill = total_fill
    
    total_qty_cell = ws_summary.cell(row=row_num, column=4, value=total_qty)
    total_qty_cell.font = black_bold_font
    total_qty_cell.fill = total_fill
    total_qty_cell.number_format = '#,##0.00'
    
    grand_total_cell = ws_summary.cell(row=row_num, column=5, value=total_value)
    grand_total_cell.font = black_bold_font
    grand_total_cell.fill = total_fill
    grand_total_cell.number_format = '₹#,##0.00'
    
    no_price_total = sum(c["no_price"] for c in category_summary.values())
    ws_summary.cell(row=row_num, column=6, value=no_price_total).font = black_bold_font
    ws_summary.cell(row=row_num, column=6).fill = total_fill
    
    for col in range(1, 7):
        ws_summary.cell(row=row_num, column=col).border = thin_border
    
    # Column widths for summary
    ws_summary.column_dimensions['A'].width = 5
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 18
    ws_summary.column_dimensions['F'].width = 15
    
    # ============ SHEET 2: Item Details ============
    ws_items = wb.create_sheet(title="Item Details")
    
    # Title
    ws_items['A1'] = f"Stock in Hand - Item Details"
    ws_items['A1'].font = Font(bold=True, size=14, color="000000")
    ws_items.merge_cells('A1:H1')
    
    ws_items['A2'] = f"Store: {main_store.get('name', 'Main Store')}"
    ws_items['A2'].font = black_font
    ws_items['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_items['A3'].font = black_font
    
    # Item headers - added Vendor column
    item_headers = ["#", "Item Name", "Category", "Unit", "Quantity", "Price (₹)", "Amount (₹)", "Vendor"]
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Item data rows
    row_num = 6
    current_category = None
    category_subtotal = 0
    
    for idx, item in enumerate(items_data, 1):
        # Add category subtotal row when category changes
        if current_category and item["category"] != current_category:
            ws_items.cell(row=row_num, column=2, value=f"Subtotal - {current_category}").font = black_bold_font
            ws_items.cell(row=row_num, column=2).fill = total_fill
            subtotal_cell = ws_items.cell(row=row_num, column=7, value=category_subtotal)
            subtotal_cell.font = black_bold_font
            subtotal_cell.fill = total_fill
            subtotal_cell.number_format = '₹#,##0.00'
            for col in range(1, 9):
                ws_items.cell(row=row_num, column=col).fill = total_fill
                ws_items.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            category_subtotal = 0
        
        current_category = item["category"]
        category_subtotal += item["amount"]
        
        ws_items.cell(row=row_num, column=1, value=idx).font = black_font
        ws_items.cell(row=row_num, column=2, value=item["item_name"]).font = black_font
        ws_items.cell(row=row_num, column=3, value=item["category"]).font = black_font
        ws_items.cell(row=row_num, column=4, value=item["unit"]).font = black_font
        
        qty_cell = ws_items.cell(row=row_num, column=5, value=item["quantity"])
        qty_cell.font = black_font
        qty_cell.number_format = '#,##0.00'
        
        price_cell = ws_items.cell(row=row_num, column=6, value=item["price"])
        if item["price"] == 0:
            price_cell.font = Font(color="FF0000")
        else:
            price_cell.font = black_font
        price_cell.number_format = '₹#,##0.00'
        
        amount_cell = ws_items.cell(row=row_num, column=7, value=item["amount"])
        amount_cell.font = black_font
        amount_cell.number_format = '₹#,##0.00'
        
        # Vendor column
        ws_items.cell(row=row_num, column=8, value=item.get("vendor", "-")).font = black_font
        
        for col in range(1, 9):
            ws_items.cell(row=row_num, column=col).border = thin_border
        
        row_num += 1
    
    # Add final category subtotal
    if current_category:
        ws_items.cell(row=row_num, column=2, value=f"Subtotal - {current_category}").font = black_bold_font
        ws_items.cell(row=row_num, column=2).fill = total_fill
        subtotal_cell = ws_items.cell(row=row_num, column=7, value=category_subtotal)
        subtotal_cell.font = black_bold_font
        subtotal_cell.fill = total_fill
        subtotal_cell.number_format = '₹#,##0.00'
        for col in range(1, 9):
            ws_items.cell(row=row_num, column=col).fill = total_fill
            ws_items.cell(row=row_num, column=col).border = thin_border
        row_num += 1
    
    # Grand total row
    row_num += 1
    ws_items.cell(row=row_num, column=6, value="GRAND TOTAL:").font = black_bold_font
    grand_cell = ws_items.cell(row=row_num, column=7, value=total_value)
    grand_cell.font = Font(bold=True, size=12, color="000000")
    grand_cell.number_format = '₹#,##0.00'
    
    # Column widths for items
    ws_items.column_dimensions['A'].width = 6
    ws_items.column_dimensions['B'].width = 45
    ws_items.column_dimensions['C'].width = 20
    ws_items.column_dimensions['D'].width = 8
    ws_items.column_dimensions['E'].width = 12
    ws_items.column_dimensions['F'].width = 12
    ws_items.column_dimensions['G'].width = 15
    ws_items.column_dimensions['H'].width = 25  # Vendor column
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    today = datetime.now().strftime('%Y-%m-%d')
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Stock_in_Hand_{today}.xlsx"}
    )


# ============ Daily Perishables Vendor Ledger ============
@app.get("/api/reports/daily-perishables-vendor-ledger")
async def get_daily_perishables_vendor_ledger(
    vendor_id: str = Query(None, description="Filter by specific vendor ID"),
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    kitchen_id: str = Query(None, description="Filter by specific kitchen ID"),
    current_user = Depends(get_current_user)
):
    """
    Get Daily Perishables Vendor Ledger - shows items received by each kitchen from vendors.
    Groups data by: Vendor → Kitchen → Items
    """
    kitchen_receivables = db["kitchen_receivables"]
    
    # Build query
    query = {"type": "daily_perishable"}
    
    if vendor_id:
        query["vendor_id"] = vendor_id
    
    if kitchen_id:
        query["kitchen_id"] = kitchen_id
    
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        if date_filter:
            query["receive_date"] = date_filter
    
    # Fetch all matching records
    records = list(kitchen_receivables.find(query).sort([("receive_date", -1), ("vendor_name", 1), ("kitchen_name", 1)]))
    
    # Group by vendor → kitchen → items
    vendor_data = {}
    
    for rec in records:
        v_id = rec.get("vendor_id", "unknown")
        v_name = rec.get("vendor_name", "Unknown Vendor")
        k_id = rec.get("kitchen_id", "unknown")
        k_name = rec.get("kitchen_name", "Unknown Kitchen")
        
        if v_id not in vendor_data:
            vendor_data[v_id] = {
                "vendor_id": v_id,
                "vendor_name": v_name,
                "total_amount": 0,
                "total_items": 0,
                "kitchens": {}
            }
        
        if k_id not in vendor_data[v_id]["kitchens"]:
            vendor_data[v_id]["kitchens"][k_id] = {
                "kitchen_id": k_id,
                "kitchen_name": k_name,
                "total_amount": 0,
                "items": []
            }
        
        item_amount = float(rec.get("amount", 0) or 0)
        
        vendor_data[v_id]["kitchens"][k_id]["items"].append({
            "item_name": rec.get("item_name", ""),
            "category": rec.get("category", ""),
            "quantity": float(rec.get("quantity", 0) or 0),
            "unit": rec.get("unit", ""),
            "rate": float(rec.get("rate", 0) or 0),
            "amount": item_amount,
            "receive_date": rec.get("receive_date", ""),
            "invoice_number": rec.get("invoice_number", ""),
            "po_number": rec.get("po_number", "")
        })
        
        vendor_data[v_id]["kitchens"][k_id]["total_amount"] += item_amount
        vendor_data[v_id]["total_amount"] += item_amount
        vendor_data[v_id]["total_items"] += 1
    
    # Convert to list and aggregate items by item_name within each kitchen
    result = []
    grand_total = 0
    
    for v_id, v_data in vendor_data.items():
        kitchens_list = []
        for k_id, k_data in v_data["kitchens"].items():
            # Aggregate items by item_name
            item_aggregates = {}
            for item in k_data["items"]:
                item_key = item["item_name"]
                if item_key not in item_aggregates:
                    item_aggregates[item_key] = {
                        "item_name": item["item_name"],
                        "category": item["category"],
                        "unit": item["unit"],
                        "total_quantity": 0,
                        "total_amount": 0,
                        "avg_rate": 0,
                        "entries": []
                    }
                item_aggregates[item_key]["total_quantity"] += item["quantity"]
                item_aggregates[item_key]["total_amount"] += item["amount"]
                item_aggregates[item_key]["entries"].append({
                    "date": item["receive_date"],
                    "quantity": item["quantity"],
                    "rate": item["rate"],
                    "amount": item["amount"],
                    "invoice": item["invoice_number"]
                })
            
            # Calculate avg rate
            for item_key, item_data in item_aggregates.items():
                if item_data["total_quantity"] > 0:
                    item_data["avg_rate"] = round(item_data["total_amount"] / item_data["total_quantity"], 2)
                item_data["total_quantity"] = round(item_data["total_quantity"], 2)
                item_data["total_amount"] = round(item_data["total_amount"], 2)
            
            kitchens_list.append({
                "kitchen_id": k_id,
                "kitchen_name": k_data["kitchen_name"],
                "total_amount": round(k_data["total_amount"], 2),
                "item_count": len(item_aggregates),
                "items": sorted(item_aggregates.values(), key=lambda x: x["total_amount"], reverse=True)
            })
        
        result.append({
            "vendor_id": v_id,
            "vendor_name": v_data["vendor_name"],
            "total_amount": round(v_data["total_amount"], 2),
            "total_items": v_data["total_items"],
            "kitchens": sorted(kitchens_list, key=lambda x: x["total_amount"], reverse=True)
        })
        grand_total += v_data["total_amount"]
    
    # Sort vendors by total amount
    result.sort(key=lambda x: x["total_amount"], reverse=True)
    
    return {
        "success": True,
        "filters": {
            "vendor_id": vendor_id,
            "kitchen_id": kitchen_id,
            "start_date": start_date,
            "end_date": end_date
        },
        "summary": {
            "total_vendors": len(result),
            "grand_total": round(grand_total, 2)
        },
        "vendors": result
    }


@app.get("/api/export/daily-perishables-vendor-ledger")
async def export_daily_perishables_vendor_ledger(
    vendor_id: str = Query(None, description="Filter by specific vendor ID"),
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    kitchen_id: str = Query(None, description="Filter by specific kitchen ID"),
    current_user = Depends(get_current_user)
):
    """Export Daily Perishables Vendor Ledger as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    
    kitchen_receivables = db["kitchen_receivables"]
    
    # Build query
    query = {"type": "daily_perishable"}
    
    if vendor_id:
        query["vendor_id"] = vendor_id
    
    if kitchen_id:
        query["kitchen_id"] = kitchen_id
    
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        if date_filter:
            query["receive_date"] = date_filter
    
    records = list(kitchen_receivables.find(query).sort([("vendor_name", 1), ("kitchen_name", 1), ("item_name", 1)]))
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Ledger"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    black_font = Font(color="000000")
    black_bold_font = Font(bold=True, color="000000")
    vendor_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    kitchen_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Daily Perishables - Vendor Ledger"
    ws['A1'].font = Font(bold=True, size=14, color="000000")
    ws.merge_cells('A1:G1')
    
    date_range_text = ""
    if start_date and end_date:
        date_range_text = f"Period: {start_date} to {end_date}"
    elif start_date:
        date_range_text = f"From: {start_date}"
    elif end_date:
        date_range_text = f"Until: {end_date}"
    else:
        date_range_text = "All Time"
    
    ws['A2'] = date_range_text
    ws['A2'].font = black_font
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = black_font
    
    # Headers
    headers = ["Vendor", "Kitchen", "Item Name", "Qty", "Unit", "Rate (₹)", "Amount (₹)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Group records by vendor → kitchen
    vendor_kitchen_items = {}
    for rec in records:
        v_name = rec.get("vendor_name", "Unknown")
        k_name = rec.get("kitchen_name", "Unknown")
        key = (v_name, k_name)
        
        if key not in vendor_kitchen_items:
            vendor_kitchen_items[key] = []
        
        vendor_kitchen_items[key].append(rec)
    
    row_num = 6
    current_vendor = None
    vendor_total = 0
    kitchen_total = 0
    current_kitchen = None
    grand_total = 0
    
    for (v_name, k_name), items in sorted(vendor_kitchen_items.items()):
        # Vendor change - add vendor total
        if current_vendor and current_vendor != v_name:
            # Kitchen subtotal
            ws.cell(row=row_num, column=2, value=f"Subtotal - {current_kitchen}").font = black_bold_font
            ws.cell(row=row_num, column=7, value=kitchen_total).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=7).font = black_bold_font
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = subtotal_fill
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            
            # Vendor total
            ws.cell(row=row_num, column=1, value=f"TOTAL - {current_vendor}").font = Font(bold=True, color="000000", size=11)
            ws.cell(row=row_num, column=7, value=vendor_total).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=7).font = Font(bold=True, color="000000", size=11)
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = vendor_fill
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 2
            vendor_total = 0
            kitchen_total = 0
        elif current_kitchen and current_kitchen != k_name:
            # Kitchen subtotal
            ws.cell(row=row_num, column=2, value=f"Subtotal - {current_kitchen}").font = black_bold_font
            ws.cell(row=row_num, column=7, value=kitchen_total).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=7).font = black_bold_font
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = subtotal_fill
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            kitchen_total = 0
        
        current_vendor = v_name
        current_kitchen = k_name
        
        # Aggregate items by item_name for this kitchen
        item_aggregates = {}
        for rec in items:
            item_name = rec.get("item_name", "")
            if item_name not in item_aggregates:
                item_aggregates[item_name] = {
                    "item_name": item_name,
                    "unit": rec.get("unit", ""),
                    "total_qty": 0,
                    "total_amount": 0
                }
            item_aggregates[item_name]["total_qty"] += float(rec.get("quantity", 0) or 0)
            item_aggregates[item_name]["total_amount"] += float(rec.get("amount", 0) or 0)
        
        # Write aggregated items
        first_item = True
        for item_name, item_data in sorted(item_aggregates.items()):
            qty = item_data["total_qty"]
            amount = item_data["total_amount"]
            rate = amount / qty if qty > 0 else 0
            
            ws.cell(row=row_num, column=1, value=v_name if first_item else "").font = black_font
            ws.cell(row=row_num, column=2, value=k_name if first_item else "").font = black_font
            ws.cell(row=row_num, column=3, value=item_name).font = black_font
            ws.cell(row=row_num, column=4, value=round(qty, 2)).font = black_font
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5, value=item_data["unit"]).font = black_font
            ws.cell(row=row_num, column=6, value=round(rate, 2)).font = black_font
            ws.cell(row=row_num, column=6).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=7, value=round(amount, 2)).font = black_font
            ws.cell(row=row_num, column=7).number_format = '₹#,##0.00'
            
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = thin_border
            
            kitchen_total += amount
            vendor_total += amount
            grand_total += amount
            first_item = False
            row_num += 1
    
    # Final kitchen subtotal
    if current_kitchen:
        ws.cell(row=row_num, column=2, value=f"Subtotal - {current_kitchen}").font = black_bold_font
        ws.cell(row=row_num, column=7, value=kitchen_total).number_format = '₹#,##0.00'
        ws.cell(row=row_num, column=7).font = black_bold_font
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).fill = subtotal_fill
            ws.cell(row=row_num, column=col).border = thin_border
        row_num += 1
    
    # Final vendor total
    if current_vendor:
        ws.cell(row=row_num, column=1, value=f"TOTAL - {current_vendor}").font = Font(bold=True, color="000000", size=11)
        ws.cell(row=row_num, column=7, value=vendor_total).number_format = '₹#,##0.00'
        ws.cell(row=row_num, column=7).font = Font(bold=True, color="000000", size=11)
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).fill = vendor_fill
            ws.cell(row=row_num, column=col).border = thin_border
        row_num += 2
    
    # Grand total
    ws.cell(row=row_num, column=1, value="GRAND TOTAL").font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=row_num, column=1).fill = header_fill
    ws.cell(row=row_num, column=7, value=grand_total).number_format = '₹#,##0.00'
    ws.cell(row=row_num, column=7).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=row_num, column=7).fill = header_fill
    for col in range(1, 8):
        ws.cell(row=row_num, column=col).border = thin_border
        if col not in [1, 7]:
            ws.cell(row=row_num, column=col).fill = header_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Daily_Perishables_Vendor_Ledger"
    if start_date:
        filename += f"_{start_date}"
    if end_date:
        filename += f"_to_{end_date}"
    filename += ".xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ============ Consumption Analysis Report ============
@app.get("/api/reports/consumption-analysis")
async def get_consumption_analysis(
    days_of_data: int = Query(10, description="Number of days of historical data to analyze"),
    par_stock_days: int = Query(10, description="Number of days for PAR stock calculation"),
    kitchen_id: str = Query(None, description="Filter by specific kitchen"),
    category: str = Query(None, description="Filter by category"),
    current_user = Depends(get_current_user)
):
    """
    Analyze stock movements (IN/OUT) for PAR stock planning.
    
    Data Source:
    - OUT = Transactions with type='issue' (items issued from Main Store to kitchens)
    - IN = Lots created (GRN)
    
    PAR Stock Calculation:
    - Daily Average = Total OUT / number of days
    - PAR Stock (10 days) = Daily Average × 10
    """
    from datetime import timedelta
    
    # Calculate date range
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=days_of_data)
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")
    
    # Get Main Store location
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Initialize aggregations
    item_consumption = {}
    category_consumption = {}
    daily_movements = {}  # Track daily IN/OUT for each item
    
    # ===== QUERY TRANSACTIONS FOR OUT MOVEMENTS =====
    # Include ALL transaction types that represent items leaving Main Store
    # This matches the Stock Movement report: dispatch, issue, transfer
    txn_query = {
        "type": {"$in": ["dispatch", "issue", "transfer"]},
        "created_at": {
            "$gte": start_date_str,
            "$lte": end_date_str + "T23:59:59"
        }
    }
    if kitchen_id:
        txn_query["destination_location_id"] = kitchen_id
    
    total_out = 0
    total_out_value = 0
    days_with_data = set()
    
    for txn in transactions_collection.find(txn_query):
        lot_id = txn.get("lot_id")
        qty = float(txn.get("quantity", 0) or 0)
        txn_date = txn.get("created_at", "")[:10]  # Get date portion
        
        if not lot_id or qty <= 0:
            continue
        
        # Get lot details for item info
        lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
        if not lot:
            continue
        
        item_id = lot.get("item_id", "")
        item_name = lot.get("item_name", "Unknown")
        
        # Get item details
        db_item = None
        if item_id:
            try:
                db_item = items_collection.find_one({"_id": ObjectId(item_id)})
            except:
                pass
        
        if not db_item:
            db_item = items_collection.find_one({"name": item_name})
        
        item_category = db_item.get("category", "Uncategorized") if db_item else "Uncategorized"
        unit = db_item.get("unit", lot.get("unit", "")) if db_item else lot.get("unit", "")
        
        # Get rate with multiple fallbacks:
        # 1. Transaction's value field (already calculated, most reliable)
        # 2. Lot's purchase_rate
        # 3. Item's standard_price or purchase_price
        txn_value = float(txn.get("value", 0) or 0)
        
        if txn_value > 0:
            # Use pre-calculated value from transaction
            value = txn_value
            rate = value / qty if qty > 0 else 0
        else:
            # Calculate from rate
            rate = lot.get("purchase_rate") or 0
            if rate == 0:
                rate = txn.get("rate") or 0  # Some transactions store rate directly
            if rate == 0 and db_item:
                rate = db_item.get("standard_price") or db_item.get("purchase_price") or 0
            value = qty * float(rate)
        
        # Apply category filter
        if category and item_category.lower() != category.lower():
            continue
        
        total_out += qty
        total_out_value += value
        days_with_data.add(txn_date)
        
        # Aggregate by item
        item_key = item_id or item_name
        if item_key not in item_consumption:
            item_consumption[item_key] = {
                "item_id": item_id,
                "item_name": item_name,
                "category": item_category,
                "unit": unit,
                "rate": rate,
                "total_out": 0,
                "total_out_value": 0,
                "total_in": 0,
                "total_in_value": 0,
                "daily_out": {},
                "current_stock": 0,
                "current_par_stock": 0
            }
        item_consumption[item_key]["total_out"] += qty
        item_consumption[item_key]["total_out_value"] += value
        
        # Track daily OUT
        if txn_date not in item_consumption[item_key]["daily_out"]:
            item_consumption[item_key]["daily_out"][txn_date] = 0
        item_consumption[item_key]["daily_out"][txn_date] += qty
        
        # Aggregate by category
        if item_category not in category_consumption:
            category_consumption[item_category] = {
                "total_out": 0,
                "total_out_value": 0,
                "item_count": set()
            }
        category_consumption[item_category]["total_out"] += qty
        category_consumption[item_category]["total_out_value"] += value
        category_consumption[item_category]["item_count"].add(item_key)
    
    # ===== GET CURRENT STOCK AND PAR STOCK FOR EACH ITEM =====
    for item_key, item_data in item_consumption.items():
        item_id = item_data["item_id"]
        if item_id:
            # Get current stock from lots
            stock_query = {"item_id": item_id, "current_quantity": {"$gt": 0}}
            if main_store_id:
                stock_query["location_id"] = main_store_id
            lots = list(lots_collection.find(stock_query))
            current_stock = sum(lot.get("current_quantity", 0) for lot in lots)
            item_data["current_stock"] = round(current_stock, 2)
            
            # Get current PAR stock setting
            try:
                db_item = items_collection.find_one({"_id": ObjectId(item_id)})
                if db_item:
                    item_data["current_par_stock"] = db_item.get("par_stock", 0) or 0
            except:
                pass
    
    # ===== CALCULATE PAR STOCK =====
    actual_days = len(days_with_data) if days_with_data else days_of_data
    
    # Build item summary with PAR stock calculation
    item_summary = []
    for item_key, item_data in item_consumption.items():
        daily_avg_out = item_data["total_out"] / actual_days if actual_days > 0 else 0
        calculated_par = daily_avg_out * par_stock_days
        par_stock_value = calculated_par * (item_data["rate"] or 0)
        
        item_summary.append({
            "item_id": item_data["item_id"],
            "item_name": item_data["item_name"],
            "category": item_data["category"],
            "unit": item_data["unit"],
            "rate": round(item_data["rate"], 2),
            "total_quantity": round(item_data["total_out"], 2),  # Frontend expects this name
            "total_out": round(item_data["total_out"], 2),
            "total_value": round(item_data["total_out_value"], 2),
            "total_out_value": round(item_data["total_out_value"], 2),
            "daily_avg_qty": round(daily_avg_out, 2),  # Frontend expects this name
            "daily_avg_out": round(daily_avg_out, 2),
            "par_stock_10_days_qty": round(calculated_par, 2),  # Frontend expects this name
            "calculated_par_stock": round(calculated_par, 2),
            "calculated_par_value": round(par_stock_value, 2),
            "current_stock": item_data["current_stock"],
            "current_par_stock": item_data["current_par_stock"],
            "stock_status": "OK" if item_data["current_stock"] >= calculated_par else "LOW",
            "shortage": round(max(0, calculated_par - item_data["current_stock"]), 2)
        })
    
    item_summary.sort(key=lambda x: x["total_quantity"], reverse=True)
    
    # Build category summary
    category_summary = []
    for cat_name, cat_data in category_consumption.items():
        daily_avg = cat_data["total_out"] / actual_days if actual_days > 0 else 0
        par_qty = daily_avg * par_stock_days
        
        category_summary.append({
            "category": cat_name,
            "total_quantity": round(cat_data["total_out"], 2),  # Frontend expects this name
            "total_out": round(cat_data["total_out"], 2),
            "total_value": round(cat_data["total_out_value"], 2),
            "total_out_value": round(cat_data["total_out_value"], 2),
            "unique_items": len(cat_data["item_count"]),
            "daily_avg_qty": round(daily_avg, 2),  # Frontend expects this name
            "daily_avg_out": round(daily_avg, 2),
            "par_stock_10_days_qty": round(par_qty, 2),  # Frontend expects this name
            f"par_stock_{par_stock_days}_days": round(par_qty, 2)
        })
    
    category_summary.sort(key=lambda x: x["total_quantity"], reverse=True)
    
    return {
        "success": True,
        "report_type": "PAR Stock Analysis - Based on Stock Movements (Issues)",
        "data_source": "transactions (type=dispatch/issue/transfer) - Items leaving Main Store",
        "analysis_period": {
            "start_date": start_date_str,
            "end_date": end_date_str,
            "days_analyzed": actual_days,  # Frontend expects this name
            "days_with_data": actual_days,
            "par_stock_days": par_stock_days
        },
        "filters": {
            "kitchen_id": kitchen_id,
            "category": category
        },
        "summary": {
            "total_quantity": round(total_out, 2),  # Frontend expects this name
            "total_out_quantity": round(total_out, 2),
            "total_value": round(total_out_value, 2),
            "total_out_value": round(total_out_value, 2),
            "daily_avg_qty": round(total_out / actual_days if actual_days > 0 else 0, 2),  # Frontend expects this
            "daily_avg_out": round(total_out / actual_days if actual_days > 0 else 0, 2),
            "daily_avg_value": round(total_out_value / actual_days if actual_days > 0 else 0, 2),
            "par_stock_10_days_qty": round((total_out / actual_days if actual_days > 0 else 0) * par_stock_days, 2),  # Frontend expects this
            f"par_stock_{par_stock_days}_days_qty": round((total_out / actual_days if actual_days > 0 else 0) * par_stock_days, 2),
            f"par_stock_{par_stock_days}_days_value": round((total_out_value / actual_days if actual_days > 0 else 0) * par_stock_days, 2),
            "unique_items": len(item_consumption),
            "categories_count": len(category_consumption)
        },
        "by_category": category_summary,
        "by_item": item_summary[:100]
    }


def _add_to_consumption(item_consumption, category_consumption, kitchen_consumption, daily_totals,
                       item_id, item_name, item_category, unit, rate, qty, value, date_key, k_id, kitchen_name):
    """Helper function to add consumption data to aggregations"""
    
    # Item-level aggregation
    item_key = item_id or item_name
    if item_key not in item_consumption:
        item_consumption[item_key] = {
            "item_id": item_id,
            "item_name": item_name,
            "category": item_category,
            "unit": unit,
            "rate": rate,
            "total_quantity": 0,
            "total_value": 0
        }
    item_consumption[item_key]["total_quantity"] += qty
    item_consumption[item_key]["total_value"] += value
    # Update rate if we found a better one
    if item_consumption[item_key]["rate"] == 0 and rate > 0:
        item_consumption[item_key]["rate"] = rate
    
    # Category-level aggregation
    if item_category not in category_consumption:
        category_consumption[item_category] = {
            "category": item_category,
            "total_quantity": 0,
            "total_value": 0,
            "item_count": set()
        }
    category_consumption[item_category]["total_quantity"] += qty
    category_consumption[item_category]["total_value"] += value
    category_consumption[item_category]["item_count"].add(item_key)
    
    # Kitchen-level aggregation - use kitchen_name as key to avoid duplicate IDs
    kitchen_key = kitchen_name or k_id
    if kitchen_key not in kitchen_consumption:
        kitchen_consumption[kitchen_key] = {
            "kitchen_id": k_id,
            "kitchen_name": kitchen_name,
            "total_quantity": 0,
            "total_value": 0,
            "categories": {}
        }
    kitchen_consumption[kitchen_key]["total_quantity"] += qty
    kitchen_consumption[kitchen_key]["total_value"] += value
    if item_category not in kitchen_consumption[kitchen_key]["categories"]:
        kitchen_consumption[kitchen_key]["categories"][item_category] = 0
    kitchen_consumption[kitchen_key]["categories"][item_category] += value
    
    # Daily totals
    if date_key not in daily_totals:
        daily_totals[date_key] = 0
    daily_totals[date_key] += value


@app.get("/api/export/consumption-analysis")
async def export_consumption_analysis(
    days_of_data: int = Query(30, description="Number of days of historical data to analyze"),
    par_stock_days: int = Query(10, description="Number of days for PAR stock calculation"),
    kitchen_id: str = Query(None, description="Filter by specific kitchen"),
    category: str = Query(None, description="Filter by category"),
    current_user = Depends(get_current_user)
):
    """Export PAR Stock Analysis as Excel - uses same logic as report endpoint"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    from datetime import timedelta
    
    # Get the analysis data using the same logic as the report endpoint
    report_data = await get_consumption_analysis(days_of_data, par_stock_days, kitchen_id, category, current_user)
    
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=days_of_data)
    actual_days = report_data.get("analysis_period", {}).get("days_analyzed", 1)
    
    # Create Excel
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    black_font = Font(color="000000")
    black_bold_font = Font(bold=True, color="000000")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Category Summary
    ws_cat = wb.active
    ws_cat.title = "Category PAR Stock"
    
    ws_cat['A1'] = "PAR Stock Analysis - Kitchen Requisitions Only"
    ws_cat['A1'].font = Font(bold=True, size=14, color="000000")
    ws_cat.merge_cells('A1:F1')
    
    ws_cat['A2'] = f"Analysis Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} ({actual_days} days)"
    ws_cat['A2'].font = black_font
    ws_cat['A3'] = f"PAR Stock: {par_stock_days} days | Based on Kitchen Requisitions to Main Store"
    ws_cat['A3'].font = black_font
    
    ws_cat['A4'] = f"Total Requisitions Analyzed: {report_data.get('summary', {}).get('total_requisitions', 0)}"
    ws_cat['A4'].font = Font(italic=True, color="666666")
    
    cat_headers = ["Category", "Items", "Total Qty", "Daily Avg Qty", f"{par_stock_days}-Day PAR Qty", f"{par_stock_days}-Day PAR Value (₹)"]
    for col, header in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_num = 7
    for cat in report_data.get("by_category", []):
        ws_cat.cell(row=row_num, column=1, value=cat["category"]).font = black_font
        ws_cat.cell(row=row_num, column=2, value=cat["unique_items"]).font = black_font
        ws_cat.cell(row=row_num, column=3, value=cat["total_quantity"]).font = black_font
        ws_cat.cell(row=row_num, column=4, value=cat["daily_avg_qty"]).font = black_font
        ws_cat.cell(row=row_num, column=5, value=cat.get(f"par_stock_{par_stock_days}_days_qty", 0)).font = black_bold_font
        ws_cat.cell(row=row_num, column=6, value=cat.get(f"par_stock_{par_stock_days}_days_value", 0)).number_format = '₹#,##0.00'
        
        for col in range(1, 7):
            ws_cat.cell(row=row_num, column=col).border = thin_border
        
        row_num += 1
    
    # Grand total row
    summary = report_data.get("summary", {})
    ws_cat.cell(row=row_num, column=1, value="GRAND TOTAL").font = black_bold_font
    ws_cat.cell(row=row_num, column=2, value=summary.get("unique_items", 0)).font = black_bold_font
    ws_cat.cell(row=row_num, column=3, value=summary.get("total_quantity", 0)).font = black_bold_font
    ws_cat.cell(row=row_num, column=4, value=summary.get("daily_avg_qty", 0)).font = black_bold_font
    ws_cat.cell(row=row_num, column=5, value=summary.get(f"par_stock_{par_stock_days}_days_qty", 0)).font = black_bold_font
    ws_cat.cell(row=row_num, column=6, value=summary.get(f"par_stock_{par_stock_days}_days_value", 0)).number_format = '₹#,##0.00'
    
    for col in range(1, 7):
        ws_cat.cell(row=row_num, column=col).fill = total_fill
        ws_cat.cell(row=row_num, column=col).border = thin_border
    
    # Adjust column widths
    ws_cat.column_dimensions['A'].width = 25
    ws_cat.column_dimensions['B'].width = 10
    ws_cat.column_dimensions['C'].width = 12
    ws_cat.column_dimensions['D'].width = 14
    ws_cat.column_dimensions['E'].width = 16
    ws_cat.column_dimensions['F'].width = 20
    
    # Sheet 2: Item Details
    ws_items = wb.create_sheet(title="Item PAR Stock")
    
    ws_items['A1'] = "PAR Stock Analysis - Item Details"
    ws_items['A1'].font = Font(bold=True, size=14, color="000000")
    ws_items.merge_cells('A1:H1')
    
    ws_items['A2'] = f"Analysis Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} ({actual_days} days)"
    ws_items['A2'].font = black_font
    
    item_headers = ["Item Name", "Category", "Unit", "Total Qty", "Daily Avg Qty", f"{par_stock_days}-Day PAR Qty", "Rate (₹)", f"{par_stock_days}-Day PAR Value (₹)"]
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_num = 5
    for item in report_data.get("by_item", [])[:200]:  # Top 200 items
        ws_items.cell(row=row_num, column=1, value=item["item_name"]).font = black_font
        ws_items.cell(row=row_num, column=2, value=item["category"]).font = black_font
        ws_items.cell(row=row_num, column=3, value=item["unit"]).font = black_font
        ws_items.cell(row=row_num, column=4, value=item["total_quantity"]).number_format = '#,##0.00'
        ws_items.cell(row=row_num, column=5, value=item["daily_avg_qty"]).number_format = '#,##0.00'
        ws_items.cell(row=row_num, column=6, value=item.get(f"par_stock_{par_stock_days}_days_qty", 0)).number_format = '#,##0.00'
        ws_items.cell(row=row_num, column=6).font = black_bold_font
        ws_items.cell(row=row_num, column=7, value=item["rate"]).number_format = '₹#,##0.00'
        ws_items.cell(row=row_num, column=8, value=item.get(f"par_stock_{par_stock_days}_days_value", 0)).number_format = '₹#,##0.00'
        
        for col in range(1, 9):
            ws_items.cell(row=row_num, column=col).border = thin_border
        
        row_num += 1
    
    # Column widths
    ws_items.column_dimensions['A'].width = 40
    ws_items.column_dimensions['B'].width = 20
    ws_items.column_dimensions['C'].width = 8
    ws_items.column_dimensions['D'].width = 12
    ws_items.column_dimensions['E'].width = 14
    ws_items.column_dimensions['F'].width = 16
    ws_items.column_dimensions['G'].width = 12
    ws_items.column_dimensions['H'].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"PAR_Stock_Analysis_{par_stock_days}days.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/reports/kitchen-requisition-analysis")
async def get_kitchen_requisition_analysis(
    kitchen_name: str = Query(..., description="Kitchen name to analyze (e.g., 'Sticky Rice GCR')"),
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user = Depends(get_current_user)
):
    """
    Analyze requisitions from a specific kitchen for a date range.
    Returns category-wise and item-wise breakdown of quantities requisitioned.
    """
    from collections import defaultdict
    
    # Parse dates
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Find kitchen
    kitchen = locations_collection.find_one({"name": {"$regex": kitchen_name, "$options": "i"}})
    kitchen_id = str(kitchen["_id"]) if kitchen else None
    actual_kitchen_name = kitchen.get("name") if kitchen else kitchen_name
    
    # Get requisitions
    req_query = {
        "$or": [
            {"kitchen_id": kitchen_id} if kitchen_id else {"kitchen_id": "none"},
            {"location_name": {"$regex": kitchen_name, "$options": "i"}}
        ],
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "status": {"$in": ["dispatched", "partial", "completed", "received"]}
    }
    
    requisitions = list(requisitions_collection.find(req_query).sort("created_at", 1))
    
    # Aggregate data
    category_data = defaultdict(lambda: {
        "items": defaultdict(lambda: {"qty": 0, "unit": "", "requisitions": []}),
        "total_qty": 0
    })
    
    requisition_list = []
    
    for req in requisitions:
        req_date = req.get("created_at")
        if isinstance(req_date, datetime):
            date_str = req_date.strftime("%Y-%m-%d")
        else:
            date_str = str(req_date)[:10] if req_date else "Unknown"
        
        req_info = {
            "serial_number": req.get("serial_number"),
            "date": date_str,
            "status": req.get("status"),
            "items_count": len(req.get("items", []))
        }
        requisition_list.append(req_info)
        
        for item in req.get("items", []):
            item_name = item.get("item_name", "Unknown")
            category = item.get("category", "Uncategorized") or "Uncategorized"
            qty = float(item.get("quantity", 0) or 0)
            unit = item.get("unit", "")
            
            category_data[category]["items"][item_name]["qty"] += qty
            category_data[category]["items"][item_name]["unit"] = unit
            category_data[category]["items"][item_name]["requisitions"].append({
                "serial": req.get("serial_number"),
                "date": date_str,
                "qty": qty
            })
            category_data[category]["total_qty"] += qty
    
    # Build response
    category_summary = []
    category_details = {}
    
    for cat_name, cat_data in sorted(category_data.items(), key=lambda x: x[1]["total_qty"], reverse=True):
        category_summary.append({
            "category": cat_name,
            "unique_items": len(cat_data["items"]),
            "total_qty": round(cat_data["total_qty"], 2)
        })
        
        items_list = []
        for item_name, item_data in sorted(cat_data["items"].items(), key=lambda x: x[1]["qty"], reverse=True):
            items_list.append({
                "item_name": item_name,
                "qty": round(item_data["qty"], 2),
                "unit": item_data["unit"],
                "requisition_details": item_data["requisitions"]
            })
        
        category_details[cat_name] = items_list
    
    # Calculate days
    days_in_range = (end_dt - start_dt).days + 1
    
    return {
        "success": True,
        "kitchen_name": actual_kitchen_name,
        "analysis_period": {
            "start_date": start_date,
            "end_date": end_date,
            "days": days_in_range
        },
        "summary": {
            "total_requisitions": len(requisitions),
            "total_categories": len(category_data),
            "grand_total_qty": round(sum(c["total_qty"] for c in category_data.values()), 2)
        },
        "requisitions": requisition_list,
        "category_summary": category_summary,
        "category_details": category_details
    }


@app.get("/api/export/kitchen-requisition-analysis")
async def export_kitchen_requisition_analysis(
    kitchen_name: str = Query(..., description="Kitchen name to analyze"),
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user = Depends(get_current_user)
):
    """Export Kitchen Requisition Analysis as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    
    # Get the analysis data
    report_data = await get_kitchen_requisition_analysis(kitchen_name, start_date, end_date, current_user)
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    category_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    black_font = Font(color="000000")
    black_bold = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Requisition Analysis"
    
    ws['A1'] = f"Kitchen Requisition Analysis - {report_data['kitchen_name']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Period: {start_date} to {end_date} ({report_data['analysis_period']['days']} days)"
    ws['A3'] = f"Total Requisitions: {report_data['summary']['total_requisitions']}"
    
    # Category Summary
    row = 5
    ws.cell(row=row, column=1, value="CATEGORY SUMMARY").font = black_bold
    row += 1
    
    headers = ["Category", "Unique Items", "Total Qty"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row += 1
    for cat in report_data['category_summary']:
        ws.cell(row=row, column=1, value=cat['category']).border = thin_border
        ws.cell(row=row, column=2, value=cat['unique_items']).border = thin_border
        ws.cell(row=row, column=3, value=cat['total_qty']).border = thin_border
        row += 1
    
    # Item Details by Category
    row += 2
    ws.cell(row=row, column=1, value="ITEM-WISE DETAILS BY CATEGORY").font = black_bold
    row += 1
    
    target_categories = ["Chinese Grocery", "Beverage", "Housekeeping", "Indian Grocery", "Packaging"]
    
    for cat_name in target_categories:
        if cat_name in report_data['category_details']:
            row += 1
            ws.cell(row=row, column=1, value=cat_name.upper()).font = black_bold
            ws.cell(row=row, column=1).fill = category_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Headers
            item_headers = ["#", "Item Name", "Qty", "Unit"]
            for col, h in enumerate(item_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1
            
            items = report_data['category_details'][cat_name]
            for idx, item in enumerate(items, 1):
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=item['item_name']).border = thin_border
                ws.cell(row=row, column=3, value=item['qty']).border = thin_border
                ws.cell(row=row, column=4, value=item['unit']).border = thin_border
                row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    safe_kitchen = kitchen_name.replace(" ", "_")
    filename = f"Requisition_Analysis_{safe_kitchen}_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@app.get("/api/reports/par-stock-calculation")
async def calculate_par_stock(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    par_days: int = Query(10, description="Number of days for PAR stock calculation"),
    current_user = Depends(get_current_user)
):
    """
    Calculate PAR Stock based on ACTUAL CONSUMPTION (challans/issues) from Main Store to Kitchens.
    This tracks what was actually dispatched via challans, not what was requested.
    Example: If Salt is sent to 8 kitchens (1kg each), it shows 8kg total consumption.
    """
    from collections import defaultdict
    
    # Parse dates
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    days_in_range = (end_dt - start_dt).days + 1
    
    # Build location map for kitchen names
    locations = list(locations_collection.find({}, {"_id": 1, "name": 1, "type": 1}))
    location_map = {str(loc["_id"]): loc for loc in locations}
    
    # Get Main Store ID
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Build item info map for categories and units
    items_list = list(items_collection.find({}, {"_id": 1, "name": 1, "category": 1, "unit": 1, "standard_price": 1}))
    items_map = {str(item["_id"]): item for item in items_list}
    
    # Get ISSUE transactions (challans) from Main Store to Kitchens in date range
    # These are the actual dispatches via challans
    issue_query = {
        "type": "issue",
        "created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"}
    }
    
    # If we have main_store_id, filter by source_location_id
    if main_store_id:
        issue_query["source_location_id"] = main_store_id
    
    issue_transactions = list(transactions_collection.find(issue_query))
    
    # Aggregate consumption by item
    item_consumption = defaultdict(lambda: {
        "item_id": "",
        "item_name": "",
        "category": "",
        "unit": "",
        "total_qty": 0,
        "total_value": 0,
        "rate": 0,
        "kitchens": defaultdict(float),  # kitchen_name -> qty
        "challan_count": 0
    })
    
    category_totals = defaultdict(lambda: {"qty": 0, "value": 0, "items": set()})
    
    for txn in issue_transactions:
        item_id = txn.get("item_id", "")
        qty = float(txn.get("quantity", 0) or 0)
        rate = float(txn.get("rate", 0) or 0)
        value = float(txn.get("value", 0) or 0) or (qty * rate)
        
        # Get destination kitchen name
        dest_loc_id = txn.get("destination_location_id", "")
        dest_loc = location_map.get(dest_loc_id, {})
        kitchen_name = dest_loc.get("name", "Unknown Kitchen")
        
        # Get item details from items_map
        item_info = items_map.get(item_id, {})
        item_name = item_info.get("name", "Unknown Item")
        category = item_info.get("category", "Uncategorized") or "Uncategorized"
        unit = item_info.get("unit", "")
        
        # Use rate from transaction, or standard_price as fallback
        if rate == 0:
            rate = float(item_info.get("standard_price", 0) or 0)
            value = qty * rate
        
        # Use item_id as key
        key = item_id or item_name
        
        item_consumption[key]["item_id"] = item_id
        item_consumption[key]["item_name"] = item_name
        item_consumption[key]["category"] = category
        item_consumption[key]["unit"] = unit
        item_consumption[key]["total_qty"] += qty
        item_consumption[key]["total_value"] += value
        if rate > 0:
            item_consumption[key]["rate"] = rate
        item_consumption[key]["kitchens"][kitchen_name] += qty
        item_consumption[key]["challan_count"] += 1
        
        category_totals[category]["qty"] += qty
        category_totals[category]["value"] += value
        category_totals[category]["items"].add(key)
    
    # Calculate PAR stock for each item
    item_par_stock = []
    for key, data in item_consumption.items():
        daily_avg = data["total_qty"] / days_in_range if days_in_range > 0 else 0
        par_stock_qty = daily_avg * par_days
        par_stock_value = par_stock_qty * data["rate"]
        
        item_par_stock.append({
            "item_id": data["item_id"],
            "item_name": data["item_name"],
            "category": data["category"],
            "unit": data["unit"],
            "rate": round(data["rate"], 2),
            "consumption_qty": round(data["total_qty"], 2),
            "consumption_value": round(data["total_value"], 2),
            "daily_avg_qty": round(daily_avg, 2),
            f"par_stock_{par_days}_days_qty": round(par_stock_qty, 2),
            f"par_stock_{par_days}_days_value": round(par_stock_value, 2),
            "kitchen_breakdown": {k: round(v, 2) for k, v in data["kitchens"].items()},
            "challan_count": data["challan_count"]
        })
    
    # Sort by category then by qty
    item_par_stock.sort(key=lambda x: (x["category"], -x["consumption_qty"]))
    
    # Calculate category summary
    category_summary = []
    for cat, data in sorted(category_totals.items(), key=lambda x: x[1]["qty"], reverse=True):
        daily_avg = data["qty"] / days_in_range if days_in_range > 0 else 0
        par_qty = daily_avg * par_days
        category_summary.append({
            "category": cat,
            "items_count": len(data["items"]),
            "consumption_qty": round(data["qty"], 2),
            "consumption_value": round(data["value"], 2),
            "daily_avg_qty": round(daily_avg, 2),
            f"par_stock_{par_days}_days_qty": round(par_qty, 2)
        })
    
    grand_total_qty = sum(c["consumption_qty"] for c in category_summary)
    grand_total_value = sum(c["consumption_value"] for c in category_summary)
    
    return {
        "success": True,
        "report_type": f"PAR Stock Calculation ({par_days} days) - Based on Challans/Issues",
        "analysis_period": {
            "start_date": start_date,
            "end_date": end_date,
            "days_analyzed": days_in_range,
            "par_days": par_days
        },
        "summary": {
            "total_challans": len(issue_transactions),
            "unique_items": len(item_consumption),
            "categories": len(category_totals),
            "total_consumption_qty": round(grand_total_qty, 2),
            "total_consumption_value": round(grand_total_value, 2)
        },
        "category_summary": category_summary,
        "items": item_par_stock
    }


@app.get("/api/export/par-stock-calculation")
async def export_par_stock_calculation(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    par_days: int = Query(10, description="Number of days for PAR stock calculation"),
    current_user = Depends(get_current_user)
):
    """
    Export PAR Stock Calculation as Excel with item-wise breakdown by category.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    
    # Get the calculation data
    report_data = await calculate_par_stock(start_date, end_date, par_days, current_user)
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Green
    category_fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")  # Light green
    subtotal_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Very light green
    highlight_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Light orange for PAR
    black_font = Font(color="000000")
    bold_font = Font(bold=True, color="000000")
    par_font = Font(bold=True, color="E65100")  # Orange for PAR values
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Category Summary
    ws_summary = wb.active
    ws_summary.title = "Category Summary"
    
    ws_summary['A1'] = f"PAR Stock Calculation - {par_days} Days"
    ws_summary['A1'].font = Font(bold=True, size=16, color="2E7D32")
    ws_summary.merge_cells('A1:F1')
    
    ws_summary['A2'] = f"Consumption Period: {start_date} to {end_date} ({report_data['analysis_period']['days_analyzed']} days)"
    ws_summary['A2'].font = black_font
    
    ws_summary['A3'] = f"Total Challans: {report_data['summary']['total_challans']} | Unique Items: {report_data['summary']['unique_items']}"
    ws_summary['A3'].font = Font(italic=True, color="666666")
    
    # Category headers
    cat_headers = ["Category", "Items", "Total Consumed", "Daily Avg", f"{par_days}-Day PAR Stock", "Unit"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    row = 6
    for cat in report_data['category_summary']:
        ws_summary.cell(row=row, column=1, value=cat['category']).border = thin_border
        ws_summary.cell(row=row, column=2, value=cat['items_count']).border = thin_border
        ws_summary.cell(row=row, column=3, value=cat['consumption_qty']).border = thin_border
        ws_summary.cell(row=row, column=4, value=cat['daily_avg_qty']).border = thin_border
        
        par_cell = ws_summary.cell(row=row, column=5, value=cat[f'par_stock_{par_days}_days_qty'])
        par_cell.font = par_font
        par_cell.fill = highlight_fill
        par_cell.border = thin_border
        
        ws_summary.cell(row=row, column=6, value="Mixed").border = thin_border
        row += 1
    
    # Grand total
    ws_summary.cell(row=row, column=1, value="GRAND TOTAL").font = bold_font
    ws_summary.cell(row=row, column=1).fill = subtotal_fill
    ws_summary.cell(row=row, column=3, value=report_data['summary']['total_consumption_qty']).font = bold_font
    ws_summary.cell(row=row, column=3).fill = subtotal_fill
    
    daily_avg_total = report_data['summary']['total_consumption_qty'] / report_data['analysis_period']['days_analyzed']
    ws_summary.cell(row=row, column=4, value=round(daily_avg_total, 2)).font = bold_font
    ws_summary.cell(row=row, column=4).fill = subtotal_fill
    
    par_total = daily_avg_total * par_days
    ws_summary.cell(row=row, column=5, value=round(par_total, 2)).font = par_font
    ws_summary.cell(row=row, column=5).fill = highlight_fill
    
    for col in range(1, 7):
        ws_summary.cell(row=row, column=col).border = thin_border
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 18
    ws_summary.column_dimensions['F'].width = 10
    
    # Sheet 2: Item-wise PAR Stock
    ws_items = wb.create_sheet(title=f"Item PAR Stock ({par_days} days)")
    
    ws_items['A1'] = f"Item-wise PAR Stock - {par_days} Days"
    ws_items['A1'].font = Font(bold=True, size=16, color="2E7D32")
    ws_items.merge_cells('A1:H1')
    
    ws_items['A2'] = f"Based on consumption from {start_date} to {end_date}"
    ws_items['A2'].font = black_font
    
    # Group items by category
    items_by_category = {}
    for item in report_data['items']:
        cat = item['category']
        if cat not in items_by_category:
            items_by_category[cat] = []
        items_by_category[cat].append(item)
    
    row = 4
    
    for category in sorted(items_by_category.keys()):
        items = items_by_category[category]
        
        # Category header
        ws_items.cell(row=row, column=1, value=category.upper()).font = bold_font
        ws_items.cell(row=row, column=1).fill = category_fill
        ws_items.merge_cells(f'A{row}:H{row}')
        for col in range(1, 9):
            ws_items.cell(row=row, column=col).border = thin_border
        row += 1
        
        # Item headers
        item_headers = ["#", "Item Name", "Unit", "Rate (₹)", "Consumed", "Daily Avg", f"{par_days}-Day PAR", "PAR Value (₹)"]
        for col, h in enumerate(item_headers, 1):
            cell = ws_items.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        
        # Items
        cat_total_consumed = 0
        cat_total_par = 0
        cat_total_par_value = 0
        
        for idx, item in enumerate(items, 1):
            ws_items.cell(row=row, column=1, value=idx).border = thin_border
            ws_items.cell(row=row, column=2, value=item['item_name']).border = thin_border
            ws_items.cell(row=row, column=3, value=item['unit']).border = thin_border
            ws_items.cell(row=row, column=4, value=item['rate']).number_format = '₹#,##0.00'
            ws_items.cell(row=row, column=4).border = thin_border
            ws_items.cell(row=row, column=5, value=item['consumption_qty']).border = thin_border
            ws_items.cell(row=row, column=6, value=item['daily_avg_qty']).border = thin_border
            
            par_qty = item[f'par_stock_{par_days}_days_qty']
            par_value = item[f'par_stock_{par_days}_days_value']
            
            par_cell = ws_items.cell(row=row, column=7, value=par_qty)
            par_cell.font = par_font
            par_cell.fill = highlight_fill
            par_cell.border = thin_border
            
            ws_items.cell(row=row, column=8, value=par_value).number_format = '₹#,##0.00'
            ws_items.cell(row=row, column=8).border = thin_border
            
            cat_total_consumed += item['consumption_qty']
            cat_total_par += par_qty
            cat_total_par_value += par_value
            
            row += 1
        
        # Category subtotal
        ws_items.cell(row=row, column=2, value=f"{category} TOTAL").font = bold_font
        ws_items.cell(row=row, column=2).fill = subtotal_fill
        ws_items.cell(row=row, column=5, value=round(cat_total_consumed, 2)).font = bold_font
        ws_items.cell(row=row, column=5).fill = subtotal_fill
        ws_items.cell(row=row, column=7, value=round(cat_total_par, 2)).font = par_font
        ws_items.cell(row=row, column=7).fill = highlight_fill
        ws_items.cell(row=row, column=8, value=round(cat_total_par_value, 2)).font = bold_font
        ws_items.cell(row=row, column=8).number_format = '₹#,##0.00'
        
        for col in range(1, 9):
            ws_items.cell(row=row, column=col).border = thin_border
        
        row += 2  # Gap before next category
    
    # Column widths
    ws_items.column_dimensions['A'].width = 6
    ws_items.column_dimensions['B'].width = 45
    ws_items.column_dimensions['C'].width = 10
    ws_items.column_dimensions['D'].width = 12
    ws_items.column_dimensions['E'].width = 12
    ws_items.column_dimensions['F'].width = 12
    ws_items.column_dimensions['G'].width = 15
    ws_items.column_dimensions['H'].width = 15
    
    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"PAR_Stock_{par_days}days_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )






# Debug endpoint to verify Kitchen GRN data for a specific kitchen and date
@app.get("/api/debug/kitchen-grn-details")
async def debug_kitchen_grn_details(
    kitchen_name: str = Query(..., description="Kitchen name to search"),
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    category: str = Query(None, description="Optional category filter: Non Veg, Vegetables, Dairy")
):
    """Debug endpoint to see detailed breakdown of GRN records for a kitchen"""
    kitchen_receivables = db["kitchen_receivables"]
    
    query = {"receive_date": date}
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    # Find all records for the date
    all_records = list(kitchen_receivables.find(query))
    
    # Filter by kitchen name (case-insensitive partial match)
    matched_records = []
    for rec in all_records:
        rec_kitchen = rec.get("kitchen_name", "")
        if kitchen_name.lower() in rec_kitchen.lower() or rec_kitchen.lower() in kitchen_name.lower():
            matched_records.append({
                "id": str(rec.get("_id", "")),
                "kitchen_name": rec_kitchen,
                "category": rec.get("category", ""),
                "item_name": rec.get("item_name", ""),
                "amount": float(rec.get("amount", 0) or 0),
                "quantity": rec.get("quantity", 0),
                "rate": rec.get("rate", 0),
                "receive_date": rec.get("receive_date", ""),
                "grn_date": rec.get("grn_date", ""),
                "po_date": rec.get("po_date", ""),
                "vendor_name": rec.get("vendor_name", ""),
                "created_at": rec.get("created_at", "")
            })
    
    # Calculate totals by category
    totals_by_category = {}
    for rec in matched_records:
        cat = rec["category"]
        if cat not in totals_by_category:
            totals_by_category[cat] = {"count": 0, "total_amount": 0}
        totals_by_category[cat]["count"] += 1
        totals_by_category[cat]["total_amount"] += rec["amount"]
    
    return {
        "kitchen_search": kitchen_name,
        "date": date,
        "category_filter": category,
        "total_records_found": len(matched_records),
        "totals_by_category": totals_by_category,
        "grand_total": sum(rec["amount"] for rec in matched_records),
        "records": matched_records
    }



@app.post("/api/admin/update-par-stock-from-consumption")
async def api_update_par_stock_from_consumption(
    start_date: str = Query(None, description="Start date of consumption period (YYYY-MM-DD). If not provided, uses previous month."),
    end_date: str = Query(None, description="End date of consumption period (YYYY-MM-DD). If not provided, uses previous month."),
    par_days: int = Query(10, description="Number of days for PAR stock"),
    dry_run: bool = Query(True, description="If true, only preview changes without updating"),
    current_user = Depends(require_role(["admin"]))
):
    """
    Calculate and update PAR stock for all items based on consumption data.
    PAR Stock = (Total Consumed / Days in Period) * PAR Days
    
    Uses TRANSACTIONS (type='issue') - items actually issued from Main Store to kitchens.
    This matches the Kitchen GRN Summary data source.
    """
    from collections import defaultdict
    from calendar import monthrange
    
    # If dates not provided, use previous month
    if not start_date or not end_date:
        today = datetime.now(timezone.utc)
        first_day_current_month = today.replace(day=1)
        last_day_prev_month = first_day_current_month - timedelta(days=1)
        first_day_prev_month = last_day_prev_month.replace(day=1)
        start_date = first_day_prev_month.strftime("%Y-%m-%d")
        end_date = last_day_prev_month.strftime("%Y-%m-%d")
    
    # Parse dates
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    days_in_period = (end_dt - start_dt).days + 1
    
    # Get all OUT transactions in date range (dispatch, issue, transfer)
    # This matches the Stock Movement report
    txn_query = {
        "type": {"$in": ["dispatch", "issue", "transfer"]},
        "created_at": {
            "$gte": start_date + "T00:00:00",
            "$lte": end_date + "T23:59:59"
        }
    }
    
    # Aggregate consumption by item
    item_consumption = defaultdict(lambda: {
        "item_id": "",
        "item_name": "",
        "category": "",
        "unit": "",
        "total_qty": 0,
        "total_value": 0
    })
    
    total_transactions = 0
    for txn in transactions_collection.find(txn_query):
        lot_id = txn.get("lot_id")
        qty = float(txn.get("quantity", 0) or 0)
        value = float(txn.get("value", 0) or 0)
        
        if not lot_id or qty <= 0:
            continue
        
        total_transactions += 1
        
        # Get lot info
        lot = None
        try:
            lot = lots_collection.find_one({"_id": ObjectId(lot_id)})
        except:
            pass
        
        if not lot:
            continue
        
        item_id = lot.get("item_id", "")
        item_name = lot.get("item_name", "Unknown")
        
        # Get item details
        db_item = None
        if item_id:
            try:
                db_item = items_collection.find_one({"_id": ObjectId(item_id)})
            except:
                pass
        
        category = db_item.get("category", "Uncategorized") if db_item else "Uncategorized"
        unit = db_item.get("unit", lot.get("unit", "")) if db_item else lot.get("unit", "")
        
        key = item_id or item_name
        item_consumption[key]["item_id"] = item_id
        item_consumption[key]["item_name"] = db_item.get("name", item_name) if db_item else item_name
        item_consumption[key]["category"] = category
        item_consumption[key]["unit"] = unit
        item_consumption[key]["total_qty"] += qty
        item_consumption[key]["total_value"] += value
    
    # Calculate PAR stock and prepare updates
    updates = []
    updated_count = 0
    errors = []
    
    for key, data in item_consumption.items():
        item_id = data["item_id"]
        if not item_id:
            continue
        
        # Calculate daily average and PAR stock
        daily_avg = data["total_qty"] / days_in_period if days_in_period > 0 else 0
        par_stock = round(daily_avg * par_days, 2)
        
        # Get current item from database
        try:
            db_item = items_collection.find_one({"_id": ObjectId(item_id)})
        except:
            db_item = None
        
        if not db_item:
            errors.append(f"Item not found: {data['item_name']} (ID: {item_id})")
            continue
        
        current_par = db_item.get("par_stock", 0) or 0
        
        update_info = {
            "item_id": item_id,
            "item_name": data["item_name"],
            "category": data["category"],
            "unit": data["unit"],
            "consumed_qty": round(data["total_qty"], 2),
            "consumed_value": round(data["total_value"], 2),
            "daily_avg": round(daily_avg, 2),
            "new_par_stock": par_stock,
            "old_par_stock": current_par,
            "changed": par_stock != current_par
        }
        updates.append(update_info)
        
        # Update database if not dry run
        if not dry_run and par_stock > 0:
            items_collection.update_one(
                {"_id": ObjectId(item_id)},
                {
                    "$set": {
                        "par_stock": par_stock,
                        "par_stock_days": par_days,
                        "par_stock_updated_at": datetime.now(timezone.utc).isoformat(),
                        "par_stock_based_on": f"{start_date} to {end_date}",
                        "par_stock_updated_by": current_user.get("email", "admin")
                    }
                }
            )
            updated_count += 1
    
    # Save history if not dry run
    if not dry_run and updated_count > 0:
        history_record = {
            "run_date": datetime.now(timezone.utc).isoformat(),
            "triggered_by": current_user.get("email", "admin"),
            "period_start": start_date,
            "period_end": end_date,
            "days_analyzed": days_in_period,
            "par_days": par_days,
            "items_updated": updated_count,
            "update_details": updates[:100]  # Store first 100 for reference
        }
        par_stock_history_collection.insert_one(history_record)
    
    # Sort by category then by consumed qty
    updates.sort(key=lambda x: (x["category"], -x["consumed_qty"]))
    
    # Group by category for summary
    category_summary = defaultdict(lambda: {"items": 0, "total_consumed": 0, "total_value": 0, "total_par": 0})
    for u in updates:
        cat = u["category"]
        category_summary[cat]["items"] += 1
        category_summary[cat]["total_consumed"] += u["consumed_qty"]
        category_summary[cat]["total_value"] += u.get("consumed_value", 0)
        category_summary[cat]["total_par"] += u["new_par_stock"]
    
    return {
        "success": True,
        "dry_run": dry_run,
        "data_source": "transactions (type=dispatch/issue/transfer) - Items leaving Main Store",
        "message": f"{'Preview of ' if dry_run else ''}PAR stock updates based on {days_in_period} days of consumption",
        "period": {
            "start_date": start_date,
            "end_date": end_date,
            "days": days_in_period,
            "par_days": par_days
        },
        "summary": {
            "total_transactions": total_transactions,
            "items_with_consumption": len(item_consumption),
            "items_to_update": len(updates),
            "items_actually_updated": updated_count if not dry_run else 0
        },
        "category_summary": [
            {
                "category": cat,
                "items": data["items"],
                "total_consumed": round(data["total_consumed"], 2),
                "total_value": round(data["total_value"], 2),
                "total_par_stock": round(data["total_par"], 2)
            }
            for cat, data in sorted(category_summary.items(), key=lambda x: x[1]["total_consumed"], reverse=True)
        ],
        "updates": updates[:200],
        "errors": errors
    }


@app.get("/api/admin/par-stock-history")
async def get_par_stock_history(
    limit: int = Query(10, description="Number of records to return"),
    current_user = Depends(require_role(["admin"]))
):
    """Get history of PAR stock updates"""
    history = list(par_stock_history_collection.find().sort("run_date", -1).limit(limit))
    
    return [{
        "id": str(h["_id"]),
        "run_date": h.get("run_date"),
        "triggered_by": h.get("triggered_by"),
        "period": f"{h.get('period_start')} to {h.get('period_end')}",
        "days_analyzed": h.get("days_analyzed"),
        "par_days": h.get("par_days", 10),
        "items_updated": h.get("items_updated"),
        "update_count": len(h.get("update_details", []))
    } for h in history]


@app.post("/api/admin/trigger-monthly-par-update")
async def trigger_monthly_par_update(
    current_user = Depends(require_role(["admin"]))
):
    """
    Manually trigger the monthly PAR stock update.
    Uses previous month's consumption data.
    """
    result = update_par_stock_from_consumption(triggered_by=current_user.get("email", "admin_manual"))
    return result


# ==================== AUTO PO GENERATION ====================

@app.get("/api/auto-po/suggestions")
async def get_auto_po_suggestions(
    days_forecast: int = Query(3, description="Days to forecast stock depletion"),
    par_days: int = Query(10, description="PAR stock days to maintain"),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Generate Auto PO suggestions based on:
    1. Items that will go below PAR stock in next N days (based on daily consumption)
    2. Group by vendor
    3. Calculate quantities needed to reach PAR stock
    
    Logic: If (Current Stock - (Daily Consumption * days_forecast)) < PAR Stock
           Then suggest ordering: PAR Stock - Current Stock + buffer
    """
    from collections import defaultdict
    
    # Get all items with PAR stock and vendor
    items_with_par = list(items_collection.find({
        "par_stock": {"$exists": True, "$gt": 0}
    }))
    
    # Get current stock for all items (aggregate from lots at main store)
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Aggregate current stock by item
    stock_pipeline = [
        {"$match": {"location_id": main_store_id, "current_quantity": {"$gt": 0}}},
        {"$group": {
            "_id": "$item_id",
            "current_stock": {"$sum": "$current_quantity"}
        }}
    ]
    stock_by_item = {str(s["_id"]): s["current_stock"] for s in lots_collection.aggregate(stock_pipeline)}
    
    # Group suggestions by vendor
    vendor_suggestions = defaultdict(lambda: {
        "vendor_id": "",
        "vendor_name": "",
        "items": [],
        "total_value": 0,
        "total_items": 0
    })
    
    items_without_vendor = []
    items_needing_reorder = []
    
    for item in items_with_par:
        item_id = str(item["_id"])
        item_name = item.get("name", "Unknown")
        category = item.get("category", "Uncategorized")
        unit = item.get("unit", "")
        par_stock = float(item.get("par_stock", 0) or 0)
        daily_consumption = par_stock / par_days if par_days > 0 else 0  # Estimated daily consumption
        current_stock = stock_by_item.get(item_id, 0)
        price = float(item.get("standard_price", 0) or 0)
        
        # Calculate projected stock after N days
        projected_stock = current_stock - (daily_consumption * days_forecast)
        
        # Check if item will go below PAR
        if projected_stock < par_stock:
            # Calculate order quantity (enough to reach PAR + buffer for lead time)
            order_qty = round(par_stock - current_stock + (daily_consumption * days_forecast), 2)
            if order_qty <= 0:
                order_qty = round(par_stock - current_stock, 2)
            if order_qty <= 0:
                continue
            
            order_value = round(order_qty * price, 2)
            
            item_suggestion = {
                "item_id": item_id,
                "item_name": item_name,
                "category": category,
                "unit": unit,
                "current_stock": round(current_stock, 2),
                "par_stock": par_stock,
                "daily_consumption": round(daily_consumption, 2),
                "projected_stock_after_days": round(projected_stock, 2),
                "suggested_order_qty": order_qty,
                "price": price,
                "order_value": order_value,
                "urgency": "high" if projected_stock <= 0 else "medium" if projected_stock < par_stock * 0.5 else "low"
            }
            
            items_needing_reorder.append(item_suggestion)
            
            # Get vendor for this item
            vendor_id = item.get("vendor_id")
            if vendor_id:
                vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id) if isinstance(vendor_id, str) else vendor_id})
                vendor_name = vendor.get("name", "Unknown") if vendor else "Unknown"
                
                vendor_suggestions[str(vendor_id)]["vendor_id"] = str(vendor_id)
                vendor_suggestions[str(vendor_id)]["vendor_name"] = vendor_name
                vendor_suggestions[str(vendor_id)]["items"].append(item_suggestion)
                vendor_suggestions[str(vendor_id)]["total_value"] += order_value
                vendor_suggestions[str(vendor_id)]["total_items"] += 1
            else:
                items_without_vendor.append(item_suggestion)
    
    # Sort vendors by total value
    sorted_vendors = sorted(
        vendor_suggestions.values(),
        key=lambda x: x["total_value"],
        reverse=True
    )
    
    # Sort items within each vendor by urgency then value
    urgency_order = {"high": 0, "medium": 1, "low": 2}
    for vendor in sorted_vendors:
        vendor["items"].sort(key=lambda x: (urgency_order.get(x["urgency"], 3), -x["order_value"]))
    
    return {
        "success": True,
        "forecast_days": days_forecast,
        "par_days": par_days,
        "summary": {
            "total_items_with_par": len(items_with_par),
            "items_needing_reorder": len(items_needing_reorder),
            "items_with_vendor": len(items_needing_reorder) - len(items_without_vendor),
            "items_without_vendor": len(items_without_vendor),
            "vendors_count": len(sorted_vendors),
            "total_suggested_value": round(sum(v["total_value"] for v in sorted_vendors), 2)
        },
        "vendor_suggestions": sorted_vendors,
        "items_without_vendor": items_without_vendor[:50]  # Limit to 50
    }


@app.post("/api/auto-po/create-po")
async def create_po_from_suggestion(
    vendor_id: str = Query(..., description="Vendor ID"),
    item_ids: str = Query(..., description="Comma-separated item IDs to include"),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Create an actual PO from auto-suggestion.
    """
    # Get vendor
    try:
        vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
    except:
        vendor = None
    
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Parse item IDs
    item_id_list = [id.strip() for id in item_ids.split(",") if id.strip()]
    
    if not item_id_list:
        raise HTTPException(status_code=400, detail="No items provided")
    
    # Get suggestions for these items
    suggestions_response = await get_auto_po_suggestions(3, 10, current_user)
    
    # Find the vendor's items from suggestions
    vendor_data = None
    for v in suggestions_response["vendor_suggestions"]:
        if v["vendor_id"] == vendor_id:
            vendor_data = v
            break
    
    if not vendor_data:
        raise HTTPException(status_code=404, detail="No suggestions found for this vendor")
    
    # Build PO items from selected items
    po_items = []
    total_amount = 0
    
    for item_suggestion in vendor_data["items"]:
        if item_suggestion["item_id"] in item_id_list:
            po_items.append({
                "item_id": item_suggestion["item_id"],
                "item_name": item_suggestion["item_name"],
                "category": item_suggestion["category"],
                "unit": item_suggestion["unit"],
                "quantity": item_suggestion["suggested_order_qty"],
                "rate": item_suggestion["price"],
                "amount": item_suggestion["order_value"],
                "received_qty": 0,
                "status": "pending"
            })
            total_amount += item_suggestion["order_value"]
    
    if not po_items:
        raise HTTPException(status_code=400, detail="No matching items found")
    
    # Generate PO number
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    last_po = purchase_orders_collection.find_one(
        {"po_number": {"$regex": f"^PO-{today}"}},
        sort=[("po_number", -1)]
    )
    if last_po:
        last_num = int(last_po["po_number"].split("-")[-1])
        po_number = f"PO-{today}-{str(last_num + 1).zfill(4)}"
    else:
        po_number = f"PO-{today}-0001"
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    
    # Create PO document
    po_doc = {
        "po_number": po_number,
        "vendor_id": vendor_id,
        "vendor_name": vendor.get("name", "Unknown"),
        "items": po_items,
        "total_amount": round(total_amount, 2),
        "status": "pending",
        "created_by": current_user.get("user_id"),
        "created_by_email": current_user.get("email"),
        "created_by_location_id": str(main_store["_id"]) if main_store else None,
        "created_by_location_name": main_store.get("name", "Main Store") if main_store else "Main Store",
        "created_at": datetime.now(timezone.utc),
        "source": "auto_po",
        "notes": f"Auto-generated PO based on PAR stock analysis"
    }
    
    result = purchase_orders_collection.insert_one(po_doc)
    
    return {
        "success": True,
        "message": f"PO {po_number} created successfully",
        "po_id": str(result.inserted_id),
        "po_number": po_number,
        "vendor_name": vendor.get("name"),
        "items_count": len(po_items),
        "total_amount": round(total_amount, 2)
    }


@app.get("/api/auto-po/items-without-vendor")
async def get_items_without_vendor(
    category: Optional[str] = None,
    current_user = Depends(require_role(["admin"]))
):
    """
    Get items that don't have a vendor assigned.
    """
    query = {"$or": [{"vendor_id": {"$exists": False}}, {"vendor_id": None}, {"vendor_id": ""}]}
    if category:
        query["category"] = category
    
    items = list(items_collection.find(query).sort("category", 1))
    
    # Get all vendors for assignment dropdown
    vendors = list(vendors_collection.find({}).sort("name", 1))
    
    return {
        "success": True,
        "total_items": len(items),
        "items": [
            {
                "item_id": str(item["_id"]),
                "name": item.get("name"),
                "category": item.get("category"),
                "unit": item.get("unit"),
                "standard_price": item.get("standard_price", 0)
            }
            for item in items
        ],
        "vendors": [
            {"id": str(v["_id"]), "name": v.get("name")}
            for v in vendors
        ]
    }


@app.post("/api/auto-po/assign-vendor")
async def assign_vendor_to_items(
    vendor_id: str = Query(..., description="Vendor ID to assign"),
    item_ids: str = Query(..., description="Comma-separated item IDs"),
    current_user = Depends(require_role(["admin"]))
):
    """
    Assign a vendor to multiple items.
    """
    item_id_list = [id.strip() for id in item_ids.split(",") if id.strip()]
    
    if not item_id_list:
        raise HTTPException(status_code=400, detail="No items provided")
    
    # Verify vendor exists
    try:
        vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
    except:
        vendor = None
    
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Update items
    updated_count = 0
    for item_id in item_id_list:
        try:
            result = items_collection.update_one(
                {"_id": ObjectId(item_id)},
                {"$set": {"vendor_id": vendor_id, "vendor_name": vendor.get("name")}}
            )
            if result.modified_count > 0:
                updated_count += 1
        except:
            pass
    
    return {
        "success": True,
        "message": f"Assigned {updated_count} items to vendor {vendor.get('name')}",
        "updated_count": updated_count
    }


# Debug endpoint to analyze Kitchen Ledger data discrepancy
@app.get("/api/debug/kitchen-ledger-analysis")
async def debug_kitchen_ledger_analysis(
    kitchen_name: str = Query(..., description="Kitchen name to search")
):
    """Debug endpoint to analyze why Kitchen Ledger amounts are low"""
    
    # Find the kitchen
    kitchen = locations_collection.find_one({"name": {"$regex": kitchen_name, "$options": "i"}, "type": "kitchen"})
    if not kitchen:
        return {"error": f"Kitchen '{kitchen_name}' not found"}
    
    k_id = str(kitchen["_id"])
    k_name = kitchen["name"]
    
    analysis = {
        "kitchen_id": k_id,
        "kitchen_name": k_name,
        "data_sources": {}
    }
    
    # 1. Check transactions collection - issue type
    issue_txns = list(transactions_collection.find({
        "type": "issue",
        "$or": [
            {"destination_location_id": k_id},
            {"kitchen_id": k_id},
            {"kitchen_name": {"$regex": k_name, "$options": "i"}}
        ]
    }).limit(100))
    
    issue_total = 0
    for txn in issue_txns:
        value = txn.get("value") or 0
        if not value:
            qty = txn.get("quantity", 0) or 0
            rate = txn.get("rate", 0) or 0
            value = qty * rate
        issue_total += value
    
    analysis["data_sources"]["transactions_issue"] = {
        "count": len(issue_txns),
        "total_value": issue_total,
        "sample_fields": list(issue_txns[0].keys()) if issue_txns else [],
        "sample_record": {k: str(v) for k, v in (issue_txns[0].items() if issue_txns else {})} if issue_txns else None
    }
    
    # 2. Check transactions collection - daily_perishable type
    perishable_txns = list(transactions_collection.find({
        "type": "daily_perishable",
        "$or": [
            {"destination_location_id": k_id},
            {"kitchen_id": k_id},
            {"kitchen_name": {"$regex": k_name, "$options": "i"}}
        ]
    }).limit(100))
    
    perishable_total = 0
    for txn in perishable_txns:
        value = txn.get("value") or txn.get("amount") or 0
        perishable_total += value
    
    analysis["data_sources"]["transactions_daily_perishable"] = {
        "count": len(perishable_txns),
        "total_value": perishable_total,
        "sample_fields": list(perishable_txns[0].keys()) if perishable_txns else []
    }
    
    # 3. Check kitchen_receivables collection
    kr_data = list(db["kitchen_receivables"].find({
        "kitchen_name": {"$regex": k_name, "$options": "i"}
    }).limit(100))
    
    kr_total = 0
    for rec in kr_data:
        kr_total += float(rec.get("amount", 0) or 0)
    
    analysis["data_sources"]["kitchen_receivables"] = {
        "count": len(kr_data),
        "total_value": kr_total,
        "sample_fields": list(kr_data[0].keys()) if kr_data else [],
        "sample_record": {k: str(v) for k, v in (kr_data[0].items() if kr_data else {})} if kr_data else None
    }
    
    # 4. Check requisitions for this kitchen
    reqs = list(requisitions_collection.find({
        "$or": [
            {"kitchen_id": k_id},
            {"location_name": k_name},
            {"kitchen_name": k_name}
        ]
    }).limit(50))
    
    req_total = 0
    for req in reqs:
        for item in req.get("items", []):
            qty = item.get("quantity_sent", 0) or item.get("quantity", 0) or 0
            rate = item.get("rate", 0) or 0
            req_total += qty * rate
    
    analysis["data_sources"]["requisitions"] = {
        "count": len(reqs),
        "total_value": req_total,
        "sample_fields": list(reqs[0].keys()) if reqs else []
    }
    
    # 5. Check all transaction types in the system
    all_txn_types = list(transactions_collection.aggregate([
        {"$group": {"_id": "$type", "count": {"$sum": 1}}}
    ]))
    analysis["all_transaction_types"] = {t["_id"]: t["count"] for t in all_txn_types}
    
    # 6. Check destination_location_id formats
    sample_destinations = list(transactions_collection.aggregate([
        {"$match": {"type": "issue"}},
        {"$group": {"_id": "$destination_location_id", "count": {"$sum": 1}}},
        {"$limit": 10}
    ]))
    analysis["sample_destination_ids"] = [{"id": str(d["_id"]), "count": d["count"]} for d in sample_destinations]
    
    # Recommendation
    analysis["recommendation"] = (
        "Compare the transaction types and data sources. "
        "If transactions_issue shows 0 records but kitchen_receivables has data, "
        "we need to include kitchen_receivables in the Kitchen Ledger calculation."
    )
    
    return analysis

# Debug endpoint to analyze Vendor Ledger data discrepancy
@app.get("/api/debug/vendor-ledger-analysis")
async def debug_vendor_ledger_analysis(
    vendor_name: str = Query(..., description="Vendor name to search")
):
    """Debug endpoint to analyze why Vendor Ledger amounts don't match"""
    
    # Find the vendor
    vendor = vendors_collection.find_one({"name": {"$regex": vendor_name, "$options": "i"}})
    if not vendor:
        return {"error": f"Vendor '{vendor_name}' not found"}
    
    v_id = str(vendor["_id"])
    v_name = vendor["name"]
    
    analysis = {
        "vendor_id": v_id,
        "vendor_name": v_name,
        "data_sources": {}
    }
    
    # 1. Check POs
    pos = list(purchase_orders_collection.find({"vendor_id": v_id}))
    po_total = 0
    for po in pos:
        for item in po.get("items", []):
            po_total += item.get("quantity", 0) * item.get("rate", 0)
    
    analysis["data_sources"]["purchase_orders"] = {
        "count": len(pos),
        "total_value": po_total
    }
    
    # 2. Check lots with this vendor
    lots = list(lots_collection.find({"vendor_id": v_id}))
    lots_total = 0
    for lot in lots:
        qty = lot.get("initial_quantity", 0) or lot.get("quantity", 0) or 0
        rate = lot.get("purchase_rate", 0) or 0
        lots_total += qty * rate
    
    analysis["data_sources"]["lots_with_vendor"] = {
        "count": len(lots),
        "total_value": lots_total
    }
    
    # 3. Check GRN transactions
    grn_txns = list(transactions_collection.find({"type": "grn", "vendor_id": v_id}))
    grn_total = 0
    for txn in grn_txns:
        grn_total += txn.get("value", 0) or (txn.get("quantity", 0) * txn.get("rate", 0))
    
    analysis["data_sources"]["grn_transactions"] = {
        "count": len(grn_txns),
        "total_value": grn_total
    }
    
    # 4. Check lots WITHOUT vendor_id but with po_reference_id
    po_ids = [str(po["_id"]) for po in pos]
    lots_via_po = list(lots_collection.find({
        "po_reference_id": {"$in": po_ids},
        "$or": [{"vendor_id": {"$exists": False}}, {"vendor_id": None}, {"vendor_id": ""}]
    }))
    lots_via_po_total = 0
    for lot in lots_via_po:
        qty = lot.get("initial_quantity", 0) or lot.get("quantity", 0) or 0
        rate = lot.get("purchase_rate", 0) or 0
        lots_via_po_total += qty * rate
    
    analysis["data_sources"]["lots_linked_via_po"] = {
        "count": len(lots_via_po),
        "total_value": lots_via_po_total,
        "note": "These lots don't have vendor_id but are linked via PO"
    }
    
    return analysis

# Debug endpoint to analyze Stock discrepancy
@app.get("/api/debug/stock-analysis")
async def debug_stock_analysis():
    """Debug endpoint to analyze why Stock values are too high"""
    
    # Get main store
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        return {"error": "Main Store not found"}
    
    main_store_id = str(main_store["_id"])
    
    analysis = {
        "main_store_id": main_store_id,
        "data": {}
    }
    
    # 1. Total current stock
    current_stock = list(lots_collection.find({
        "location_id": main_store_id,
        "current_quantity": {"$gt": 0}
    }))
    
    stock_value = 0
    for lot in current_stock:
        qty = lot.get("current_quantity", 0) or 0
        rate = lot.get("purchase_rate", 0) or 0
        stock_value += qty * rate
    
    analysis["data"]["current_stock"] = {
        "lot_count": len(current_stock),
        "total_value": stock_value
    }
    
    # 2. Total purchased (all lots initial quantity)
    all_lots = list(lots_collection.find({"location_id": main_store_id}))
    purchased_value = 0


@app.get("/api/debug/packaging-stock")
async def debug_packaging_stock():
    """Debug endpoint to analyze packaging stock values"""
    
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Get all packaging items
    packaging_items = {str(item["_id"]): item for item in items_collection.find({"category": "Packaging"})}
    
    # Get lots for packaging items
    high_value_lots = []
    total_value = 0
    
    for lot in lots_collection.find({"location_id": main_store_id, "current_quantity": {"$gt": 0}}):
        item_id = str(lot.get("item_id", ""))
        if item_id in packaging_items:
            qty = lot.get("current_quantity", 0) or 0
            rate = lot.get("purchase_rate", 0) or 0
            value = qty * rate
            total_value += value
            
            item = packaging_items[item_id]
            high_value_lots.append({
                "item_name": item.get("name", "Unknown"),
                "quantity": qty,
                "rate": rate,
                "value": round(value, 2),
                "standard_price": item.get("standard_price", 0) or 0
            })
    
    # Sort by value
    high_value_lots.sort(key=lambda x: x["value"], reverse=True)
    
    return {
        "total_packaging_items": len(packaging_items),
        "total_stock_value": round(total_value, 2),
        "top_items_by_value": high_value_lots[:30],
        "items_with_rate_mismatch": [
            lot for lot in high_value_lots 
            if lot["standard_price"] > 0 and abs(lot["rate"] - lot["standard_price"]) > 1
        ][:20]
    }


# Debug endpoint to find items with zero or abnormal prices
@app.get("/api/debug/price-analysis")
async def debug_price_analysis():
    """Analyze items and lots to find pricing issues"""
    
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    analysis = {
        "items_analysis": {},
        "lots_analysis": {},
        "stock_calculation": {}
    }
    
    # 1. Check items with zero standard_price
    all_items = list(items_collection.find({}))
    zero_price_items = [i for i in all_items if not i.get("standard_price")]
    analysis["items_analysis"] = {
        "total_items": len(all_items),
        "items_with_zero_price": len(zero_price_items),
        "sample_zero_price": [{"name": i["name"], "category": i.get("category", "")} for i in zero_price_items[:10]]
    }
    
    # 2. Check lots with pricing issues
    all_lots = list(lots_collection.find({"location_id": main_store_id, "current_quantity": {"$gt": 0}}))
    
    zero_rate_lots = []
    abnormal_rate_lots = []
    normal_lots = []
    
    for lot in all_lots:
        rate = lot.get("purchase_rate", 0) or 0
        qty = lot.get("current_quantity", 0) or 0
        item_name = lot.get("item_name", "Unknown")
        
        if rate == 0:
            zero_rate_lots.append({
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "category": lot.get("category", "")
            })
        elif rate > 10000:  # Abnormally high rate
            abnormal_rate_lots.append({
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "value": qty * rate,
                "category": lot.get("category", "")
            })
        else:
            normal_lots.append({
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "value": qty * rate
            })
    
    analysis["lots_analysis"] = {
        "total_lots_in_stock": len(all_lots),
        "zero_rate_lots": {
            "count": len(zero_rate_lots),
            "total_qty": sum(l["qty"] for l in zero_rate_lots),
            "samples": zero_rate_lots[:10]
        },
        "abnormal_rate_lots": {
            "count": len(abnormal_rate_lots),
            "total_value": sum(l["value"] for l in abnormal_rate_lots),
            "samples": sorted(abnormal_rate_lots, key=lambda x: x["value"], reverse=True)[:10]
        },
        "normal_lots": {
            "count": len(normal_lots),
            "total_value": sum(l["value"] for l in normal_lots)
        }
    }
    
    # 3. Recalculate stock value properly
    correct_total = 0
    category_breakdown = {}
    
    for lot in all_lots:
        rate = lot.get("purchase_rate", 0) or 0
        qty = lot.get("current_quantity", 0) or 0
        value = qty * rate
        correct_total += value
        
        cat = lot.get("category", "Uncategorized")
        if cat not in category_breakdown:
            category_breakdown[cat] = {"qty": 0, "value": 0, "lot_count": 0}
        category_breakdown[cat]["qty"] += qty
        category_breakdown[cat]["value"] += value
        category_breakdown[cat]["lot_count"] += 1
    
    analysis["stock_calculation"] = {
        "correct_total_value": round(correct_total, 2),
        "category_breakdown": category_breakdown
    }
    
    return analysis

# Debug endpoint to check kitchen receivables data for a kitchen
@app.get("/api/debug/kitchen-receivables-check")
async def debug_kitchen_receivables_check(
    kitchen_name: str = Query(..., description="Kitchen name")
):
    """Check kitchen receivables data for anomalies"""
    
    kitchen_receivables = db["kitchen_receivables"]
    
    # Find all records for this kitchen
    records = list(kitchen_receivables.find({
        "kitchen_name": {"$regex": kitchen_name, "$options": "i"}
    }))
    
    # Group by category
    category_data = {}
    for rec in records:
        cat = rec.get("category", "Unknown")
        if cat not in category_data:
            category_data[cat] = {
                "count": 0,
                "total_qty": 0,
                "total_amount": 0,
                "samples": []
            }
        
        qty = float(rec.get("quantity", 0) or 0)
        amount = float(rec.get("amount", 0) or 0)
        
        category_data[cat]["count"] += 1
        category_data[cat]["total_qty"] += qty
        category_data[cat]["total_amount"] += amount
        
        if len(category_data[cat]["samples"]) < 5:
            category_data[cat]["samples"].append({
                "item_name": rec.get("item_name", ""),
                "qty": qty,
                "rate": rec.get("rate", 0),
                "amount": amount,
                "date": rec.get("receive_date", ""),
                "vendor": rec.get("vendor_name", "")
            })
    
    # Check for anomalies (very high amounts per record)
    anomalies = []
    for rec in records:
        amount = float(rec.get("amount", 0) or 0)
        if amount > 50000:  # Flag records over 50k
            anomalies.append({
                "item_name": rec.get("item_name", ""),
                "category": rec.get("category", ""),
                "qty": rec.get("quantity", 0),
                "rate": rec.get("rate", 0),
                "amount": amount,
                "date": rec.get("receive_date", ""),
                "vendor": rec.get("vendor_name", "")
            })
    
    return {
        "kitchen_name": kitchen_name,
        "total_records": len(records),
        "category_breakdown": category_data,
        "high_value_anomalies": sorted(anomalies, key=lambda x: x["amount"], reverse=True)[:20],
        "grand_total": sum(cat["total_amount"] for cat in category_data.values())
    }


# Reports
@app.get("/api/reports/stock-on-hand")
async def get_stock_on_hand(
    location_id: Optional[str] = None,
    category: Optional[str] = None
):
    query = {"current_quantity": {"$gt": 0}}
    if location_id:
        query["location_id"] = location_id
    
    lots = list(lots_collection.find(query).sort("expiry_date", 1))
    
    result = []
    for lot in lots:
        serialized = serialize_lot(lot)
        if category and serialized["category"] != category:
            continue
        result.append(serialized)
    
    by_location = {}
    for item in result:
        loc = item["location_name"]
        if loc not in by_location:
            by_location[loc] = []
        by_location[loc].append(item)
    
    total_items = len(result)
    total_value = sum(
        (item.get("purchase_rate") or 0) * item["current_quantity"]
        for item in result
    )
    
    return {
        "items": result,
        "by_location": by_location,
        "summary": {
            "total_lots": total_items,
            "total_value": total_value
        }
    }

@app.get("/api/reports/movement-ledger")
async def get_movement_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    type: Optional[str] = None,
    item_id: Optional[str] = None
):
    query = {}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if type:
        query["type"] = type
    
    transactions = list(transactions_collection.find(query).sort("created_at", -1))
    
    result = []
    for txn in transactions:
        serialized = serialize_transaction(txn)
        
        if item_id:
            lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
            if lot and lot["item_id"] != item_id:
                continue
        
        result.append(serialized)
    
    summary = {"grn": 0, "issue": 0, "transfer": 0, "waste": 0}
    for txn in result:
        if txn["type"] in summary:
            summary[txn["type"]] += 1
    
    return {
        "transactions": result,
        "summary": summary
    }

@app.get("/api/reports/wastage")
async def get_wastage_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {"type": "waste"}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    transactions = list(transactions_collection.find(query).sort("created_at", -1))
    
    result = []
    total_waste_value = 0
    
    for txn in transactions:
        serialized = serialize_transaction(txn)
        lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
        
        waste_value = 0
        if lot and lot.get("purchase_rate"):
            waste_value = lot["purchase_rate"] * txn["quantity"]
        
        serialized["waste_value"] = waste_value
        total_waste_value += waste_value
        result.append(serialized)
    
    return {
        "transactions": result,
        "summary": {
            "total_waste_entries": len(result),
            "total_waste_value": total_waste_value
        }
    }

# ============ Vendor Ledger Report ============

@app.get("/api/reports/vendor-ledger")
async def get_vendor_ledger(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    fast: bool = Query(True, description="Fast mode - skip detailed entries for speed")
):
    """
    Vendor-wise ledger with actual financial values
    Uses MongoDB aggregation with indexes for performance
    """
    import time
    start_time = time.time()
    
    try:
        # Get all vendors
        vendors_list = list(vendors_collection.find({}, {"_id": 1, "name": 1}))
        vendors_map = {str(v["_id"]): v.get("name", "Unknown") for v in vendors_list}
        
        # Build date filter for queries
        # IMPORTANT: Both POs and lots store created_at as ISO STRING, not datetime
        date_filter_po = {}
        date_filter_lots = {}
        
        if start_date:
            # Use ISO string format for comparison (strings compare lexicographically)
            date_filter_po["created_at"] = {"$gte": start_date}
            date_filter_lots["created_at"] = {"$gte": start_date}  # String comparison works for ISO dates
                
        if end_date:
            end_date_str = end_date + "T23:59:59"
            if "created_at" in date_filter_po:
                date_filter_po["created_at"]["$lte"] = end_date_str
            else:
                date_filter_po["created_at"] = {"$lte": end_date_str}
            if "created_at" in date_filter_lots:
                date_filter_lots["created_at"]["$lte"] = end_date_str
            else:
                date_filter_lots["created_at"] = {"$lte": end_date_str}
        
        # Debug logging removed for production
        
        # ===== SIMPLIFIED PO/GRN AGGREGATION =====
        # Query vendors INDIVIDUALLY to avoid timeout on large collections
        
        main_store_pos = {}
        kitchen_pos = {}
        main_store_grns = {}
        debug_info = {"method": "individual_vendor_queries", "filter": "main_store_only"}
        
        print(f"Starting PO/GRN aggregation - MAIN STORE ONLY (date range: {start_date} to {end_date})")
        
        try:
            # Get total count for debug (fast estimate)
            total_pos = purchase_orders_collection.estimated_document_count()
            debug_info["estimated_total_pos"] = total_pos
            print(f"Estimated total POs: {total_pos}")
            
            # Query each vendor individually
            vendor_ids = list(vendors_map.keys())
            debug_info["vendors_to_query"] = len(vendor_ids)
            
            success_count = 0
            error_count = 0
            
            # Build date filter once
            date_match = {}
            if start_date and end_date:
                # Use $regex to match date prefix for faster string comparison
                date_match["created_at"] = {
                    "$gte": start_date + "T00:00:00",
                    "$lte": end_date + "T23:59:59"
                }
            
            for vid in vendor_ids:
                try:
                    # Build query for this vendor - MAIN STORE POs ONLY
                    # Kitchen POs are tracked separately via Daily Perishables (kitchen_receivables)
                    query = {
                        "vendor_id": vid,
                        "created_by_location_name": "Main Store"  # Filter to Main Store POs only
                    }
                    if date_match:
                        query.update(date_match)
                    
                    # Use aggregation with timeout per vendor
                    # Longer timeout (8s) when date filter is applied
                    timeout_ms = 8000 if date_match else 5000
                    
                    # Calculate PO value from items array (quantity * rate) if total_amount is missing
                    # This ensures consistency with Excel export
                    result = list(purchase_orders_collection.aggregate([
                        {"$match": query},
                        {"$addFields": {
                            "calculated_po_value": {
                                "$cond": {
                                    "if": {"$and": [
                                        {"$gt": ["$total_amount", 0]},
                                        {"$ne": ["$total_amount", None]}
                                    ]},
                                    "then": "$total_amount",
                                    "else": {
                                        "$reduce": {
                                            "input": {"$ifNull": ["$items", []]},
                                            "initialValue": 0,
                                            "in": {"$add": [
                                                "$$value",
                                                {"$multiply": [
                                                    {"$ifNull": ["$$this.quantity", 0]},
                                                    {"$ifNull": ["$$this.rate", 0]}
                                                ]}
                                            ]}
                                        }
                                    }
                                }
                            }
                        }},
                        {"$group": {
                            "_id": None,
                            "po_count": {"$sum": 1},
                            "total_po_value": {"$sum": "$calculated_po_value"},
                            "grn_count": {"$sum": {"$cond": [{"$gt": ["$grn_amount", 0]}, 1, 0]}},
                            "total_grn_value": {"$sum": {"$ifNull": ["$grn_amount", 0]}}
                        }}
                    ], maxTimeMS=timeout_ms))
                    
                    if result and len(result) > 0:
                        r = result[0]
                        po_count = r.get("po_count", 0)
                        po_value = r.get("total_po_value", 0)
                        grn_count = r.get("grn_count", 0)
                        grn_value = r.get("total_grn_value", 0)
                        
                        if po_count > 0:
                            main_store_pos[vid] = {"count": po_count, "value": po_value, "pos": []}
                        if grn_count > 0:
                            main_store_grns[vid] = {"count": grn_count, "value": grn_value}
                        
                        success_count += 1
                except Exception as ve:
                    error_count += 1
                    # On timeout, try a simpler count query
                    try:
                        query = {
                            "vendor_id": vid,
                            "created_by_location_name": "Main Store"  # Filter to Main Store POs only
                        }
                        if date_match:
                            query.update(date_match)
                        
                        # Simple find with projection for faster query
                        pos = list(purchase_orders_collection.find(
                            query,
                            {"total_amount": 1, "grn_amount": 1, "items": 1}
                        ).limit(500))
                        
                        if pos:
                            po_count = len(pos)
                            # Calculate PO value from items if total_amount is missing (consistent with Excel)
                            po_value = 0
                            for p in pos:
                                if p.get("total_amount") and p.get("total_amount") > 0:
                                    po_value += p.get("total_amount")
                                else:
                                    # Calculate from items array
                                    for item in p.get("items", []):
                                        po_value += (item.get("quantity", 0) or 0) * (item.get("rate", 0) or 0)
                            grn_count = sum(1 for p in pos if (p.get("grn_amount") or 0) > 0)
                            grn_value = sum(p.get("grn_amount", 0) or 0 for p in pos)
                            
                            if po_count > 0:
                                main_store_pos[vid] = {"count": po_count, "value": po_value, "pos": []}
                            if grn_count > 0:
                                main_store_grns[vid] = {"count": grn_count, "value": grn_value}
                            
                            success_count += 1
                            error_count -= 1  # Recovered
                    except:
                        pass
                    continue
            
            debug_info["vendors_queried_success"] = success_count
            debug_info["vendors_queried_error"] = error_count
            debug_info["vendors_with_pos"] = len(main_store_pos)
            debug_info["vendors_with_grns"] = len(main_store_grns)
            debug_info["total_po_value"] = sum(p['value'] for p in main_store_pos.values())
            debug_info["total_grn_value"] = sum(g['value'] for g in main_store_grns.values())
            
            print(f"PO/GRN complete: {len(main_store_pos)} vendors with POs, {len(main_store_grns)} with GRNs")
            
        except Exception as e:
            debug_info["error"] = str(e)
            print(f"PO/GRN aggregation error: {e}")
        
        # ===== Kitchen GRNs (from kitchen_receivables collection) - Daily Perishables =====
        # Kitchen GRNs go to kitchen_receivables, not lots
        dp_grns = {}
        kitchen_receivables = db["kitchen_receivables"]
        dp_match_stage = {**date_filter_lots, "vendor_id": {"$exists": True, "$ne": None, "$ne": ""}}
        if not date_filter_lots:
            dp_match_stage = {"vendor_id": {"$exists": True, "$ne": None, "$ne": ""}}
        
        dp_pipeline = [
            {"$match": dp_match_stage},
            {"$group": {
                "_id": "$vendor_id",
                "count": {"$sum": 1},
                "total_value": {"$sum": {"$ifNull": ["$amount", 0]}}
            }}
        ]
        try:
            for doc in kitchen_receivables.aggregate(dp_pipeline, maxTimeMS=60000, allowDiskUse=True):
                v_id = doc["_id"]
                if not v_id:
                    continue
                dp_grns[v_id] = {
                    "count": doc["count"],
                    "value": doc.get("total_value", 0) or 0
                }
        except Exception as e:
            print(f"Kitchen receivables aggregation error: {e}")
        
        # ===== Build result =====
        result = []
        grand_total_main_store_po = 0
        grand_total_main_store_grn = 0
        grand_total_kitchen_po = 0
        grand_total_dp_grn = 0
        total_main_store_po_count = 0
        total_main_store_grn_count = 0
        total_kitchen_po_count = 0
        total_dp_count = 0
        
        # Filter vendors if vendor_id is specified
        if vendor_id:
            vendors_to_process = [(v_id, v_name) for v_id, v_name in vendors_map.items() if v_id == vendor_id]
        else:
            vendors_to_process = list(vendors_map.items())
        
        print(f"Building result for {len(vendors_to_process)} vendors")
        
        result = []  # Initialize result array here to ensure it's empty
        
        for v_id, v_name in vendors_to_process:
            ms_po = main_store_pos.get(v_id, {"count": 0, "value": 0, "pos": []})
            k_po = kitchen_pos.get(v_id, {"count": 0, "value": 0})
            ms_grn = main_store_grns.get(v_id, {"count": 0, "value": 0})
            dp = dp_grns.get(v_id, {"count": 0, "value": 0})
            
            # Update grand totals
            grand_total_main_store_po += ms_po["value"]
            grand_total_main_store_grn += ms_grn["value"]
            grand_total_kitchen_po += k_po["value"]
            grand_total_dp_grn += dp["value"]
            total_main_store_po_count += ms_po["count"]
            total_main_store_grn_count += ms_grn["count"]
            total_kitchen_po_count += k_po["count"]
            total_dp_count += dp["count"]
            
            result.append({
                "vendor_id": v_id,
                "vendor_name": v_name,
                # Main store fields - using names frontend expects
                "po_count": ms_po["count"],
                "main_store_po_count": ms_po["count"],
                "total_po_value": round(ms_po["value"], 2),
                "main_store_po_value": round(ms_po["value"], 2),
                "main_store_grn_count": ms_grn["count"],
                "total_main_store_grn_value": round(ms_grn["value"], 2),
                "main_store_grn_value": round(ms_grn["value"], 2),
                "total_grn_value": round(ms_grn["value"], 2),  # For GRN total display
                # Kitchen/DP fields
                "kitchen_po_count": k_po["count"],
                "kitchen_po_value": round(k_po["value"], 2),
                "daily_perishable_count": dp["count"],
                "total_daily_perishable_value": round(dp["value"], 2),
                "daily_perishable_value": round(dp["value"], 2),
                # Detail arrays
                "purchase_orders": ms_po.get("pos", []) if not fast else [],
                "grn_entries": [],
                "dp_entries": []
            })
        
        # Sort by total activity (PO + GRN counts)
        result.sort(key=lambda x: x["main_store_po_count"] + x["main_store_grn_count"] + x["kitchen_po_count"], reverse=True)
        
        query_time = time.time() - start_time
        
        return {
            "vendors": result,
            "summary": {
                "total_vendors": len(vendors_map),
                "total_po_count": total_main_store_po_count,
                "grand_total_po_value": round(grand_total_main_store_po, 2),
                "total_main_store_grn_count": total_main_store_grn_count,
                "grand_total_main_store_grn_value": round(grand_total_main_store_grn, 2),
                "total_kitchen_po_count": total_kitchen_po_count,
                "grand_total_kitchen_po_value": round(grand_total_kitchen_po, 2),
                "total_daily_perishable_count": total_dp_count,
                "grand_total_daily_perishable_value": round(grand_total_dp_grn, 2),
                "query_time_seconds": round(query_time, 2)
            },
            "debug": debug_info  # Include debug info in response
        }
    except Exception as e:
        print(f"Vendor ledger error: {e}")
        import traceback
        traceback.print_exc()
        return {
            "vendors": [],
            "summary": {
                "total_vendors": 0,
                "total_po_count": 0,
                "grand_total_po_value": 0,
                "total_main_store_grn_count": 0,
                "grand_total_main_store_grn_value": 0,
                "total_kitchen_po_count": 0,
                "grand_total_kitchen_po_value": 0,
                "total_daily_perishable_count": 0,
                "grand_total_daily_perishable_value": 0,
                "error": str(e)
            }
        }
    
# ============ Kitchen Ledger Report ============

@app.get("/api/reports/kitchen-ledger")
async def get_kitchen_ledger(
    kitchen_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_daily_perishables: bool = Query(True, description="Include daily perishables from vendors")
):
    """
    Kitchen-wise ledger showing:
    - Requisitions raised by kitchen
    - Goods dispatched/issued to each kitchen (Main Store + optionally Daily Perishables)
    - Category-wise breakdown
    """
    
    try:
        # Get all kitchens or specific kitchen
        if kitchen_id:
            try:
                kitchens = [locations_collection.find_one({"_id": ObjectId(kitchen_id), "type": "kitchen"})]
            except:
                raise HTTPException(status_code=404, detail="Invalid kitchen ID format")
            if not kitchens[0]:
                raise HTTPException(status_code=404, detail="Kitchen not found")
        else:
            kitchens = list(locations_collection.find({"type": "kitchen"}).sort("name", 1))
        
        result = []
        grand_total_value = 0
        grand_total_qty = 0
        
        # Category mapping for standardization
        category_map = {
            "Indian Grocery": "Groceries",
            "Chinese Grocery": "Groceries", 
            "Continental Grocery": "Groceries",
            "Continental grocery": "Groceries",
            "MALA": "Groceries",
            "Grocery": "Groceries",
            "Seasoning": "Groceries",
            "Spices": "Groceries",
            "Beverage": "Beverages",
            "Beverages": "Beverages",
            "Packaging": "Packaging",
            "Seafood": "Seafood",
            "Sea Food": "Seafood",
            "Housekeeping": "Housekeeping",
            "House Keeping": "Housekeeping",
            "Non Veg": "Non Veg",
            "NON Veg": "Non Veg",
            "NON VEG": "Non Veg",
            "Dairy Product": "Dairy",
            "Dairy": "Dairy",
            "Vegetables": "Vegetables",
            "Veg": "Vegetables"
        }
        
        for kitchen in kitchens:
            if not kitchen:
                continue
                
            k_id = str(kitchen["_id"])
            k_name = kitchen["name"]
        
            # Build date query for transactions
            date_query = {}
            date_query_dt = {}  # For datetime fields
            if start_date:
                date_query["$gte"] = start_date
                try:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    date_query_dt["$gte"] = start_dt
                except:
                    pass
            if end_date:
                date_query["$lte"] = end_date
                try:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                    date_query_dt["$lte"] = end_dt
                except:
                    pass
            
            # 1. Get requisitions for this kitchen
            try:
                k_id_obj = ObjectId(k_id)
            except:
                k_id_obj = None
            
            req_query = {
                "$or": [
                    {"kitchen_id": k_id},
                    {"location_name": k_name},
                    {"kitchen_name": k_name}
                ],
                "status": {"$in": ["dispatched", "partial", "completed", "received"]}
            }
            if k_id_obj:
                req_query["$or"].insert(0, {"location_id": k_id_obj})
            # Use datetime query for requisitions (created_at is datetime)
            if date_query_dt:
                req_query["created_at"] = date_query_dt
            
            requisitions = list(requisitions_collection.find(req_query).sort("created_at", -1))
            
            # Calculate requisition totals
            requisition_entries = []
            total_requisition_value = 0
            total_requisition_qty = 0
            
            # OPTIMIZATION: Batch fetch all items from requisitions
            all_req_item_ids = []
            for req in requisitions:
                for item in req.get("items", []):
                    item_id = item.get("item_id")
                    if item_id:
                        all_req_item_ids.append(item_id)
            
            req_items_map = {}
            if all_req_item_ids:
                try:
                    unique_ids = list(set(all_req_item_ids))
                    items_data = list(items_collection.find(
                        {"_id": {"$in": [ObjectId(i) for i in unique_ids[:500]]}},
                        {"_id": 1, "standard_price": 1}
                    ))
                    req_items_map = {str(item["_id"]): item for item in items_data}
                except:
                    pass
            
            for req in requisitions:
                req_total = 0
                req_qty = 0
                for item in req.get("items", []):
                    qty = item.get("quantity_sent", 0) or item.get("quantity", 0) or 0
                    item_id = item.get("item_id")
                    db_item = req_items_map.get(item_id) if item_id else None
                    rate = item.get("rate") or (db_item.get("standard_price", 0) if db_item else 0) or 0
                    req_total += (qty or 0) * (rate or 0)
                    req_qty += (qty or 0)
                
                total_requisition_value += req_total
                total_requisition_qty += req_qty
                
                created_at = req.get("created_at")
                if created_at:
                    if isinstance(created_at, str):
                        date_str = created_at[:10]
                    else:
                        date_str = str(created_at)[:10]
                else:
                    date_str = ""
                
                requisition_entries.append({
                    "id": str(req["_id"]),
                    "requisition_number": req.get("serial_number", ""),
                    "date": date_str,
                    "items_count": len(req.get("items", [])),
                    "total_quantity": req_qty,
                    "total_value": req_total,
                    "status": req.get("status", "pending")
                })
            
            # 2. Get issue transactions from Main Store (transactions collection)
            txn_query = {
                "type": {"$in": ["issue"]},
                "$or": [
                    {"destination_location_id": k_id},
                    {"kitchen_id": k_id},
                    {"kitchen_name": k_name}
                ]
            }
            if k_id_obj:
                txn_query["$or"].insert(0, {"destination_location_id": k_id_obj})
            if date_query:
                txn_query["created_at"] = date_query
            
            transactions = list(transactions_collection.find(txn_query).sort("created_at", -1).limit(2000))
            
            # Batch fetch all lots and items to avoid N+1 queries
            lot_ids = [ObjectId(t["lot_id"]) for t in transactions if t.get("lot_id")]
            item_ids = [ObjectId(t["item_id"]) for t in transactions if t.get("item_id")]
            
            # Batch fetch lots
            lots_map = {}
            if lot_ids:
                try:
                    lots_data = list(lots_collection.find(
                        {"_id": {"$in": lot_ids[:1000]}},  # Limit to 1000
                        {"_id": 1, "item_id": 1, "purchase_rate": 1}
                    ))
                    lots_map = {str(l["_id"]): l for l in lots_data}
                    # Add item_ids from lots
                    for l in lots_data:
                        if l.get("item_id"):
                            try:
                                item_ids.append(ObjectId(l["item_id"]))
                            except:
                                pass
                except Exception as e:
                    print(f"Lot batch fetch error: {e}")
            
            # Batch fetch items
            items_map = {}
            if item_ids:
                try:
                    items_data = list(items_collection.find(
                        {"_id": {"$in": list(set(item_ids))[:1000]}},  # Dedupe and limit
                        {"_id": 1, "name": 1, "category": 1, "unit": 1, "standard_price": 1}
                    ))
                    items_map = {str(i["_id"]): i for i in items_data}
                except Exception as e:
                    print(f"Item batch fetch error: {e}")
            
            # Group all dispatches by category
            category_data = {}
            dispatch_entries = []
            total_issued_value = 0
            total_issued_qty = 0
            
            # Process Main Store issue transactions
            for txn in transactions:
                item = None
                lot = None
                
                # Always try to get lot first (for purchase_rate)
                if txn.get("lot_id"):
                    lot = lots_map.get(str(txn["lot_id"]))
                
                # Get item - from direct item_id, or from lot's item_id
                if txn.get("item_id"):
                    item = items_map.get(str(txn["item_id"]))
                elif lot and lot.get("item_id"):
                    item = items_map.get(str(lot["item_id"]))
                
                if not item:
                    continue
                
                raw_category = item.get("category", "Uncategorized")
                category = category_map.get(raw_category, raw_category)
                
                qty = float(txn.get("quantity") or 0)
                rate = float(txn.get("rate") or (lot.get("purchase_rate") if lot else 0) or item.get("standard_price") or 0)
                value = float(txn.get("value") or (qty * rate))
                
                if category not in category_data:
                    category_data[category] = {"category": category, "total_quantity": 0, "total_value": 0, "source": "Main Store"}
                
                category_data[category]["total_quantity"] += qty
                category_data[category]["total_value"] += value
                
                txn_created = txn.get("created_at")
                if txn_created:
                    txn_date = txn_created[:10] if isinstance(txn_created, str) else str(txn_created)[:10]
                else:
                    txn_date = ""
                
                dispatch_entries.append({
                    "date": txn_date,
                    "item_name": item["name"],
                    "category": category,
                    "quantity": qty,
                    "unit": item.get("unit", ""),
                    "rate": rate,
                    "value": value,
                    "type": "Main Store",
                    "challan_number": txn.get("challan_number", "")
                })
                
                total_issued_value += value
                total_issued_qty += qty
            
            # 3. Get Daily Perishables from kitchen_receivables collection (if enabled)
            kr_records = []
            if include_daily_perishables:
                kr_query = {"kitchen_name": {"$regex": f"^{k_name}$", "$options": "i"}}
                # Also try partial match
                kr_query_alt = {"kitchen_name": {"$regex": k_name, "$options": "i"}}
                
                # Build date query for kitchen_receivables (uses receive_date)
                if start_date or end_date:
                    kr_date_query = {}
                    if start_date:
                        kr_date_query["$gte"] = start_date
                    if end_date:
                        kr_date_query["$lte"] = end_date
                    kr_query["receive_date"] = kr_date_query
                    kr_query_alt["receive_date"] = kr_date_query
                
                # Try exact match first, then partial
                kitchen_receivables = db["kitchen_receivables"]
                kr_records = list(kitchen_receivables.find(kr_query).sort("receive_date", -1))
                if not kr_records:
                    kr_records = list(kitchen_receivables.find(kr_query_alt).sort("receive_date", -1))
            
            # Process Daily Perishables (kitchen_receivables)
            for rec in kr_records:
                raw_category = rec.get("category", "Uncategorized")
                category = category_map.get(raw_category, raw_category)
                
                qty = float(rec.get("quantity") or 0)
                rate = float(rec.get("rate") or 0)
                value = float(rec.get("amount") or (qty * rate))
                
                if category not in category_data:
                    category_data[category] = {"category": category, "total_quantity": 0, "total_value": 0, "source": "Daily Perishables"}
                
                category_data[category]["total_quantity"] += qty
                category_data[category]["total_value"] += value
                
                rec_date = rec.get("receive_date", "")
                
                dispatch_entries.append({
                    "date": rec_date,
                    "item_name": rec.get("item_name", "Unknown"),
                    "category": category,
                    "quantity": qty,
                    "unit": rec.get("unit", ""),
                    "rate": rate,
                    "value": value,
                    "type": "Daily Perishables",
                    "challan_number": rec.get("invoice_number", ""),
                    "vendor_name": rec.get("vendor_name", "")
                })
                
                total_issued_value += value
                total_issued_qty += qty
            
            # Sort dispatch entries by date (newest first)
            dispatch_entries.sort(key=lambda x: x["date"], reverse=True)
            
            categories_list = sorted(category_data.values(), key=lambda x: x["total_value"], reverse=True)
            
            grand_total_value += total_issued_value
            grand_total_qty += total_issued_qty
            
            result.append({
                "kitchen_id": k_id,
                "kitchen_name": k_name,
                "kitchen_address": kitchen.get("address", ""),
                "requisition_count": len(requisition_entries),
                "total_requisition_value": total_requisition_value,
                "total_requisition_qty": total_requisition_qty,
                "requisition_entries": requisition_entries[:50],
                "issued_count": len(dispatch_entries),
                "dispatch_count": len(dispatch_entries),
                "total_issued_value": total_issued_value,
                "total_issued_qty": total_issued_qty,
                "categories": categories_list,
                "dispatch_entries": dispatch_entries[:100],  # Increased limit
                "received_count": 0,
                "total_received_value": 0
            })
        
        return {
            "kitchens": result,
            "summary": {
                "total_kitchens": len(result),
                "grand_total_value": grand_total_value,
                "grand_total_quantity": grand_total_qty,
                "total_dispatches": sum(k["issued_count"] for k in result),
                "total_requisitions": sum(k["requisition_count"] for k in result)
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Kitchen ledger error: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error generating kitchen ledger: {str(e)}")


@app.get("/api/reports/kitchen-consumption/download")
async def download_kitchen_consumption(
    kitchen_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Download Excel report of item consumption for a specific kitchen.
    Format: Date | Item Name | Category | Quantity | Unit | Rate | Value | Lot Number
    """
    from io import BytesIO
    
    # Validate kitchen
    try:
        kitchen = locations_collection.find_one({"_id": ObjectId(kitchen_id), "type": "kitchen"})
    except:
        raise HTTPException(status_code=400, detail="Invalid kitchen ID format")
    
    if not kitchen:
        raise HTTPException(status_code=404, detail="Kitchen not found")
    
    # Build query for transactions
    txn_query = {
        "type": "issue",
        "destination_location_id": kitchen_id
    }
    
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date + "T23:59:59"  # Include full end date
        if date_filter:
            txn_query["created_at"] = date_filter
    
    # Get all issue transactions to this kitchen
    transactions = list(transactions_collection.find(txn_query).sort("created_at", 1))
    
    # Build data for Excel
    rows = []
    for txn in transactions:
        lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
        if not lot:
            continue
        
        item = items_collection.find_one({"_id": ObjectId(lot["item_id"])})
        if not item:
            continue
        
        qty = txn.get("quantity", 0)
        rate = lot.get("purchase_rate") or item.get("standard_price") or 0
        value = qty * rate
        
        rows.append({
            "Date": txn["created_at"][:10] if txn.get("created_at") else "",
            "Item Name": item["name"],
            "Category": item.get("category", "Uncategorized"),
            "Quantity": qty,
            "Unit": item.get("unit", ""),
            "Rate (₹)": round(rate, 2),
            "Value (₹)": round(value, 2),
            "Lot Number": lot.get("lot_number", "")
        })
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumption Report"
    
    # Header style
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    kitchen_name = kitchen["name"]
    date_range = ""
    if start_date and end_date:
        date_range = f" ({start_date} to {end_date})"
    elif start_date:
        date_range = f" (from {start_date})"
    elif end_date:
        date_range = f" (until {end_date})"
    
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Consumption Report - {kitchen_name}{date_range}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Date", "Item Name", "Category", "Quantity", "Unit", "Rate (₹)", "Value (₹)", "Lot Number"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    total_value = 0
    for row_idx, row_data in enumerate(rows, 4):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.border = thin_border
            if header in ["Rate (₹)", "Value (₹)"]:
                cell.alignment = Alignment(horizontal='right')
            elif header in ["Quantity"]:
                cell.alignment = Alignment(horizontal='center')
        total_value += row_data.get("Value (₹)", 0)
    
    # Total row
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=5, value="Total:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(total_value, 2)).font = Font(bold=True)
    
    # Adjust column widths
    col_widths = [12, 35, 20, 10, 8, 12, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = width
    
    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    safe_kitchen_name = kitchen_name.replace(" ", "_").replace("/", "-")
    filename = f"Consumption_{safe_kitchen_name}"
    if start_date:
        filename += f"_{start_date}"
    if end_date:
        filename += f"_to_{end_date}"
    filename += ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ Kitchen Receivables (Daily Perishables) Report ============

@app.get("/api/reports/kitchen-receivables")
async def get_kitchen_receivables(
    kitchen_id: Optional[str] = None,
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    view: str = "daily"  # daily, monthly, vendor
):
    """
    Kitchen receivables report for main store dashboard.
    Shows daily perishables received by kitchens (NOT added to main stock).
    
    Views:
    - daily: Day-wise breakdown of items received
    - monthly: Month-wise aggregated data
    - vendor: Vendor-wise breakdown
    """
    kitchen_receivables = db["kitchen_receivables"]
    
    # Build query
    query = {}
    if kitchen_id and kitchen_id != "all":
        query["kitchen_id"] = kitchen_id
    if vendor_id and vendor_id != "all":
        query["vendor_id"] = vendor_id
    if start_date:
        query["receive_date"] = {"$gte": start_date}
    if end_date:
        if "receive_date" in query:
            query["receive_date"]["$lte"] = end_date
        else:
            query["receive_date"] = {"$lte": end_date}
    
    receivables = list(kitchen_receivables.find(query).sort("receive_date", -1))
    
    # Get all kitchens and vendors for reference
    kitchens_map = {str(k["_id"]): k["name"] for k in locations_collection.find({"type": "kitchen"})}
    vendors_map = {str(v["_id"]): v["name"] for v in vendors_collection.find()}
    
    if view == "daily":
        # Group by date and kitchen
        daily_data = {}
        po_ids_by_key = {}  # Track unique PO IDs per date/kitchen
        
        for rec in receivables:
            date_key = rec.get("receive_date", "Unknown")
            kitchen_key = rec.get("kitchen_id", "Unknown")
            key = f"{date_key}_{kitchen_key}"
            
            if key not in daily_data:
                daily_data[key] = {
                    "date": date_key,
                    "kitchen_id": kitchen_key,
                    "kitchen_name": rec.get("kitchen_name", kitchens_map.get(kitchen_key, "Unknown")),
                    "items": [],
                    "total_qty": 0,
                    "total_value": 0,
                    "vendor_count": set(),
                    "po_ids": set()
                }
                po_ids_by_key[key] = set()
            
            daily_data[key]["items"].append({
                "item_name": rec.get("item_name"),
                "category": rec.get("category"),
                "quantity": rec.get("quantity", 0),
                "unit": rec.get("unit"),
                "rate": rec.get("rate", 0),
                "amount": rec.get("amount", 0),
                "vendor_name": rec.get("vendor_name"),
                "invoice_number": rec.get("invoice_number"),
                "po_number": rec.get("po_number"),
                "po_id": rec.get("po_id")
            })
            daily_data[key]["total_qty"] += rec.get("quantity", 0)
            daily_data[key]["total_value"] += rec.get("amount", 0)
            daily_data[key]["vendor_count"].add(rec.get("vendor_id"))
            if rec.get("po_id"):
                po_ids_by_key[key].add(rec.get("po_id"))
        
        # Fetch verification photos for each PO
        result = []
        for key, data in daily_data.items():
            data["vendor_count"] = len(data["vendor_count"])
            data["items_count"] = len(data["items"])
            data["po_ids"] = list(po_ids_by_key.get(key, []))
            
            # Get verification photos from POs
            verification_photos = []
            for po_id in data["po_ids"]:
                try:
                    po = purchase_orders_collection.find_one({"_id": ObjectId(po_id)})
                    if po and po.get("grn_verification"):
                        verification_photos.append({
                            "po_id": po_id,
                            "po_number": po.get("po_number"),
                            "photo": po.get("grn_verification", {}).get("photo"),
                            "gps_location": po.get("grn_verification", {}).get("gps_location"),
                            "capture_time": po.get("grn_verification", {}).get("capture_time"),
                            "verified_at": po.get("grn_verification", {}).get("verified_at")
                        })
                except:
                    pass
            data["verification_photos"] = verification_photos
            result.append(data)
        
        # Sort by date descending
        result.sort(key=lambda x: x["date"], reverse=True)
        
        # Also build flat receivables list for frontend compatibility
        flat_receivables = []
        for rec in receivables:
            flat_receivables.append({
                "kitchen_id": rec.get("kitchen_id"),
                "kitchen_name": rec.get("kitchen_name", kitchens_map.get(rec.get("kitchen_id"), "Unknown")),
                "item_id": rec.get("item_id"),
                "item_name": rec.get("item_name"),
                "category": rec.get("category"),
                "unit": rec.get("unit"),
                "quantity": rec.get("quantity", 0),
                "rate": rec.get("rate", 0),
                "amount": rec.get("amount", 0),
                "vendor_id": rec.get("vendor_id"),
                "vendor_name": rec.get("vendor_name", vendors_map.get(rec.get("vendor_id"), "Unknown")),
                "receive_date": rec.get("receive_date"),
                "po_number": rec.get("po_number"),
                "po_id": rec.get("po_id"),
                "invoice_number": rec.get("invoice_number")
            })
        
        return {
            "view": "daily",
            "data": result,
            "receivables": flat_receivables,  # Flat list for frontend
            "summary": {
                "total_days": len(set(r["date"] for r in result)),
                "total_kitchens": len(set(r["kitchen_id"] for r in result)),
                "total_entries": len(receivables),
                "total_value": sum(r.get("amount", 0) for r in receivables)
            }
        }
    
    elif view == "monthly":
        # Group by month and kitchen
        monthly_data = {}
        for rec in receivables:
            date_str = rec.get("receive_date", "")
            month_key = date_str[:7] if date_str else "Unknown"  # YYYY-MM
            kitchen_key = rec.get("kitchen_id", "Unknown")
            key = f"{month_key}_{kitchen_key}"
            
            if key not in monthly_data:
                monthly_data[key] = {
                    "month": month_key,
                    "kitchen_id": kitchen_key,
                    "kitchen_name": rec.get("kitchen_name", kitchens_map.get(kitchen_key, "Unknown")),
                    "total_qty": 0,
                    "total_value": 0,
                    "items_count": 0,
                    "days_count": set(),
                    "vendors": set(),
                    "categories": {}
                }
            
            monthly_data[key]["total_qty"] += rec.get("quantity", 0)
            monthly_data[key]["total_value"] += rec.get("amount", 0)
            monthly_data[key]["items_count"] += 1
            monthly_data[key]["days_count"].add(rec.get("receive_date"))
            monthly_data[key]["vendors"].add(rec.get("vendor_name"))
            
            # Category breakdown
            cat = rec.get("category", "Uncategorized")
            if cat not in monthly_data[key]["categories"]:
                monthly_data[key]["categories"][cat] = {"qty": 0, "value": 0}
            monthly_data[key]["categories"][cat]["qty"] += rec.get("quantity", 0)
            monthly_data[key]["categories"][cat]["value"] += rec.get("amount", 0)
        
        # Convert to list
        result = []
        for key, data in monthly_data.items():
            data["days_count"] = len(data["days_count"])
            data["vendors"] = list(data["vendors"])
            data["vendor_count"] = len(data["vendors"])
            result.append(data)
        
        result.sort(key=lambda x: x["month"], reverse=True)
        
        return {
            "view": "monthly",
            "data": result,
            "summary": {
                "total_months": len(set(r["month"] for r in result)),
                "total_kitchens": len(set(r["kitchen_id"] for r in result)),
                "total_value": sum(r["total_value"] for r in result)
            }
        }
    
    elif view == "vendor":
        # Group by vendor and kitchen
        vendor_data = {}
        for rec in receivables:
            vendor_key = rec.get("vendor_id", "Unknown")
            kitchen_key = rec.get("kitchen_id", "Unknown")
            key = f"{vendor_key}_{kitchen_key}"
            
            if key not in vendor_data:
                vendor_data[key] = {
                    "vendor_id": vendor_key,
                    "vendor_name": rec.get("vendor_name", vendors_map.get(vendor_key, "Unknown")),
                    "kitchen_id": kitchen_key,
                    "kitchen_name": rec.get("kitchen_name", kitchens_map.get(kitchen_key, "Unknown")),
                    "total_qty": 0,
                    "total_value": 0,
                    "items_count": 0,
                    "dates": set(),
                    "categories": {}
                }
            
            vendor_data[key]["total_qty"] += rec.get("quantity", 0)
            vendor_data[key]["total_value"] += rec.get("amount", 0)
            vendor_data[key]["items_count"] += 1
            vendor_data[key]["dates"].add(rec.get("receive_date"))
            
            cat = rec.get("category", "Uncategorized")
            if cat not in vendor_data[key]["categories"]:
                vendor_data[key]["categories"][cat] = {"qty": 0, "value": 0}
            vendor_data[key]["categories"][cat]["qty"] += rec.get("quantity", 0)
            vendor_data[key]["categories"][cat]["value"] += rec.get("amount", 0)
        
        result = []
        for key, data in vendor_data.items():
            data["dates_count"] = len(data["dates"])
            del data["dates"]
            result.append(data)
        
        result.sort(key=lambda x: x["total_value"], reverse=True)
        
        return {
            "view": "vendor",
            "data": result,
            "summary": {
                "total_vendors": len(set(r["vendor_id"] for r in result)),
                "total_kitchens": len(set(r["kitchen_id"] for r in result)),
                "total_value": sum(r["total_value"] for r in result)
            }
        }
    
    return {"error": "Invalid view parameter"}


# Dashboard
@app.get("/api/dashboard/stats")
async def get_dashboard_stats():
    """Optimized dashboard stats - uses estimated counts and timeouts for speed"""
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    seven_days_from_now = (now + timedelta(days=7)).strftime("%Y-%m-%d")
    
    # Auto-cleanup: Delete pending requisitions older than 2 days
    try:
        two_days_ago = (now - timedelta(days=2)).isoformat()
        deleted_old_reqs = requisitions_collection.delete_many({
            "status": "pending",
            "created_at": {"$lt": two_days_ago}
        })
        if deleted_old_reqs.deleted_count > 0:
            print(f"Auto-deleted {deleted_old_reqs.deleted_count} old pending requisitions")
    except:
        pass
    
    # Use estimated counts for global stats (much faster)
    stats = {}
    
    try:
        stats["total_lots"] = lots_collection.estimated_document_count()
    except:
        stats["total_lots"] = 0
        
    try:
        stats["total_items"] = items_collection.estimated_document_count()
    except:
        stats["total_items"] = 0
        
    try:
        stats["total_locations"] = locations_collection.estimated_document_count()
    except:
        stats["total_locations"] = 0
    
    # Expiry counts with timeout
    try:
        stats["expired_count"] = lots_collection.count_documents(
            {"current_quantity": {"$gt": 0}, "expiry_date": {"$lt": today_str}},
            maxTimeMS=5000
        )
    except:
        stats["expired_count"] = 0
        
    try:
        stats["expiring_soon_count"] = lots_collection.count_documents(
            {"current_quantity": {"$gt": 0}, "expiry_date": {"$gte": today_str, "$lte": seven_days_from_now}},
            maxTimeMS=5000
        )
    except:
        stats["expiring_soon_count"] = 0
    
    # Today's activity with timeout
    try:
        stats["today_grn"] = transactions_collection.count_documents(
            {"type": "grn", "created_at": {"$gte": today_start}},
            maxTimeMS=3000
        )
    except:
        stats["today_grn"] = 0
        
    try:
        stats["today_issues"] = transactions_collection.count_documents(
            {"type": "issue", "created_at": {"$gte": today_start}},
            maxTimeMS=3000
        )
    except:
        stats["today_issues"] = 0
        
    try:
        stats["today_transfers"] = transactions_collection.count_documents(
            {"type": "transfer", "created_at": {"$gte": today_start}},
            maxTimeMS=3000
        )
    except:
        stats["today_transfers"] = 0
    
    # Requisition counts with timeout
    try:
        stats["pending_requisitions"] = requisitions_collection.count_documents(
            {"status": "pending"},
            maxTimeMS=3000
        )
    except:
        stats["pending_requisitions"] = 0
    
    # Kitchen stats - use simpler query with timeout
    stats["today_kitchen_pos"] = 0
    stats["pending_kitchen_pos"] = 0
    
    try:
        # Count kitchen POs directly from PO collection with location filter
        stats["today_kitchen_pos"] = purchase_orders_collection.count_documents(
            {"created_at": {"$gte": today_start}, "created_by_location_name": {"$ne": "Main Store"}},
            maxTimeMS=5000
        )
    except:
        pass
        
    try:
        # Simplified pending count - just count all pending POs
        stats["pending_kitchen_pos"] = purchase_orders_collection.count_documents(
            {"status": {"$in": ["pending", "partial"]}},
            maxTimeMS=5000
        )
    except:
        pass
    
    # Kitchen receivables today
    try:
        kitchen_receivables = db["kitchen_receivables"]
        stats["today_kitchen_grns"] = kitchen_receivables.count_documents(
            {"receive_date": today_str},
            maxTimeMS=3000
        )
    except:
        stats["today_kitchen_grns"] = 0
    
    try:
        kitchen_receivables = db["kitchen_receivables"]
        value_result = list(kitchen_receivables.aggregate([
            {"$match": {"receive_date": today_str}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ], maxTimeMS=5000))
        stats["today_kitchen_value"] = value_result[0]["total"] if value_result else 0
    except:
        stats["today_kitchen_value"] = 0
    
    return {
        "total_lots": stats["total_lots"],
        "total_items": stats["total_items"],
        "total_locations": stats["total_locations"],
        "expired_count": stats["expired_count"],
        "expiring_soon_count": stats["expiring_soon_count"],
        "today_grn": stats["today_grn"],
        "today_issues": stats["today_issues"],
        "today_transfers": stats["today_transfers"],
        "pending_requisitions": stats["pending_requisitions"],
        "kitchen_activity": {
            "today_pos": stats["today_kitchen_pos"],
            "today_grns": stats["today_kitchen_grns"],
            "today_value": stats["today_kitchen_value"],
            "pending_pos": stats["pending_kitchen_pos"],
            "recent_pos": []  # Removed to speed up - load separately if needed
        }
    }
@app.get("/api/pending-alerts")
async def get_pending_alerts(current_user = Depends(get_current_user)):
    """Get pending alerts for dashboard reminders"""
    now = datetime.now(timezone.utc)
    alerts = {
        "pending_pos_count": 0,
        "urgent_pos_count": 0,
        "pending_requisitions_count": 0,
        "dispatched_requisitions_count": 0,
        "pending_kitchen_grn_count": 0
    }
    
    # Pending POs (all)
    pending_pos = list(purchase_orders_collection.find({"status": "pending"}))
    alerts["pending_pos_count"] = len(pending_pos)
    
    # Urgent POs (more than 24 hours old)
    for po in pending_pos:
        try:
            created = datetime.fromisoformat(po["created_at"].replace("Z", "+00:00"))
            hours_old = (now - created).total_seconds() / 3600
            if hours_old > 24:
                alerts["urgent_pos_count"] += 1
        except:
            pass
    
    # Pending requisitions (waiting to dispatch)
    alerts["pending_requisitions_count"] = requisitions_collection.count_documents({"status": "pending"})
    
    # Dispatched requisitions (waiting for kitchen confirmation)
    alerts["dispatched_requisitions_count"] = requisitions_collection.count_documents({"status": "dispatched"})
    
    # If user is kitchen, filter by their location
    if current_user.get("role") == "kitchen":
        location_id = current_user.get("location_id")
        if location_id:
            alerts["pending_kitchen_grn_count"] = purchase_orders_collection.count_documents({
                "status": "pending",
                "created_by_location_id": location_id
            })
            alerts["dispatched_requisitions_count"] = requisitions_collection.count_documents({
                "status": "dispatched",
                "kitchen_id": location_id
            })
    
    return alerts


# Excel Export Endpoints
@app.get("/api/export/kitchen-ledger")
async def export_kitchen_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    kitchen_id: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Export Kitchen Ledger data to Excel"""
    # Build query
    query = {"type": {"$in": ["issue", "daily_perishable"]}}
    if kitchen_id:
        query["destination_location_id"] = kitchen_id
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        query["created_at"] = date_filter
    
    transactions = list(transactions_collection.find(query).sort("created_at", -1))
    
    # Build data for Excel
    rows = []
    for txn in transactions:
        item = None
        lot = None
        
        if txn.get("item_id"):
            item = items_collection.find_one({"_id": ObjectId(txn["item_id"])})
        elif txn.get("lot_id"):
            lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
            if lot:
                item = items_collection.find_one({"_id": ObjectId(lot["item_id"])})
        
        if not item:
            continue
        
        kitchen = locations_collection.find_one({"_id": ObjectId(txn.get("destination_location_id", ""))})
        
        qty = txn.get("quantity", 0)
        rate = txn.get("rate") or (lot.get("purchase_rate") if lot else 0) or item.get("standard_price") or 0
        value = txn.get("value") or (qty * rate)
        
        rows.append({
            "Date": txn.get("created_at", "")[:10],
            "Kitchen": kitchen.get("name", "Unknown") if kitchen else "Unknown",
            "Kitchen Code": kitchen.get("code", "") if kitchen else "",
            "Item": item.get("name", ""),
            "Category": item.get("category", ""),
            "Quantity": qty,
            "Unit": item.get("unit", ""),
            "Rate": rate,
            "Value": value,
            "Type": "Perishable" if txn.get("type") == "daily_perishable" else "Dispatch",
            "PO Number": txn.get("po_number", ""),
            "Challan": txn.get("challan_number", "")
        })
    
    # Create Excel file
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Kitchen Ledger')
    output.seek(0)
    
    filename = f"kitchen_ledger_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/export/kitchen-ledger-itemwise")
async def export_kitchen_ledger_itemwise(
    kitchen_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Export Kitchen Ledger with Item-wise breakdown by Category.
    Creates a multi-sheet Excel with:
    - Summary sheet (kitchen-wise totals)
    - One sheet per kitchen with category-wise item breakdown
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from collections import defaultdict
    import io
    
    # Get kitchens
    if kitchen_id:
        kitchens = [locations_collection.find_one({"_id": ObjectId(kitchen_id), "type": "kitchen"})]
    else:
        kitchens = list(locations_collection.find({"type": "kitchen"}).sort("name", 1))
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    category_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    black_font = Font(color="000000")
    bold_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = "Kitchen Ledger - Item-wise Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:E1')
    
    if start_date or end_date:
        ws_summary['A2'] = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
    else:
        ws_summary['A2'] = "Period: All Time"
    ws_summary['A2'].font = black_font
    
    # Summary headers
    summary_headers = ["Kitchen", "Categories", "Items", "Total Qty", "Total Value (₹)"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    summary_row = 5
    grand_total_qty = 0
    grand_total_value = 0
    
    for kitchen in kitchens:
        if not kitchen:
            continue
        
        k_id = str(kitchen["_id"])
        k_name = kitchen["name"]
        
        # Get requisitions for this kitchen
        req_query = {
            "$or": [
                {"kitchen_id": k_id},
                {"location_name": k_name},
                {"kitchen_name": k_name}
            ],
            "status": {"$in": ["dispatched", "partial", "completed", "received"]}
        }
        
        if start_date or end_date:
            date_filter = {}
            if start_date:
                try:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    date_filter["$gte"] = start_dt
                except:
                    pass
            if end_date:
                try:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                    date_filter["$lte"] = end_dt
                except:
                    pass
            if date_filter:
                req_query["created_at"] = date_filter
        
        requisitions = list(requisitions_collection.find(req_query))
        
        # Aggregate by category and item
        category_items = defaultdict(lambda: defaultdict(lambda: {"qty": 0, "value": 0, "unit": ""}))
        
        for req in requisitions:
            for item in req.get("items", []):
                item_name = item.get("item_name", "Unknown")
                category = item.get("category", "Uncategorized") or "Uncategorized"
                qty = float(item.get("quantity", 0) or 0)
                unit = item.get("unit", "")
                
                # Get price from items collection
                item_id = item.get("item_id")
                rate = 0
                if item_id:
                    try:
                        db_item = items_collection.find_one({"_id": ObjectId(item_id)})
                        if db_item:
                            rate = float(db_item.get("standard_price", 0) or 0)
                    except:
                        pass
                
                value = qty * rate
                
                category_items[category][item_name]["qty"] += qty
                category_items[category][item_name]["value"] += value
                category_items[category][item_name]["unit"] = unit
        
        # Calculate totals for this kitchen
        total_qty = sum(
            item_data["qty"] 
            for cat_items in category_items.values() 
            for item_data in cat_items.values()
        )
        total_value = sum(
            item_data["value"] 
            for cat_items in category_items.values() 
            for item_data in cat_items.values()
        )
        
        # Add to summary
        ws_summary.cell(row=summary_row, column=1, value=k_name).border = thin_border
        ws_summary.cell(row=summary_row, column=2, value=len(category_items)).border = thin_border
        ws_summary.cell(row=summary_row, column=3, value=sum(len(items) for items in category_items.values())).border = thin_border
        ws_summary.cell(row=summary_row, column=4, value=round(total_qty, 2)).border = thin_border
        ws_summary.cell(row=summary_row, column=5, value=round(total_value, 2)).number_format = '₹#,##0.00'
        ws_summary.cell(row=summary_row, column=5).border = thin_border
        summary_row += 1
        
        grand_total_qty += total_qty
        grand_total_value += total_value
        
        # Create sheet for this kitchen
        safe_name = k_name[:28].replace("/", "-").replace("\\", "-")  # Excel sheet name limit
        ws_kitchen = wb.create_sheet(title=safe_name)
        
        ws_kitchen['A1'] = f"Kitchen Ledger - {k_name}"
        ws_kitchen['A1'].font = Font(bold=True, size=14)
        ws_kitchen.merge_cells('A1:E1')
        
        if start_date or end_date:
            ws_kitchen['A2'] = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
        ws_kitchen['A2'].font = black_font
        
        ws_kitchen['A3'] = f"Total Requisitions: {len(requisitions)} | Total Items: {sum(len(items) for items in category_items.values())}"
        ws_kitchen['A3'].font = Font(italic=True)
        
        row = 5
        
        # Sort categories
        sorted_categories = sorted(category_items.items(), key=lambda x: sum(i["value"] for i in x[1].values()), reverse=True)
        
        for category, items in sorted_categories:
            # Category header
            ws_kitchen.cell(row=row, column=1, value=category.upper()).font = bold_font
            ws_kitchen.cell(row=row, column=1).fill = category_fill
            ws_kitchen.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Item headers
            item_headers = ["#", "Item Name", "Qty", "Unit", "Value (₹)"]
            for col, h in enumerate(item_headers, 1):
                cell = ws_kitchen.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1
            
            # Items sorted by qty
            sorted_items = sorted(items.items(), key=lambda x: x[1]["qty"], reverse=True)
            cat_total_qty = 0
            cat_total_value = 0
            
            for idx, (item_name, item_data) in enumerate(sorted_items, 1):
                ws_kitchen.cell(row=row, column=1, value=idx).border = thin_border
                ws_kitchen.cell(row=row, column=2, value=item_name).border = thin_border
                ws_kitchen.cell(row=row, column=3, value=round(item_data["qty"], 2)).border = thin_border
                ws_kitchen.cell(row=row, column=4, value=item_data["unit"]).border = thin_border
                ws_kitchen.cell(row=row, column=5, value=round(item_data["value"], 2)).number_format = '₹#,##0.00'
                ws_kitchen.cell(row=row, column=5).border = thin_border
                
                cat_total_qty += item_data["qty"]
                cat_total_value += item_data["value"]
                row += 1
            
            # Category subtotal
            ws_kitchen.cell(row=row, column=1, value="").fill = subtotal_fill
            ws_kitchen.cell(row=row, column=2, value=f"{category} Total").font = bold_font
            ws_kitchen.cell(row=row, column=2).fill = subtotal_fill
            ws_kitchen.cell(row=row, column=3, value=round(cat_total_qty, 2)).font = bold_font
            ws_kitchen.cell(row=row, column=3).fill = subtotal_fill
            ws_kitchen.cell(row=row, column=4, value="").fill = subtotal_fill
            ws_kitchen.cell(row=row, column=5, value=round(cat_total_value, 2)).font = bold_font
            ws_kitchen.cell(row=row, column=5).fill = subtotal_fill
            ws_kitchen.cell(row=row, column=5).number_format = '₹#,##0.00'
            
            for col in range(1, 6):
                ws_kitchen.cell(row=row, column=col).border = thin_border
            
            row += 2  # Gap before next category
        
        # Column widths for kitchen sheet
        ws_kitchen.column_dimensions['A'].width = 6
        ws_kitchen.column_dimensions['B'].width = 45
        ws_kitchen.column_dimensions['C'].width = 12
        ws_kitchen.column_dimensions['D'].width = 10
        ws_kitchen.column_dimensions['E'].width = 15
    
    # Grand total in summary
    ws_summary.cell(row=summary_row, column=1, value="GRAND TOTAL").font = bold_font
    ws_summary.cell(row=summary_row, column=1).fill = subtotal_fill
    ws_summary.cell(row=summary_row, column=4, value=round(grand_total_qty, 2)).font = bold_font
    ws_summary.cell(row=summary_row, column=4).fill = subtotal_fill
    ws_summary.cell(row=summary_row, column=5, value=round(grand_total_value, 2)).font = bold_font
    ws_summary.cell(row=summary_row, column=5).fill = subtotal_fill
    ws_summary.cell(row=summary_row, column=5).number_format = '₹#,##0.00'
    
    for col in range(1, 6):
        ws_summary.cell(row=summary_row, column=col).border = thin_border
    
    # Column widths for summary
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 18
    
    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Kitchen_Ledger_Itemwise_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/export/vendor-ledger")
async def export_vendor_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vendor_id: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Export Vendor Ledger data to Excel.
    Uses the same data source as dashboard: purchase_orders.grn_amount
    OPTIMIZED: Uses batch processing to avoid timeouts.
    """
    try:
        # Build PO query - MAIN STORE ONLY
        po_query = {"created_by_location_name": "Main Store"}
        
        if vendor_id:
            po_query["vendor_id"] = vendor_id
            
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date + "T00:00:00"
            if end_date:
                date_filter["$lte"] = end_date + "T23:59:59"
            if date_filter:
                po_query["created_at"] = date_filter
        
        # Get all POs with pagination to avoid timeout
        all_pos = []
        batch_size = 100
        skip = 0
        max_docs = 500  # Limit total docs for performance
        
        while skip < max_docs:
            try:
                batch = list(purchase_orders_collection.find(
                    po_query,
                    {"po_number": 1, "created_at": 1, "total_amount": 1, "grn_amount": 1, 
                     "grn_date": 1, "grn_invoice_number": 1, "status": 1, "items": 1, "vendor_id": 1}
                ).skip(skip).limit(batch_size).max_time_ms(10000))
                
                if not batch:
                    break
                all_pos.extend(batch)
                skip += batch_size
                
                if len(batch) < batch_size:
                    break
            except Exception as batch_error:
                print(f"Batch {skip} error: {batch_error}")
                break
        
        print(f"Export: Retrieved {len(all_pos)} POs for vendor ledger")
        
        # Get vendor names in one query
        vendor_ids = list(set(po.get("vendor_id") for po in all_pos if po.get("vendor_id")))
        vendors_map = {}
        if vendor_ids:
            try:
                for v in vendors_collection.find(
                    {"_id": {"$in": [ObjectId(vid) for vid in vendor_ids[:50]]}},  # Limit vendors
                    {"name": 1}
                ).max_time_ms(5000):
                    vendors_map[str(v["_id"])] = v.get("name", "Unknown")
            except:
                pass
        
        rows = []
        vendor_totals = {}
        
        for po in all_pos:
            v_id = po.get("vendor_id", "")
            vendor_name = vendors_map.get(v_id, "Unknown Vendor")
            
            # Calculate PO amount from items if total_amount is 0
            po_amount = po.get("total_amount", 0) or 0
            if po_amount == 0:
                for item in po.get("items", []):
                    po_amount += (item.get("quantity", 0) or 0) * (item.get("rate", 0) or 0)
            
            grn_amount = po.get("grn_amount", 0) or 0
            
            # Track vendor totals
            if vendor_name not in vendor_totals:
                vendor_totals[vendor_name] = {"po_value": 0, "grn_value": 0}
            vendor_totals[vendor_name]["po_value"] += po_amount
            vendor_totals[vendor_name]["grn_value"] += grn_amount
            
            # Add row for each PO with GRN
            if grn_amount > 0:
                rows.append({
                    "Vendor": vendor_name,
                    "PO Number": po.get("po_number", ""),
                    "PO Date": po.get("created_at", "")[:10] if po.get("created_at") else "",
                    "PO Amount": round(po_amount, 2),
                    "GRN Date": po.get("grn_date", "")[:10] if po.get("grn_date") else "",
                    "GRN Amount": round(grn_amount, 2),
                    "Invoice Number": po.get("grn_invoice_number", ""),
                    "Status": po.get("status", "")
                })
        
        # Build summary rows
        summary_rows = []
        grand_total_po = 0
        grand_total_grn = 0
        
        for vendor_name, totals in sorted(vendor_totals.items()):
            if totals["grn_value"] > 0 or totals["po_value"] > 0:
                summary_rows.append({
                    "Vendor": vendor_name,
                    "Total PO Amount": round(totals["po_value"], 2),
                    "Total GRN Amount": round(totals["grn_value"], 2),
                    "Difference": round(totals["po_value"] - totals["grn_value"], 2)
                })
                grand_total_po += totals["po_value"]
                grand_total_grn += totals["grn_value"]
        
        summary_rows.append({
            "Vendor": "GRAND TOTAL",
            "Total PO Amount": round(grand_total_po, 2),
            "Total GRN Amount": round(grand_total_grn, 2),
            "Difference": round(grand_total_po - grand_total_grn, 2)
        })
        
        # Create Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if rows:
                df_detail = pd.DataFrame(rows)
                df_detail.to_excel(writer, index=False, sheet_name='GRN Details')
            else:
                df_detail = pd.DataFrame(columns=["Vendor", "PO Number", "PO Date", "PO Amount", "GRN Date", "GRN Amount", "Invoice Number", "Status"])
                df_detail.to_excel(writer, index=False, sheet_name='GRN Details')
            
            df_summary = pd.DataFrame(summary_rows)
            df_summary.to_excel(writer, index=False, sheet_name='Vendor Summary')
        
        output.seek(0)
        
        filename = f"Vendor_Ledger_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Vendor Ledger Export Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)[:100]}")


@app.get("/api/export/items")
async def export_items(
    category: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Export Items list to Excel"""
    
    # Build query
    query = {}
    if category and category != 'all':
        query["category"] = category
    
    # Get all items
    items = list(items_collection.find(query).sort("name", 1))
    
    # Get vendors for mapping
    vendors_map = {str(v["_id"]): v["name"] for v in vendors_collection.find()}
    
    # Build rows
    rows = []
    for idx, item in enumerate(items, 1):
        vendor_ids = item.get("vendor_ids", [])
        vendor_names = [vendors_map.get(vid, "") for vid in vendor_ids if vid in vendors_map]
        
        rows.append({
            "S.No": idx,
            "Item Name": item.get("name", ""),
            "Category": item.get("category", ""),
            "Unit": item.get("unit", ""),
            "Standard Price": item.get("standard_price", 0) or 0,
            "HSN Code": item.get("hsn_code", ""),
            "GST Rate": item.get("gst_rate", ""),
            "Vendors": ", ".join(vendor_names)
        })
    
    # Create Excel
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Items')
        
        # Auto-fit columns
        worksheet = writer.sheets['Items']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    output.seek(0)
    
    today = datetime.now().strftime("%Y%m%d")
    filename = f"items_list_{today}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/export/current-stock")
async def export_current_stock(
    category: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Export Current Stock to Excel for physical inventory verification"""
    
    # Define perishable categories that don't go to main store stock
    perishable_categories = ["Vegetables", "Non-Veg", "Dairy", "Daily Perishables"]
    
    # Get all items excluding perishables
    items_query = {"category": {"$nin": perishable_categories}}
    if category and category != 'all':
        items_query["category"] = category
    
    items = list(items_collection.find(items_query).sort([("category", 1), ("name", 1)]))
    
    # Get main store location
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Get today's date range for GRN calculation
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    today_start_str = today_start.isoformat()
    today_end_str = today_end.isoformat()
    
    # Build stock data
    rows = []
    for idx, item in enumerate(items, 1):
        item_id = str(item["_id"])
        
        # Calculate current stock from lots at main store
        stock_query = {"item_id": item_id, "current_quantity": {"$gt": 0}}
        if main_store_id:
            stock_query["location_id"] = main_store_id
        
        lots = list(lots_collection.find(stock_query))
        current_qty = sum(lot.get("current_quantity", 0) for lot in lots)
        
        # Calculate today's GRN (lots created today)
        todays_grn_query = {
            "item_id": item_id,
            "created_at": {"$gte": today_start_str, "$lt": today_end_str}
        }
        if main_store_id:
            todays_grn_query["location_id"] = main_store_id
        
        todays_grn_lots = list(lots_collection.find(todays_grn_query))
        todays_grn = sum(lot.get("initial_quantity", 0) for lot in todays_grn_lots)
        
        # Get par stock level
        par_stock = item.get("par_stock", 0) or 0
        
        # Determine status
        if par_stock > 0:
            status = "Below Par" if current_qty < par_stock else "OK"
        else:
            status = "No Par Set"
        
        rows.append({
            "S.No": idx,
            "Item Name": item.get("name", ""),
            "Category": item.get("category", ""),
            "Unit": item.get("unit", ""),
            "Current Stock": current_qty - todays_grn,  # Stock before today's GRN
            "Today's GRN": todays_grn if todays_grn > 0 else "NIL",
            "Total": current_qty,  # Current stock includes today's GRN
            "Par Stock": par_stock,
            "Variance": current_qty - par_stock if par_stock > 0 else "",
            "Status": status,
            "Physical Count": "",  # Empty column for user to fill
            "Difference": "",  # Empty column for user to calculate
            "Remarks": ""  # Empty column for notes
        })
    
    # Create Excel with formatting
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Current Stock')
        
        workbook = writer.book
        worksheet = writer.sheets['Current Stock']
        
        # Style the header
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        # Highlight Today's GRN column with green if has value
        green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        for row_idx in range(2, len(rows) + 2):
            grn_cell = worksheet.cell(row=row_idx, column=6)  # Today's GRN column
            if grn_cell.value and grn_cell.value != "NIL":
                grn_cell.fill = green_fill
        
        # Highlight "Below Par" rows
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(rows)+1), 2):
            status_cell = worksheet.cell(row=row_idx, column=10)  # Status column (shifted)
            if status_cell.value == "Below Par":
                for cell in row:
                    if cell.column != 6:  # Don't override GRN highlight
                        cell.fill = red_fill
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in worksheet.iter_rows(min_row=1, max_row=len(rows)+1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
    
    output.seek(0)
    
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"Current_Stock_{today}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/export/stock-movement")
async def export_stock_movement(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Export IN/OUT stock movement report in ledger format.
    Creates sheets per category with columns:
    S.No | Item Name | Unit | Current Stock | Par Stock | Day1 IN | Day1 OUT | Day2 IN | Day2 OUT | ... | Closing Balance
    """
    
    # Default to last 7 days if no dates provided
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start_dt = datetime.now(timezone.utc) - timedelta(days=6)
        start_date = start_dt.strftime("%Y-%m-%d")
    
    # Parse dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    
    # Generate list of dates in range
    dates = []
    current_date = start_dt
    while current_date <= end_dt:
        dates.append(current_date.strftime("%Y-%m-%d"))
        current_date += timedelta(days=1)
    
    # Get main store
    main_store = locations_collection.find_one({"type": "main_store"})
    main_store_id = str(main_store["_id"]) if main_store else None
    
    # Get all items sorted by category and name
    perishable_categories = ["Vegetables", "Non-Veg", "Dairy", "Daily Perishables", "Fruits"]
    items = list(items_collection.find({"category": {"$nin": perishable_categories}}).sort([("category", 1), ("name", 1)]))
    
    # Build IN data (GRN received) by item and date
    in_by_item_date = {}  # {item_id: {date: quantity}}
    
    for date_str in dates:
        day_start = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        day_end = day_start + timedelta(days=1)
        
        lots_query = {
            "created_at": {"$gte": day_start.isoformat(), "$lt": day_end.isoformat()}
        }
        if main_store_id:
            lots_query["location_id"] = main_store_id
        
        lots = list(lots_collection.find(lots_query))
        
        for lot in lots:
            item_id = lot.get("item_id")
            qty = lot.get("initial_quantity", 0)
            
            if item_id not in in_by_item_date:
                in_by_item_date[item_id] = {}
            if date_str not in in_by_item_date[item_id]:
                in_by_item_date[item_id][date_str] = 0
            in_by_item_date[item_id][date_str] += qty
    
    # Build OUT data (dispatched) by item and date
    out_by_item_date = {}  # {item_id: {date: quantity}}
    
    for date_str in dates:
        day_start = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        day_end = day_start + timedelta(days=1)
        
        txn_query = {
            "type": {"$in": ["dispatch", "issue", "transfer"]},
            "created_at": {"$gte": day_start.isoformat(), "$lt": day_end.isoformat()}
        }
        
        transactions = list(transactions_collection.find(txn_query))
        
        for txn in transactions:
            item_id = txn.get("item_id")
            qty = txn.get("quantity", 0)
            
            if item_id not in out_by_item_date:
                out_by_item_date[item_id] = {}
            if date_str not in out_by_item_date[item_id]:
                out_by_item_date[item_id][date_str] = 0
            out_by_item_date[item_id][date_str] += qty
    
    # Get current stock by item
    stock_by_item = {}
    if main_store_id:
        pipeline = [
            {"$match": {"location_id": main_store_id, "current_quantity": {"$gt": 0}}},
            {"$group": {"_id": "$item_id", "total": {"$sum": "$current_quantity"}}}
        ]
        for doc in lots_collection.aggregate(pipeline):
            stock_by_item[doc["_id"]] = doc["total"]
    
    # Group items by category
    items_by_category = {}
    for item in items:
        cat = item.get("category", "Other") or "Other"
        if cat not in items_by_category:
            items_by_category[cat] = []
        items_by_category[cat].append(item)
    
    # Create Excel workbook
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styles
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    in_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green for IN
    out_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Light red for OUT
    balance_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue for balance
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Create a sheet for each category
    for category, cat_items in items_by_category.items():
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = category[:31] if len(category) > 31 else category
        ws = wb.create_sheet(title=sheet_name)
        
        # Build headers
        headers = ["S.No", "Item Name", "Unit", "Opening Stock", "Par Stock"]
        for date_str in dates:
            # Format date as DD-MMM (e.g., 12-Feb)
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_label = dt.strftime("%d-%b")
            headers.append(f"{date_label} IN")
            headers.append(f"{date_label} Out")
        headers.append("Closing Balance")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # Write data rows
        for row_idx, item in enumerate(cat_items, 2):
            item_id = str(item["_id"])
            current_stock = stock_by_item.get(item_id, 0)
            par_stock = item.get("par_stock", 0) or 0
            
            # Calculate total IN and OUT for the period
            total_in = sum(in_by_item_date.get(item_id, {}).values())
            total_out = sum(out_by_item_date.get(item_id, {}).values())
            
            # Opening balance = Current Stock - Total IN + Total OUT
            # (because current_stock already includes all IN and excludes all OUT)
            opening_balance = current_stock - total_in + total_out
            
            # Build row data - show Opening Stock in column 4
            row_data = [
                row_idx - 1,  # S.No
                item.get("name", ""),
                item.get("unit", ""),
                opening_balance,  # Opening Stock (stock at START of period)
                par_stock if par_stock > 0 else ""
            ]
            
            # Add daily IN/OUT
            for date_str in dates:
                day_in = in_by_item_date.get(item_id, {}).get(date_str, 0)
                day_out = out_by_item_date.get(item_id, {}).get(date_str, 0)
                
                row_data.append(day_in if day_in > 0 else "")
                row_data.append(day_out if day_out > 0 else "")
            
            # Closing balance = Opening Stock + Total IN - Total OUT
            closing_balance = opening_balance + total_in - total_out
            row_data.append(closing_balance)
            
            # Write row
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = center_align if col > 2 else Alignment(horizontal="left", vertical="center")
                
                # Apply conditional formatting
                if col > 5:  # Date columns
                    col_in_date_section = col - 5
                    if col_in_date_section % 2 == 1:  # IN column
                        if value and value != "":
                            cell.fill = in_fill
                    elif col_in_date_section % 2 == 0:  # OUT column
                        if value and value != "":
                            cell.fill = out_fill
                
                # Closing balance column
                if col == len(row_data):
                    cell.fill = balance_fill
                    cell.font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 6  # S.No
        ws.column_dimensions['B'].width = 40  # Item Name
        ws.column_dimensions['C'].width = 8  # Unit
        ws.column_dimensions['D'].width = 12  # Current Stock
        ws.column_dimensions['E'].width = 10  # Par Stock
        
        # Date columns
        for col_idx in range(6, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        
        # Closing balance column
        ws.column_dimensions[get_column_letter(len(headers))].width = 14
    
    # Save workbook
    wb.save(output)
    output.seek(0)
    
    filename = f"Stock_Ledger_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/export/daily-perishables")
async def export_daily_perishables(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    kitchen_id: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Export Daily Perishables data to Excel"""
    kitchen_receivables = db["kitchen_receivables"]
    
    query = {}
    if kitchen_id:
        query["kitchen_id"] = kitchen_id
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        if date_filter:
            query["receive_date"] = date_filter
    
    receivables = list(kitchen_receivables.find(query).sort("receive_date", -1))
    
    rows = []
    for rec in receivables:
        kitchen = locations_collection.find_one({"_id": ObjectId(rec.get("kitchen_id", ""))})
        rows.append({
            "Date": rec.get("receive_date", ""),
            "Kitchen": rec.get("kitchen_name", kitchen.get("name", "Unknown") if kitchen else "Unknown"),
            "Kitchen Code": kitchen.get("code", "") if kitchen else "",
            "Item": rec.get("item_name", ""),
            "Category": rec.get("category", ""),
            "Quantity": rec.get("quantity", 0),
            "Unit": rec.get("unit", ""),
            "Rate": rec.get("rate", 0),
            "Amount": rec.get("amount", 0),
            "Vendor": rec.get("vendor_name", ""),
            "PO Number": rec.get("po_number", ""),
            "Invoice": rec.get("invoice_number", "")
        })
    
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Daily Perishables')
    output.seek(0)
    
    filename = f"daily_perishables_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/export/purchase-orders")
async def export_purchase_orders(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    vendor_id: Optional[str] = None,
    vendor_ids: Optional[str] = None,  # Comma-separated list of vendor IDs (max 6)
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Export Purchase Orders data to Excel - includes vendor filter and multi-vendor support"""
    try:
        query = {}
        if status:
            query["status"] = status
        
        # Handle multiple vendors (comma-separated, max 6)
        vendor_id_list = []
        if vendor_ids:
            vendor_id_list = [v.strip() for v in vendor_ids.split(",") if v.strip()][:6]  # Max 6 vendors
            if vendor_id_list:
                query["vendor_id"] = {"$in": vendor_id_list}
        elif vendor_id:
            query["vendor_id"] = vendor_id
            vendor_id_list = [vendor_id]
        
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date + "T23:59:59"
            if date_filter:
                query["created_at"] = date_filter
        
        # Require at least one filter (vendor or date range) to prevent timeout on large datasets
        if not query:
            raise HTTPException(
                status_code=400, 
                detail="Please select at least one vendor or specify a date range to export. Exporting all data without filters may timeout."
            )
        
        # Use aggregation pipeline for better performance on production Atlas
        pipeline = [
            {"$match": query},
            {"$sort": {"created_at": -1}},
            {"$limit": 200},  # Strict limit to prevent timeout
            {"$project": {
                "_id": 0,
                "po_number": 1,
                "created_at": 1,
                "vendor_id": 1,
                "vendor_name": 1,
                "delivery_address": 1,
                "created_by_location_name": 1,  # Include location that raised the PO
                "status": 1,
                "items": 1,
                "grn_amount": 1,
                "grn_date": 1,
                "grn_invoice_number": 1,
                "grn_verification": {"$ifNull": ["$grn_verification", {}]}
            }}
        ]
        
        pos = list(purchase_orders_collection.aggregate(pipeline, maxTimeMS=20000, allowDiskUse=True))
        
        if len(pos) == 200:
            print(f"PO Export: Hit 200 limit, query had more results. Query: {query}")
        
        # Batch fetch all vendors at once instead of one-by-one (N+1 query fix)
        unique_vendor_ids = list(set(po.get("vendor_id") for po in pos if po.get("vendor_id")))
        vendor_map = {}
        if unique_vendor_ids:
            try:
                vendor_oids = [ObjectId(vid) for vid in unique_vendor_ids if vid]
                vendors_list = list(vendors_collection.find(
                    {"_id": {"$in": vendor_oids}},
                    {"_id": 1, "name": 1}
                ))
                vendor_map = {str(v["_id"]): v.get("name", "") for v in vendors_list}
            except Exception as e:
                print(f"Error fetching vendors: {e}")
        
        rows = []
        for po in pos:
            # Get vendor name from map (or use stored vendor_name as fallback)
            vendor_name = vendor_map.get(po.get("vendor_id", ""), po.get("vendor_name", ""))
            
            # Get items summary
            items_list = []
            total_qty = 0
            total_value = 0
            for item in po.get("items", []):
                items_list.append(item.get("item_name", ""))
                total_qty += item.get("quantity", 0) or 0
                total_value += (item.get("quantity", 0) or 0) * (item.get("rate", 0) or 0)
            
            created_at = po.get("created_at", "")
            created_date = created_at[:10] if created_at and len(created_at) >= 10 else ""
            
            rows.append({
                "PO Number": po.get("po_number", ""),
                "Date": created_date,
                "Vendor": vendor_name,
                "Raised By Location": po.get("created_by_location_name", "Main Store"),  # Show who raised the PO
                "Delivery Address": po.get("delivery_address", ""),
                "Status": po.get("status", ""),
                "Items Count": len(po.get("items", [])),
                "Total Quantity": round(total_qty, 2),
                "PO Value": round(total_value, 2),
                "GRN Amount": po.get("grn_amount", ""),
                "GRN Date": po.get("grn_date", "")[:10] if po.get("grn_date") else "",
                "Invoice Number": po.get("grn_invoice_number", ""),
                "Items": ", ".join(items_list[:5]) + ("..." if len(items_list) > 5 else ""),
                "Has Photo": "Yes" if (po.get("grn_verification") or {}).get("photo") else "No"
            })
        
        # Return error if no data found
        if not rows:
            raise HTTPException(
                status_code=404,
                detail=f"No purchase orders found for the selected filters. Check: 1) Selected vendors have POs, 2) Date range contains PO dates. Query: vendor_ids={vendor_id_list}, dates={start_date} to {end_date}"
            )
        
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Purchase Orders')
        output.seek(0)
        
        # Build descriptive filename
        vendor_str = f"_{len(vendor_id_list)}vendors" if vendor_id_list else ""
        date_str = f"_{start_date or 'all'}_to_{end_date or 'all'}" if start_date or end_date else ""
        filename = f"purchase_orders{vendor_str}{date_str}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"PO Export error: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)[:100]}")

@app.get("/api/export/all-data")
async def export_all_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user = Depends(require_role(["admin"]))
):
    """Export all data to Excel (Admin only) - Data Backup"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Locations/Kitchens
        locations = list(locations_collection.find({}))
        loc_rows = [{
            "Code": loc.get("code", ""),
            "Name": loc.get("name", ""),
            "Type": loc.get("type", ""),
            "Address": loc.get("address", ""),
            "Contact Person": loc.get("contact_person", ""),
            "Contact Phone": loc.get("contact_phone", "")
        } for loc in locations]
        pd.DataFrame(loc_rows).to_excel(writer, index=False, sheet_name='Locations')
        
        # Items
        items = list(items_collection.find({}))
        items_rows = [{
            "Name": item.get("name", ""),
            "Category": item.get("category", ""),
            "Unit": item.get("unit", ""),
            "Standard Price": item.get("standard_price", 0),
            "HSN Code": item.get("hsn_code", "")
        } for item in items]
        pd.DataFrame(items_rows).to_excel(writer, index=False, sheet_name='Items')
        
        # Vendors
        vendors = list(vendors_collection.find({}))
        vendor_rows = [{
            "Name": v.get("name", ""),
            "Contact": v.get("contact", ""),
            "Email": v.get("email", ""),
            "Phone": v.get("phone", ""),
            "Address": v.get("address", ""),
            "GST": v.get("gst_number", "")
        } for v in vendors]
        pd.DataFrame(vendor_rows).to_excel(writer, index=False, sheet_name='Vendors')
        
        # Purchase Orders (last 3 months)
        three_months_ago = (datetime.now(timezone.utc) - timedelta(days=90)).isoformat()
        po_query = {"created_at": {"$gte": three_months_ago}}
        if start_date:
            po_query["created_at"]["$gte"] = start_date
        if end_date:
            po_query["created_at"]["$lte"] = end_date + "T23:59:59"
        
        pos = list(purchase_orders_collection.find(po_query).sort("created_at", -1))
        po_rows = []
        for po in pos:
            vendor = vendors_collection.find_one({"_id": ObjectId(po.get("vendor_id", ""))})
            total = sum(i.get("quantity", 0) * i.get("rate", 0) for i in po.get("items", []))
            po_rows.append({
                "PO Number": po.get("po_number", ""),
                "Date": po.get("created_at", "")[:10],
                "Vendor": vendor.get("name", "") if vendor else "",
                "Delivery Address": po.get("delivery_address", ""),
                "Status": po.get("status", ""),
                "PO Value": total,
                "GRN Amount": po.get("grn_amount", ""),
                "GRN Date": po.get("grn_date", "")[:10] if po.get("grn_date") else ""
            })
        pd.DataFrame(po_rows).to_excel(writer, index=False, sheet_name='Purchase Orders')
        
        # Kitchen Receivables
        kr_query = {}
        if start_date:
            kr_query["receive_date"] = {"$gte": start_date}
        if end_date:
            if "receive_date" not in kr_query:
                kr_query["receive_date"] = {}
            kr_query["receive_date"]["$lte"] = end_date
        
        kitchen_receivables = db["kitchen_receivables"]
        kr = list(kitchen_receivables.find(kr_query).sort("receive_date", -1))
        kr_rows = [{
            "Date": r.get("receive_date", ""),
            "Kitchen": r.get("kitchen_name", ""),
            "Item": r.get("item_name", ""),
            "Category": r.get("category", ""),
            "Quantity": r.get("quantity", 0),
            "Rate": r.get("rate", 0),
            "Amount": r.get("amount", 0),
            "Vendor": r.get("vendor_name", "")
        } for r in kr]
        pd.DataFrame(kr_rows).to_excel(writer, index=False, sheet_name='Daily Perishables')
        
        # Transactions Summary
        txn_query = {}
        if start_date:
            txn_query["created_at"] = {"$gte": start_date}
        if end_date:
            if "created_at" not in txn_query:
                txn_query["created_at"] = {}
            txn_query["created_at"]["$lte"] = end_date + "T23:59:59"
        
        txns = list(transactions_collection.find(txn_query).sort("created_at", -1).limit(5000))
        txn_rows = [{
            "Date": t.get("created_at", "")[:10],
            "Type": t.get("type", ""),
            "Quantity": t.get("quantity", 0),
            "Rate": t.get("rate", 0),
            "Value": t.get("value", 0),
            "Notes": t.get("notes", "")
        } for t in txns]
        pd.DataFrame(txn_rows).to_excel(writer, index=False, sheet_name='Transactions')
    
    output.seek(0)
    
    filename = f"kinfolk_data_backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Seed Data
@app.post("/api/seed")
async def seed_data():
    locations_collection.delete_many({})
    categories_collection.delete_many({})
    vendors_collection.delete_many({})
    
    locations = [
        {"name": "Main Store", "type": "main_store", "address": "Main Store Address, City"},
        {"name": "Sticky Rice GCR", "type": "kitchen", "address": "GCR Mall, Gurugram"},
        {"name": "Kalaunji GCR", "type": "kitchen", "address": "GCR Mall, Gurugram"},
        {"name": "Sticky Rice Bestech", "type": "kitchen", "address": "Bestech Mall, Gurugram"},
        {"name": "Kalaunji Bestech", "type": "kitchen", "address": "Bestech Mall, Gurugram"},
        {"name": "Mala Project", "type": "kitchen", "address": "Mala Project, Gurugram"},
        {"name": "Tanyan", "type": "kitchen", "address": "Tanyan, Gurugram"},
        {"name": "Sticky Rice Elan Epic", "type": "kitchen", "address": "Elan Epic Mall, Gurugram"},
        {"name": "Super Salsa", "type": "kitchen", "address": "Super Salsa, Gurugram"},
    ]
    
    for loc in locations:
        loc["created_at"] = datetime.now(timezone.utc).isoformat()
    locations_collection.insert_many(locations)
    
    categories = [
        "Beverage", "Indian Grocery", "Chinese Grocery", 
        "Continental Grocery", "Housekeeping", "Dairy", 
        "Seafood", "Packaging"
    ]
    
    for cat in categories:
        categories_collection.insert_one({
            "name": cat,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    vendors = [
        "CP", "FOOD MONDE", "SHREE JEE ENTERPRISES", "ECO TEGRITY",
        "MK HOSPITALITY SOLUTION", "hyperpure", "Beeta Machines Pvt. Ltd.",
        "SELICO PLASTICS", "INSTAPACK ENTERPRISE", "HONETPRINTER",
        "PL ENTERPRISES", "PRIME FOODS", "PACKEDGE", "KRISHAN BHANDAR",
        "Y K TRADING COMPANY"
    ]
    
    for vendor in vendors:
        vendors_collection.insert_one({
            "name": vendor,
            "contact": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    sample_items = [
        {"name": "COCA COLA CAN 300 ML", "category": "Beverage", "unit": "CASE", "gst_rate": 18, "vendor": "CP", "standard_price": 650},
        {"name": "TATA TEA AGNI 500 G", "category": "Indian Grocery", "unit": "PKT", "gst_rate": 5, "vendor": "CP", "standard_price": 180},
        {"name": "MDH WHITE PEPPER POWDER 100 G", "category": "Indian Grocery", "unit": "PKT", "gst_rate": 5, "vendor": "CP", "standard_price": 95},
        {"name": "GOLDEN CROWN TOMATO KETCHUP 6 KG", "category": "Indian Grocery", "unit": "CAN", "gst_rate": 18, "vendor": "CP", "standard_price": 520},
        {"name": "GAINDA WHITE DISINFECTANT FLOOR CLEANER 5 L", "category": "Housekeeping", "unit": "CAN", "gst_rate": 18, "vendor": "CP", "standard_price": 280},
        {"name": "FOOD MONDE BAMBOO SHOOT 560 GRM", "category": "Chinese Grocery", "unit": "TIN", "gst_rate": 18, "vendor": "FOOD MONDE", "standard_price": 165},
        {"name": "AASHIRVAAD ATTA 10 KG", "category": "Indian Grocery", "unit": "BAG", "gst_rate": 0, "vendor": "CP", "standard_price": 480},
        {"name": "NB 3611 DUBAR RICE 30 KG", "category": "Indian Grocery", "unit": "BAG", "gst_rate": 0, "vendor": "CP", "standard_price": 1450},
        {"name": "REAL JUICE MIXED FRUIT 1 L", "category": "Beverage", "unit": "BTL", "gst_rate": 18, "vendor": "CP", "standard_price": 95},
        {"name": "HARPIC TOILET CLEANER 500 ML", "category": "Housekeeping", "unit": "BTL", "gst_rate": 18, "vendor": "CP", "standard_price": 145},
    ]
    
    for item in sample_items:
        item["hsn_code"] = None
        item["created_at"] = datetime.now(timezone.utc).isoformat()
    
    if items_collection.count_documents({}) == 0:
        items_collection.insert_many(sample_items)
    
    # Create default admin user if none exists
    if users_collection.count_documents({"role": "admin"}) == 0:
        users_collection.insert_one({
            "email": "adreamoven@gmail.com",
            "password": hash_password("Store@123"),
            "name": "Admin",
            "role": "admin",
            "location_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return {
        "message": "Seed data created successfully",
        "locations": len(locations),
        "categories": len(categories),
        "vendors": len(vendors),
        "sample_items": len(sample_items),
        "default_admin": "adreamoven@gmail.com / Store@123"
    }

# ============ Daywise GRN Reports ============

@app.get("/api/export/daywise-kitchen-grn")
async def export_daywise_kitchen_grn(
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Export daywise GRN summary for all kitchens.
    Data sources:
    1. Items received from Main Store (issue transactions)
    2. Daily perishables received directly from vendors (kitchen_receivables)
    """
    
    # Fixed kitchen order
    kitchen_order = [
        "Sticky Rice GCR", "Sticky Rice Bestech", "Sticky Rice Elan Epic",
        "Kalaunji GCR", "Kalaunji Bestech", "Mala Project", "Tanyan", "Super Salsa"
    ]
    
    # Get all kitchens
    all_kitchens = {}
    for k in locations_collection.find({"type": "kitchen"}):
        all_kitchens[str(k["_id"])] = k["name"]
    
    # Initialize kitchen data with correct categories and counts
    kitchen_data = {}
    for name in kitchen_order:
        kitchen_data[name] = {
            "main_store": {"Groceries": 0, "Beverages": 0, "Packaging": 0, "Seafood": 0, "Housekeeping": 0, "Non Veg": 0},
            "perishables": {"Vegetables": 0, "Dairy": 0},
            "perishable_count": 0,
            "main_store_count": 0,
            "total": 0,
            # Track GRN dates per category for remarks
            "category_grn_dates": {
                "Non Veg": [],
                "Vegetables": [],
                "Dairy": []
            },
            # Track actual GRN record count per category
            "category_grn_counts": {
                "Non Veg": 0,
                "Vegetables": 0,
                "Dairy": 0
            }
        }
    
    # 1. Get Daily Perishables from kitchen_receivables
    # Filter by receive_date which is the actual date of GRN receipt
    kitchen_receivables = db["kitchen_receivables"]
    kr_data = list(kitchen_receivables.find({"receive_date": date}))
    
    for rec in kr_data:
        kitchen_name = rec.get("kitchen_name", "")
        # Match kitchen name
        matched_kitchen = None
        for k_name in kitchen_order:
            if k_name.lower() in kitchen_name.lower() or kitchen_name.lower() in k_name.lower():
                matched_kitchen = k_name
                break
        
        if matched_kitchen:
            cat = rec.get("category", "Vegetables")
            item_name = rec.get("item_name", "").lower()
            amount = float(rec.get("amount", 0) or 0)
            
            # Get the PO/GRN date (could be different from receive_date if entered later)
            grn_date = rec.get("grn_date") or rec.get("po_date") or rec.get("receive_date", date)
            # Format date as d/m
            if grn_date and len(grn_date) >= 10:
                try:
                    day = grn_date[8:10].lstrip('0')
                    month = grn_date[5:7].lstrip('0')
                    formatted_date = f"{day}/{month}"
                except:
                    formatted_date = grn_date[:10]
            else:
                formatted_date = grn_date
            
            # Increment perishable count
            kitchen_data[matched_kitchen]["perishable_count"] += 1
            
            # Map category for perishables and track dates
            if "Veg" in cat and "Non" not in cat and "NON" not in cat:
                kitchen_data[matched_kitchen]["perishables"]["Vegetables"] += amount
                kitchen_data[matched_kitchen]["category_grn_counts"]["Vegetables"] += 1
                if formatted_date not in kitchen_data[matched_kitchen]["category_grn_dates"]["Vegetables"]:
                    kitchen_data[matched_kitchen]["category_grn_dates"]["Vegetables"].append(formatted_date)
            elif "Non" in cat or "NON" in cat:
                kitchen_data[matched_kitchen]["main_store"]["Non Veg"] += amount
                kitchen_data[matched_kitchen]["category_grn_counts"]["Non Veg"] += 1
                if formatted_date not in kitchen_data[matched_kitchen]["category_grn_dates"]["Non Veg"]:
                    kitchen_data[matched_kitchen]["category_grn_dates"]["Non Veg"].append(formatted_date)
            elif "Dairy" in cat:
                kitchen_data[matched_kitchen]["perishables"]["Dairy"] += amount
                kitchen_data[matched_kitchen]["category_grn_counts"]["Dairy"] += 1
                if formatted_date not in kitchen_data[matched_kitchen]["category_grn_dates"]["Dairy"]:
                    kitchen_data[matched_kitchen]["category_grn_dates"]["Dairy"].append(formatted_date)
            else:
                kitchen_data[matched_kitchen]["perishables"]["Vegetables"] += amount
                kitchen_data[matched_kitchen]["category_grn_counts"]["Vegetables"] += 1
                if formatted_date not in kitchen_data[matched_kitchen]["category_grn_dates"]["Vegetables"]:
                    kitchen_data[matched_kitchen]["category_grn_dates"]["Vegetables"].append(formatted_date)
            
            kitchen_data[matched_kitchen]["total"] += amount
    
    # 2. Get Main Store dispatches (issue transactions)
    dispatch_data = list(transactions_collection.find({
        "type": "issue",
        "created_at": {"$regex": f"^{date}"}
    }))
    
    for txn in dispatch_data:
        dest_id = txn.get("destination_location_id")
        kitchen_name = all_kitchens.get(dest_id, "")
        
        matched_kitchen = None
        for k_name in kitchen_order:
            if k_name.lower() in kitchen_name.lower() or kitchen_name.lower() in k_name.lower():
                matched_kitchen = k_name
                break
        
        if matched_kitchen:
            item_id = txn.get("item_id")
            value = float(txn.get("value", 0) or 0)
            
            if item_id:
                item = items_collection.find_one({"_id": ObjectId(item_id)})
                if item:
                    cat = item.get("category", "Grocery")
                    
                    # Map to standard categories
                    if "Grocery" in cat or "MALA" in cat or "Spice" in cat or "Season" in cat:
                        kitchen_data[matched_kitchen]["main_store"]["Groceries"] += value
                    elif "Beverage" in cat:
                        kitchen_data[matched_kitchen]["main_store"]["Beverages"] += value
                    elif "Pack" in cat:
                        kitchen_data[matched_kitchen]["main_store"]["Packaging"] += value
                    elif "Sea" in cat or "fish" in cat.lower():
                        kitchen_data[matched_kitchen]["main_store"]["Seafood"] += value
                    elif "House" in cat:
                        kitchen_data[matched_kitchen]["main_store"]["Housekeeping"] += value
                    elif "Non" in cat or "NON" in cat:
                        kitchen_data[matched_kitchen]["main_store"]["Non Veg"] += value
                    else:
                        kitchen_data[matched_kitchen]["main_store"]["Groceries"] += value
                    
                    kitchen_data[matched_kitchen]["total"] += value
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Kitchen GRN"
    
    # Define black font for visibility
    black_font = Font(color="000000")
    black_bold_font = Font(bold=True, color="000000")
    
    # Row 2: Headers
    ws['B2'] = f"Date = {date}"
    ws['B2'].font = black_bold_font
    
    ws.merge_cells('C2:G2')
    ws['C2'] = "Main Store"
    ws['C2'].font = black_bold_font
    ws['C2'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('H2:K2')
    ws['H2'] = "Daily Perishable"
    ws['H2'].font = black_bold_font
    ws['H2'].alignment = Alignment(horizontal="center")
    
    # Row 3: Column headers - added Remarks column
    headers = ["", "Kitchens", "Groceries", "Beverages", "Packaging", "Seafood", "Housekeeping", "Non Veg",
               "Vegetables", "Dairy", "Total", "Remarks"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = black_bold_font
    
    # Data rows
    row_num = 4
    grand_total = 0
    total_perishable_count = 0
    
    for idx, kitchen_name in enumerate(kitchen_order, 1):
        data = kitchen_data.get(kitchen_name, {"main_store": {}, "perishables": {}, "perishable_count": 0, "total": 0, "category_grn_dates": {}, "category_grn_counts": {}})
        ms = data["main_store"]
        ps = data["perishables"]
        total = data["total"]
        perishable_count = data.get("perishable_count", 0)
        category_dates = data.get("category_grn_dates", {})
        category_counts = data.get("category_grn_counts", {})
        grand_total += total
        total_perishable_count += perishable_count
        
        # Build remarks string: "Non Veg: 5 items (3 dates: 24/2, 25/2, 26/2)"
        # Now shows actual record count AND date count for transparency
        remarks_parts = []
        for cat_name in ["Non Veg", "Vegetables", "Dairy"]:
            dates = category_dates.get(cat_name, [])
            record_count = category_counts.get(cat_name, 0)
            if dates or record_count > 0:
                dates_str = ", ".join(sorted(dates, key=lambda x: (int(x.split('/')[1]) if '/' in x else 0, int(x.split('/')[0]) if '/' in x else 0)))
                # Show "X items on Y dates: date1, date2"
                remarks_parts.append(f"{cat_name}: {record_count} items ({len(dates)} dates: {dates_str})")
        
        remarks = ", ".join(remarks_parts) if remarks_parts else ""
        
        row_data = [
            idx, kitchen_name,
            ms.get("Groceries", 0), ms.get("Beverages", 0), ms.get("Packaging", 0), 
            ms.get("Seafood", 0), ms.get("Housekeeping", 0), ms.get("Non Veg", 0),
            ps.get("Vegetables", 0), ps.get("Dairy", 0),
            total, remarks
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value if value else "")
            cell.font = black_font
            if col >= 3 and col <= 10 and value and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            elif col == 11 and value and isinstance(value, (int, float)):  # Total column
                cell.number_format = '#,##0.00'
        
        row_num += 1
    
    # Total row
    ws.cell(row=row_num, column=2, value="TOTAL").font = black_bold_font
    total_cell = ws.cell(row=row_num, column=11, value=grand_total)
    total_cell.font = black_bold_font
    total_cell.number_format = '#,##0.00'
    
    # Column widths - updated for remarks column
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    for i in range(3, 12):
        ws.column_dimensions[chr(64 + i)].width = 14
    ws.column_dimensions['L'].width = 50  # Wider for remarks
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=kitchen_grn_{date}.xlsx"}
    )


@app.get("/api/export/daywise-mainstore-input")
async def export_daywise_mainstore_input(
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Export daywise Main Store GRN summary.
    Simple format: Category | Vendor | Amount
    """
    
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        raise HTTPException(status_code=404, detail="Main Store not configured")
    
    main_store_id = str(main_store["_id"])
    
    # Get all vendors and items
    vendors_map = {str(v["_id"]): v["name"] for v in vendors_collection.find()}
    items_map = {str(i["_id"]): {"name": i["name"], "category": i.get("category", "Other")} for i in items_collection.find()}
    
    # Data collection: category -> vendor -> amount
    category_vendor_amounts = {}
    
    # 1. From lots created on this date
    lots_data = list(lots_collection.find({
        "location_id": main_store_id,
        "created_at": {"$regex": f"^{date}"}
    }))
    
    for lot in lots_data:
        vendor_id = lot.get("vendor_id")
        vendor_name = vendors_map.get(vendor_id, "Unknown Vendor")
        item_id = lot.get("item_id")
        item_info = items_map.get(item_id, {"category": "Other"})
        category = item_info.get("category", "Other")
        
        qty = lot.get("initial_quantity", 0)
        rate = lot.get("purchase_rate", 0) or 0
        amount = qty * rate
        
        if category not in category_vendor_amounts:
            category_vendor_amounts[category] = {}
        if vendor_name not in category_vendor_amounts[category]:
            category_vendor_amounts[category][vendor_name] = 0
        category_vendor_amounts[category][vendor_name] += amount
    
    # 2. From completed POs with GRN on this date
    completed_pos = list(purchase_orders_collection.find({
        "status": {"$in": ["received", "completed"]},
        "grn_verification.capture_time": {"$regex": f"^{date}"}
    }))
    
    for po in completed_pos:
        vendor_name = po.get("vendor_name", "Unknown Vendor")
        
        for item in po.get("items", []):
            item_id = item.get("item_id")
            item_info = items_map.get(item_id, {"category": "Other"})
            category = item_info.get("category", "Other")
            
            qty = item.get("received_quantity") or item.get("quantity", 0)
            rate = item.get("rate", 0) or 0
            amount = qty * rate
            
            if category not in category_vendor_amounts:
                category_vendor_amounts[category] = {}
            if vendor_name not in category_vendor_amounts[category]:
                category_vendor_amounts[category][vendor_name] = 0
            category_vendor_amounts[category][vendor_name] += amount
    
    # Create Excel - Simple format: Category | Vendor | Amount
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Store GRN"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Main Store GRN Summary - Date: {date}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ["Category", "Vendor", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row_num = 4
    grand_total = 0
    
    # Sort categories
    sorted_categories = sorted(category_vendor_amounts.keys())
    
    for category in sorted_categories:
        vendors = category_vendor_amounts[category]
        for vendor_name in sorted(vendors.keys()):
            amount = vendors[vendor_name]
            if amount > 0:
                ws.cell(row=row_num, column=1, value=category).border = thin_border
                ws.cell(row=row_num, column=2, value=vendor_name).border = thin_border
                cell = ws.cell(row=row_num, column=3, value=amount)
                cell.border = thin_border
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
                grand_total += amount
                row_num += 1
    
    # Grand total
    row_num += 1
    ws.cell(row=row_num, column=2, value="GRAND TOTAL").font = Font(bold=True)
    cell = ws.cell(row=row_num, column=3, value=grand_total)
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mainstore_grn_{date}.xlsx"}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)

# Direct fix for specific requisition
@app.post("/api/setup/fix-requisition-kitchen")
async def fix_requisition_kitchen(secret_key: str = "", req_id: str = "", kitchen_name: str = "", kitchen_id: str = ""):
    """Fix a specific requisition's kitchen info"""
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    if not req_id:
        raise HTTPException(status_code=400, detail="req_id required")
    
    update = {}
    if kitchen_name:
        update["kitchen_name"] = kitchen_name
    if kitchen_id:
        update["kitchen_id"] = kitchen_id
    
    if update:
        from bson import ObjectId
        result = requisitions_collection.update_one(
            {"_id": ObjectId(req_id)},
            {"$set": update}
        )
        return {"success": True, "modified": result.modified_count}
    
    return {"success": False, "message": "No updates provided"}

# Fix POs with missing location - assign to Main Store
@app.post("/api/setup/fix-po-locations")
async def fix_po_locations(secret_key: str = ""):
    """Fix POs that have missing/null location by assigning to Main Store"""
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Get Main Store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        main_store = locations_collection.find_one({"name": {"$regex": "main", "$options": "i"}})
    
    if not main_store:
        raise HTTPException(status_code=404, detail="Main Store location not found")
    
    main_store_id = str(main_store["_id"])
    main_store_name = main_store["name"]
    
    # Update all POs with null/missing location_id
    result = purchase_orders_collection.update_many(
        {"$or": [
            {"location_id": None},
            {"location_id": ""},
            {"location_id": {"$exists": False}}
        ]},
        {"$set": {
            "location_id": main_store_id,
            "location_name": main_store_name
        }}
    )
    
    return {
        "success": True,
        "updated": result.modified_count,
        "main_store_id": main_store_id,
        "main_store_name": main_store_name
    }

# Fix requisitions with wrong kitchen - use created_by user's actual location
@app.post("/api/setup/fix-requisition-kitchens-v2")
async def fix_requisition_kitchens_v2(secret_key: str = ""):
    """
    Fix ALL requisitions by looking up the creating user's actual kitchen location.
    This ensures each requisition shows the correct originating kitchen.
    """
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Build lookup maps
    locations = {str(loc["_id"]): loc for loc in locations_collection.find({})}
    users = {}
    for u in users_collection.find({}):
        users[str(u["_id"])] = u
    
    fixed_count = 0
    failed_list = []
    
    # Process ALL requisitions to ensure correct kitchen assignment
    for req in requisitions_collection.find({}):
        req_id = req["_id"]
        serial = req.get("serial_number", str(req_id))
        
        # Get created_by user
        created_by = req.get("created_by")
        if created_by:
            created_by_str = str(created_by)
            user = users.get(created_by_str)
            
            if user and user.get("location_id"):
                user_loc_id = str(user["location_id"])
                location = locations.get(user_loc_id)
                
                if location:
                    # Update requisition with correct kitchen info
                    requisitions_collection.update_one(
                        {"_id": req_id},
                        {"$set": {
                            "location_id": ObjectId(user_loc_id),
                            "location_name": location["name"],
                            "kitchen_id": user_loc_id,
                            "kitchen_name": location["name"],
                            "kitchen_address": location.get("address", "")
                        }}
                    )
                    fixed_count += 1
                else:
                    failed_list.append(f"{serial}: user location {user_loc_id} not found")
            else:
                failed_list.append(f"{serial}: user {created_by_str} has no location")
        else:
            # Try kitchen_id as fallback
            kitchen_id = req.get("kitchen_id") or req.get("location_id")
            if kitchen_id:
                kitchen_id_str = str(kitchen_id)
                location = locations.get(kitchen_id_str)
                if location:
                    requisitions_collection.update_one(
                        {"_id": req_id},
                        {"$set": {
                            "location_id": ObjectId(kitchen_id_str),
                            "location_name": location["name"],
                            "kitchen_name": location["name"],
                            "kitchen_address": location.get("address", "")
                        }}
                    )
                    fixed_count += 1
                else:
                    failed_list.append(f"{serial}: no created_by and kitchen_id {kitchen_id_str} not found")
            else:
                failed_list.append(f"{serial}: no created_by or kitchen_id")
    
    return {
        "success": True,
        "fixed": fixed_count,
        "failed_count": len(failed_list),
        "failed_details": failed_list[:20]  # Limit to first 20 failures
    }

# Fix users with old/invalid location IDs by matching email to kitchen name
@app.post("/api/setup/fix-user-locations")
async def fix_user_locations(secret_key: str = ""):
    """
    Fix users that have invalid/old location_ids by mapping their email to the correct kitchen.
    """
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Email to kitchen name mapping
    email_to_kitchen = {
        "stickyricereceiving@gmail.com": "Sticky Rice GCR",
        "asianbbqtanyan@gmail.com": "Tanyan",
        "malaproject@kinfolk.com": "Mala Project",
        "kalaunjibestech@gmail.com": "Kalaunji Bestech",
        "kalaunji43@gmail.com": "Kalaunji GCR",
        "sstickyricebestech@gmail.com": "Sticky Rice Bestech",
        "supersalsa5325@gmail.com": "Super Salsa",
        "stickyriceelan@gmail.com": "Sticky Rice Elan Epic",
        "adreamoven@gmail.com": "Main Store"
    }
    
    # Build location lookup by name
    locations_by_name = {}
    for loc in locations_collection.find({}):
        locations_by_name[loc["name"].lower()] = loc
    
    fixed_users = []
    failed_users = []
    
    for email, kitchen_name in email_to_kitchen.items():
        user = users_collection.find_one({"email": email})
        if not user:
            failed_users.append(f"{email}: user not found")
            continue
        
        location = locations_by_name.get(kitchen_name.lower())
        if not location:
            failed_users.append(f"{email}: kitchen '{kitchen_name}' not found")
            continue
        
        # Update user with correct location
        users_collection.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "location_id": str(location["_id"]),
                "location_name": location["name"]
            }}
        )
        fixed_users.append(f"{email} -> {location['name']}")
    
    return {
        "success": True,
        "fixed_users": fixed_users,
        "failed_users": failed_users
    }

# Fix POs with old/invalid vendor IDs by matching vendor names
@app.post("/api/setup/fix-po-vendor-ids")
async def fix_po_vendor_ids(secret_key: str = ""):
    """
    Fix POs that have invalid/old vendor_ids by mapping vendor names to current vendor IDs.
    """
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Build vendor name to ID mapping (case-insensitive)
    vendors_by_name = {}
    for v in vendors_collection.find({}):
        name_lower = v["name"].lower().strip()
        vendors_by_name[name_lower] = str(v["_id"])
        # Also add variations
        if "pvt" in name_lower:
            variants = [
                name_lower.replace("pvt ltd", "private limited"),
                name_lower.replace("private limited", "pvt ltd"),
                name_lower.replace(" pvt ltd", ""),
                name_lower.replace(" private limited", "")
            ]
            for var in variants:
                vendors_by_name[var] = str(v["_id"])
    
    # Manual mappings for known variations
    manual_mappings = {
        "cp wholesale india private limited": "CP Wholesale PVT LTD",
        "vipin kumar vegetable": "VIPIN",
        "hindustan chemix": "VIPIN",  # Fallback
        "jvm overseas private limited": "VIPIN",  # Fallback
        "kap сones pvt. ltd.": "VIPIN",  # Fallback
        "kapoor enterprises": "VIPIN",  # Fallback
        "azoic water pvt ltd": "hyperpure"  # Fallback
    }
    
    # Get current valid vendor IDs
    valid_vendor_ids = set(str(v["_id"]) for v in vendors_collection.find({}))
    
    fixed_count = 0
    failed_list = []
    
    for po in purchase_orders_collection.find({}):
        vendor_id = po.get("vendor_id")
        
        # Check if vendor_id is valid
        if vendor_id and str(vendor_id) in valid_vendor_ids:
            continue  # Already valid
        
        # Try to find correct vendor by name
        vendor_name = po.get("vendor_name", "").lower().strip()
        
        # Check manual mappings first
        if vendor_name in manual_mappings:
            target_name = manual_mappings[vendor_name].lower()
            if target_name in vendors_by_name:
                new_vendor_id = vendors_by_name[target_name]
                purchase_orders_collection.update_one(
                    {"_id": po["_id"]},
                    {"$set": {"vendor_id": new_vendor_id}}
                )
                fixed_count += 1
                continue
        
        # Try exact match
        if vendor_name in vendors_by_name:
            new_vendor_id = vendors_by_name[vendor_name]
            purchase_orders_collection.update_one(
                {"_id": po["_id"]},
                {"$set": {"vendor_id": new_vendor_id}}
            )
            fixed_count += 1
            continue
        
        # Try fuzzy match - check if vendor name contains any known vendor
        matched = False
        for known_name, vid in vendors_by_name.items():
            if known_name in vendor_name or vendor_name in known_name:
                purchase_orders_collection.update_one(
                    {"_id": po["_id"]},
                    {"$set": {"vendor_id": vid}}
                )
                fixed_count += 1
                matched = True
                break
        
        if not matched:
            failed_list.append(f"{po.get('po_number')}: vendor '{po.get('vendor_name')}' not found")
    
    return {
        "success": True,
        "fixed": fixed_count,
        "failed_count": len(failed_list),
        "failed_details": failed_list[:30]
    }

# Initialize stock lots for items that don't have any
@app.post("/api/setup/init-missing-lots")
async def init_missing_lots(secret_key: str = "", category: str = ""):
    """
    Create lot records for items that don't have any lots.
    This makes them visible in current stock with 0 quantity.
    """
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Get Main Store location
    main_store = locations_collection.find_one({"type": "main_store"})
    if not main_store:
        main_store = locations_collection.find_one({"name": {"$regex": "main", "$options": "i"}})
    
    if not main_store:
        raise HTTPException(status_code=404, detail="Main Store location not found")
    
    main_store_id = main_store["_id"]
    
    # Get all items (optionally filtered by category)
    query = {}
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    items = list(items_collection.find(query))
    
    # Get existing lots to find items without lots
    existing_lots = set()
    for lot in lots_collection.find({"location_id": main_store_id}):
        existing_lots.add(str(lot.get("item_id")))
    
    created_count = 0
    created_items = []
    
    for item in items:
        item_id = str(item["_id"])
        if item_id not in existing_lots:
            # Create a lot with 0 quantity
            lot_doc = {
                "item_id": item["_id"],
                "location_id": main_store_id,
                "lot_number": f"INIT-{item_id[-8:]}",
                "current_quantity": 0,
                "expiry_date": None,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
            lots_collection.insert_one(lot_doc)
            created_count += 1
            created_items.append(f"{item['name']} ({item.get('category', 'N/A')})")
    
    return {
        "success": True,
        "created_lots": created_count,
        "main_store_id": str(main_store_id),
        "category_filter": category or "All",
        "created_items": created_items[:50]  # Show first 50
    }

# Fix lots missing vendor_id by looking up from PO
@app.post("/api/setup/fix-lot-vendor-ids")
async def fix_lot_vendor_ids(secret_key: str = "", force_update: bool = False):
    """
    Fix lots that are missing vendor_id by looking up the vendor from the associated PO.
    If force_update=True, also update lots where vendor_id exists but doesn't match PO.
    """
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Get all POs with their vendor_id
    po_vendor_map = {}
    for po in purchase_orders_collection.find({}, {"po_number": 1, "vendor_id": 1}):
        if po.get("po_number") and po.get("vendor_id"):
            po_vendor_map[po["po_number"]] = po["vendor_id"]
    
    # Also map by _id
    po_id_vendor_map = {}
    for po in purchase_orders_collection.find({}, {"_id": 1, "vendor_id": 1}):
        if po.get("vendor_id"):
            po_id_vendor_map[str(po["_id"])] = po["vendor_id"]
    
    fixed_count = 0
    updated_count = 0
    failed_count = 0
    
    # Query: either missing vendor_id, or force_update all lots with PO reference
    if force_update:
        query = {"$or": [
            {"po_id": {"$exists": True, "$ne": None}},
            {"po_number": {"$exists": True, "$ne": None, "$ne": ""}}
        ]}
    else:
        query = {"$or": [{"vendor_id": None}, {"vendor_id": {"$exists": False}}, {"vendor_id": ""}]}
    
    for lot in lots_collection.find(query):
        vendor_id = None
        
        # Try to get vendor from PO number
        po_number = lot.get("po_number")
        if po_number and po_number in po_vendor_map:
            vendor_id = po_vendor_map[po_number]
        
        # Try to get vendor from po_id
        if not vendor_id and lot.get("po_id"):
            vendor_id = po_id_vendor_map.get(str(lot["po_id"]))
        
        if vendor_id:
            current_vendor = lot.get("vendor_id")
            if not current_vendor or current_vendor == "":
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"vendor_id": vendor_id}}
                )
                fixed_count += 1
            elif force_update and current_vendor != vendor_id:
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"vendor_id": vendor_id}}
                )
                updated_count += 1
        else:
            failed_count += 1
    
    return {
        "success": True,
        "fixed": fixed_count,
        "updated": updated_count,
        "failed": failed_count,
        "message": f"Fixed {fixed_count} lots, updated {updated_count} lots, {failed_count} could not be fixed (no PO reference)"
    }

@app.post("/api/setup/fix-lot-rates")
async def fix_lot_rates(secret_key: str = ""):
    """
    Fix lots that have purchase_rate=0 by looking up the rate from the PO items.
    """
    if secret_key != "KINFOLK-MIGRATE-2026":
        raise HTTPException(status_code=403, detail="Invalid secret key")
    
    # Build PO item rates map: {po_number: {item_id: rate}}
    po_rates = {}
    for po in purchase_orders_collection.find({}, {"po_number": 1, "items": 1}):
        po_num = po.get("po_number")
        if po_num:
            po_rates[po_num] = {}
            for item in po.get("items", []):
                item_id = item.get("item_id")
                rate = item.get("rate", 0)
                if item_id and rate:
                    po_rates[po_num][item_id] = rate
    
    fixed_count = 0
    skipped_count = 0
    
    # Find lots with missing or zero purchase_rate
    for lot in lots_collection.find({"$or": [
        {"purchase_rate": {"$exists": False}},
        {"purchase_rate": None},
        {"purchase_rate": 0},
        {"purchase_rate": ""}
    ]}):
        po_number = lot.get("po_number")
        item_id = lot.get("item_id")
        
        rate = None
        
        # Try to get rate from PO
        if po_number and item_id and po_number in po_rates:
            rate = po_rates[po_number].get(item_id)
        
        # If still no rate, try item standard price
        if not rate and item_id:
            try:
                item = items_collection.find_one({"_id": ObjectId(item_id)})
                if item:
                    rate = item.get("standard_price", 0)
            except:
                pass
        
        if rate and rate > 0:
            lots_collection.update_one(
                {"_id": lot["_id"]},
                {"$set": {"purchase_rate": rate}}
            )
            fixed_count += 1
        else:
            skipped_count += 1
    
    return {
        "success": True,
        "fixed": fixed_count,
        "skipped": skipped_count,
        "message": f"Fixed {fixed_count} lots with rates, {skipped_count} skipped (no rate found)"
    }

# ============ Bulk Update Item Prices ============
@app.post("/api/admin/update-item-prices")
async def update_item_prices(
    items: list,
    current_user = Depends(require_role(["admin"]))
):
    """
    Bulk update item standard_price based on name matching.
    Expects: [{"name": "ITEM NAME", "standard_price": 123.45}, ...]
    """
    updated = 0
    not_found = []
    errors = []
    
    for item_data in items:
        name = item_data.get("name", "").strip()
        price = item_data.get("standard_price")
        
        if not name:
            continue
        
        try:
            # Try exact match first
            result = items_collection.update_one(
                {"name": {"$regex": f"^{name}$", "$options": "i"}},
                {"$set": {"standard_price": float(price) if price else 0}}
            )
            
            if result.modified_count > 0:
                updated += 1
            elif result.matched_count == 0:
                not_found.append(name)
        except Exception as e:
            errors.append({"name": name, "error": str(e)})
    
    return {
        "updated": updated,
        "not_found_count": len(not_found),
        "not_found": not_found[:20],  # First 20 not found
        "errors": errors[:10]  # First 10 errors
    }

# ============ Safe Price Fix - Only for items with HSN codes as prices ============
@app.get("/api/admin/preview-price-fix")
async def preview_price_fix(
    current_user = Depends(require_role(["admin"]))
):
    """
    Preview items that have abnormally high prices (likely HSN codes stored as prices).
    Only shows items with standard_price > 100,000 (likely HSN codes).
    """
    # Find items with prices over 100,000 (likely HSN codes)
    abnormal_items = list(items_collection.find({
        "standard_price": {"$gt": 100000}
    }))
    
    result = []
    for item in abnormal_items:
        result.append({
            "id": str(item["_id"]),
            "name": item.get("name", ""),
            "category": item.get("category", ""),
            "current_price": item.get("standard_price", 0),
            "likely_hsn_code": True if item.get("standard_price", 0) > 1000000 else False
        })
    
    return {
        "message": "These items have prices > ₹100,000 which are likely HSN codes stored as prices",
        "count": len(result),
        "items": sorted(result, key=lambda x: x["current_price"], reverse=True)
    }

@app.post("/api/admin/fix-hsn-prices")
async def fix_hsn_prices(
    price_corrections: list,
    current_user = Depends(require_role(["admin"]))
):
    """
    Fix specific items that have HSN codes stored as prices.
    Only updates standard_price field, nothing else.
    
    Expects: [{"name": "ITEM NAME", "correct_price": 123.45}, ...]
    """
    updated = []
    not_found = []
    skipped = []
    
    for correction in price_corrections:
        name = correction.get("name", "").strip()
        correct_price = correction.get("correct_price")
        
        if not name:
            continue
        
        if correct_price is None:
            skipped.append({"name": name, "reason": "No price provided"})
            continue
        
        try:
            # Find the item first
            item = items_collection.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
            
            if not item:
                not_found.append(name)
                continue
            
            old_price = item.get("standard_price", 0)
            
            # Only update if current price is abnormally high (likely HSN code)
            if old_price > 10000:  # Only fix items with prices over 10,000
                result = items_collection.update_one(
                    {"_id": item["_id"]},
                    {"$set": {"standard_price": float(correct_price)}}
                )
                
                if result.modified_count > 0:
                    updated.append({
                        "name": name,
                        "old_price": old_price,
                        "new_price": float(correct_price)
                    })
            else:
                skipped.append({
                    "name": name,
                    "current_price": old_price,
                    "reason": "Price is normal, not updating"
                })
                
        except Exception as e:
            skipped.append({"name": name, "reason": str(e)})
    
    return {
        "success": True,
        "updated_count": len(updated),
        "updated_items": updated,
        "not_found_count": len(not_found),
        "not_found": not_found[:20],
        "skipped_count": len(skipped),
        "skipped": skipped[:20]
    }


# ============ Database Index Management ============
@app.get("/api/admin/db-indexes")
async def get_db_indexes(
    current_user = Depends(require_role(["admin"]))
):
    """Get current database indexes and recommendations"""
    indexes_info = {}
    
    collections_to_check = [
        ("purchase_orders", purchase_orders_collection),
        ("lots", lots_collection),
        ("transactions", transactions_collection),
        ("items", items_collection),
        ("requisitions", requisitions_collection)
    ]
    
    for name, coll in collections_to_check:
        try:
            indexes = list(coll.list_indexes())
            indexes_info[name] = [idx.get("key") for idx in indexes]
        except Exception as e:
            indexes_info[name] = f"Error: {str(e)}"
    
    recommended = {
        "purchase_orders": ["created_at DESC", "status", "vendor_id"],
        "lots": ["created_at DESC", "item_id", "location_id", "po_reference_id"],
        "transactions": ["timestamp DESC", "lot_id", "item_id", "location_id"],
        "requisitions": ["created_at DESC", "from_location_id", "status"]
    }
    
    return {
        "current_indexes": indexes_info,
        "recommended_indexes": recommended,
        "message": "Run POST /api/admin/create-indexes to create recommended indexes"
    }


@app.post("/api/admin/create-indexes")
async def create_db_indexes(
    current_user = Depends(require_role(["admin"]))
):
    """Create recommended indexes to improve query performance"""
    from pymongo import DESCENDING, ASCENDING
    
    results = []
    
    # Purchase Orders indexes
    try:
        purchase_orders_collection.create_index([("created_at", DESCENDING)], background=True)
        results.append({"collection": "purchase_orders", "index": "created_at_-1", "status": "created"})
    except Exception as e:
        results.append({"collection": "purchase_orders", "index": "created_at_-1", "status": f"error: {str(e)}"})
    
    try:
        purchase_orders_collection.create_index([("status", ASCENDING)], background=True)
        results.append({"collection": "purchase_orders", "index": "status_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "purchase_orders", "index": "status_1", "status": f"error: {str(e)}"})
    
    try:
        purchase_orders_collection.create_index([("vendor_id", ASCENDING)], background=True)
        results.append({"collection": "purchase_orders", "index": "vendor_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "purchase_orders", "index": "vendor_id_1", "status": f"error: {str(e)}"})
    
    # Lots indexes
    try:
        lots_collection.create_index([("created_at", DESCENDING)], background=True)
        results.append({"collection": "lots", "index": "created_at_-1", "status": "created"})
    except Exception as e:
        results.append({"collection": "lots", "index": "created_at_-1", "status": f"error: {str(e)}"})
    
    try:
        lots_collection.create_index([("item_id", ASCENDING)], background=True)
        results.append({"collection": "lots", "index": "item_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "lots", "index": "item_id_1", "status": f"error: {str(e)}"})
    
    try:
        lots_collection.create_index([("location_id", ASCENDING)], background=True)
        results.append({"collection": "lots", "index": "location_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "lots", "index": "location_id_1", "status": f"error: {str(e)}"})
    
    try:
        lots_collection.create_index([("po_reference_id", ASCENDING)], background=True)
        results.append({"collection": "lots", "index": "po_reference_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "lots", "index": "po_reference_id_1", "status": f"error: {str(e)}"})
    
    # Transactions indexes
    try:
        transactions_collection.create_index([("timestamp", DESCENDING)], background=True)
        results.append({"collection": "transactions", "index": "timestamp_-1", "status": "created"})
    except Exception as e:
        results.append({"collection": "transactions", "index": "timestamp_-1", "status": f"error: {str(e)}"})
    
    try:
        transactions_collection.create_index([("lot_id", ASCENDING)], background=True)
        results.append({"collection": "transactions", "index": "lot_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "transactions", "index": "lot_id_1", "status": f"error: {str(e)}"})
    
    try:
        transactions_collection.create_index([("item_id", ASCENDING)], background=True)
        results.append({"collection": "transactions", "index": "item_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "transactions", "index": "item_id_1", "status": f"error: {str(e)}"})
    
    # Requisitions indexes  
    try:
        requisitions_collection.create_index([("created_at", DESCENDING)], background=True)
        results.append({"collection": "requisitions", "index": "created_at_-1", "status": "created"})
    except Exception as e:
        results.append({"collection": "requisitions", "index": "created_at_-1", "status": f"error: {str(e)}"})
    
    try:
        requisitions_collection.create_index([("from_location_id", ASCENDING)], background=True)
        results.append({"collection": "requisitions", "index": "from_location_id_1", "status": "created"})
    except Exception as e:
        results.append({"collection": "requisitions", "index": "from_location_id_1", "status": f"error: {str(e)}"})
    
    return {
        "success": True,
        "message": "Index creation completed",
        "results": results
    }


# ============ Safe Price Update with Excel Upload ============
@app.post("/api/admin/preview-price-update")
async def preview_price_update_from_excel(
    file: UploadFile = File(...),
    current_user = Depends(require_role(["admin"]))
):
    """
    Preview price updates from an Excel file before applying.
    Excel must have columns: name (or Item Name), standard_price (or Rate)
    Returns what would change without modifying anything.
    """
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()
        
        # Map common column names
        name_cols = ["name", "item name", "item_name", "itemname"]
        price_cols = ["standard_price", "rate", "price", "standard price"]
        
        name_col = None
        price_col = None
        
        for col in name_cols:
            if col in df.columns:
                name_col = col
                break
        
        for col in price_cols:
            if col in df.columns:
                price_col = col
                break
        
        if not name_col:
            return {"error": f"Could not find name column. Available: {list(df.columns)}"}
        if not price_col:
            return {"error": f"Could not find price column. Available: {list(df.columns)}"}
        
        preview_changes = []
        not_found = []
        no_change = []
        invalid_price = []
        
        for _, row in df.iterrows():
            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
            new_price = row[price_col]
            
            if not name:
                continue
            
            # Validate price
            try:
                new_price = float(new_price) if pd.notna(new_price) else None
            except:
                invalid_price.append({"name": name, "value": str(row[price_col])})
                continue
            
            if new_price is None or new_price < 0:
                invalid_price.append({"name": name, "value": str(row[price_col])})
                continue
            
            # Find item in DB
            item = items_collection.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}})
            
            if not item:
                not_found.append(name)
                continue
            
            current_price = item.get("standard_price", 0) or 0
            
            if abs(current_price - new_price) < 0.01:
                no_change.append({"name": name, "price": current_price})
                continue
            
            preview_changes.append({
                "id": str(item["_id"]),
                "name": item.get("name", ""),
                "category": item.get("category", ""),
                "current_price": current_price,
                "new_price": new_price,
                "change": round(new_price - current_price, 2),
                "change_percent": round((new_price - current_price) / current_price * 100, 1) if current_price > 0 else None
            })
        
        # Sort by change amount
        preview_changes.sort(key=lambda x: abs(x.get("change", 0)), reverse=True)
        
        return {
            "success": True,
            "message": "Preview generated. Review changes and call /api/admin/apply-price-update to apply.",
            "summary": {
                "will_update": len(preview_changes),
                "not_found": len(not_found),
                "no_change": len(no_change),
                "invalid_price": len(invalid_price)
            },
            "changes": preview_changes,
            "not_found_items": not_found[:30],
            "invalid_prices": invalid_price[:10]
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/apply-price-update")
async def apply_price_update_from_excel(
    file: UploadFile = File(...),
    current_user = Depends(require_role(["admin"]))
):
    """
    Apply price updates from an Excel file.
    Excel must have columns: name (or Item Name), standard_price (or Rate)
    """
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()
        
        # Map common column names
        name_cols = ["name", "item name", "item_name", "itemname"]
        price_cols = ["standard_price", "rate", "price", "standard price"]
        
        name_col = None
        price_col = None
        
        for col in name_cols:
            if col in df.columns:
                name_col = col
                break
        
        for col in price_cols:
            if col in df.columns:
                price_col = col
                break
        
        if not name_col:
            return {"error": f"Could not find name column. Available: {list(df.columns)}"}
        if not price_col:
            return {"error": f"Could not find price column. Available: {list(df.columns)}"}
        
        updated = []
        not_found = []
        errors = []
        
        for _, row in df.iterrows():
            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
            new_price = row[price_col]
            
            if not name:
                continue
            
            try:
                new_price = float(new_price) if pd.notna(new_price) else None
            except:
                errors.append({"name": name, "error": f"Invalid price: {row[price_col]}"})
                continue
            
            if new_price is None or new_price < 0:
                errors.append({"name": name, "error": f"Invalid price: {new_price}"})
                continue
            
            try:
                result = items_collection.update_one(
                    {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
                    {"$set": {"standard_price": new_price}}
                )
                
                if result.modified_count > 0:
                    updated.append({"name": name, "new_price": new_price})
                elif result.matched_count == 0:
                    not_found.append(name)
            except Exception as e:
                errors.append({"name": name, "error": str(e)})
        
        return {
            "success": True,
            "message": f"Price update completed. {len(updated)} items updated.",
            "updated_count": len(updated),
            "not_found_count": len(not_found),
            "error_count": len(errors),
            "updated_items": updated[:50],
            "not_found_items": not_found[:30],
            "errors": errors[:10]
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/bulk-update-items")
async def bulk_update_items_from_excel(
    file: UploadFile = File(...),
    update_price: bool = True,
    update_gst: bool = True,
    clear_hsn: bool = True,
    current_user = Depends(require_role(["admin"]))
):
    """
    Comprehensive item update from Excel file.
    Updates: standard_price, gst_rate, and optionally clears hsn_code.
    
    Expected columns: Item Name, Standard Price, GST Rate
    """
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()
        
        # Map column names
        name_col = next((c for c in ["item name", "name", "item_name"] if c in df.columns), None)
        price_col = next((c for c in ["standard price", "standard_price", "price", "rate"] if c in df.columns), None)
        gst_col = next((c for c in ["gst rate", "gst_rate", "gst"] if c in df.columns), None)
        
        if not name_col:
            return {"error": f"Could not find name column. Available: {list(df.columns)}"}
        
        results = {
            "price_updated": 0,
            "gst_updated": 0,
            "hsn_cleared": 0,
            "not_found": [],
            "errors": []
        }
        
        for _, row in df.iterrows():
            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
            if not name:
                continue
            
            update_doc = {}
            
            # Price update
            if update_price and price_col:
                try:
                    price = float(row[price_col]) if pd.notna(row[price_col]) else None
                    if price is not None and price >= 0:
                        update_doc["standard_price"] = price
                except:
                    pass
            
            # GST update
            if update_gst and gst_col:
                try:
                    gst = float(row[gst_col]) if pd.notna(row[gst_col]) else None
                    if gst is not None and gst >= 0:
                        update_doc["gst_rate"] = gst
                except:
                    pass
            
            # Clear HSN code
            if clear_hsn:
                update_doc["hsn_code"] = None
            
            if not update_doc:
                continue
            
            try:
                result = items_collection.update_one(
                    {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
                    {"$set": update_doc}
                )
                
                if result.modified_count > 0:
                    if "standard_price" in update_doc:
                        results["price_updated"] += 1
                    if "gst_rate" in update_doc:
                        results["gst_updated"] += 1
                    if "hsn_code" in update_doc:
                        results["hsn_cleared"] += 1
                elif result.matched_count == 0:
                    results["not_found"].append(name)
            except Exception as e:
                results["errors"].append({"name": name, "error": str(e)})
        
        return {
            "success": True,
            "message": "Bulk update completed",
            "price_updated": results["price_updated"],
            "gst_updated": results["gst_updated"],
            "hsn_cleared": results["hsn_cleared"],
            "not_found_count": len(results["not_found"]),
            "not_found_items": results["not_found"][:30],
            "errors": results["errors"][:10]
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}



# ============ Fix Historical Prices in Lots and Transactions ============
@app.get("/api/admin/preview-historical-price-fix")
async def preview_historical_price_fix(
    current_user = Depends(require_role(["admin"]))
):
    """
    Preview what will change when we fix historical prices in lots and transactions.
    Uses current item standard_price to update purchase_rate in lots and rate/value in transactions.
    """
    try:
        # Get all items with their standard prices
        items = {str(item["_id"]): item for item in items_collection.find({})}
        
        lots_to_fix = []
        transactions_to_fix = []
        
        # Check lots with mismatched prices
        for lot in lots_collection.find({}):
            item_id = str(lot.get("item_id", ""))
            item = items.get(item_id)
            
            if not item:
                continue
            
            current_rate = lot.get("purchase_rate", 0) or 0
            standard_price = item.get("standard_price", 0) or 0
            
            # Skip if prices match or standard price is 0
            if standard_price == 0 or abs(current_rate - standard_price) < 0.01:
                continue
            
            lots_to_fix.append({
                "lot_id": str(lot["_id"]),
                "item_name": item.get("name", "Unknown"),
                "current_rate": current_rate,
                "new_rate": standard_price,
                "difference": round(standard_price - current_rate, 2),
                "quantity": lot.get("quantity", 0)
            })
        
        # Check transactions with mismatched prices
        for txn in transactions_collection.find({"type": {"$in": ["issue", "grn", "receive"]}}):
            item_id = str(txn.get("item_id", ""))
            
            # Try to get item from lot if not directly available
            if not item_id and txn.get("lot_id"):
                lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
                if lot:
                    item_id = str(lot.get("item_id", ""))
            
            item = items.get(item_id)
            if not item:
                continue
            
            current_rate = txn.get("rate", 0) or 0
            standard_price = item.get("standard_price", 0) or 0
            
            # Skip if prices match or standard price is 0
            if standard_price == 0 or abs(current_rate - standard_price) < 0.01:
                continue
            
            qty = txn.get("quantity", 0) or 0
            old_value = txn.get("value", 0) or (current_rate * qty)
            new_value = standard_price * qty
            
            transactions_to_fix.append({
                "txn_id": str(txn["_id"]),
                "item_name": item.get("name", "Unknown"),
                "type": txn.get("type", ""),
                "quantity": qty,
                "current_rate": current_rate,
                "new_rate": standard_price,
                "old_value": round(old_value, 2),
                "new_value": round(new_value, 2),
                "value_change": round(new_value - old_value, 2)
            })
        
        # Sort by absolute value change
        lots_to_fix.sort(key=lambda x: abs(x["difference"]), reverse=True)
        transactions_to_fix.sort(key=lambda x: abs(x.get("value_change", 0)), reverse=True)
        
        total_lots_value_change = sum(l["difference"] * l.get("quantity", 0) for l in lots_to_fix)
        total_txn_value_change = sum(t.get("value_change", 0) for t in transactions_to_fix)
        
        return {
            "success": True,
            "message": "Preview of historical price corrections",
            "summary": {
                "lots_to_update": len(lots_to_fix),
                "transactions_to_update": len(transactions_to_fix),
                "estimated_lots_value_change": round(total_lots_value_change, 2),
                "estimated_txn_value_change": round(total_txn_value_change, 2)
            },
            "lots_sample": lots_to_fix[:50],
            "transactions_sample": transactions_to_fix[:50]
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/apply-historical-price-fix")
async def apply_historical_price_fix(
    current_user = Depends(require_role(["admin"]))
):
    """
    Apply price corrections to historical lots and transactions.
    Updates purchase_rate in lots and rate/value in transactions based on current item standard_price.
    """
    try:
        # Get all items with their standard prices
        items = {str(item["_id"]): item for item in items_collection.find({})}
        
        lots_updated = 0
        lots_skipped = 0
        transactions_updated = 0
        transactions_skipped = 0
        errors = []
        
        # Fix lots
        for lot in lots_collection.find({}):
            item_id = str(lot.get("item_id", ""))
            item = items.get(item_id)
            
            if not item:
                continue
            
            current_rate = lot.get("purchase_rate", 0) or 0
            standard_price = item.get("standard_price", 0) or 0
            
            # Skip if prices match or standard price is 0
            if standard_price == 0:
                lots_skipped += 1
                continue
            
            if abs(current_rate - standard_price) < 0.01:
                lots_skipped += 1
                continue
            
            try:
                lots_collection.update_one(
                    {"_id": lot["_id"]},
                    {"$set": {"purchase_rate": standard_price}}
                )
                lots_updated += 1
            except Exception as e:
                errors.append({"type": "lot", "id": str(lot["_id"]), "error": str(e)})
        
        # Fix transactions
        for txn in transactions_collection.find({"type": {"$in": ["issue", "grn", "receive"]}}):
            item_id = str(txn.get("item_id", ""))
            
            # Try to get item from lot if not directly available
            if not item_id and txn.get("lot_id"):
                try:
                    lot = lots_collection.find_one({"_id": ObjectId(txn["lot_id"])})
                    if lot:
                        item_id = str(lot.get("item_id", ""))
                except:
                    pass
            
            item = items.get(item_id)
            if not item:
                continue
            
            current_rate = txn.get("rate", 0) or 0
            standard_price = item.get("standard_price", 0) or 0
            
            # Skip if prices match or standard price is 0
            if standard_price == 0:
                transactions_skipped += 1
                continue
            
            if abs(current_rate - standard_price) < 0.01:
                transactions_skipped += 1
                continue
            
            qty = txn.get("quantity", 0) or 0
            new_value = standard_price * qty
            
            try:
                transactions_collection.update_one(
                    {"_id": txn["_id"]},
                    {"$set": {
                        "rate": standard_price,
                        "value": new_value
                    }}
                )
                transactions_updated += 1
            except Exception as e:
                errors.append({"type": "transaction", "id": str(txn["_id"]), "error": str(e)})
        
        return {
            "success": True,
            "message": "Historical price correction completed",
            "lots_updated": lots_updated,
            "lots_skipped": lots_skipped,
            "transactions_updated": transactions_updated,
            "transactions_skipped": transactions_skipped,
            "errors_count": len(errors),
            "errors": errors[:20]
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}



# ============ Vendor Ledger Diagnostic ============

@app.get("/api/admin/diagnose-vendor-ledger")
async def diagnose_vendor_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Diagnostic endpoint to identify why some vendors show ₹0 in the vendor ledger.
    Checks:
    1. Lots without vendor_id
    2. Lots with vendor_id that doesn't exist in vendors collection
    3. Date filtering issues
    4. Vendors with POs but no linked GRNs
    """
    try:
        results = {
            "summary": {},
            "lots_without_vendor_id": [],
            "lots_with_invalid_vendor_id": [],
            "vendors_with_po_no_grn": [],
            "date_filter_issues": [],
            "fix_recommendations": []
        }
        
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        # 1. Count lots without vendor_id
        lots_no_vendor_query = {"$or": [
            {"vendor_id": {"$exists": False}},
            {"vendor_id": None},
            {"vendor_id": ""}
        ]}
        lots_no_vendor = lots_collection.count_documents(lots_no_vendor_query)
        
        # Sample lots without vendor_id
        for lot in lots_collection.find(lots_no_vendor_query).limit(10):
            item = items_collection.find_one({"_id": ObjectId(lot.get("item_id"))}) if lot.get("item_id") else None
            results["lots_without_vendor_id"].append({
                "lot_id": str(lot["_id"]),
                "lot_number": lot.get("lot_number"),
                "po_number": lot.get("po_number"),
                "po_id": str(lot.get("po_id")) if lot.get("po_id") else None,
                "item_name": item["name"] if item else "Unknown",
                "created_at": str(lot.get("created_at", ""))[:10],
                "quantity": lot.get("initial_quantity", 0)
            })
        
        # 2. Check vendors with POs but no GRNs/lots in the date range
        all_vendors = list(vendors_collection.find({}, {"_id": 1, "name": 1}))
        
        for vendor in all_vendors:
            v_id = str(vendor["_id"])
            v_name = vendor["name"]
            
            # Count POs for this vendor in date range
            po_query = {"vendor_id": v_id}
            if date_query:
                po_query["created_at"] = date_query
            po_count = purchase_orders_collection.count_documents(po_query)
            
            # Count lots for this vendor in date range  
            lot_query = {"vendor_id": v_id}
            if date_query:
                lot_query["created_at"] = date_query
            lot_count = lots_collection.count_documents(lot_query)
            
            # Count lots linked via PO but not directly by vendor_id
            # Get PO IDs for this vendor
            po_ids = [str(po["_id"]) for po in purchase_orders_collection.find({"vendor_id": v_id}, {"_id": 1})]
            po_numbers = [po["po_number"] for po in purchase_orders_collection.find({"vendor_id": v_id}, {"po_number": 1}) if po.get("po_number")]
            
            # Count lots that should belong to this vendor (via PO) but have wrong/missing vendor_id
            orphaned_lots_query = {
                "$and": [
                    {"$or": [
                        {"po_number": {"$in": po_numbers}},
                        {"po_id": {"$in": [ObjectId(pid) for pid in po_ids] if po_ids else []}}
                    ]},
                    {"$or": [
                        {"vendor_id": {"$ne": v_id}},
                        {"vendor_id": {"$exists": False}},
                        {"vendor_id": None},
                        {"vendor_id": ""}
                    ]}
                ]
            }
            orphaned_count = lots_collection.count_documents(orphaned_lots_query) if po_numbers or po_ids else 0
            
            if po_count > 0 and lot_count == 0 and orphaned_count > 0:
                results["vendors_with_po_no_grn"].append({
                    "vendor_id": v_id,
                    "vendor_name": v_name,
                    "po_count": po_count,
                    "grn_count_by_vendor_id": lot_count,
                    "orphaned_lots_count": orphaned_count,
                    "issue": "Lots exist but not linked to this vendor"
                })
        
        # 3. Summary stats
        total_lots = lots_collection.count_documents({})
        lots_with_vendor = lots_collection.count_documents({"vendor_id": {"$exists": True, "$ne": None, "$ne": ""}})
        
        results["summary"] = {
            "total_lots": total_lots,
            "lots_with_vendor_id": lots_with_vendor,
            "lots_without_vendor_id": lots_no_vendor,
            "total_vendors": len(all_vendors),
            "vendors_with_linkage_issues": len(results["vendors_with_po_no_grn"]),
            "date_range": f"{start_date or 'all'} to {end_date or 'all'}"
        }
        
        # 4. Recommendations
        if lots_no_vendor > 0:
            results["fix_recommendations"].append({
                "issue": f"{lots_no_vendor} lots missing vendor_id",
                "fix_endpoint": "POST /api/setup/fix-lot-vendor-ids?secret_key=KINFOLK-MIGRATE-2026",
                "description": "This will link lots to vendors via their PO reference"
            })
        
        if results["vendors_with_po_no_grn"]:
            results["fix_recommendations"].append({
                "issue": f"{len(results['vendors_with_po_no_grn'])} vendors have POs but no linked GRNs",
                "fix_endpoint": "POST /api/setup/fix-lot-vendor-ids?secret_key=KINFOLK-MIGRATE-2026&force_update=true",
                "description": "Force update all lots with PO reference to correct vendor_id"
            })
        
        return results
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/fix-vendor-linkage")
async def fix_vendor_linkage():
    """
    One-click fix for vendor linkage issues in lots.
    Links lots to vendors via their PO reference.
    """
    try:
        # Get all POs with their vendor_id - map by po_number
        po_vendor_map = {}
        for po in purchase_orders_collection.find({}, {"po_number": 1, "vendor_id": 1, "_id": 1}):
            if po.get("po_number") and po.get("vendor_id"):
                po_vendor_map[po["po_number"]] = po["vendor_id"]
            if po.get("vendor_id"):
                po_vendor_map[str(po["_id"])] = po["vendor_id"]
        
        fixed_count = 0
        already_correct = 0
        no_po_ref = 0
        
        # Find all lots and fix vendor_id
        for lot in lots_collection.find({}):
            current_vendor = lot.get("vendor_id")
            new_vendor = None
            
            # Try to get vendor from PO number
            po_number = lot.get("po_number")
            if po_number and po_number in po_vendor_map:
                new_vendor = po_vendor_map[po_number]
            
            # Try to get vendor from po_id
            if not new_vendor and lot.get("po_id"):
                new_vendor = po_vendor_map.get(str(lot["po_id"]))
            
            if new_vendor:
                if not current_vendor or current_vendor == "" or current_vendor != new_vendor:
                    lots_collection.update_one(
                        {"_id": lot["_id"]},
                        {"$set": {"vendor_id": new_vendor}}
                    )
                    fixed_count += 1
                else:
                    already_correct += 1
            else:
                no_po_ref += 1
        
        return {
            "success": True,
            "message": f"Fixed {fixed_count} lots, {already_correct} were already correct, {no_po_ref} have no PO reference",
            "fixed": fixed_count,
            "already_correct": already_correct,
            "no_po_reference": no_po_ref
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}



@app.get("/api/admin/po-status-check")
async def po_status_check(
    vendor_name: str = Query(..., description="Vendor name to check"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Check PO statuses for a vendor - are they received or pending?
    """
    try:
        vendor = vendors_collection.find_one({"name": {"$regex": vendor_name, "$options": "i"}})
        if not vendor:
            return {"error": f"Vendor not found: {vendor_name}"}
        
        v_id = str(vendor["_id"])
        
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        po_query = {"vendor_id": v_id}
        if date_query:
            po_query["created_at"] = date_query
        
        pos = list(purchase_orders_collection.find(po_query))
        
        status_counts = {}
        po_details = []
        
        for po in pos:
            status = po.get("status", "unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
            
            # Check received_items
            received_items = po.get("received_items", [])
            
            po_details.append({
                "po_number": po.get("po_number"),
                "date": str(po.get("created_at", ""))[:10],
                "status": status,
                "items_count": len(po.get("items", [])),
                "received_items_count": len(received_items),
                "total_value": sum(item.get("quantity", 0) * item.get("rate", 0) for item in po.get("items", []))
            })
        
        return {
            "vendor_name": vendor["name"],
            "total_pos": len(pos),
            "status_breakdown": status_counts,
            "po_details": po_details[:20]
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.get("/api/admin/orphan-lots-analysis")
async def orphan_lots_analysis(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Find lots that have no vendor_id AND no po_number AND no po_id.
    These are completely orphaned and need manual investigation.
    """
    try:
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        # Find completely orphaned lots
        orphan_query = {
            "$and": [
                {"$or": [{"vendor_id": {"$exists": False}}, {"vendor_id": None}, {"vendor_id": ""}]},
                {"$or": [{"po_number": {"$exists": False}}, {"po_number": None}, {"po_number": ""}]},
                {"$or": [{"po_id": {"$exists": False}}, {"po_id": None}]}
            ]
        }
        if date_query:
            orphan_query["created_at"] = date_query
        
        orphan_count = lots_collection.count_documents(orphan_query)
        
        # Get samples with item details
        samples = []
        for lot in lots_collection.find(orphan_query).limit(20):
            item = items_collection.find_one({"_id": ObjectId(lot.get("item_id"))}) if lot.get("item_id") else None
            samples.append({
                "lot_id": str(lot["_id"]),
                "lot_number": lot.get("lot_number"),
                "item_name": item["name"] if item else lot.get("item_name", "Unknown"),
                "category": item.get("category", "") if item else lot.get("category", ""),
                "quantity": lot.get("initial_quantity", 0),
                "rate": lot.get("purchase_rate", 0),
                "value": lot.get("initial_quantity", 0) * lot.get("purchase_rate", 0),
                "created_at": str(lot.get("created_at", ""))[:10],
                "source": lot.get("source", "unknown"),
                "has_vendor_id": bool(lot.get("vendor_id")),
                "has_po_number": bool(lot.get("po_number")),
                "has_po_id": bool(lot.get("po_id"))
            })
        
        # Group by category
        category_breakdown = {}
        for lot in lots_collection.find(orphan_query):
            item = items_collection.find_one({"_id": ObjectId(lot.get("item_id"))}) if lot.get("item_id") else None
            cat = item.get("category", "Unknown") if item else lot.get("category", "Unknown")
            if cat not in category_breakdown:
                category_breakdown[cat] = {"count": 0, "value": 0}
            category_breakdown[cat]["count"] += 1
            category_breakdown[cat]["value"] += lot.get("initial_quantity", 0) * lot.get("purchase_rate", 0)
        
        return {
            "total_orphan_lots": orphan_count,
            "date_range": f"{start_date or 'all'} to {end_date or 'all'}",
            "category_breakdown": category_breakdown,
            "samples": samples
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}



@app.get("/api/admin/vendor-grn-debug")
async def vendor_grn_debug(
    vendor_name: str = Query(..., description="Vendor name to debug (partial match)"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Debug why a specific vendor shows incorrect/zero GRN values.
    Shows detailed breakdown of lots and their linkage.
    """
    try:
        # Find vendor by name (partial match)
        vendor = vendors_collection.find_one({"name": {"$regex": vendor_name, "$options": "i"}})
        if not vendor:
            return {"error": f"Vendor not found with name containing: {vendor_name}"}
        
        v_id = str(vendor["_id"])
        v_name = vendor["name"]
        
        result = {
            "vendor_id": v_id,
            "vendor_name": v_name,
            "analysis": {}
        }
        
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        # 1. Get POs for this vendor
        po_query = {"vendor_id": v_id}
        if date_query:
            po_query["created_at"] = date_query
        
        pos = list(purchase_orders_collection.find(po_query))
        po_numbers = [po.get("po_number") for po in pos if po.get("po_number")]
        po_ids = [str(po["_id"]) for po in pos]
        
        result["analysis"]["purchase_orders"] = {
            "count": len(pos),
            "po_numbers": po_numbers[:20],
            "total_value": sum(sum(item.get("quantity", 0) * item.get("rate", 0) for item in po.get("items", [])) for po in pos)
        }
        
        # 2. Check lots directly by vendor_id
        lot_query_direct = {"vendor_id": v_id}
        if date_query:
            lot_query_direct["created_at"] = date_query
        lots_by_vendor_id = lots_collection.count_documents(lot_query_direct)
        
        # 3. Check lots by po_number
        lots_by_po_number = 0
        if po_numbers:
            lot_query_po_num = {"po_number": {"$in": po_numbers}}
            if date_query:
                lot_query_po_num["created_at"] = date_query
            lots_by_po_number = lots_collection.count_documents(lot_query_po_num)
        
        # 4. Check lots by po_id
        lots_by_po_id = 0
        if po_ids:
            lot_query_po_id = {"po_id": {"$in": po_ids}}
            if date_query:
                lot_query_po_id["created_at"] = date_query
            lots_by_po_id = lots_collection.count_documents(lot_query_po_id)
        
        result["analysis"]["lots_found"] = {
            "by_vendor_id": lots_by_vendor_id,
            "by_po_number": lots_by_po_number,
            "by_po_id": lots_by_po_id
        }
        
        # 5. Sample lots to understand structure
        sample_lots = []
        
        # Get sample lot by vendor_id
        sample = lots_collection.find_one({"vendor_id": v_id})
        if sample:
            sample_lots.append({
                "source": "by_vendor_id",
                "lot_id": str(sample["_id"]),
                "lot_number": sample.get("lot_number"),
                "vendor_id": sample.get("vendor_id"),
                "po_number": sample.get("po_number"),
                "po_id": str(sample.get("po_id")) if sample.get("po_id") else None,
                "created_at": str(sample.get("created_at", ""))[:10]
            })
        
        # Get sample lot by po_number
        if po_numbers:
            sample = lots_collection.find_one({"po_number": {"$in": po_numbers}})
            if sample:
                sample_lots.append({
                    "source": "by_po_number",
                    "lot_id": str(sample["_id"]),
                    "lot_number": sample.get("lot_number"),
                    "vendor_id": sample.get("vendor_id"),
                    "po_number": sample.get("po_number"),
                    "po_id": str(sample.get("po_id")) if sample.get("po_id") else None,
                    "created_at": str(sample.get("created_at", ""))[:10]
                })
        
        result["analysis"]["sample_lots"] = sample_lots
        
        # 6. Check if lots exist but have wrong/missing vendor_id
        if po_numbers:
            orphaned_query = {
                "po_number": {"$in": po_numbers},
                "$or": [
                    {"vendor_id": {"$ne": v_id}},
                    {"vendor_id": {"$exists": False}},
                    {"vendor_id": None},
                    {"vendor_id": ""}
                ]
            }
            orphaned_count = lots_collection.count_documents(orphaned_query)
            
            orphaned_samples = []
            for lot in lots_collection.find(orphaned_query).limit(5):
                orphaned_samples.append({
                    "lot_id": str(lot["_id"]),
                    "lot_number": lot.get("lot_number"),
                    "po_number": lot.get("po_number"),
                    "current_vendor_id": lot.get("vendor_id"),
                    "expected_vendor_id": v_id,
                    "created_at": str(lot.get("created_at", ""))[:10]
                })
            
            result["analysis"]["orphaned_lots"] = {
                "count": orphaned_count,
                "samples": orphaned_samples
            }
        
        # 7. Diagnosis
        if lots_by_vendor_id > 0:
            result["diagnosis"] = "Lots exist and are linked correctly. Check date filter or calculation logic."
        elif lots_by_po_number > 0 or lots_by_po_id > 0:
            result["diagnosis"] = "Lots exist but have incorrect/missing vendor_id. Run fix-vendor-linkage endpoint."
            result["fix_url"] = "POST /api/admin/fix-vendor-linkage"
        else:
            result["diagnosis"] = "No lots found for this vendor's POs. Either GRNs were not done, or lots have no PO reference."
        
        return result
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/fix-vendor-linkage-v2")
async def fix_vendor_linkage_v2():
    """
    Enhanced fix for vendor linkage - also tries to match by item_id from PO items.
    """
    try:
        stats = {
            "fixed_by_po_number": 0,
            "fixed_by_po_id": 0,
            "fixed_by_item_match": 0,
            "already_correct": 0,
            "no_match_found": 0,
            "errors": []
        }
        
        # Build PO lookup maps
        po_vendor_by_number = {}
        po_vendor_by_id = {}
        po_items_by_vendor = {}  # vendor_id -> set of item_ids
        
        for po in purchase_orders_collection.find({}):
            vendor_id = po.get("vendor_id")
            if not vendor_id:
                continue
            
            po_number = po.get("po_number")
            if po_number:
                po_vendor_by_number[po_number] = vendor_id
            
            po_vendor_by_id[str(po["_id"])] = vendor_id
            
            # Build item->vendor mapping from PO items
            for item in po.get("items", []):
                item_id = item.get("item_id")
                if item_id:
                    if vendor_id not in po_items_by_vendor:
                        po_items_by_vendor[vendor_id] = set()
                    po_items_by_vendor[vendor_id].add(item_id)
        
        # Process all lots
        for lot in lots_collection.find({}):
            current_vendor = lot.get("vendor_id")
            new_vendor = None
            fix_method = None
            
            # Method 1: By PO number
            po_number = lot.get("po_number")
            if po_number and po_number in po_vendor_by_number:
                new_vendor = po_vendor_by_number[po_number]
                fix_method = "po_number"
            
            # Method 2: By PO ID
            if not new_vendor:
                po_id = lot.get("po_id")
                if po_id:
                    new_vendor = po_vendor_by_id.get(str(po_id))
                    if new_vendor:
                        fix_method = "po_id"
            
            # Method 3: By item_id (if item only belongs to one vendor's POs)
            if not new_vendor:
                item_id = lot.get("item_id")
                if item_id:
                    matching_vendors = [v for v, items in po_items_by_vendor.items() if item_id in items]
                    if len(matching_vendors) == 1:
                        new_vendor = matching_vendors[0]
                        fix_method = "item_match"
            
            # Apply fix
            if new_vendor:
                if not current_vendor or current_vendor == "" or current_vendor != new_vendor:
                    lots_collection.update_one(
                        {"_id": lot["_id"]},
                        {"$set": {"vendor_id": new_vendor}}
                    )
                    if fix_method == "po_number":
                        stats["fixed_by_po_number"] += 1
                    elif fix_method == "po_id":
                        stats["fixed_by_po_id"] += 1
                    else:
                        stats["fixed_by_item_match"] += 1
                else:
                    stats["already_correct"] += 1
            else:
                stats["no_match_found"] += 1
        
        total_fixed = stats["fixed_by_po_number"] + stats["fixed_by_po_id"] + stats["fixed_by_item_match"]
        
        return {
            "success": True,
            "message": f"Fixed {total_fixed} lots total",
            "stats": stats
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}



# ============ Outlet-wise PO vs DP Comparison Report ============

@app.get("/api/reports/outlet-wise-po-dp")
async def get_outlet_wise_po_dp(
    vendor_id: Optional[str] = None,
    vendor_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Outlet/Kitchen-wise comparison of PO vs DP deliveries.
    Shows:
    1. Which outlets are receiving goods without PO
    2. Quantities for each outlet
    3. Whether outlets are doing GRN
    """
    try:
        # Find vendor
        if vendor_name:
            vendor = vendors_collection.find_one({"name": {"$regex": vendor_name, "$options": "i"}})
            if vendor:
                vendor_id = str(vendor["_id"])
        
        if not vendor_id:
            return {"error": "Please provide vendor_id or vendor_name"}
        
        vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
        if not vendor:
            return {"error": "Vendor not found"}
        
        v_name = vendor["name"]
        
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        # Get all kitchens/locations
        locations = list(db["locations"].find({"type": {"$in": ["kitchen", "outlet"]}}))
        location_map = {str(loc["_id"]): loc for loc in locations}
        
        # Get PO items grouped by location (created_by_location_id)
        # Use aggregation with limit to avoid timeout
        po_pipeline = [
            {"$match": {"vendor_id": vendor_id, **({"created_at": date_query} if date_query else {})}},
            {"$sort": {"created_at": -1}},
            {"$limit": 500},  # Limit to most recent 500 POs to avoid timeout
            {"$project": {
                "_id": 0,
                "po_number": 1,
                "created_by_location_id": 1,
                "created_by_location_name": 1,
                "items": 1
            }}
        ]
        
        try:
            pos = list(purchase_orders_collection.aggregate(po_pipeline, maxTimeMS=20000, allowDiskUse=True))
        except Exception as e:
            pos = []
            print(f"PO query error: {e}")
        
        # Build PO data by location NAME (not ID, for better matching)
        # Using location name as key because IDs might not match between collections
        po_by_location = {}  # location_name_normalized -> {item_name -> {qty, value, po_numbers}}
        
        for po in pos:
            location_id = po.get("created_by_location_id") or "main_store"
            location_name = po.get("created_by_location_name") or "Main Store"
            # Normalize location name for matching
            location_key = location_name.upper().strip()
            
            if location_key not in po_by_location:
                po_by_location[location_key] = {
                    "location_id": location_id,
                    "location_name": location_name,
                    "items": {},
                    "total_po_value": 0,
                    "po_count": 0,
                    "po_numbers": []
                }
            
            po_by_location[location_key]["po_count"] += 1
            po_by_location[location_key]["po_numbers"].append(po.get("po_number", ""))
            
            for item in po.get("items", []):
                item_name = (item.get("item_name", "Unknown") or "Unknown").upper().strip()
                qty = float(item.get("quantity", 0))
                rate = float(item.get("rate", 0))
                value = qty * rate
                
                if item_name not in po_by_location[location_key]["items"]:
                    po_by_location[location_key]["items"][item_name] = {
                        "ordered_qty": 0,
                        "ordered_value": 0,
                        "rate": rate
                    }
                
                po_by_location[location_key]["items"][item_name]["ordered_qty"] += qty
                po_by_location[location_key]["items"][item_name]["ordered_value"] += value
                po_by_location[location_key]["total_po_value"] += value
        
        # Get DP items from kitchen_receivables grouped by kitchen
        kitchen_receivables = db["kitchen_receivables"]
        
        kr_query = {"type": "daily_perishable"}
        kr_query["$or"] = [
            {"vendor_id": vendor_id},
            {"vendor_name": {"$regex": f"^{re.escape(v_name)}$", "$options": "i"}}
        ]
        
        if start_date or end_date:
            kr_date_query = {}
            if start_date:
                kr_date_query["$gte"] = start_date
            if end_date:
                kr_date_query["$lte"] = end_date
            kr_query["receive_date"] = kr_date_query
        
        # Use aggregation with limit to avoid timeout
        try:
            kr_pipeline = [
                {"$match": kr_query},
                {"$sort": {"receive_date": -1}},
                {"$limit": 1000},  # Limit to avoid timeout
                {"$project": {
                    "_id": 0,
                    "kitchen_id": 1,
                    "kitchen_name": 1,
                    "item_name": 1,
                    "quantity": 1,
                    "rate": 1,
                    "amount": 1,
                    "receive_date": 1
                }}
            ]
            kr_records = list(kitchen_receivables.aggregate(kr_pipeline, maxTimeMS=20000, allowDiskUse=True))
        except Exception as e:
            kr_records = []
            print(f"KR query error: {e}")
        
        # Build DP data by kitchen NAME (not ID, for better matching with PO data)
        dp_by_kitchen = {}  # kitchen_name_normalized -> {item_name -> {qty, value, entries}}
        
        for rec in kr_records:
            kitchen_id = rec.get("kitchen_id") or "unknown"
            kitchen_name = rec.get("kitchen_name") or "Unknown Kitchen"
            # Normalize kitchen name for matching with PO location names
            kitchen_key = kitchen_name.upper().strip()
            
            if kitchen_key not in dp_by_kitchen:
                dp_by_kitchen[kitchen_key] = {
                    "kitchen_id": kitchen_id,
                    "kitchen_name": kitchen_name,
                    "items": {},
                    "total_dp_value": 0,
                    "dp_entries": 0,
                    "dates": set()
                }
            
            dp_by_kitchen[kitchen_key]["dp_entries"] += 1
            dp_by_kitchen[kitchen_key]["dates"].add(str(rec.get("receive_date", ""))[:10])
            
            item_name = (rec.get("item_name", "Unknown") or "Unknown").upper().strip()
            qty = float(rec.get("quantity", 0))
            rate = float(rec.get("rate", 0))
            amount = float(rec.get("amount", 0)) or (qty * rate)
            
            if item_name not in dp_by_kitchen[kitchen_key]["items"]:
                dp_by_kitchen[kitchen_key]["items"][item_name] = {
                    "delivered_qty": 0,
                    "delivered_value": 0,
                    "rate": rate,
                    "entry_count": 0
                }
            
            dp_by_kitchen[kitchen_key]["items"][item_name]["delivered_qty"] += qty
            dp_by_kitchen[kitchen_key]["items"][item_name]["delivered_value"] += amount
            dp_by_kitchen[kitchen_key]["items"][item_name]["entry_count"] += 1
            dp_by_kitchen[kitchen_key]["total_dp_value"] += amount
        
        # Check GRN status for each kitchen (from lots collection)
        lots_query = {"vendor_id": vendor_id}
        if date_query:
            lots_query["created_at"] = date_query
        
        grn_by_location = {}
        for lot in lots_collection.find(lots_query):
            loc_id = lot.get("location_id") or lot.get("kitchen_id") or "main_store"
            if loc_id not in grn_by_location:
                grn_by_location[loc_id] = {"count": 0, "value": 0}
            grn_by_location[loc_id]["count"] += 1
            grn_by_location[loc_id]["value"] += float(lot.get("initial_quantity", 0)) * float(lot.get("purchase_rate", 0))
        
        # Combine all locations/kitchens using normalized names
        all_location_keys = set(list(po_by_location.keys()) + list(dp_by_kitchen.keys()))
        
        outlets = []
        total_po_value = 0
        total_dp_value = 0
        total_short_value = 0
        outlets_without_po = 0
        outlets_with_excess = 0
        
        for loc_key in all_location_keys:
            po_data = po_by_location.get(loc_key, {
                "location_id": "",
                "location_name": dp_by_kitchen.get(loc_key, {}).get("kitchen_name", loc_key),
                "items": {},
                "total_po_value": 0,
                "po_count": 0,
                "po_numbers": []
            })
            
            dp_data = dp_by_kitchen.get(loc_key, {
                "kitchen_id": "",
                "kitchen_name": po_data.get("location_name", loc_key),
                "items": {},
                "total_dp_value": 0,
                "dp_entries": 0,
                "dates": set()
            })
            
            # GRN lookup still uses location ID if available
            loc_id = po_data.get("location_id") or dp_data.get("kitchen_id") or ""
            grn_data = grn_by_location.get(loc_id, {"count": 0, "value": 0})
            
            # Get location name
            location_name = po_data.get("location_name") or dp_data.get("kitchen_name") or "Unknown"
            
            # Combine items
            all_items = set(list(po_data.get("items", {}).keys()) + list(dp_data.get("items", {}).keys()))
            
            items_comparison = []
            outlet_po_value = po_data.get("total_po_value", 0)
            outlet_dp_value = dp_data.get("total_dp_value", 0)
            
            for item_name in sorted(all_items):
                po_item = po_data.get("items", {}).get(item_name, {"ordered_qty": 0, "ordered_value": 0, "rate": 0})
                dp_item = dp_data.get("items", {}).get(item_name, {"delivered_qty": 0, "delivered_value": 0, "rate": 0})
                
                ordered_qty = po_item.get("ordered_qty", 0)
                delivered_qty = dp_item.get("delivered_qty", 0)
                short_qty = ordered_qty - delivered_qty
                
                ordered_value = po_item.get("ordered_value", 0)
                delivered_value = dp_item.get("delivered_value", 0)
                short_value = ordered_value - delivered_value
                
                # Determine status
                if ordered_qty == 0 and delivered_qty > 0:
                    status = "NO PO"
                elif delivered_qty == 0 and ordered_qty > 0:
                    status = "NOT DELIVERED"
                elif short_qty > 0.1:
                    status = "SHORT"
                elif short_qty < -0.1:
                    status = "EXCESS"
                else:
                    status = "OK"
                
                items_comparison.append({
                    "item_name": item_name,
                    "ordered_qty": round(ordered_qty, 2),
                    "delivered_qty": round(delivered_qty, 2),
                    "short_qty": round(short_qty, 2),
                    "ordered_value": round(ordered_value, 2),
                    "delivered_value": round(delivered_value, 2),
                    "short_value": round(short_value, 2),
                    "status": status
                })
            
            # Determine outlet status
            has_po = po_data.get("po_count", 0) > 0
            has_dp = dp_data.get("dp_entries", 0) > 0
            has_grn = grn_data["count"] > 0
            
            if has_dp and not has_po:
                outlet_status = "RECEIVING WITHOUT PO"
                outlets_without_po += 1
            elif outlet_dp_value > outlet_po_value * 1.1:  # More than 10% excess
                outlet_status = "EXCESS DELIVERY"
                outlets_with_excess += 1
            elif outlet_po_value > outlet_dp_value * 1.1:
                outlet_status = "SHORT DELIVERY"
            else:
                outlet_status = "OK"
            
            total_po_value += outlet_po_value
            total_dp_value += outlet_dp_value
            if outlet_po_value > outlet_dp_value:
                total_short_value += (outlet_po_value - outlet_dp_value)
            
            outlets.append({
                "location_id": loc_id,
                "location_name": location_name,
                "location_key": loc_key,  # Normalized name used for matching
                "po_count": po_data.get("po_count", 0),
                "po_numbers": po_data.get("po_numbers", []),
                "total_po_value": round(outlet_po_value, 2),
                "dp_entries": dp_data.get("dp_entries", 0),
                "dp_dates": sorted(list(dp_data.get("dates", set()))),
                "total_dp_value": round(outlet_dp_value, 2),
                "grn_count": grn_data["count"],
                "grn_value": round(grn_data["value"], 2),
                "has_po": has_po,
                "has_dp": has_dp,
                "has_grn": has_grn,
                "outlet_status": outlet_status,
                "variance": round(outlet_dp_value - outlet_po_value, 2),
                "items": items_comparison
            })
        
        # Sort outlets by status (problematic ones first)
        status_order = {"RECEIVING WITHOUT PO": 0, "EXCESS DELIVERY": 1, "SHORT DELIVERY": 2, "OK": 3}
        outlets.sort(key=lambda x: (status_order.get(x["outlet_status"], 4), -x["total_dp_value"]))
        
        return {
            "success": True,
            "vendor_id": vendor_id,
            "vendor_name": v_name,
            "date_range": f"{start_date or 'all'} to {end_date or 'all'}",
            "summary": {
                "total_outlets": len(outlets),
                "total_po_value": round(total_po_value, 2),
                "total_dp_value": round(total_dp_value, 2),
                "total_variance": round(total_dp_value - total_po_value, 2),
                "outlets_without_po": outlets_without_po,
                "outlets_with_excess": outlets_with_excess,
                "outlets_with_grn": len([o for o in outlets if o["has_grn"]])
            },
            "outlets": outlets
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.get("/api/export/outlet-wise-po-dp")
async def export_outlet_wise_po_dp(
    vendor_id: Optional[str] = None,
    vendor_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Export Outlet-wise PO vs DP comparison to Excel"""
    try:
        data = await get_outlet_wise_po_dp(vendor_id, vendor_name, start_date, end_date)
        
        if "error" in data:
            raise HTTPException(status_code=400, detail=data["error"])
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Report Info sheet with date range
            info_rows = [{
                "Report": "Outlet-wise PO vs DP Comparison",
                "Vendor": data["vendor_name"],
                "Date Range": data.get("date_range", f"{start_date or 'all'} to {end_date or 'all'}"),
                "Generated At": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "Total Outlets": data.get("summary", {}).get("total_outlets", 0),
                "Total PO Value": data.get("summary", {}).get("total_po_value", 0),
                "Total DP Value": data.get("summary", {}).get("total_dp_value", 0),
                "Total Variance": data.get("summary", {}).get("total_variance", 0)
            }]
            info_df = pd.DataFrame(info_rows)
            info_df.to_excel(writer, index=False, sheet_name='Report Info')
            
            # Summary sheet
            summary_rows = []
            for outlet in data.get("outlets", []):
                summary_rows.append({
                    "Outlet Name": outlet["location_name"],
                    "Status": outlet["outlet_status"],
                    "PO Count": outlet["po_count"],
                    "PO Value": outlet["total_po_value"],
                    "DP Entries": outlet["dp_entries"],
                    "DP Value": outlet["total_dp_value"],
                    "Variance": outlet["variance"],
                    "Has GRN": "Yes" if outlet["has_grn"] else "No",
                    "GRN Count": outlet["grn_count"],
                    "GRN Value": outlet["grn_value"]
                })
            
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, index=False, sheet_name='Outlet Summary')
            
            # Detail sheet - all items by outlet
            detail_rows = []
            for outlet in data.get("outlets", []):
                for item in outlet.get("items", []):
                    detail_rows.append({
                        "Outlet": outlet["location_name"],
                        "Outlet Status": outlet["outlet_status"],
                        "Item Name": item["item_name"],
                        "Ordered Qty": item["ordered_qty"],
                        "Delivered Qty": item["delivered_qty"],
                        "Short Qty": item["short_qty"],
                        "Ordered Value": item["ordered_value"],
                        "Delivered Value": item["delivered_value"],
                        "Short Value": item["short_value"],
                        "Item Status": item["status"]
                    })
            
            detail_df = pd.DataFrame(detail_rows)
            detail_df.to_excel(writer, index=False, sheet_name='Item Details')
            
            # Issues sheet - only problematic outlets
            issues_rows = []
            for outlet in data.get("outlets", []):
                if outlet["outlet_status"] != "OK":
                    for item in outlet.get("items", []):
                        if item["status"] in ["NO PO", "SHORT", "EXCESS", "NOT DELIVERED"]:
                            issues_rows.append({
                                "Outlet": outlet["location_name"],
                                "Issue Type": outlet["outlet_status"],
                                "Item Name": item["item_name"],
                                "Ordered Qty": item["ordered_qty"],
                                "Delivered Qty": item["delivered_qty"],
                                "Difference": item["short_qty"],
                                "Value Impact": item["short_value"],
                                "Item Status": item["status"]
                            })
            
            if issues_rows:
                issues_df = pd.DataFrame(issues_rows)
                issues_df.to_excel(writer, index=False, sheet_name='Issues Only')
        
        output.seek(0)
        
        vendor_slug = data["vendor_name"].replace(" ", "_")[:20]
        filename = f"Outlet_Analysis_{vendor_slug}_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ============ PO vs DP Item-wise Comparison Report ============

@app.get("/api/reports/po-vs-dp-comparison")
async def get_po_vs_dp_comparison(
    vendor_id: Optional[str] = None,
    vendor_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Compare PO ordered items vs DP delivered items to identify mismatches.
    Shows item-wise: ordered qty, delivered qty, shortage, value difference.
    """
    try:
        # Find vendor
        if vendor_name:
            vendor = vendors_collection.find_one({"name": {"$regex": vendor_name, "$options": "i"}})
            if vendor:
                vendor_id = str(vendor["_id"])
        
        if not vendor_id:
            return {"error": "Please provide vendor_id or vendor_name"}
        
        vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
        if not vendor:
            return {"error": "Vendor not found"}
        
        v_name = vendor["name"]
        
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        # Get PO items with limit to avoid timeout
        po_pipeline = [
            {"$match": {"vendor_id": vendor_id, **({"created_at": date_query} if date_query else {})}},
            {"$sort": {"created_at": -1}},
            {"$limit": 500},
            {"$project": {"_id": 0, "po_number": 1, "created_at": 1, "items": 1}}
        ]
        
        try:
            pos = list(purchase_orders_collection.aggregate(po_pipeline, maxTimeMS=20000, allowDiskUse=True))
        except Exception as e:
            pos = []
            print(f"PO vs DP - PO query error: {e}")
        
        po_items = {}  # item_name -> {qty, value, rate, po_numbers}
        for po in pos:
            po_date = str(po.get("created_at", ""))[:10]
            po_number = po.get("po_number", "")
            
            for item in po.get("items", []):
                name = item.get("item_name", "Unknown").upper().strip()
                qty = float(item.get("quantity", 0))
                rate = float(item.get("rate", 0))
                value = qty * rate
                
                if name not in po_items:
                    po_items[name] = {
                        "ordered_qty": 0,
                        "ordered_value": 0,
                        "avg_rate": rate,
                        "po_count": 0,
                        "po_numbers": []
                    }
                
                po_items[name]["ordered_qty"] += qty
                po_items[name]["ordered_value"] += value
                po_items[name]["po_count"] += 1
                if po_number not in po_items[name]["po_numbers"]:
                    po_items[name]["po_numbers"].append(po_number)
        
        # Get DP items from kitchen_receivables
        kitchen_receivables = db["kitchen_receivables"]
        
        kr_query = {"type": "daily_perishable"}
        # Match by vendor_id or vendor_name
        kr_query["$or"] = [
            {"vendor_id": vendor_id},
            {"vendor_name": {"$regex": f"^{re.escape(v_name)}$", "$options": "i"}}
        ]
        
        if start_date or end_date:
            kr_date_query = {}
            if start_date:
                kr_date_query["$gte"] = start_date
            if end_date:
                kr_date_query["$lte"] = end_date
            kr_query["receive_date"] = kr_date_query
        
        # Use aggregation with limit to avoid timeout
        try:
            kr_pipeline = [
                {"$match": kr_query},
                {"$sort": {"receive_date": -1}},
                {"$limit": 1000},
                {"$project": {"_id": 0, "item_name": 1, "quantity": 1, "rate": 1, "amount": 1}}
            ]
            kr_records = list(kitchen_receivables.aggregate(kr_pipeline, maxTimeMS=20000, allowDiskUse=True))
        except Exception as e:
            kr_records = []
            print(f"PO vs DP - KR query error: {e}")
        
        dp_items = {}  # item_name -> {qty, value, rate, entries}
        for rec in kr_records:
            name = (rec.get("item_name", "Unknown") or "Unknown").upper().strip()
            qty = float(rec.get("quantity", 0))
            rate = float(rec.get("rate", 0))
            amount = float(rec.get("amount", 0)) or (qty * rate)
            
            if name not in dp_items:
                dp_items[name] = {
                    "delivered_qty": 0,
                    "delivered_value": 0,
                    "avg_rate": rate,
                    "entry_count": 0
                }
            
            dp_items[name]["delivered_qty"] += qty
            dp_items[name]["delivered_value"] += amount
            dp_items[name]["entry_count"] += 1
        
        # Combine and compare
        all_items = set(list(po_items.keys()) + list(dp_items.keys()))
        
        comparison = []
        total_po_qty = 0
        total_dp_qty = 0
        total_po_value = 0
        total_dp_value = 0
        total_short_qty = 0
        total_short_value = 0
        
        for item_name in sorted(all_items):
            po_data = po_items.get(item_name, {"ordered_qty": 0, "ordered_value": 0, "avg_rate": 0, "po_count": 0})
            dp_data = dp_items.get(item_name, {"delivered_qty": 0, "delivered_value": 0, "avg_rate": 0, "entry_count": 0})
            
            ordered_qty = po_data["ordered_qty"]
            delivered_qty = dp_data["delivered_qty"]
            ordered_value = po_data["ordered_value"]
            delivered_value = dp_data["delivered_value"]
            
            short_qty = ordered_qty - delivered_qty
            short_value = ordered_value - delivered_value
            
            # Determine status
            if ordered_qty == 0 and delivered_qty > 0:
                status = "EXTRA (No PO)"
            elif delivered_qty == 0 and ordered_qty > 0:
                status = "NOT DELIVERED"
            elif short_qty > 0.1:  # Small tolerance for rounding
                status = "SHORT"
            elif short_qty < -0.1:
                status = "EXCESS"
            else:
                status = "MATCHED"
            
            total_po_qty += ordered_qty
            total_dp_qty += delivered_qty
            total_po_value += ordered_value
            total_dp_value += delivered_value
            if short_qty > 0:
                total_short_qty += short_qty
                total_short_value += short_value
            
            comparison.append({
                "item_name": item_name,
                "ordered_qty": round(ordered_qty, 2),
                "delivered_qty": round(delivered_qty, 2),
                "short_qty": round(short_qty, 2),
                "po_rate": po_data["avg_rate"],
                "dp_rate": dp_data["avg_rate"],
                "ordered_value": round(ordered_value, 2),
                "delivered_value": round(delivered_value, 2),
                "short_value": round(short_value, 2),
                "status": status,
                "po_count": po_data.get("po_count", 0),
                "dp_entries": dp_data.get("entry_count", 0)
            })
        
        # Sort by short value (descending) to show biggest shortages first
        comparison.sort(key=lambda x: x["short_value"], reverse=True)
        
        return {
            "success": True,
            "vendor_id": vendor_id,
            "vendor_name": v_name,
            "date_range": f"{start_date or 'all'} to {end_date or 'all'}",
            "summary": {
                "total_items": len(comparison),
                "total_po_qty": round(total_po_qty, 2),
                "total_dp_qty": round(total_dp_qty, 2),
                "total_short_qty": round(total_short_qty, 2),
                "total_po_value": round(total_po_value, 2),
                "total_dp_value": round(total_dp_value, 2),
                "total_short_value": round(total_short_value, 2),
                "items_short": len([c for c in comparison if c["status"] == "SHORT"]),
                "items_excess": len([c for c in comparison if c["status"] == "EXCESS"]),
                "items_not_delivered": len([c for c in comparison if c["status"] == "NOT DELIVERED"]),
                "items_extra_no_po": len([c for c in comparison if c["status"] == "EXTRA (No PO)"])
            },
            "items": comparison
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.get("/api/export/po-vs-dp-comparison")
async def export_po_vs_dp_comparison(
    vendor_id: Optional[str] = None,
    vendor_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Export PO vs DP comparison to Excel"""
    try:
        # Get comparison data
        data = await get_po_vs_dp_comparison(vendor_id, vendor_name, start_date, end_date)
        
        if "error" in data:
            raise HTTPException(status_code=400, detail=data["error"])
        
        # Create Excel
        rows = []
        for item in data.get("items", []):
            rows.append({
                "Item Name": item["item_name"],
                "Ordered Qty": item["ordered_qty"],
                "Delivered Qty": item["delivered_qty"],
                "Short Qty": item["short_qty"],
                "PO Rate": item["po_rate"],
                "DP Rate": item["dp_rate"],
                "Ordered Value": item["ordered_value"],
                "Delivered Value": item["delivered_value"],
                "Short Value": item["short_value"],
                "Status": item["status"],
                "PO Count": item["po_count"],
                "DP Entries": item["dp_entries"]
            })
        
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Report Info sheet with date range
            info_rows = [{
                "Report": "PO vs DP Comparison",
                "Vendor": data["vendor_name"],
                "Date Range": data.get("date_range", f"{start_date or 'all'} to {end_date or 'all'}"),
                "Generated At": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "Total Items": data.get("summary", {}).get("total_items", 0),
                "Total PO Value": data.get("summary", {}).get("total_po_value", 0),
                "Total DP Value": data.get("summary", {}).get("total_dp_value", 0),
                "Total Short Value": data.get("summary", {}).get("total_short_value", 0)
            }]
            info_df = pd.DataFrame(info_rows)
            info_df.to_excel(writer, index=False, sheet_name='Report Info')
            
            # Main comparison sheet
            df.to_excel(writer, index=False, sheet_name='PO vs DP Comparison')
            
            # Summary sheet
            summary_data = {
                "Metric": [
                    "Vendor", "Date Range", "Total Items",
                    "Total PO Qty", "Total DP Qty", "Total Short Qty",
                    "Total PO Value", "Total DP Value", "Total Short Value",
                    "Items Short", "Items Excess", "Items Not Delivered", "Items Extra (No PO)"
                ],
                "Value": [
                    data["vendor_name"], data["date_range"], data["summary"]["total_items"],
                    data["summary"]["total_po_qty"], data["summary"]["total_dp_qty"], data["summary"]["total_short_qty"],
                    f"₹{data['summary']['total_po_value']:,.2f}", f"₹{data['summary']['total_dp_value']:,.2f}", f"₹{data['summary']['total_short_value']:,.2f}",
                    data["summary"]["items_short"], data["summary"]["items_excess"], 
                    data["summary"]["items_not_delivered"], data["summary"]["items_extra_no_po"]
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name='Summary')
        
        output.seek(0)
        
        vendor_slug = data["vendor_name"].replace(" ", "_")[:20]
        filename = f"PO_vs_DP_Comparison_{vendor_slug}_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ============ Duplicate Detection for Daily Perishables ============

@app.get("/api/admin/detect-dp-duplicates")
async def detect_dp_duplicates(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Detect duplicate entries in kitchen_receivables.
    Duplicates are identified by same: vendor_id, kitchen_id, item_name, receive_date, quantity, rate
    """
    try:
        kitchen_receivables = db["kitchen_receivables"]
        
        # Build query
        query = {"type": "daily_perishable"}
        if vendor_id:
            query["vendor_id"] = vendor_id
        
        if start_date or end_date:
            date_query = {}
            if start_date:
                date_query["$gte"] = start_date
            if end_date:
                date_query["$lte"] = end_date
            query["receive_date"] = date_query
        
        records = list(kitchen_receivables.find(query))
        
        # Group by key fields to find duplicates
        groups = {}
        for rec in records:
            key = (
                rec.get("vendor_id", ""),
                rec.get("vendor_name", ""),
                rec.get("kitchen_id", ""),
                rec.get("item_name", ""),
                str(rec.get("receive_date", ""))[:10],
                float(rec.get("quantity", 0)),
                float(rec.get("rate", 0))
            )
            
            if key not in groups:
                groups[key] = []
            groups[key].append(rec)
        
        # Find duplicates (groups with more than 1 record)
        duplicates = []
        total_duplicate_value = 0
        total_duplicate_count = 0
        
        for key, recs in groups.items():
            if len(recs) > 1:
                vendor_name, kitchen_id, item_name, date, qty, rate = key[1], key[2], key[3], key[4], key[5], key[6]
                duplicate_count = len(recs) - 1  # Extra copies beyond first
                duplicate_value = duplicate_count * qty * rate
                
                total_duplicate_count += duplicate_count
                total_duplicate_value += duplicate_value
                
                duplicates.append({
                    "vendor_name": vendor_name,
                    "item_name": item_name,
                    "date": date,
                    "quantity": qty,
                    "rate": rate,
                    "occurrences": len(recs),
                    "duplicate_count": duplicate_count,
                    "duplicate_value": duplicate_value,
                    "record_ids": [str(r["_id"]) for r in recs]
                })
        
        # Sort by duplicate value
        duplicates.sort(key=lambda x: x["duplicate_value"], reverse=True)
        
        return {
            "success": True,
            "summary": {
                "total_records_checked": len(records),
                "duplicate_groups": len(duplicates),
                "total_duplicate_records": total_duplicate_count,
                "total_duplicate_value": total_duplicate_value
            },
            "duplicates": duplicates[:100]  # Top 100 duplicates
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/remove-dp-duplicates")
async def remove_dp_duplicates(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    dry_run: bool = True
):
    """
    Remove duplicate entries in kitchen_receivables, keeping only the first occurrence.
    Set dry_run=false to actually delete.
    """
    try:
        kitchen_receivables = db["kitchen_receivables"]
        
        # Build query
        query = {"type": "daily_perishable"}
        if vendor_id:
            query["vendor_id"] = vendor_id
        
        if start_date or end_date:
            date_query = {}
            if start_date:
                date_query["$gte"] = start_date
            if end_date:
                date_query["$lte"] = end_date
            query["receive_date"] = date_query
        
        records = list(kitchen_receivables.find(query).sort("created_at", 1))  # Sort by created_at to keep oldest
        
        # Group by key fields
        groups = {}
        for rec in records:
            key = (
                rec.get("vendor_id", ""),
                rec.get("kitchen_id", ""),
                rec.get("item_name", ""),
                str(rec.get("receive_date", ""))[:10],
                float(rec.get("quantity", 0)),
                float(rec.get("rate", 0))
            )
            
            if key not in groups:
                groups[key] = []
            groups[key].append(rec)
        
        # Find records to delete (all but the first in each duplicate group)
        to_delete = []
        for key, recs in groups.items():
            if len(recs) > 1:
                # Keep the first, delete the rest
                for rec in recs[1:]:
                    to_delete.append(rec["_id"])
        
        deleted_count = 0
        deleted_value = 0
        
        if not dry_run and to_delete:
            # Actually delete
            for rec_id in to_delete:
                rec = kitchen_receivables.find_one({"_id": rec_id})
                if rec:
                    deleted_value += float(rec.get("quantity", 0)) * float(rec.get("rate", 0))
                    kitchen_receivables.delete_one({"_id": rec_id})
                    deleted_count += 1
        else:
            # Calculate what would be deleted
            for rec_id in to_delete:
                rec = kitchen_receivables.find_one({"_id": rec_id})
                if rec:
                    deleted_value += float(rec.get("quantity", 0)) * float(rec.get("rate", 0))
                    deleted_count += 1
        
        return {
            "success": True,
            "dry_run": dry_run,
            "records_to_delete": deleted_count,
            "value_to_remove": deleted_value,
            "message": f"{'Would delete' if dry_run else 'Deleted'} {deleted_count} duplicate records worth ₹{deleted_value:,.2f}"
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.post("/api/admin/cleanup-feb-test-data")
async def cleanup_feb_test_data(dry_run: bool = True):
    """
    Delete pending POs and requisitions from February (test data cleanup)
    Set dry_run=false to actually delete
    """
    try:
        # February date range
        feb_start = "2026-02-01"
        feb_end = "2026-02-29"
        
        # Count pending POs from Feb
        pending_pos_query = {
            "status": "pending",
            "created_at": {"$gte": feb_start, "$lt": "2026-03-01"}
        }
        pending_pos_count = purchase_orders_collection.count_documents(pending_pos_query)
        
        # Count pending requisitions from Feb
        pending_reqs_query = {
            "status": "pending",
            "created_at": {"$gte": f"{feb_start}T00:00:00", "$lt": "2026-03-01T00:00:00"}
        }
        pending_reqs_count = requisitions_collection.count_documents(pending_reqs_query)
        
        deleted_pos = 0
        deleted_reqs = 0
        
        if not dry_run:
            # Delete pending POs
            result_pos = purchase_orders_collection.delete_many(pending_pos_query)
            deleted_pos = result_pos.deleted_count
            
            # Delete pending requisitions
            result_reqs = requisitions_collection.delete_many(pending_reqs_query)
            deleted_reqs = result_reqs.deleted_count
        
        return {
            "success": True,
            "dry_run": dry_run,
            "pending_pos_found": pending_pos_count,
            "pending_reqs_found": pending_reqs_count,
            "deleted_pos": deleted_pos if not dry_run else 0,
            "deleted_reqs": deleted_reqs if not dry_run else 0,
            "message": f"{'Would delete' if dry_run else 'Deleted'} {pending_pos_count} pending POs and {pending_reqs_count} pending requisitions from February"
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


# ============ Short Items / PO vs GRN Variance Report ============

@app.get("/api/reports/po-grn-variance")
async def get_po_grn_variance(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    show_only_short: bool = True
):
    """
    Report showing PO ordered vs GRN received variance.
    Useful for identifying short deliveries and vendor payment deductions.
    """
    try:
        # Build date query
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        
        # Get POs
        po_query = {}
        if vendor_id:
            po_query["vendor_id"] = vendor_id
        if date_query:
            po_query["created_at"] = date_query
        
        pos = list(purchase_orders_collection.find(po_query).sort("created_at", -1))
        
        # Get all lots in date range for lookup
        lot_query = {}
        if date_query:
            lot_query["created_at"] = date_query
        all_lots = list(lots_collection.find(lot_query))
        
        # Build lot lookup by item_id
        lots_by_item = {}
        for lot in all_lots:
            item_id = lot.get("item_id")
            if item_id:
                if item_id not in lots_by_item:
                    lots_by_item[item_id] = []
                lots_by_item[item_id].append(lot)
        
        result = []
        total_po_value = 0
        total_grn_value = 0
        total_short_value = 0
        
        for po in pos:
            vendor = vendors_collection.find_one({"_id": ObjectId(po["vendor_id"])}) if po.get("vendor_id") else None
            vendor_name = vendor["name"] if vendor else "Unknown"
            
            po_items_variance = []
            po_total = 0
            grn_total = 0
            
            for item in po.get("items", []):
                item_id = item.get("item_id")
                item_name = item.get("item_name", "Unknown")
                ordered_qty = float(item.get("quantity", 0))
                ordered_rate = float(item.get("rate", 0))
                ordered_value = ordered_qty * ordered_rate
                po_total += ordered_value
                
                # Find matching lots for this item (received after PO date)
                received_qty = 0
                received_value = 0
                received_lots = []
                
                po_date = po.get("created_at", "")[:10] if po.get("created_at") else ""
                
                if item_id and item_id in lots_by_item:
                    for lot in lots_by_item[item_id]:
                        lot_date = str(lot.get("created_at", ""))[:10]
                        # Only count lots created on or after PO date
                        if lot_date >= po_date:
                            lot_qty = float(lot.get("initial_quantity", 0))
                            lot_rate = float(lot.get("purchase_rate", 0)) or ordered_rate
                            lot_value = lot_qty * lot_rate
                            received_qty += lot_qty
                            received_value += lot_value
                            received_lots.append({
                                "lot_number": lot.get("lot_number"),
                                "date": lot_date,
                                "quantity": lot_qty,
                                "rate": lot_rate,
                                "value": lot_value
                            })
                
                grn_total += received_value
                
                short_qty = ordered_qty - received_qty
                short_value = ordered_value - received_value
                
                # Only include if there's a shortage (or show all if show_only_short=False)
                if not show_only_short or short_qty > 0:
                    po_items_variance.append({
                        "item_id": item_id,
                        "item_name": item_name,
                        "unit": item.get("unit", ""),
                        "ordered_qty": ordered_qty,
                        "ordered_rate": ordered_rate,
                        "ordered_value": ordered_value,
                        "received_qty": received_qty,
                        "received_value": received_value,
                        "short_qty": short_qty,
                        "short_value": short_value,
                        "received_lots": received_lots,
                        "status": "FULL" if short_qty <= 0 else ("PARTIAL" if received_qty > 0 else "NOT RECEIVED")
                    })
            
            short_value = po_total - grn_total
            
            total_po_value += po_total
            total_grn_value += grn_total
            total_short_value += max(0, short_value)
            
            # Only include POs with shortages (or all if show_only_short=False)
            if not show_only_short or short_value > 0:
                result.append({
                    "po_id": str(po["_id"]),
                    "po_number": po.get("po_number", ""),
                    "po_date": po.get("created_at", "")[:10] if po.get("created_at") else "",
                    "vendor_id": po.get("vendor_id", ""),
                    "vendor_name": vendor_name,
                    "po_status": po.get("status", ""),
                    "po_total": po_total,
                    "grn_total": grn_total,
                    "short_value": max(0, short_value),
                    "items": po_items_variance
                })
        
        return {
            "success": True,
            "filters": {
                "vendor_id": vendor_id,
                "start_date": start_date,
                "end_date": end_date,
                "show_only_short": show_only_short
            },
            "summary": {
                "total_pos": len(result),
                "total_po_value": total_po_value,
                "total_grn_value": total_grn_value,
                "total_short_value": total_short_value
            },
            "purchase_orders": result
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.get("/api/reports/daily-perishables-short")
async def get_daily_perishables_short(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Report showing Daily Perishables items that were short or had issues.
    Provides date-wise proof for vendor payment deductions.
    """
    try:
        kitchen_receivables = db["kitchen_receivables"]
        
        # Build query
        query = {"type": "daily_perishable"}
        if vendor_id:
            query["vendor_id"] = vendor_id
        
        if start_date or end_date:
            date_query = {}
            if start_date:
                date_query["$gte"] = start_date
            if end_date:
                date_query["$lte"] = end_date
            query["receive_date"] = date_query
        
        # Get all records
        records = list(kitchen_receivables.find(query).sort("receive_date", -1))
        
        # Group by vendor and date
        vendor_data = {}
        
        for rec in records:
            v_id = rec.get("vendor_id", "unknown")
            v_name = rec.get("vendor_name", "Unknown")
            date = rec.get("receive_date", "")[:10] if rec.get("receive_date") else ""
            
            # Check for short/issue indicators
            ordered_qty = float(rec.get("ordered_quantity", 0) or rec.get("quantity", 0))
            received_qty = float(rec.get("quantity", 0))
            short_qty = float(rec.get("short_quantity", 0))
            rejected_qty = float(rec.get("rejected_quantity", 0))
            remarks = rec.get("remarks", "") or rec.get("notes", "")
            
            # Calculate short if not explicitly provided
            if short_qty == 0 and ordered_qty > received_qty:
                short_qty = ordered_qty - received_qty
            
            has_issue = short_qty > 0 or rejected_qty > 0 or "short" in remarks.lower() or "reject" in remarks.lower()
            
            if v_id not in vendor_data:
                vendor_data[v_id] = {
                    "vendor_id": v_id,
                    "vendor_name": v_name,
                    "dates": {},
                    "total_ordered_value": 0,
                    "total_received_value": 0,
                    "total_short_value": 0,
                    "total_entries": 0,
                    "short_entries": 0
                }
            
            if date not in vendor_data[v_id]["dates"]:
                vendor_data[v_id]["dates"][date] = {
                    "date": date,
                    "items": [],
                    "total_value": 0,
                    "short_value": 0
                }
            
            qty = float(rec.get("quantity", 0))
            rate = float(rec.get("rate", 0))
            amount = float(rec.get("amount", 0)) or (qty * rate)
            short_value = short_qty * rate
            
            entry = {
                "id": str(rec["_id"]),
                "item_name": rec.get("item_name", ""),
                "category": rec.get("category", ""),
                "kitchen_name": rec.get("kitchen_name", ""),
                "ordered_qty": ordered_qty,
                "received_qty": received_qty,
                "short_qty": short_qty,
                "rejected_qty": rejected_qty,
                "rate": rate,
                "amount": amount,
                "short_value": short_value,
                "remarks": remarks,
                "has_issue": has_issue,
                "invoice_number": rec.get("invoice_number", "")
            }
            
            vendor_data[v_id]["dates"][date]["items"].append(entry)
            vendor_data[v_id]["dates"][date]["total_value"] += amount
            vendor_data[v_id]["dates"][date]["short_value"] += short_value
            
            vendor_data[v_id]["total_ordered_value"] += ordered_qty * rate
            vendor_data[v_id]["total_received_value"] += amount
            vendor_data[v_id]["total_short_value"] += short_value
            vendor_data[v_id]["total_entries"] += 1
            if has_issue:
                vendor_data[v_id]["short_entries"] += 1
        
        # Convert to list and sort
        result = []
        for v_id, data in vendor_data.items():
            data["dates"] = sorted(data["dates"].values(), key=lambda x: x["date"], reverse=True)
            result.append(data)
        
        result.sort(key=lambda x: x["vendor_name"])
        
        return {
            "success": True,
            "filters": {
                "vendor_id": vendor_id,
                "start_date": start_date,
                "end_date": end_date
            },
            "summary": {
                "total_vendors": len(result),
                "total_entries": sum(v["total_entries"] for v in result),
                "short_entries": sum(v["short_entries"] for v in result),
                "total_ordered_value": sum(v["total_ordered_value"] for v in result),
                "total_received_value": sum(v["total_received_value"] for v in result),
                "total_short_value": sum(v["total_short_value"] for v in result)
            },
            "vendors": result
        }
        
    except Exception as e:
        import traceback
        return {"error": str(e), "details": traceback.format_exc()}


@app.get("/api/export/po-grn-variance")
async def export_po_grn_variance(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    show_only_short: bool = True
):
    """Export PO vs GRN variance report to Excel"""
    try:
        # Get variance data
        variance_data = await get_po_grn_variance(vendor_id, start_date, end_date, show_only_short)
        
        if "error" in variance_data:
            raise HTTPException(status_code=500, detail=variance_data["error"])
        
        # Create Excel
        rows = []
        for po in variance_data.get("purchase_orders", []):
            for item in po.get("items", []):
                rows.append({
                    "PO Number": po["po_number"],
                    "PO Date": po["po_date"],
                    "Vendor": po["vendor_name"],
                    "PO Status": po["po_status"],
                    "Item Name": item["item_name"],
                    "Unit": item.get("unit", ""),
                    "Ordered Qty": item["ordered_qty"],
                    "Ordered Rate": item["ordered_rate"],
                    "Ordered Value": item["ordered_value"],
                    "Received Qty": item["received_qty"],
                    "Received Value": item["received_value"],
                    "Short Qty": item["short_qty"],
                    "Short Value": item["short_value"],
                    "Status": item["status"]
                })
        
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PO vs GRN Variance')
            
            # Add summary sheet
            summary_df = pd.DataFrame([{
                "Total POs": variance_data["summary"]["total_pos"],
                "Total PO Value": variance_data["summary"]["total_po_value"],
                "Total GRN Value": variance_data["summary"]["total_grn_value"],
                "Total Short Value": variance_data["summary"]["total_short_value"]
            }])
            summary_df.to_excel(writer, index=False, sheet_name='Summary')
        
        output.seek(0)
        
        filename = f"po_grn_variance_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/daily-perishables-short")
async def export_daily_perishables_short(
    vendor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Export Daily Perishables short items report to Excel - for vendor payment deduction proof"""
    try:
        # Get short items data
        short_data = await get_daily_perishables_short(vendor_id, start_date, end_date)
        
        if "error" in short_data:
            raise HTTPException(status_code=500, detail=short_data["error"])
        
        # Create Excel with date-wise breakdown
        rows = []
        for vendor in short_data.get("vendors", []):
            for date_data in vendor.get("dates", []):
                for item in date_data.get("items", []):
                    rows.append({
                        "Date": date_data["date"],
                        "Vendor": vendor["vendor_name"],
                        "Kitchen": item.get("kitchen_name", ""),
                        "Item Name": item["item_name"],
                        "Category": item.get("category", ""),
                        "Ordered Qty": item.get("ordered_qty", 0),
                        "Received Qty": item.get("received_qty", 0),
                        "Short Qty": item.get("short_qty", 0),
                        "Rejected Qty": item.get("rejected_qty", 0),
                        "Rate": item.get("rate", 0),
                        "Amount Received": item.get("amount", 0),
                        "Short Value": item.get("short_value", 0),
                        "Invoice #": item.get("invoice_number", ""),
                        "Remarks": item.get("remarks", ""),
                        "Has Issue": "Yes" if item.get("has_issue") else "No"
                    })
        
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Daily Perishables Short')
            
            # Add vendor-wise summary
            vendor_summary = []
            for vendor in short_data.get("vendors", []):
                vendor_summary.append({
                    "Vendor": vendor["vendor_name"],
                    "Total Entries": vendor["total_entries"],
                    "Short Entries": vendor["short_entries"],
                    "Total Ordered Value": vendor["total_ordered_value"],
                    "Total Received Value": vendor["total_received_value"],
                    "Total Short Value": vendor["total_short_value"]
                })
            
            summary_df = pd.DataFrame(vendor_summary)
            summary_df.to_excel(writer, index=False, sheet_name='Vendor Summary')
        
        output.seek(0)
        
        filename = f"daily_perishables_short_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ============================================
# KITCHEN CLOSING STOCK FEATURE
# ============================================

@app.get("/api/reports/kitchen-closing-stock")
async def get_kitchen_closing_stock_summary(
    kitchen_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Get summary of items dispatched to each kitchen for closing stock tracking.
    Returns list of kitchens with their dispatched items.
    """
    
    # Get all kitchens
    kitchens = list(locations_collection.find({"type": "kitchen"}))
    
    # Build date query
    date_query = {}
    if start_date:
        try:
            date_query["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        except:
            pass
    if end_date:
        try:
            date_query["$lte"] = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except:
            pass
    
    result = []
    
    for kitchen in kitchens:
        k_id = str(kitchen["_id"])
        k_name = kitchen.get("name", "Unknown")
        
        if kitchen_id and k_id != kitchen_id:
            continue
        
        # Get requisitions dispatched to this kitchen
        req_query = {
            "$or": [
                {"kitchen_id": k_id},
                {"location_id": kitchen["_id"]},
                {"location_name": k_name}
            ],
            "status": {"$in": ["dispatched", "completed", "partial", "received"]}
        }
        
        if date_query:
            req_query["created_at"] = date_query
        
        requisitions = list(requisitions_collection.find(req_query))
        
        # Aggregate items by item_id
        item_totals = {}
        for req in requisitions:
            for item in req.get("items", []):
                qty_sent = item.get("quantity_sent", 0) or 0
                if qty_sent <= 0:
                    continue
                    
                item_id = item.get("item_id", "")
                if item_id not in item_totals:
                    item_totals[item_id] = {
                        "item_id": item_id,
                        "item_name": item.get("item_name", "Unknown"),
                        "category": item.get("category", "Uncategorized"),
                        "unit": item.get("unit", ""),
                        "total_sent": 0
                    }
                item_totals[item_id]["total_sent"] += qty_sent
        
        # Sort by category and name
        items_list = sorted(item_totals.values(), key=lambda x: (x["category"], x["item_name"]))
        
        result.append({
            "kitchen_id": k_id,
            "kitchen_name": k_name,
            "kitchen_code": kitchen.get("code", ""),
            "total_items": len(items_list),
            "items": items_list
        })
    
    return {
        "kitchens": result,
        "total_kitchens": len(result),
        "date_range": {
            "start": start_date,
            "end": end_date
        }
    }


@app.get("/api/export/kitchen-closing-stock")
async def export_kitchen_closing_stock(
    kitchen_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """
    Export Kitchen Closing Stock Excel for a specific kitchen.
    Shows items dispatched from Main Store with empty closing stock column.
    Organized by Category.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    # Get kitchen details
    try:
        kitchen = locations_collection.find_one({"_id": ObjectId(kitchen_id)})
    except:
        kitchen = locations_collection.find_one({"code": kitchen_id})
    
    if not kitchen:
        raise HTTPException(status_code=404, detail="Kitchen not found")
    
    k_name = kitchen.get("name", "Unknown")
    k_code = kitchen.get("code", "")
    
    # Build date query
    date_query = {}
    if start_date:
        try:
            date_query["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        except:
            pass
    if end_date:
        try:
            date_query["$lte"] = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except:
            pass
    
    # Get requisitions dispatched to this kitchen
    req_query = {
        "$or": [
            {"kitchen_id": kitchen_id},
            {"location_id": kitchen["_id"]},
            {"location_name": k_name}
        ],
        "status": {"$in": ["dispatched", "completed", "partial", "received"]}
    }
    
    if date_query:
        req_query["created_at"] = date_query
    
    requisitions = list(requisitions_collection.find(req_query))
    
    # OPTIMIZATION: Batch fetch all items to get prices
    all_item_ids = set()
    for req in requisitions:
        for item in req.get("items", []):
            item_id = item.get("item_id", "")
            if item_id:
                all_item_ids.add(item_id)
    
    items_price_map = {}
    if all_item_ids:
        try:
            items_data = list(items_collection.find(
                {"_id": {"$in": [ObjectId(i) for i in all_item_ids]}},
                {"_id": 1, "standard_price": 1}
            ))
            items_price_map = {str(item["_id"]): float(item.get("standard_price", 0) or 0) for item in items_data}
        except:
            pass
    
    # Aggregate items by category and item_id
    category_items = {}
    
    for req in requisitions:
        for item in req.get("items", []):
            qty_sent = item.get("quantity_sent", 0) or 0
            if qty_sent <= 0:
                continue
            
            item_id = item.get("item_id", "")
            category = item.get("category", "Uncategorized") or "Uncategorized"
            
            if category not in category_items:
                category_items[category] = {}
            
            if item_id not in category_items[category]:
                # Get price from items collection
                price = items_price_map.get(item_id, 0)
                category_items[category][item_id] = {
                    "item_name": item.get("item_name", "Unknown"),
                    "unit": item.get("unit", ""),
                    "price": price,
                    "total_sent": 0
                }
            
            category_items[category][item_id]["total_sent"] += qty_sent
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Closing Stock"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True, size=11, color="1F4E79")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    date_range_str = ""
    if start_date and end_date:
        date_range_str = f" ({start_date} to {end_date})"
    elif start_date:
        date_range_str = f" (From {start_date})"
    elif end_date:
        date_range_str = f" (Until {end_date})"
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"CLOSING STOCK - {k_name} ({k_code}){date_range_str}"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "Please fill the 'Closing Stock' column with your current inventory"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["S.No", "Item Name", "Unit", "Rate (₹)", "Opening Stock (Sent)", "Closing Stock", "Value (₹)"]
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row += 1
    serial = 1
    
    # Sort categories
    for category in sorted(category_items.keys()):
        items = category_items[category]
        
        # Category header
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"📦 {category}"
        ws[f'A{row}'].fill = category_fill
        ws[f'A{row}'].font = category_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        row += 1
        
        # Sort items by name
        for item_id, item_data in sorted(items.items(), key=lambda x: x[1]["item_name"]):
            price = item_data.get("price", 0)
            total_sent = round(item_data["total_sent"], 2)
            opening_value = round(price * total_sent, 2)
            
            ws.cell(row=row, column=1, value=serial).border = thin_border
            ws.cell(row=row, column=2, value=item_data["item_name"]).border = thin_border
            ws.cell(row=row, column=3, value=item_data["unit"]).border = thin_border
            ws.cell(row=row, column=4, value=price).border = thin_border
            ws.cell(row=row, column=5, value=total_sent).border = thin_border
            
            # Empty closing stock cell (highlighted for user input)
            closing_cell = ws.cell(row=row, column=6, value="")
            closing_cell.border = thin_border
            closing_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # Value column (Opening Value)
            ws.cell(row=row, column=7, value=opening_value).border = thin_border
            
            serial += 1
            row += 1
        
        # Empty row after category
        row += 1
    
    # Add totals row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL ITEMS:")
    ws.cell(row=row, column=2, value=serial - 1)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    today = datetime.now().strftime("%Y%m%d")
    filename = f"Closing_Stock_{k_code}_{today}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/kitchens/list")
async def get_kitchens_list(
    current_user = Depends(require_role(["admin", "main_store"]))
):
    """Get list of all kitchens for dropdown selection"""
    kitchens = list(locations_collection.find({"type": "kitchen"}))
    return [
        {
            "id": str(k["_id"]),
            "name": k.get("name", "Unknown"),
            "code": k.get("code", "")
        }
        for k in sorted(kitchens, key=lambda x: x.get("name", ""))
    ]
